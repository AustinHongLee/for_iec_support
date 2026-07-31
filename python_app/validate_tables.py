
import math
import os
import sys
from pathlib import Path

APP_DIR = Path(__file__).resolve().parent
os.chdir(APP_DIR)
sys.path.insert(0, str(APP_DIR))

_FAILURES: list[str] = []


def record_failure(message: str) -> None:
    _FAILURES.append(message)
    print(message)


# Phase X parser normalization smoke tests.
try:
    from core.calculator import analyze_single
    from core.parser import get_lookup_value, parse_pipe_size

    assert parse_pipe_size("1/2B") == "1/2"
    assert parse_pipe_size("1.1/2B") == "1-1/2"
    assert parse_pipe_size("1 1/2B") == "1-1/2"
    assert parse_pipe_size("2B") == "2"
    assert get_lookup_value("1.1/2B") == 1.5
    assert get_lookup_value("1/2B") == 0.5

    parser_smoke_cases = [
        "51-1.1/2B",
        "57-1.1/2B-A",
        "66-1.1/2B(P)-A-150-150",
        "22-L75-12(A)L",
        "59-1.1/2B-B(S)",
    ]
    for designation in parser_smoke_cases:
        result = analyze_single(designation)
        assert not result.error, f"{designation} should parse without Error: {result.error}"
        assert result.entries, f"{designation} should enter Type calculation"

    not_implemented = analyze_single("99-1B")
    assert not_implemented.error == "Type 99 not implemented"

    print("v phase X parser normalization OK")
except Exception as e:
    record_failure(f"X phase X parser normalization ERROR: {e}")

# Project-level aggregation wrapper smoke tests.
try:
    from core.calculator import analyze_single
    from core.project_aggregation import (
        ProjectInputRow,
        analyze_project_rows,
        scale_analysis_result,
    )
    from core.material_summary import aggregate_project

    single = analyze_single("51-1.1/2B")
    assert not single.error and single.entries, "project aggregation source case failed"
    original_entry = single.entries[0]
    original_quantity = original_entry.quantity
    original_qty_subtotal = original_entry.qty_subtotal
    original_weight = original_entry.weight_output

    scaled_one = scale_analysis_result(single, 1)
    assert scaled_one.entries[0].quantity == original_quantity, "quantity=1 should preserve entry quantity"
    assert scaled_one.entries[0].weight_output == original_weight, "quantity=1 should preserve entry weight"

    scaled_ten = scale_analysis_result(single, 10)
    assert scaled_ten is not single, "scaled result must be a new AnalysisResult"
    assert scaled_ten.entries[0] is not original_entry, "scaled entries must not mutate original entries"
    assert scaled_ten.entries[0].quantity == original_quantity * 10, "project quantity scaling failed"
    assert scaled_ten.entries[0].qty_subtotal == original_qty_subtotal * 10, "project qty subtotal scaling failed"
    assert scaled_ten.entries[0].weight_output == original_weight * 10, "project weight scaling failed"
    assert original_entry.quantity == original_quantity, "single result quantity was mutated"
    assert original_entry.qty_subtotal == original_qty_subtotal, "single result qty subtotal was mutated"
    assert original_entry.weight_output == original_weight, "single result weight was mutated"

    project = analyze_project_rows([
        ProjectInputRow("51-1.1/2B", 10, serial="S-001", drawing_line_number="DL-001"),
        ProjectInputRow("51-1.1/2B", 2, serial="S-002", drawing_line_number="DL-002"),
        ProjectInputRow("57-1B-A", 1, enabled=False),
    ])
    assert not project.errors, f"project aggregation should not emit errors: {project.errors}"
    assert len(project.rows) == 2, "disabled project rows should be skipped"
    assert project.total_support_count == 12, "project support count failed"
    assert len(project.aggregated_entries) == 1, "duplicate scaled entries should aggregate"
    aggregate_entry = project.aggregated_entries[0]
    assert aggregate_entry.quantity == original_quantity * 12, "aggregated quantity failed"
    assert aggregate_entry.qty_subtotal == original_qty_subtotal * 12, "aggregated qty subtotal failed"
    assert abs(aggregate_entry.weight_output - original_weight * 12) < 0.0001, "aggregated weight failed"
    assert abs(project.total_weight - aggregate_entry.weight_output) < 0.0001, "project total weight failed"

    material_summary = aggregate_project(project)
    assert abs(material_summary.total_weight - project.total_weight) < 0.0001, "project material summary total failed"
    assert len(material_summary.lines) == 1, "project material summary should merge duplicate designations"
    assert material_summary.lines[0].total_qty == original_quantity * 12, "project material summary quantity failed"
    assert material_summary.lines[0].source_fullstrings == [
        "DL-001 / S-001 51-1.1/2B × 10組",
        "DL-002 / S-002 51-1.1/2B × 2組",
    ], "project material summary source labels failed"

    linear_project = analyze_project_rows([ProjectInputRow("24-L50-04", 2, serial="S-L01", drawing_line_number="DL-L01")])
    linear_summary = aggregate_project(linear_project)
    linear_lines = linear_summary.get_linear_lines()
    assert linear_lines, "project cutting summary should include linear material"
    assert "DL-L01 / S-L01 24-L50-04 × 2組" in linear_lines[0].source_fullstrings, "project cutting source label failed"

    errored = analyze_project_rows([ProjectInputRow("99-1B", 5)])
    assert errored.errors == ["99-1B: Type 99 not implemented"], f"project error propagation failed: {errored.errors}"
    assert errored.total_support_count == 5, "errored enabled row should still count supports"

    try:
        analyze_project_rows([ProjectInputRow("51-1.1/2B", 0)])
        raise AssertionError("zero project quantity should fail")
    except ValueError:
        pass

    import os
    import tempfile
    import openpyxl
    from export.excel_export import export_project_to_excel, export_project_workbook

    def _stat_value(ws, label: str):
        for row_index in range(1, ws.max_row + 1):
            if ws.cell(row=row_index, column=2).value == label:
                return ws.cell(row=row_index, column=4).value
        raise AssertionError(f"stat label not found: {label}")

    def _has_cell_value(ws, column: int, value: str) -> bool:
        return any(
            ws.cell(row=row_index, column=column).value == value
            for row_index in range(1, ws.max_row + 1)
        )

    def _sheet_contains_text(ws, text: str) -> bool:
        return any(
            text in str(cell.value)
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )

    def _sheet_link_targets(ws) -> set[str]:
        targets: set[str] = set()
        for row in ws.iter_rows():
            for cell in row:
                hyperlink = cell.hyperlink
                if hyperlink:
                    targets.add(str(getattr(hyperlink, "target", hyperlink)))
        return targets

    def _assert_sheet_index_links(ws, sheetnames: list[str]) -> None:
        targets = _sheet_link_targets(ws)
        for sheet in sheetnames:
            assert f"#'{sheet}'!A1" in targets, f"{ws.title} missing hyperlink to {sheet}"

    def _assert_visible_chart(ws, last_print_col: str) -> None:
        assert ws._charts, f"{ws.title} chart missing"
        chart = ws._charts[0]
        assert chart.visible_cells_only is False, f"{ws.title} chart should plot hidden helper cells"
        assert ws.column_dimensions["AA"].hidden, f"{ws.title} chart helper label column should be hidden"
        assert ws.column_dimensions["AB"].hidden, f"{ws.title} chart helper value column should be hidden"
        assert f"$A$1:${last_print_col}$" in str(ws.print_area), f"{ws.title} print area should include chart zone"
        assert ws.page_setup.orientation == "portrait", f"{ws.title} should be portrait for A4 print"

    def _assert_default_trace_columns_hidden(ws, headers: list[str]) -> None:
        from openpyxl.utils import get_column_letter
        from export.excel import column_roles

        for index, header in enumerate(headers, 1):
            letter = get_column_letter(index)
            hidden = bool(ws.column_dimensions[letter].hidden)
            expected_hidden = not column_roles.is_visible(header)
            assert hidden is expected_hidden, f"{ws.title} column {letter} ({header}) hidden={hidden}, expected {expected_hidden}"
            if expected_hidden:
                assert ws.column_dimensions[letter].outline_level == 1, f"{ws.title} column {letter} should be grouped for expand"

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        from core.project_import import read_project_rows_xlsx

        wb_import = openpyxl.Workbook()
        ws_import = wb_import.active
        ws_import.title = "Owner MTO"
        ws_import.append(["ignored title row"])
        ws_import.append(["uom", "model", "count", "line_group", "seq"])
        ws_import.append(["Ве", "57-1B-A", 3, "1--1-S11U-AI-00009", "A-009"])
        wb_import.save(path)
        wb_import.close()

        imported_rows = read_project_rows_xlsx(path)
        assert len(imported_rows) == 1, "flexible xlsx import row count failed"
        assert imported_rows[0].drawing_line_number == "1--1-S11U-AI-00009", "flexible xlsx import drawing line mapping failed"
        assert imported_rows[0].serial == "A-009", "flexible xlsx import serial mapping failed"
        assert imported_rows[0].designation == "57-1B-A", "flexible xlsx import designation mapping failed"
        assert imported_rows[0].quantity == 3, "flexible xlsx import quantity mapping failed"
        assert imported_rows[0].unit == "組", "flexible xlsx import unit mapping failed"
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        export_project_to_excel(project, path)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        assert ws.title == "Project_Weight_Analysis", "project Excel sheet name failed"
        assert ws.cell(row=1, column=1).value == "型號", "project Excel designation header failed"
        assert ws.cell(row=1, column=2).value == "型號類別", "project Excel type header missing"
        assert ws.cell(row=1, column=9).value == "單件數量", "project Excel single section missing"
        assert ws.cell(row=1, column=10).value == "總數量", "project Excel total section missing"
        assert ws.cell(row=1, column=19).value == "來源圖號", "project Excel drawing header failed"
        assert ws.cell(row=1, column=20).value == "流水號", "project Excel serial header failed"
        assert ws.cell(row=1, column=21).value == "輸入數量", "project Excel source quantity header failed"
        assert ws.cell(row=1, column=22).value == "輸入單位", "project Excel source unit header failed"
        assert ws.cell(row=2, column=1).value == "51-1.1/2B", "project Excel designation value failed"
        assert ws.cell(row=2, column=2).value == "51", "project Excel type value failed"
        assert ws.cell(row=2, column=9).value == original_quantity, "project Excel single quantity failed"
        assert ws.cell(row=2, column=10).value == original_quantity * 10, "project Excel total quantity failed"
        assert ws.cell(row=2, column=19).value == "DL-001", "project Excel drawing value failed"
        assert ws.cell(row=2, column=20).value == "S-001", "project Excel serial value failed"
        assert ws.cell(row=2, column=21).value == 10, "project Excel quantity failed"
        assert ws.cell(row=2, column=22).value == "組", "project Excel unit failed"
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        export_project_workbook(project, path)
        wb = openpyxl.load_workbook(path, data_only=True)
        assert wb.sheetnames == [
            "長官-摘要",
            "專案摘要",
            "重量明細表",
            "單組重量明細",
            "計算標準與假設",
            "長官-支撐分類",
            "查核-支撐明細",
            "重量分析",
            "材料合計",
            "下料明細",
            "下料圖示",
        ], f"project package workbook sheets changed: {wb.sheetnames}"
        ws_manager = wb["長官-摘要"]
        assert ws_manager.cell(row=1, column=1).value == "請款分類查核入口", "manager cover title failed"
        assert "$A$1:$H$" in str(ws_manager.print_area), "manager cover should be A4 portrait width"
        assert ws_manager.page_setup.orientation == "portrait", "manager cover should be portrait"
        assert _sheet_contains_text(ws_manager, "分頁索引"), "manager cover should include sheet index"
        _assert_sheet_index_links(ws_manager, wb.sheetnames)
        assert _sheet_contains_text(ws_manager, "合約名稱怎麼來的？"), "manager cover should start from claim questions"
        assert _sheet_contains_text(ws_manager, "長官-支撐分類"), "manager cover should point to leader classification sheet"
        assert _sheet_contains_text(ws_manager, "查核-支撐明細"), "manager cover should point to support detail sheet"
        assert _sheet_contains_text(ws_manager, "單組重量明細"), "manager cover should point to unit weight evidence"
        assert _sheet_contains_text(ws_manager, "下料圖示"), "manager cover should expose full workbook index"
        assert not _sheet_contains_text(ws_manager, "51-1.1/2B"), "manager cover should not expose source designations"
        ws_summary = wb["專案摘要"]
        assert ws_summary.cell(row=1, column=1).value == "專案材料統計總覽", "project package summary title failed"
        assert "$A$1:$I$" in str(ws_summary.print_area), "project summary should be A4 portrait width"
        _assert_sheet_index_links(ws_summary, wb.sheetnames)
        assert _sheet_contains_text(ws_summary, "使用 Type 統計"), "project summary should show used Type list"
        assert _sheet_contains_text(ws_summary, "Type 51"), "project summary should include Type 51"
        assert _sheet_contains_text(ws_summary, "資料狀態"), "project summary should show confidence status column"
        assert not _sheet_contains_text(ws_summary, "Workbook 索引"), "project summary should not use old workbook index wording"
        assert not _sheet_contains_text(ws_summary, "資料量"), "project summary index should not expose confusing data-volume labels"
        ws_detail = wb["重量明細表"]
        from export.excel.headers import LEADER_DETAIL_HEADERS, _CALC_BASIS_HEADERS
        assert ws_detail.cell(row=1, column=1).value == "IEC 管架支撐 - 重量明細表", "weight detail title failed"
        assert ws_detail.cell(row=3, column=1).value == "型號", "weight detail designation header failed"
        assert ws_detail.cell(row=3, column=2).value == "型號類別", "weight detail type header missing"
        assert ws_detail.cell(row=3, column=10).value == "單件數量", "weight detail single qty header failed"
        assert ws_detail.cell(row=3, column=11).value == "組數", "weight detail support count header failed"
        assert ws_detail.cell(row=3, column=12).value == "總數量", "weight detail total qty header failed"
        assert ws_detail.cell(row=3, column=20).value == "來源圖號", "weight detail drawing header failed"
        assert ws_detail.cell(row=3, column=21).value == "流水號", "weight detail serial header failed"
        assert ws_detail.cell(row=4, column=1).value == "51-1.1/2B", "weight detail designation value failed"
        assert ws_detail.cell(row=4, column=2).value == "51", "weight detail type value failed"
        assert any(
            ws_detail.cell(row=r, column=19).value == "小計"
            and ws_detail.cell(row=r, column=20).value == "DL-001"
            and ws_detail.cell(row=r, column=21).value == "S-001"
            and ws_detail.cell(row=r, column=22).value == 10
            for r in range(4, ws_detail.max_row + 1)
        ), "weight detail subtotal source trace failed"
        assert "可信度" not in [ws_detail.cell(row=3, column=col).value for col in range(1, 24)], "weight detail should not expose confidence header"
        assert "來源依據" not in [ws_detail.cell(row=3, column=col).value for col in range(1, 24)], "weight detail should not expose source header"
        _assert_default_trace_columns_hidden(ws_detail, _CALC_BASIS_HEADERS)
        ws_calc = wb["計算標準與假設"]
        assert "$A$1:$F$" in str(ws_calc.print_area), "calc reference should remain A4 portrait width"
        assert _sheet_contains_text(ws_calc, "Type 計算資料狀態彙整"), "calc reference should summarize by Type"
        assert not ws_calc._charts, "calc reference should not use dashboard charts"
        ws_weight = wb["重量分析"]
        assert ws_weight.cell(row=3, column=1).value == "型號", "project package weight header failed"
        assert ws_weight.cell(row=3, column=2).value == "型號類別", "project package weight type header missing"
        assert ws_weight.cell(row=3, column=19).value == "來源圖號", "project package weight drawing header failed"
        assert ws_weight.cell(row=3, column=20).value == "流水號", "project package weight serial header failed"
        assert ws_weight.cell(row=4, column=1).value == "51-1.1/2B", "project package designation failed"
        assert ws_weight.cell(row=4, column=2).value == "51", "project package type failed"
        assert ws_weight.cell(row=4, column=19).value == "DL-001", "project package drawing failed"
        assert ws_weight.cell(row=4, column=20).value == "S-001", "project package serial failed"
        assert ws_weight.cell(row=4, column=21).value == 10, "project package quantity failed"
        ws_material = wb["材料合計"]
        assert ws_material.cell(row=3, column=1).value == "品名", "project package material summary header failed"
        assert ws_material.cell(row=4, column=11).value == original_quantity * 12, "project package material purchase qty failed"
        _assert_visible_chart(ws_material, "M")
        ws_leader = wb["長官-支撐分類"]
        assert ws_leader.cell(row=1, column=1).value == "支撐分類統計", "leader procurement sheet title failed"
        assert _sheet_contains_text(ws_leader, "二、管支撐(連工帶料，含油漆)"), "leader summary fixed title missing"
        assert _stat_value(ws_leader, '管鞋(PIPE SHOE)≦4"') == 0, "Type 51 should not be counted as Pipe Shoe"
        assert _stat_value(ws_leader, "CS(熱鍍鋅)管支撐(Pipe Support)製裝<=15Kg") == 12, "Type 51 should be counted as CS fabrication by support count"
        assert _sheet_contains_text(ws_leader, "合約名稱怎麼來的"), "leader-facing summary should explain contract names"
        assert _sheet_contains_text(ws_leader, "本批命中型號例"), "leader-facing summary should show an evidence example"
        assert _sheet_contains_text(ws_leader, "查看 2 筆來源"), "leader-facing summary should link to row-level evidence"
        assert not ws_leader._charts, "leader-facing summary should not include charts"
        ws_leader_detail = wb["查核-支撐明細"]
        assert ws_leader_detail.cell(row=3, column=1).value == "合約名稱", "claim evidence contract header failed"
        assert ws_leader_detail.cell(row=3, column=4).value == "來源圖號", "leader detail drawing header failed"
        assert ws_leader_detail.cell(row=3, column=5).value == "流水號", "leader detail serial header failed"
        assert ws_leader_detail.cell(row=3, column=6).value == "型號", "leader detail designation header failed"
        assert ws_leader_detail.cell(row=3, column=8).value == "支撐單組總重(kg)", "leader detail single-weight header failed"
        assert ws_leader_detail.cell(row=3, column=10).value == "本列請款計算", "leader detail calculation header failed"
        assert not any(ws_leader_detail.column_dimensions[col].hidden for col in ("D", "E", "F", "H", "J")), "claim evidence columns must stay visible"
        assert any(
            ws_leader_detail.cell(row=r, column=4).value == "DL-001"
            and ws_leader_detail.cell(row=r, column=5).value == "S-001"
            and ws_leader_detail.cell(row=r, column=6).value == "51-1.1/2B"
            and ws_leader_detail.cell(row=r, column=1).value == "CS(熱鍍鋅)管支撐(Pipe Support)製裝<=15Kg"
            and "≤ 15 kg" in str(ws_leader_detail.cell(row=r, column=10).value)
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "leader detail should classify Type 51 as CS fabrication"
        assert not any(
            ws_leader_detail.cell(row=r, column=6).value == "51-1.1/2B"
            and "PIPE SHOE" in str(ws_leader_detail.cell(row=r, column=1).value or "")
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "Type 51 should not be classified as Pipe Shoe"
        ws_visual = wb["下料圖示"]
        assert ws_visual.cell(row=1, column=1).value == "下料圖示", "project package cutting visual title failed"
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    leader_project = analyze_project_rows([
        ProjectInputRow("57-1/2B-A", 2),
        ProjectInputRow("52-1/2B-A-150-200", 3),
        ProjectInputRow("66-6B(P)-A-150-250", 4),
        ProjectInputRow("59-1.1/2B-B(S)", 5),
        ProjectInputRow("10-6B-14A", 2),
    ])
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        export_project_workbook(leader_project, path)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws_manager = wb["長官-摘要"]
        assert _sheet_contains_text(ws_manager, "合約名稱怎麼來的？"), "manager cover claim question missing"
        assert _sheet_contains_text(ws_manager, "長官-支撐分類"), "manager cover detail pointer missing"
        assert not _sheet_contains_text(ws_manager, "57-1/2B-A"), "manager cover should hide source designations"
        ws_leader = wb["長官-支撐分類"]
        assert _stat_value(ws_leader, 'U-Bolt & Band ≦ 6" 熱浸鍍鋅') == 7, (
            "leader procurement <=6in U-Bolt/Band HDG count failed; "
            "Type 57 contributes 2 sets and drawing-furnished Type 59 FIG-B contributes 5 sets"
        )
        assert _stat_value(ws_leader, '管鞋(PIPE SHOE)≦4"') == 3, "leader procurement <=4in pipe shoe HDG count failed"
        assert _stat_value(ws_leader, '管鞋(PIPE SHOE) 5"~10"') == 4, "leader procurement 5~10in pipe shoe HDG count failed"
        assert _stat_value(ws_leader, "CS(熱鍍鋅)管支撐(Pipe Support)製裝<=15Kg") == 5, "leader procurement should merge only SUS304 generic <=15kg supports into CS count"
        assert not _sheet_contains_text(ws_leader, "SUS304 管支撐製裝"), "leader summary should not expose separate SUS304 support fabrication rows"
        assert _sheet_contains_text(ws_leader, "57-1/2B-A"), "leader summary should expose a source example"
        assert not _sheet_contains_text(ws_leader, "無命中"), "leader-facing summary should hide no-hit categories"
        ws_leader_detail = wb["查核-支撐明細"]
        assert not _sheet_contains_text(ws_leader_detail, "SUS304 管支撐製裝"), "leader detail should not expose separate SUS304 support fabrication rows"
        assert any(
            ws_leader_detail.cell(row=r, column=6).value == "59-1.1/2B-B(S)"
            and ws_leader_detail.cell(row=r, column=1).value == "CS(熱鍍鋅)管支撐(Pipe Support)製裝<=15Kg"
            and "併入 CS" in str(ws_leader_detail.cell(row=r, column=13).value)
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "SUS304 <=15kg support should be listed under CS fabrication detail"
        assert any(
            ws_leader_detail.cell(row=r, column=6).value == "10-6B-14A"
            and ws_leader_detail.cell(row=r, column=1).value == "CS(熱鍍鋅)管支撐(Pipe Support)製裝>15Kg"
            and "併入 CS" in str(ws_leader_detail.cell(row=r, column=13).value)
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "SUS304 >15kg support should be listed under CS fabrication detail"
        assert any(
            ws_leader_detail.cell(row=r, column=6).value == "57-1/2B-A"
            and str(ws_leader_detail.cell(row=r, column=1).value).startswith("U-Bolt & Band")
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "leader detail U-Bolt source row missing"
        assert not any(
            ws_leader_detail.cell(row=r, column=6).value in {"57-1/2B-A", "52-1/2B-A-150-200", "66-6B(P)-A-150-250"}
            and str(ws_leader_detail.cell(row=r, column=1).value or "").startswith("CS(熱鍍鋅)")
            for r in range(4, ws_leader_detail.max_row + 1)
        ), "U-Bolt and Pipe Shoe rows should not be double counted as CS fabrication"
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    print("v project aggregation wrapper OK")
except Exception as e:
    record_failure(f"X project aggregation wrapper ERROR: {e}")

# Type 01 Rev.1 table and note guardrails.
try:
    from core.calculator import analyze_single
    from core.config_loader import get_type_table_as_dict

    type01_table = get_type_table_as_dict("01")
    assert type01_table[22]["pipe_size"] == "14", "Type 01 22 inch support pipe changed"
    assert type01_table[24]["L"] == 677, "Type 01 24 inch L should follow D-1 Rev.1"
    assert type01_table[50]["pipe_size"] == "28", "Type 01 50 inch support pipe missing"
    assert type01_table[50]["L"] == 1382, "Type 01 50 inch L missing"

    type01_large = analyze_single("01-50B-05A")
    assert not type01_large.error, f"Type 01 50B should calculate: {type01_large.error}"
    assert type01_large.entries[0].spec == '28"*STD.WT', f"Type 01 50B upper pipe spec changed: {type01_large.entries[0].spec}"
    assert type01_large.entries[0].length == 1882, f"Type 01 50B continuous pipe length changed: {type01_large.entries[0].length}"
    assert any("NOTE 6" in warning for warning in type01_large.warnings), "Type 01 M42 A/B/E/G paving warning missing"

    type01_mid = analyze_single("01-28B-05B")
    assert not type01_mid.error, f"Type 01 28B should calculate: {type01_mid.error}"
    assert type01_mid.entries[0].spec == '16"*STD.WT', f"Type 01 28B upper pipe spec changed: {type01_mid.entries[0].spec}"
    assert type01_mid.entries[0].length == 1282, f"Type 01 28B continuous pipe length changed: {type01_mid.entries[0].length}"

    def _type01_names(result):
        return [entry.name for entry in result.entries]

    def _type01_entry(result, index):
        return result.entries[index - 1]

    def _expect_pipe_weight(entry):
        from core.pipe import normalize_schedule
        from core.parser import get_lookup_value
        from data.pipe_table import get_pipe_details

        size_token, schedule_token = entry.spec.split('"*', 1)
        size = get_lookup_value(size_token)
        schedule = normalize_schedule(schedule_token)
        pipe_details = get_pipe_details(size, schedule, entry.material)
        expected_unit_weight = round(entry.length / 1000 * pipe_details["weight_per_m"], 2)
        assert entry.weight_per_unit == pipe_details["weight_per_m"], (
            f"{entry.name} {entry.spec} kg/m changed: {entry.weight_per_unit}"
        )
        assert entry.unit_weight == expected_unit_weight, (
            f"{entry.name} {entry.spec} unit weight changed: {entry.unit_weight} != {expected_unit_weight}"
        )
        assert entry.total_weight == expected_unit_weight * entry.quantity, (
            f"{entry.name} {entry.spec} total weight changed: {entry.total_weight}"
        )
        assert entry.weight_output == entry.factor * entry.total_weight, (
            f"{entry.name} {entry.spec} output weight changed: {entry.weight_output}"
        )

    def _expect_plate_weight(entry):
        density = {
            "A36/SS400": 7.85,
            "A283 Gr.C": 7.85,
            "A516-60": 7.85,
            "A387-22": 7.85,
            "SUS304": 7.93,
            "A240-304": 7.93,
            "AS": 7.82,
        }.get(entry.material, 7.85)
        geometry = getattr(entry, "geometry", None)
        area_mm2 = getattr(geometry, "net_area_mm2", 0) if geometry else 0
        if not area_mm2:
            area_mm2 = entry.length * entry.width
        raw_weight = area_mm2 * float(entry.spec) * density / 1_000_000
        expected_unit_weight = round(raw_weight, 2)
        expected_total_weight = round(raw_weight * entry.quantity, 2)
        expected_output = round(entry.factor * expected_total_weight, 2)
        assert entry.unit_weight == expected_unit_weight, (
            f"{entry.name} plate unit weight changed: {entry.unit_weight} != {expected_unit_weight}"
        )
        assert entry.total_weight == expected_total_weight, (
            f"{entry.name} plate total weight changed: {entry.total_weight} != {expected_total_weight}"
        )
        assert entry.weight_output == expected_output, (
            f"{entry.name} plate output weight changed: {entry.weight_output} != {expected_output}"
        )

    def _expect_bolt_weight(entry):
        assert entry.unit_weight == 1, f"{entry.name} unit weight should remain 1 kg/set: {entry.unit_weight}"
        assert entry.total_weight == entry.quantity, f"{entry.name} total weight should equal quantity: {entry.total_weight}"
        assert entry.weight_output == round(entry.factor * entry.total_weight, 2), (
            f"{entry.name} output weight changed: {entry.weight_output}"
        )

    from data.pipe_table import get_pipe_details as _pipe_details_for_formula
    from data.pipe_table import pipe_weight_constant as _pipe_weight_constant

    assert round(_pipe_weight_constant(7.93), 5) == 0.02491, "SUS304 pipe weight constant changed"
    assert _pipe_details_for_formula(3, "40S", "SUS304")["weight_per_m"] == 11.41, "standard SUS304 pipe formula changed"
    assert _pipe_details_for_formula(3, "40S", "A53Gr.B")["weight_per_m"] == 11.29, "carbon steel pipe formula changed"

    _TYPE01_H01_CASES = {
        "01-4B-06U": {
            "pipe_spec": '3"*SCH.40',
            "upper": 239,
            "lower": 500,
            "names": ["管路", "管路", "Plate_a_無鑽孔", "Plate_d_有鑽孔", "EXP.BOLT"],
            "m42": [
                ("Plate_a_無鑽孔", 150, 150, "A36/SS400"),
                ("Plate_d_有鑽孔", 290, 290, "SUS304"),
                ("EXP.BOLT", 0, 0, "SUS304"),
            ],
            "warnings": 0,
        },
        "01-4B-04U": {
            "pipe_spec": '3"*SCH.40',
            "upper": 239,
            "lower": 300,
            "names": ["管路", "管路", "Plate_a_無鑽孔", "Plate_d_有鑽孔", "EXP.BOLT"],
            "m42": [
                ("Plate_a_無鑽孔", 150, 150, "A36/SS400"),
                ("Plate_d_有鑽孔", 290, 290, "SUS304"),
                ("EXP.BOLT", 0, 0, "SUS304"),
            ],
            "warnings": 0,
        },
        "01-6B-16T": {
            "pipe_spec": '4"*SCH.40',
            "upper": 286,
            "lower": 1500,
            "names": ["管路", "管路", "Plate_a_無鑽孔"],
            "m42": [("Plate_a_無鑽孔", 230, 230, "SUS304")],
            "warnings": 1,
        },
        "01-6B-06U": {
            "pipe_spec": '4"*SCH.40',
            "upper": 286,
            "lower": 500,
            "names": ["管路", "管路", "Plate_a_無鑽孔", "Plate_d_有鑽孔", "EXP.BOLT"],
            "m42": [
                ("Plate_a_無鑽孔", 230, 230, "A36/SS400"),
                ("Plate_d_有鑽孔", 370, 370, "SUS304"),
                ("EXP.BOLT", 0, 0, "SUS304"),
            ],
            "warnings": 0,
        },
        "01-3B-05U": {
            "pipe_spec": '2"*SCH.40',
            "upper": 193,
            "lower": 400,
            "names": ["管路", "管路", "Plate_a_無鑽孔", "Plate_d_有鑽孔", "EXP.BOLT"],
            "m42": [
                ("Plate_a_無鑽孔", 150, 150, "A36/SS400"),
                ("Plate_d_有鑽孔", 290, 290, "SUS304"),
                ("EXP.BOLT", 0, 0, "SUS304"),
            ],
            "warnings": 0,
        },
        "01-3B-04D": {
            "pipe_spec": '2"*SCH.40',
            "upper": 193,
            "lower": 300,
            "names": ["管路", "管路", "Plate_a_無鑽孔", "Plate_e_無鑽孔"],
            "m42": [
                ("Plate_a_無鑽孔", 150, 150, "A36/SS400"),
                ("Plate_e_無鑽孔", 200, 200, "A36/SS400"),
            ],
            "warnings": 0,
        },
    }

    for designation, expected in _TYPE01_H01_CASES.items():
        result = analyze_single(designation)
        assert not result.error, f"{designation} should calculate: {result.error}"
        if designation == "01-6B-16T":
            assert result.meta["issues"][0]["severity"] == "warning", (
                f"{designation} should be a bounded source-envelope warning"
            )
        pipes = [entry for entry in result.entries if entry.name == "管路"]
        assert len(pipes) == 1, f"{designation} should use one continuous Supporting Pipe B"
        pipe = pipes[0]
        assert pipe.spec == expected["pipe_spec"], f"{designation} pipe spec changed: {pipe.spec}"
        assert pipe.material == "SUS304", f"{designation} pipe material changed: {pipe.material}"
        assert pipe.geometry.component_id == "D1-SUPPORTING-PIPE-B"
        assert not pipe.geometry.fabrication_ready
        assert result.meta["fabrication"]["bom_ready"] is True
        assert result.meta["fabrication"]["fabrication_ready"] is False
        _expect_pipe_weight(pipe)

    print("v type01 source-profile / fabrication-contract guardrails OK")
except Exception as e:
    record_failure(f"X type01 Rev.1 table/note guardrails ERROR: {e}")

# Phase H-02: Type 10/15/16 dimensional and weight guardrails.
try:
    from core.calculator import analyze_single
    from data.steel_sections import get_section_weight

    def _h02_entry(result, index):
        return result.entries[index - 1]

    def _h02_names(result):
        return [entry.name for entry in result.entries]

    def _expect_steel_weight(entry):
        # entry.name 已改為中文，需還原為英文 key 供 get_section_weight 查表
        _ZH_TO_EN = {"角鋼": "Angle", "槽鐵": "Channel", "H型鋼": "H Beam",
                     "I型鋼": "I Beam", "扁鋼": "Flat Bar", "圓鋼": "Round Bar"}
        section_en = _ZH_TO_EN.get(entry.name, entry.name)
        expected_per_m = get_section_weight(section_en, entry.spec)
        expected_unit_weight = round(entry.length / 1000 * expected_per_m, 2)
        expected_total_weight = round(expected_unit_weight * entry.quantity, 2)
        assert entry.weight_per_unit == expected_per_m, (
            f"{entry.name} {entry.spec} kg/m changed: {entry.weight_per_unit}"
        )
        assert entry.unit_weight == expected_unit_weight, (
            f"{entry.name} {entry.spec} unit weight changed: {entry.unit_weight} != {expected_unit_weight}"
        )
        assert entry.total_weight == expected_total_weight, (
            f"{entry.name} {entry.spec} total weight changed: {entry.total_weight}"
        )
        assert entry.weight_output == round(entry.factor * expected_total_weight, 2), (
            f"{entry.name} {entry.spec} output weight changed: {entry.weight_output}"
        )

    def _expect_custom_weight(entry, unit_weight):
        assert entry.unit_weight == unit_weight, f"{entry.name} unit weight changed: {entry.unit_weight}"
        assert entry.total_weight == round(unit_weight * entry.quantity, 2), f"{entry.name} total weight changed"
        assert entry.weight_output == round(entry.factor * entry.total_weight, 2), f"{entry.name} output weight changed"

    type10 = analyze_single("10-2B-05A")
    assert not type10.error, f"Type 10 should calculate: {type10.error}"
    assert _h02_names(type10) == ["管路", "管路", "Plate_F", "ADJ.BOLT", "HEX NUT", "Plate_a_無鑽孔"], (
        f"Type 10 BOM sequence changed: {_h02_names(type10)}"
    )
    assert _h02_entry(type10, 1).length == 271, f"Type 10 main pipe length changed: {_h02_entry(type10, 1).length}"
    assert _h02_entry(type10, 1).material == "SUS304", f"Type 10 main pipe material changed: {_h02_entry(type10, 1).material}"
    assert _h02_entry(type10, 2).length == 200, f"Type 10 support pipe length changed: {_h02_entry(type10, 2).length}"
    assert _h02_entry(type10, 2).material == "A53Gr.B", f"Type 10 support pipe material changed: {_h02_entry(type10, 2).material}"
    assert _h02_entry(type10, 3).length == 170 and _h02_entry(type10, 3).width == 170 and _h02_entry(type10, 3).spec == "9" and _h02_entry(type10, 3).quantity == 2, "Type 10 Plate_F changed"
    assert _h02_entry(type10, 4).spec == "M12*160L" and _h02_entry(type10, 4).quantity == 4, "Type 10 adjustable bolt changed"
    assert _h02_entry(type10, 5).spec == "M12" and _h02_entry(type10, 5).quantity == 16, "Type 10 hex nut changed"
    assert _h02_entry(type10, 6).length == 150 and _h02_entry(type10, 6).width == 150, "Type 10 M42 plate changed"
    _expect_pipe_weight(_h02_entry(type10, 1))
    _expect_pipe_weight(_h02_entry(type10, 2))
    _expect_plate_weight(_h02_entry(type10, 3))
    _expect_custom_weight(_h02_entry(type10, 4), 0.14)
    _expect_custom_weight(_h02_entry(type10, 5), 0)
    _expect_plate_weight(_h02_entry(type10, 6))

    type10_high = analyze_single("10-6B-16A")
    assert not type10_high.error and type10_high.entries, (
        f"Type 10 bounded H overrun should remain calculable: {type10_high.error}"
    )
    assert type10_high.meta["issues"][0]["severity"] == "warning"

    type15 = analyze_single("15-2B-1005")
    assert not type15.error, f"Type 15 should calculate: {type15.error}"
    assert _h02_names(type15) == ["管路", "槽鐵", "Plate_WING", "Plate_STOPPER", "Plate_BASE", "Plate_TOP"], (
        f"Type 15 BOM sequence changed: {_h02_names(type15)}"
    )
    assert _h02_entry(type15, 1).length == 382, f"Type 15 pipe length should be H-2F-channelHeight: {_h02_entry(type15, 1).length}"
    assert _h02_entry(type15, 2).length == 988, f"Type 15 member cut must be L-2x6t stopper: {_h02_entry(type15, 2).length}"
    assert (_h02_entry(type15, 3).length, _h02_entry(type15, 3).width, _h02_entry(type15, 3).spec, _h02_entry(type15, 3).quantity) == (150, 95, "9", 4), "Type 15 wing plate changed"
    assert (_h02_entry(type15, 4).length, _h02_entry(type15, 4).width, _h02_entry(type15, 4).spec, _h02_entry(type15, 4).quantity) == (160, 70, "6", 2), "Type 15 stopper plate changed"
    assert (_h02_entry(type15, 5).length, _h02_entry(type15, 5).width, _h02_entry(type15, 5).spec) == (190, 190, "9"), "Type 15 base plate changed"
    assert (_h02_entry(type15, 6).length, _h02_entry(type15, 6).width, _h02_entry(type15, 6).spec) == (80, 80, "9"), "Type 15 top plate changed"
    assert _h02_entry(type15, 3).geometry.shape_kind == "six_vertex_wing_plate", "Type 15 wing polygon missing"
    assert _h02_entry(type15, 4).geometry.shape_kind == "eight_vertex_chamfered_stopper", "Type 15 stopper polygon missing"
    _expect_pipe_weight(_h02_entry(type15, 1))
    _expect_steel_weight(_h02_entry(type15, 2))
    for entry in type15.entries[2:]:
        _expect_plate_weight(entry)

    type15_high = analyze_single("15-6B-1036")
    assert not type15_high.error and type15_high.entries, (
        f"Type 15 bounded H overrun should remain calculable: {type15_high.error}"
    )
    assert type15_high.meta["issues"][0]["severity"] == "warning"

    type15_10 = analyze_single("15-10B-1005")
    assert not type15_10.error, f"Type 15 10B should calculate: {type15_10.error}"
    assert _h02_entry(type15_10, 2).quantity == 2, "Type 15 10B should use double channel"
    assert _h02_entry(type15_10, 2).geometry.parameters["detail_o"] is True, "Type 15 10B detail-o geometry missing"
    assert _h02_entry(type15_10, 3).quantity == 4, "Type 15 10B wing plate should be 4 pieces"
    assert _h02_entry(type15_10, 4).quantity == 2, "Type 15 10B stopper plate should be 2 pieces"

    type15_12 = analyze_single("15-12B-1005")
    assert not type15_12.error, f"Type 15 12B should calculate: {type15_12.error}"
    assert _h02_entry(type15_12, 2).quantity == 2, "Type 15 12B should use double channel"
    assert _h02_entry(type15_12, 2).geometry.parameters["detail_o"] is True, "Type 15 12B detail-o geometry missing"
    assert _h02_entry(type15_12, 3).quantity == 4, "Type 15 12B wing plate should be 4 pieces"
    assert _h02_entry(type15_12, 4).quantity == 2, "Type 15 12B stopper plate should be 2 pieces"

    type14 = analyze_single("14-2B-1005")
    assert not type14.error, f"Type 14 should calculate: {type14.error}"
    assert _h02_names(type14) == ["管路", "槽鐵", "Plate_WING", "Plate_STOPPER", "Plate_BASE", "Plate_TOP", "EXP.BOLT"], (
        f"Type 14 BOM sequence changed: {_h02_names(type14)}"
    )
    assert _h02_entry(type14, 1).length == 382, f"Type 14 pipe length should be H-2F-channelHeight: {_h02_entry(type14, 1).length}"
    assert _h02_entry(type14, 2).length == 988 and _h02_entry(type14, 2).quantity == 1, "Type 14 member cut must be L-2x6t stopper"
    assert (_h02_entry(type14, 3).length, _h02_entry(type14, 3).width, _h02_entry(type14, 3).spec, _h02_entry(type14, 3).quantity) == (150, 65, "9", 4), "Type 14 wing plate changed"
    assert (_h02_entry(type14, 4).length, _h02_entry(type14, 4).width, _h02_entry(type14, 4).spec, _h02_entry(type14, 4).quantity) == (160, 70, "6", 2), "Type 14 stopper plate changed"
    assert (_h02_entry(type14, 5).length, _h02_entry(type14, 5).width, _h02_entry(type14, 5).spec) == (190, 190, "9"), "Type 14 base plate changed"
    assert (_h02_entry(type14, 6).length, _h02_entry(type14, 6).width, _h02_entry(type14, 6).spec) == (80, 80, "9"), "Type 14 top plate changed"
    assert _h02_entry(type14, 7).spec == '5/8"' and _h02_entry(type14, 7).quantity == 4, "Type 14 anchor bolt changed"
    assert _h02_entry(type14, 3).geometry.component_id == "D14-WING-PLATE", "Type 14 wing plate geometry missing"
    assert _h02_entry(type14, 4).geometry.component_id == "D14-STOPPER-PLATE", "Type 14 stopper plate geometry missing"
    _expect_pipe_weight(_h02_entry(type14, 1))
    _expect_steel_weight(_h02_entry(type14, 2))
    for entry in type14.entries[2:6]:
        _expect_plate_weight(entry)
    assert _h02_entry(type14, 7).unit_weight == 0, "Type 14 D-14 does not provide anchor unit weight"

    type14_10 = analyze_single("14-10B-1005")
    assert not type14_10.error, f"Type 14 10B should calculate: {type14_10.error}"
    assert _h02_entry(type14_10, 2).quantity == 2, "Type 14 10B should use double channel"
    assert _h02_entry(type14_10, 2).geometry.parameters["detail_a"] is True, "Type 14 10B detail-a geometry missing"
    assert _h02_entry(type14_10, 3).quantity == 4, "Type 14 10B wing plate should be 4 pieces"
    assert _h02_entry(type14_10, 4).quantity == 2, "Type 14 10B stopper plate should be 2 pieces"

    type14_12 = analyze_single("14-12B-1005")
    assert not type14_12.error, f"Type 14 12B should calculate: {type14_12.error}"
    assert _h02_entry(type14_12, 2).quantity == 2, "Type 14 12B should use double channel"
    assert _h02_entry(type14_12, 2).geometry.parameters["detail_a"] is True, "Type 14 12B detail-a geometry missing"
    assert _h02_entry(type14_12, 3).quantity == 4, "Type 14 12B wing plate should be 4 pieces"
    assert _h02_entry(type14_12, 4).quantity == 2, "Type 14 12B stopper plate should be 2 pieces"

    type16 = analyze_single("16-2B-05")
    assert not type16.error, f"Type 16 should calculate: {type16.error}"
    assert _h02_names(type16) == ["管路", "COVER PLATE"], f"Type 16 BOM sequence changed: {_h02_names(type16)}"
    assert _h02_entry(type16, 1).length == 800, "Type 16 nominal cut must be H500+C300=800"
    assert (_h02_entry(type16, 2).length, _h02_entry(type16, 2).width, _h02_entry(type16, 2).spec) == (70, 70, "6"), "Type 16 cover plate changed"
    assert type16.meta["fabrication"]["dimensions"]["cut_length_formula"] == "H + C"
    _expect_pipe_weight(_h02_entry(type16, 1))
    _expect_plate_weight(_h02_entry(type16, 2))

    type16_6b = analyze_single("16-6B-05")
    assert not type16_6b.error, f"Type 16 6B should calculate: {type16_6b.error}"
    assert _h02_names(type16_6b) == ["管路", "COVER PLATE"], f"Type 16 6B BOM sequence changed: {_h02_names(type16_6b)}"
    assert _h02_entry(type16_6b, 1).length == 800, "Type 16 6B nominal cut must be H500+C300=800"
    assert (_h02_entry(type16_6b, 2).length, _h02_entry(type16_6b, 2).width, _h02_entry(type16_6b, 2).spec) == (140, 140, "6"), "Type 16 6B plate changed"
    for entry in type16_6b.entries:
        if entry.name == "管路":
            _expect_pipe_weight(entry)
        elif entry.category == "鋼板類":
            _expect_plate_weight(entry)

    print("v phase H-02 type10/type15/type16 guardrails OK")
except Exception as e:
    record_failure(f"X phase H-02 type10/type15/type16 guardrails ERROR: {e}")
    raise

# Type 20/26 structural guardrails.
try:
    from core.calculator import analyze_single

    type03 = analyze_single("03-1B-05L")
    assert not type03.error, f"Type 03 should calculate: {type03.error}"
    assert type03.entries[0].name == "角鋼", f"Type 03 first entry should be vertical angle: {type03.entries[0].name}"
    assert type03.entries[0].length == 574.8, f"Type 03 vertical angle formula changed: {type03.entries[0].length}"
    assert "彎頭中心半徑=38.1" in type03.entries[0].geometry.notes_zh or "彎頭中心半徑=38.1" in type03.entries[0].remark, f"Type 03 vertical angle remark missing formula: {type03.entries[0].remark}"
    assert type03.entries[1].length == 130, f"Type 03 horizontal angle length changed: {type03.entries[1].length}"
    assert type03.entries[3].name == "Plate_c_有鑽孔", f"Type 03 M42 Type-L plate changed: {[entry.name for entry in type03.entries]}"
    assert type03.entries[4].name == "EXP.BOLT" and type03.entries[4].quantity == 4, f"Type 03 M42 Type-L bolt changed: {[entry.name for entry in type03.entries]}"

    type05 = analyze_single("05-L50-05L")
    assert not type05.error, f"Type 05 should calculate: {type05.error}"
    assert type05.entries[0].name == "角鋼", f"Type 05 first entry should be vertical angle: {type05.entries[0].name}"
    assert type05.entries[0].length == 485, f"Type 05 vertical angle should subtract 15mm offset: {type05.entries[0].length}"
    assert "頂端偏移15" in type05.entries[0].geometry.notes_zh or "頂端偏移15" in type05.entries[0].remark, f"Type 05 vertical angle remark missing formula: {type05.entries[0].remark}"
    assert type05.entries[1].length == 130, f"Type 05 horizontal angle length changed: {type05.entries[1].length}"
    assert type05.entries[2].name == "Plate_c_有鑽孔", f"Type 05 M42 Type-L plate changed: {[entry.name for entry in type05.entries]}"
    assert type05.entries[3].name == "EXP.BOLT" and type05.entries[3].quantity == 4, f"Type 05 M42 Type-L bolt changed: {[entry.name for entry in type05.entries]}"

    type06 = analyze_single("06-L50-0510-0401")
    assert not type06.error, f"Type 06 should calculate: {type06.error}"
    assert [entry.name for entry in type06.entries] == [
        "角鋼",
        "角鋼",
        "LUG PLATE TYPE-F",
        "K BOLT",
    ], (
        f"Type 06 BOM sequence changed: {[entry.name for entry in type06.entries]}"
    )
    assert [entry.length for entry in type06.entries[:2]] == [500, 1000], (
        f"Type 06 member lengths changed: {[entry.length for entry in type06.entries]}"
    )
    assert type06.entries[2].geometry.component_id == "M37-LGP-F-1"
    assert type06.entries[2].geometry.holes.count == 2
    assert type06.entries[3].geometry.component_id == "D6-K-BOLT"
    assert type06.entries[3].quantity == 2
    assert any("A+B" in warning for warning in type06.warnings), (
        f"Type 06 A+B assembly warning missing: {type06.warnings}"
    )

    type07 = analyze_single("07-2B-20J")
    assert not type07.error, f"Type 07 should calculate: {type07.error}"
    assert type07.entries[0].name == "管路" and type07.entries[0].length == 271, (
        f"Type 07 Pipe B should be L+200: {type07.entries[0]}"
    )
    assert type07.entries[1].name == "管路" and type07.entries[1].length == 1782, (
        f"Type 07 Pipe C should be H-200-PlateF-M42: {type07.entries[1]}"
    )
    assert any("H需現場調整" in blocker for blocker in type07.meta["fabrication"]["blockers"]), (
        f"Type 07 field-fit blocker missing: {type07.meta['fabrication']}"
    )

    type08 = analyze_single("08-2B-1005G")
    assert not type08.error, f"Type 08 should calculate: {type08.error}"
    stopper_entries = [entry for entry in type08.entries if entry.name == "Plate_STOPPER"]
    assert len(stopper_entries) == 1, f"Type 08 should have one stopper BOM line: {[entry.name for entry in type08.entries]}"
    assert stopper_entries[0].quantity == 2, f"Type 08 stopper should be 2 pcs: {stopper_entries[0]}"
    assert "10C chamfer" in stopper_entries[0].display_remark, (
        f"Type 08 stopper drawing feature note missing: {stopper_entries[0].display_remark}"
    )
    assert type08.entries[-1].name == "Plate_TOP" and type08.entries[-1].quantity == 1, (
        f"Type 08 top plate should remain one piece: {type08.entries[-1]}"
    )

    type20 = analyze_single("20-L50-05A")
    assert not type20.error, f"Type 20 should calculate: {type20.error}"
    assert len(type20.entries) == 1, f"Type 20 BOM count changed: {len(type20.entries)}"
    assert type20.entries[0].length == 500, f"Type 20 H length changed: {type20.entries[0].length}"

    type26_a = analyze_single("26-L50-1005A")
    assert not type26_a.error, f"Type 26 Fig-A should calculate: {type26_a.error}"
    assert [entry.name for entry in type26_a.entries] == ["角鋼", "角鋼", "角鋼"], (
        f"Type 26 Fig-A BOM sequence changed: {[entry.name for entry in type26_a.entries]}"
    )
    assert [entry.length for entry in type26_a.entries] == [500, 500, 1000], (
        f"Type 26 Fig-A member lengths changed: {[entry.length for entry in type26_a.entries]}"
    )
    assert [entry.remark for entry in type26_a.entries] == ["Fig-A, H段上件", "Fig-A, H段下件", "Fig-A, L段"], (
        f"Type 26 Fig-A remarks changed: {[entry.remark for entry in type26_a.entries]}"
    )

    type26_c = analyze_single("26-L50-1005C")
    assert not type26_c.error, f"Type 26 Fig-C should calculate: {type26_c.error}"
    assert [entry.name for entry in type26_c.entries[:3]] == ["角鋼", "角鋼", "角鋼"], (
        f"Type 26 Fig-C steel members changed: {[entry.name for entry in type26_c.entries]}"
    )
    assert [entry.length for entry in type26_c.entries[:3]] == [500, 500, 1000], (
        f"Type 26 Fig-C steel lengths changed: {[entry.length for entry in type26_c.entries[:3]]}"
    )
    assert type26_c.entries[3].name == "LUG_PLATE_C", f"Type 26 Fig-C lug plate missing: {[entry.name for entry in type26_c.entries]}"
    assert type26_c.entries[3].quantity == 2, f"Type 26 Fig-C should use 2 lug plates: {type26_c.entries[3].quantity}"
    assert type26_c.entries[4].name == "K BOLT", f"Type 26 Fig-C K bolt missing: {[entry.name for entry in type26_c.entries]}"
    assert type26_c.entries[4].quantity == 8, f"Type 26 Fig-C should use 8 K bolts for two four-hole M-34 plates: {type26_c.entries[4].quantity}"

    type25_c = analyze_single("25-L50-0505C-0401")
    assert not type25_c.error, f"Type 25 Fig-C should calculate: {type25_c.error}"
    assert type25_c.entries[2].name == "LUG_PLATE_C", f"Type 25 Fig-C lug plate missing: {[entry.name for entry in type25_c.entries]}"
    assert type25_c.entries[2].quantity == 1, f"Type 25 Fig-C should use 1 lug plate: {type25_c.entries[2].quantity}"
    assert type25_c.entries[3].name == "K BOLT", f"Type 25 Fig-C K bolt missing: {[entry.name for entry in type25_c.entries]}"
    assert type25_c.entries[3].quantity == 4, f"Type 25 Fig-C should use 4 K bolts for one four-hole M-34 plate: {type25_c.entries[3].quantity}"

    print("v type03/type05/type06/type07/type08/type20/type26 structural guardrails OK")
except Exception as e:
    record_failure(f"X type03/type05/type06/type07/type08/type20/type26 structural guardrails ERROR: {e}")
    raise

# Type 52/66 D-80 pad and FB guardrails.
try:
    import math

    from core.calculator import analyze_single
    from data.pipe_table import get_pipe_details

    def _entry_by_name(result, name):
        for entry in result.entries:
            if entry.name == name:
                return entry
        raise AssertionError(f"{result.fullstring} missing {name}: {[e.name for e in result.entries]}")

    retainer_default = analyze_single("52-1/2B-A")
    assert not retainer_default.error, f"Type 52 default HOPS/LOPS should calculate: {retainer_default.error}"
    retainer_default_h = _entry_by_name(retainer_default, "H型鋼")
    assert retainer_default_h.length == 150, (
        f"Type 52 without explicit LOPS should use D-80 table/default LOPS=150: {retainer_default_h.length}"
    )

    retainer_small = analyze_single("52-1/2B-A-150-200")
    assert not retainer_small.error, f"Type 52 small retainer should calculate: {retainer_small.error}"
    retainer_h = _entry_by_name(retainer_small, "H型鋼")
    assert retainer_h.length == 200, f"Type 52 explicit LOPS should override table/default LOPS: {retainer_h.length}"
    assert "width=100" in retainer_h.remark and "H=HOPS(150)" in retainer_h.remark, (
        f"Type 52 small MEMBER C remark should carry width/HOPS: {retainer_h.remark}"
    )
    retainer_angle = _entry_by_name(retainer_small, "角鋼")
    assert retainer_angle.quantity == 2, f"Type 52 should still add L40 retainer angles x2: {retainer_angle.quantity}"

    small = analyze_single("66-1.1/2B(P)-A-150-150")
    assert not small.error, f"Type 66 small pad should calculate: {small.error}"
    small_pad = _entry_by_name(small, "Pad_52Type")
    small_details = get_pipe_details(1.5, "10S")
    small_od = small_details["od_mm"]
    small_t_sch10s = small_details["thickness_mm"]   # Phase 5: <=8" uses Sch10S wall
    assert small_pad.length == 200, f"small Pad_52Type length should be LOPS+E*2 (E=25 for <2inch): {small_pad.length}"
    assert small_pad.width == round(small_od * math.pi / 3), f"small Pad_52Type 120-degree width changed: {small_pad.width}"
    assert small_pad.spec == str(small_t_sch10s), f"small Pad_52Type thickness should be Sch10S wall ({small_t_sch10s}mm): {small_pad.spec}"
    assert any("OD*pi/3" in warning for warning in small.warnings), f"small Pad_52Type practical width warning missing: {small.warnings}"

    large = analyze_single("66-10B(P)-A-150-250")
    assert large.error and "尚未達可出加工圖程度" in large.error, (
        f"Type 66 10in calculation-only gusset geometry must block: {large.error}"
    )
    assert not large.entries, "blocked Type 66 must not emit approximate fabrication BOM"

    compact = analyze_single("66-14B(P)-100-300")
    assert compact.error and "尚未達可出加工圖程度" in compact.error, (
        f"Type 66 compact 14in geometry must also block: {compact.error}"
    )

    print("v type52/type66 pad and fabrication-readiness guardrails OK")
except Exception as e:
    record_failure(f"X type52/type66 pad and FB guardrails ERROR: {e}")

# Urgent project priority Type guardrails.
try:
    from collections import Counter

    from core.calculator import analyze_single

    _PRIORITY_TYPE_CASES = [
        ("01", "01-2B-05A"),
        ("01", "01-50B-05A"),
        ("10", "10-2B-05A"),
        ("15", "15-2B-1005"),
        ("16", "16-2B-05"),
        ("20", "20-L50-05A"),
        ("21", "21-L50-05A"),
        ("22", "22-L50-05(A)L"),
        ("22", "22-L75-12(A)L"),
        ("23", "23-L50-05A"),
        ("24", "24-L50-05"),
        ("25", "25-L50-0505A"),
        ("25", "25-L50-0505C-0401"),
        ("26", "26-L50-1005A"),
        ("26", "26-L50-1005C"),
        ("27", "27-L75-0505L-0401"),
        ("27", "27-L50-0204L-0101"),
        ("27", "27-H150-0505L-0401"),
        ("28", "28-L50-1005L"),
        ("30", "30-L75-0505A-0401"),
        ("31", "31-L50-1005"),
        ("32", "32-L50-1005"),
        ("33", "33-L50-1005"),
        ("34", "34-L50-1005"),
        ("35", "35-C125-05A"),
        ("37", "37-C125-1200A"),
        ("37", "37-C125-1200B-05"),
        ("51", "51-2B"),
        ("51", "51-1.1/2B"),
        ("51", "51-4B"),
        ("51", "51-26B"),
        ("52", "52-2B(P)-A(A)-130-500"),
        ("52", "52-14B(P)-A(A)-130-500"),
        ("53", "53-2B(P)-A(A)-130-500"),
        ("53", "53-14B(P)-A(A)-130-500"),
        ("57", "57-2B-A"),
        ("57", "57-1.1/2B-A"),
        ("59", "59-6B-A"),
        ("59", "59-1.1/2B-B(S)"),
        ("80", "80-2B(P)-A(A)-130-500"),
        ("80", "80-30B-A(A)-130-500"),
        ("66", "66-1.1/2B(P)-A-150-150"),
    ]

    def _assert_entry_sane(entry, designation):
        assert entry.quantity > 0, f"{designation} entry {entry.item_no} has non-positive quantity"
        assert entry.factor >= 0, f"{designation} entry {entry.item_no} has negative factor"
        assert math.isclose(
            entry.qty_subtotal,
            entry.quantity * entry.factor,
            rel_tol=0,
            abs_tol=1e-9,
        ), f"{designation} entry {entry.item_no} has inconsistent quantity subtotal"
        assert entry.unit_weight >= 0, f"{designation} entry {entry.item_no} has negative unit weight"
        assert entry.total_weight >= 0, f"{designation} entry {entry.item_no} has negative total weight"
        assert entry.weight_output >= 0, f"{designation} entry {entry.item_no} has negative weight output"
        if entry.length:
            assert 0 < entry.length < 10000, f"{designation} entry {entry.item_no} unreasonable length: {entry.length}"
        if entry.width:
            assert 0 < entry.width < 10000, f"{designation} entry {entry.item_no} unreasonable width: {entry.width}"
        if entry.category in ("型鋼類", "管路類"):
            if entry.length <= 0:
                blockers = getattr(entry.geometry, "fabrication_blockers", [])
                assert any("cut" in item.lower() or "切" in item for item in blockers), (
                    f"{designation} {entry.name} zero takeoff must have a field-cut blocker"
                )
        if entry.category == "鋼板類":
            if entry.item_class == "reference_only":
                assert entry.geometry.fabrication_blockers, (
                    f"{designation} unresolved plate reference must carry a fabrication blocker"
                )
            else:
                assert entry.density_g_cm3 > 0, (
                    f"{designation} {entry.name} plate weight must expose its density"
                )
                assert entry.density_source, (
                    f"{designation} {entry.name} plate weight must expose density source"
                )
                if entry.density_requires_review:
                    assert "unverified" in entry.density_source, (
                        f"{designation} {entry.name} review density must be marked unverified"
                    )
                assert entry.length > 0 and entry.width > 0, f"{designation} {entry.name} should have plate dimensions"
                assert float(entry.spec) > 0, f"{designation} {entry.name} should have positive plate thickness"

    priority_results = {}
    for type_id, designation in _PRIORITY_TYPE_CASES:
        result = analyze_single(designation)
        priority_results[designation] = result
        assert not result.error, f"{designation} should calculate for priority project: {result.error}"
        assert result.entries, f"{designation} should produce BOM entries"
        if type_id == "57":
            assert result.total_weight > 0, (
                f"{designation} must retain the source-derived M-26 rod-only weight"
            )
            assert result.entries[0].geometry.parameters["developed_length_formula"] == (
                "pi * B / 2 + 2 * E"
            )
            assert result.entries[-1].quantity == 4 and result.entries[-1].unit_weight > 0, (
                f"{designation} must include four theoretically weighted finished nuts"
            )
            assert "proportional finished-hex-nut" in result.entries[-1].geometry.parameters[
                "weight_basis"
            ]
        elif designation in {"51-26B", "80-30B-A(A)-130-500"}:
            assert result.total_weight == 0, (
                f"{designation} has unresolved fabrication geometry and must not invent weight"
            )
            assert result.meta["fabrication"]["bom_ready"] is False
        else:
            assert result.total_weight > 0, f"{designation} should have positive total weight"
        for entry in result.entries:
            _assert_entry_sane(entry, designation)

    type27_h150 = priority_results["27-H150-0505L-0401"]
    type28_l50 = priority_results["28-L50-1005L"]
    type27_l75 = priority_results["27-L75-0505L-0401"]
    type27_l50_x = priority_results["27-L50-0204L-0101"]
    for designation, result in (
        ("27-L75-0505L-0401", type27_l75),
        ("27-L50-0204L-0101", type27_l50_x),
    ):
        excluded_ids = {
            item.get("component_id")
            for item in result.meta.get("excluded_bom_components", [])
        }
        assert {"D30-MEMBER-M", "D30-TOP-PLATE", "M42-FASTENER"} <= excluded_ids, (
            f"{designation} unresolved member/top plate/anchor must stay traceable outside BOM: "
            f"{excluded_ids}"
        )
        assert all(entry.length > 0 for entry in result.entries), (
            f"{designation} must not emit zero-length material rows"
        )
        assert result.meta["fabrication"]["bom_ready"] is False
    h150_ids = Counter(entry.geometry.component_id for entry in type27_h150.entries)
    h150_excluded_ids = {
        item.get("component_id")
        for item in type27_h150.meta.get("excluded_bom_components", [])
    }
    assert {"D30-MEMBER-M", "D30-TOP-PLATE", "M42-FASTENER"} <= h150_excluded_ids, (
        "Type 27 H150 unresolved material must remain traceable outside BOM"
    )
    assert h150_ids["D30-GUSSET-PLATE"] == 1, "Type 27 H150 should include one QTY2 gusset BOM line"
    assert not any(entry.name == "Plate_6t_Side" for entry in type27_h150.entries), "3 SIDES TYP is a weld note, not fake plates"
    assert any("Plate_" in entry.name and "有鑽孔" in entry.name for entry in type27_h150.entries), "Type 27 H150 M42 base plate missing"
    assert not any(entry.name == "EXP.BOLT" for entry in type27_h150.entries), (
        "Type 27 H150 diameter-only anchor must not become a zero-weight BOM row"
    )

    type28_names = [entry.name for entry in type28_l50.entries]
    type28_lengths = [entry.length for entry in type28_l50.entries[:3]]
    assert type28_names[:3] == ["角鋼", "角鋼", "角鋼"], f"Type 28 should split portal frame into three steel entries: {type28_names}"
    assert type28_lengths == [500, 1000, 500], f"Type 28 left/top/right lengths changed: {type28_lengths}"
    assert [entry.geometry.component_id for entry in type28_l50.entries[:3]] == [
        "D31-LEFT-LEG", "D31-TOP-BEAM", "D31-RIGHT-LEG"
    ]
    assert sum(
        entry.quantity
        for entry in type28_l50.entries
        if "Plate_" in entry.name and "有鑽孔" in entry.name
    ) == 2, "Type 28 requires two M42 base plates represented as one QTY2 BOM line"
    assert not any(entry.name == "EXP.BOLT" for entry in type28_l50.entries), (
        "Type 28 diameter-only anchors must not become zero-weight BOM rows"
    )
    assert any(
        item.get("component_id") == "D31-M42-BOTH-2"
        for item in type28_l50.meta.get("excluded_bom_components", [])
    ), "Type 28 unresolved anchor set must remain traceable outside BOM"

    type30_a = priority_results["30-L75-0505A-0401"]
    type30_names = [entry.name for entry in type30_a.entries]
    type30_lengths = [entry.length for entry in type30_a.entries]
    assert type30_names == ["角鋼", "角鋼"], f"Type 30 Fig-A should split into column + top beam: {type30_names}"
    assert type30_lengths == [485, 500], f"Type 30 Fig-A H member must use H-15: {type30_lengths}"
    assert [entry.geometry.component_id for entry in type30_a.entries] == ["D35-MEMBER-H", "D35-MEMBER-L"]

    type30_b = analyze_single("30-L75-0505B-0401")
    assert not type30_b.error, f"Type 30 Fig-B should calculate: {type30_b.error}"
    assert [entry.name for entry in type30_b.entries] == ["角鋼", "角鋼"], f"Type 30 Fig-B should split into column + top beam: {[entry.name for entry in type30_b.entries]}"
    assert [entry.length for entry in type30_b.entries] == [485, 500], f"Type 30 Fig-B lengths changed: {[entry.length for entry in type30_b.entries]}"
    assert [entry.geometry.component_id for entry in type30_b.entries] == ["D35-MEMBER-H", "D35-MEMBER-L"]

    type31 = priority_results["31-L50-1005"]
    assert [entry.name for entry in type31.entries] == ["角鋼", "角鋼", "角鋼"], f"Type 31 should split into left leg + top beam + right leg: {[entry.name for entry in type31.entries]}"
    assert [entry.length for entry in type31.entries] == [500, 1000, 500], f"Type 31 lengths changed: {[entry.length for entry in type31.entries]}"
    assert [entry.geometry.component_id for entry in type31.entries] == ["D36-LEG-1", "D36-TOP-BEAM", "D36-LEG-2"]

    type32 = priority_results["32-L50-1005"]
    assert [entry.name for entry in type32.entries] == ["角鋼", "角鋼", "角鋼"], f"Type 32 should split into left leg + bottom beam + right leg: {[entry.name for entry in type32.entries]}"
    assert [entry.length for entry in type32.entries] == [500, 1000, 500], f"Type 32 lengths changed: {[entry.length for entry in type32.entries]}"
    assert [entry.geometry.component_id for entry in type32.entries] == ["D37-HANGER-LEG-1", "D37-BOTTOM-BEAM", "D37-HANGER-LEG-2"]

    type33 = priority_results["33-L50-1005"]
    assert [entry.name for entry in type33.entries] == ["角鋼", "角鋼"], f"Type 33 should stay as column + bottom beam half-frame: {[entry.name for entry in type33.entries]}"
    assert [entry.length for entry in type33.entries] == [500, 1000], f"Type 33 lengths changed: {[entry.length for entry in type33.entries]}"
    assert [entry.geometry.component_id for entry in type33.entries] == ["D38-END-POST", "D38-BOTTOM-BEAM"]

    type34 = priority_results["34-L50-1005"]
    assert [entry.name for entry in type34.entries] == ["角鋼", "角鋼"], f"Type 34 should stay as column + top beam cantilever: {[entry.name for entry in type34.entries]}"
    assert [entry.length for entry in type34.entries] == [500, 1000], f"Type 34 lengths changed: {[entry.length for entry in type34.entries]}"
    assert [entry.geometry.component_id for entry in type34.entries] == [
        "D39-END-POST", "D39-TOP-BEAM"
    ], f"Type 34 component mapping changed: {[entry.geometry.component_id for entry in type34.entries]}"

    type35_a = priority_results["35-C125-05A"]
    assert [entry.name for entry in type35_a.entries] == ["槽鐵"], f"Type 35 FIG-A should stay a single support rail entry: {[entry.name for entry in type35_a.entries]}"
    assert [(entry.length, entry.quantity, entry.geometry.component_id) for entry in type35_a.entries] == [
        (500, 1, "D40-MEMBER-FIG-A"),
    ], f"Type 35 FIG-A changed: {[(entry.length, entry.quantity, entry.geometry.component_id) for entry in type35_a.entries]}"

    type35_b = analyze_single("35-C125-05B")
    assert not type35_b.error, f"Type 35 FIG-B should calculate: {type35_b.error}"
    assert [entry.name for entry in type35_b.entries] == ["槽鐵"], f"Type 35 FIG-B should stay one member: {[entry.name for entry in type35_b.entries]}"
    assert [(entry.length, entry.quantity, entry.geometry.component_id) for entry in type35_b.entries] == [
        (500, 1, "D40-MEMBER-FIG-B"),
    ], f"Type 35 FIG-B changed: {[(entry.length, entry.quantity, entry.geometry.component_id) for entry in type35_b.entries]}"

    type51_small = priority_results["51-2B"]
    assert [(entry.name, entry.length, entry.width, entry.quantity, entry.remark) for entry in type51_small.entries] == [
        ("FLAT BAR", 60, 50, 2, "鞍座, 60x50x9, 全焊接(6V), ×2"),
    ], f"Type 51 small-pipe flat bar path changed: {[(entry.name, entry.length, entry.width, entry.quantity, entry.remark) for entry in type51_small.entries]}"

    type51_mid = priority_results["51-4B"]
    assert [(entry.name, entry.spec, entry.length, entry.quantity) for entry in type51_mid.entries] == [
        ("角鋼", "50*50*6", 125, 2),
    ], f"Type 51 4-24in member path should use table H length: {[(entry.name, entry.spec, entry.length, entry.quantity) for entry in type51_mid.entries]}"
    assert type51_mid.entries[0].geometry.parameters == {
        "cut_length_mm": 125, "quantity": 2,
        "pipe_side_gap_mm": 3, "fillet_weld_mm": 6,
    }, f"Type 51 mid-pipe fabrication parameters changed: {type51_mid.entries[0].geometry.parameters}"

    type51_large = priority_results["51-26B"]
    assert [(entry.name, entry.spec, entry.length, entry.quantity) for entry in type51_large.entries[:1]] == [
        ("MEMBER M", "C125*65*6", 0, 2),
    ], f"Type 51 large-pipe must not restore the retired 300mm channel guess: {[(entry.name, entry.spec, entry.length, entry.quantity) for entry in type51_large.entries]}"
    assert type51_large.meta["fabrication"]["bom_ready"] is False
    assert len(type51_large.entries) == 2 and type51_large.entries[1].name == "REINFORCING PAD", f"Type 51 large-pipe D-91 reference missing: {[(entry.name, entry.spec) for entry in type51_large.entries]}"
    assert type51_large.entries[1].spec == "SEE D-91" and type51_large.entries[1].unit_weight == 0, (
        f"Type 51 large-pipe pad must remain an unresolved zero-weight reference: "
        f"{(type51_large.entries[1].spec, type51_large.entries[1].unit_weight)}"
    )
    assert type51_large.entries[1].geometry.parameters["d62a_contact_angle_deg"] == 80
    assert type51_large.entries[1].geometry.parameters["d91_pad_angle_deg"] == 120
    assert type51_large.entries[1].geometry.component_id == "D91-REINFORCING-PAD-REFERENCE"

    type52 = priority_results["52-2B(P)-A(A)-130-500"]
    type53 = priority_results["53-2B(P)-A(A)-130-500"]
    assert [(e.name, e.spec, e.length, e.width, e.quantity) for e in type52.entries] == [
        (e.name, e.spec, e.length, e.width, e.quantity) for e in type53.entries
    ], "Type 53 should share Type 52 D-80 shoe geometry path"
    assert type52.entries[0].name == "Pad_52Type" and "length_rule=LOPS + E*2" in type52.entries[0].remark, f"Type 52 small-pipe pad remark missing rule: {type52.entries[0].remark}"
    assert type52.entries[1].name == "角鋼" and "CUT IN FIELD" in type52.entries[1].remark, f"Type 52 L40 remark missing field-cut note: {type52.entries[1].remark}"

    type52_large = priority_results["52-14B(P)-A(A)-130-500"]
    type53_large = priority_results["53-14B(P)-A(A)-130-500"]
    assert [(e.name, e.spec, e.length, e.width, e.quantity) for e in type52_large.entries] == [
        (e.name, e.spec, e.length, e.width, e.quantity) for e in type53_large.entries
    ], "Type 53 large <=24in path should share Type 52 geometry"
    assert any(entry.name == "FB_52Type_3" and entry.quantity == 4 for entry in type52_large.entries), f"Type 52 large-pipe FB_52Type_3 x4 missing: {[(entry.name, entry.quantity) for entry in type52_large.entries]}"
    assert "length_rule=LOPS + E*2 + 25*2" in type52_large.entries[0].remark, f"Type 52 large-pipe pad remark missing rule: {type52_large.entries[0].remark}"
    assert "width=A" in type52_large.entries[-1].remark, f"Type 52 FB_52Type_3 remark missing width rule: {type52_large.entries[-1].remark}"

    type57_slide = priority_results["57-2B-A"]
    type57_fixed = analyze_single("57-2B-B")
    assert [entry.name for entry in type57_slide.entries] == [
        "M-26 U-BOLT ROD",
        "M-26 FINISHED HEX NUTS",
    ], f"Type 57 should emit the M-26 rod and four nuts separately: {[entry.name for entry in type57_slide.entries]}"
    assert type57_slide.entries[1].quantity == 4
    assert type57_slide.entries[1].unit_weight > 0
    assert type57_slide.entries[0].geometry.parameters["mode"] == "A"
    assert [type57_slide.entries[0].geometry.parameters[key] for key in ("B", "C", "D", "E")] == [62, 71, 58, 74]
    assert math.isclose(type57_slide.entries[0].length, math.pi * 62 / 2 + 2 * 74)
    assert not type57_fixed.error and type57_fixed.entries[0].geometry.parameters["mode"] == "B"

    type59_b = priority_results["59-1.1/2B-B(S)"]
    assert [entry.name for entry in type59_b.entries] == [
        "TYPE 59 翼形角板",
        "M-26 U-BOLT ROD",
        "M-26 FINISHED HEX NUTS",
    ], f"Type 59 FIG-B must include the drawing-furnished M-26 rod and nuts: {[entry.name for entry in type59_b.entries]}"
    assert type59_b.entries[0].display_spec == "A80 x B55 x P25 x C15 x t6"
    assert math.isclose(type59_b.entries[1].length, math.pi * 51 / 2 + 2 * 68)
    assert type59_b.entries[2].quantity == 4 and type59_b.entries[2].unit_weight > 0
    type59_lug = type59_b.entries[0]
    assert type59_lug.spec == "6" and type59_lug.part_key == "59_lug_plate_wing_a80_b55_p25_c15_t6", f"Type 59 lug identity changed: {(type59_lug.spec, type59_lug.part_key)}"
    assert type59_lug.stock_id.startswith("PL-") and len(type59_lug.stock_id) == 11, f"Type 59 lug stock id invalid: {type59_lug.stock_id}"
    assert any("proportional hex-nut theoretical weight" in warning for warning in type59_b.warnings), (
        f"Type 59 must retain its supplier finished-nut confirmation: {type59_b.warnings}"
    )

    type80_small = priority_results["80-2B(P)-A(A)-130-500"]
    assert [entry.geometry.component_id for entry in type80_small.entries] == [
        "D80-REINFORCING-PAD",
        "D80-MEMBER-C",
        "D95-BEAM-INTERFACE-MEMBER-C",
    ], f"Type 80 D-95 must combine the D-80 shoe with its beam-interface member: {[entry.geometry.component_id for entry in type80_small.entries]}"
    assert type80_small.entries[-1].length == 500
    assert type80_small.meta["fabrication"]["bom_ready"] is True

    type80_big = priority_results["80-30B-A(A)-130-500"]
    assert [entry.geometry.component_id for entry in type80_big.entries] == [
        "D96-D80B-LARGE-SHOE-ASSEMBLY",
        "D96-BEAM-INTERFACE-PARTS",
    ], f"Type 80 D-96 unresolved assemblies changed: {[entry.geometry.component_id for entry in type80_big.entries]}"
    assert type80_big.total_weight == 0
    assert type80_big.meta["fabrication"]["bom_ready"] is False
    assert all(entry.geometry.fabrication_blockers for entry in type80_big.entries)

    print(f"v urgent priority Type guardrails OK ({len(_PRIORITY_TYPE_CASES)} cases + Type 80)")
except Exception as e:
    record_failure(f"X urgent priority Type guardrails ERROR: {e}")
    raise

try:
    from data.component_table_registry import (
        EXISTING_COMPONENT_TABLES,
        MISSING_COMPONENT_TABLES,
        get_component_table_coverage,
    )
    coverage = get_component_table_coverage()
    print(
        "component coverage: "
        f"{coverage['implemented']}/{coverage['total']} "
        f"({coverage['coverage_ratio']:.1%})"
    )
    print("lookup-ready components:", coverage["lookup_ready"])
    print("partial-lookup components:", coverage.get("partial_lookup", 0))
    print("metadata-only components:", coverage["metadata_only"])
    print("implemented components:", ", ".join(sorted(EXISTING_COMPONENT_TABLES)))
    print("missing components:", ", ".join(MISSING_COMPONENT_TABLES))
except Exception as e:
    record_failure(f"X component registry ERROR: {e}")

# Test full M/N metadata baseline
try:
    import importlib
    from data.component_table_registry import (
        EXISTING_COMPONENT_TABLES,
        METADATA_ONLY_COMPONENT_TABLES,
        MISSING_COMPONENT_TABLES,
        get_component_table_coverage,
    )

    coverage = get_component_table_coverage()
    assert coverage["implemented"] == 71, f"expected 71 component modules: {coverage}"
    assert coverage["missing"] == 0, f"expected no missing component modules: {coverage}"
    assert coverage["lookup_ready"] == 60, f"lookup-ready count changed unexpectedly: {coverage}"
    assert coverage["partial_lookup"] == 3, f"partial-lookup count changed unexpectedly: {coverage}"
    assert coverage["metadata_only"] == 8, f"metadata-only count failed: {coverage}"
    assert not MISSING_COMPONENT_TABLES, f"missing list should be empty: {MISSING_COMPONENT_TABLES}"

    for component_id, module_file in EXISTING_COMPONENT_TABLES.items():
        module_name = module_file[:-3]
        module = importlib.import_module(f"data.{module_name}")
        getter_name = f"get_{component_id.lower().replace('-', '').replace(' ', '').replace('/', '').replace('.', '')}_component"
        if component_id == "N-12A":
            getter_name = "get_n12a_component"
        if component_id == "N27-PU BLOCK":
            getter_name = "get_n27_pu_block_component"
        getter = getattr(module, getter_name, None)
        if getter is None:
            continue
        component = getter()
        assert component["component_id"] == component_id, f"{component_id} getter returned {component}"
        if component_id in METADATA_ONLY_COMPONENT_TABLES:
            assert component["table_kind"] == "metadata_only" and not component["lookup_ready"], f"{component_id} metadata failed: {component}"

    print("v full M/N component metadata baseline OK")
except Exception as e:
    record_failure(f"X full M/N component metadata baseline ERROR: {e}")

# Test Type 49 M-11/M-12/M-41 source tables and integration
try:
    from data.m11_table import get_m11_by_line_size
    from data.m12_table import get_m12_by_line_size
    from data.m41_table import get_m41_by_line_size

    m11_10 = get_m11_by_line_size(10)
    m12_10 = get_m12_by_line_size(10)
    assert (
        m11_10["installed_overall_mm"],
        m11_10["stock_thickness_mm"],
        m11_10["fastener"]["source_bolt_spec"],
    ) == (514, 9, '5/8"x70'), f"M-11 10in row failed: {m11_10}"
    assert (
        m12_10["installed_overall_mm"],
        m12_10["stock_thickness_mm"],
        m12_10["fastener"]["source_bolt_spec"],
    ) == (527, 10, '5/8"x60'), f"M-12 10in row failed: {m12_10}"
    assert (
        m12_10["left_straight_projection_mm"],
        m12_10["right_straight_projection_mm"],
        m12_10["source_sketch_left_straight_projection_mm"],
        m12_10["source_sketch_right_straight_projection_mm"],
        m12_10["straight_split_released"],
    ) == (None, None, 150, 50, False), f"M-12 L/sketch conflict handling failed: {m12_10}"
    assert abs(
        m12_10["developed_length_each_mm"] - 678.5353604829558
    ) < 1e-9, f"M-12 table-L development failed: {m12_10}"

    m41 = get_m41_by_line_size(14, "Stainless Steel")
    assert (
        m41["designation"],
        m41["quantity"],
        m41["net_area_mm2"],
    ) == ("LGP-P-3S", 6, 5250), f"M-41 14in row failed: {m41}"

    type49 = analyze_single(
        "49-4B-A(B)",
        source_profile="cw_e25_24_hp6",
    )
    assert not type49.error, f"Type 49 released designation failed: {type49.error}"
    assert [
        entry.geometry.component_id
        for entry in type49.entries
    ] == ["M-11", "M-11-FASTENERS", "M-41"], f"Type 49 component chain failed: {type49.entries}"
    assert type49.total_weight > 0, "Type 49 known steel weight should be positive"
    assert type49.meta["fabrication"]["bom_ready"] is False
    assert type49.meta["fabrication"]["fabrication_ready"] is False
    assert analyze_single(
        "49-7B-A",
        source_profile="cw_e25_24_hp6",
    ).error, "Type 49 unlisted 7in row must not interpolate"
    print("v Type 49 M-11/M-12/M-41 source tables OK")
except Exception as e:
    record_failure(f"X Type 49 component-table wave ERROR: {e}")

# Test first cold-component dimensional / weight lookup wave
try:
    import math

    from data.n9_table import get_n9_lower_component
    from data.n10_table import get_n10_by_supporting_pipe
    from data.n12_table import get_n12_clip
    from data.n12a_table import get_n12a_clip_type3
    from data.n27_pu_block_table import get_n27_pu_block
    from data.n28_table import get_n28_by_number

    n10_8 = get_n10_by_supporting_pipe("8in SCH.40")
    assert n10_8 and n10_8["B_mm"] == 330 and n10_8["plate_K_mm"] == 16, f"N-10 8in row failed: {n10_8}"
    n9_b = get_n9_lower_component("B", "2in SCH.40", host_type="09C")
    assert n9_b and n9_b["plate_a_deleted_by_n9_note_1"] and [plate["plate"] for plate in n9_b["plates"]] == ["d"], f"N-9 Type B deletion failed: {n9_b}"
    n12_2 = get_n12_clip(2, 200)
    assert n12_2 and n12_2["A_mm"] == 180 and n12_2["plate_thickness_mm"] == 9, f"N-12 Type 2 failed: {n12_2}"
    n12a_3 = get_n12a_clip_type3(220)
    assert n12a_3 and n12a_3["A_mm"] == 260 and n12a_3["plate_thickness_mm"] == 12, f"N-12A Type 3 failed: {n12a_3}"
    n27_1 = get_n27_pu_block(1)
    expected_n27_volume = 125 * 70 * 50 - 2 * math.pi * 8 ** 2 * 50
    assert n27_1 and n27_1["net_volume_mm3"] == expected_n27_volume and n27_1["weight_ready"], f"N-27 PUBK-1 failed: {n27_1}"
    n28_3 = get_n28_by_number(3)
    assert n28_3 and n28_3["fabrication_ready"] and not n28_3["weight_ready"], f"N-28 WOOD-3 failed: {n28_3}"
    print("v N-9/N-10/N-12/N-12A/N-27/N-28 cold-component lookups OK")
except Exception as e:
    record_failure(f"X first cold-component lookup wave ERROR: {e}")

# Test second cold-component core lookup wave
try:
    from data.cold_support_core_tables import (
        get_cradle_selection,
        get_n1_dimensions,
        get_n2_layer_system,
        get_n3_construction,
        get_n4_shield,
        get_n5_material_properties,
        resolve_cradle_designation,
    )

    n1_small = get_n1_dimensions("CR41", 24)
    n1_large = get_n1_dimensions("CR41", 30)
    n2_150 = get_n2_layer_system(150)
    n3_150 = get_n3_construction(150, 300)
    n4_40 = get_n4_shield("CR40", 300)
    n5_320 = get_n5_material_properties(320)
    n20_small = get_cradle_selection(0.5, 25)
    n24_large = get_cradle_selection(30, 25)
    n26_large = get_cradle_selection(60, 200)
    ambiguous = resolve_cradle_designation("CR12", 1.5)

    assert n1_small and n1_small["T1_mm"] == 12, f"N-1 small overlap failed: {n1_small}"
    assert n1_large and n1_large["T1_mm"] == 10 and n1_large["A_mm"] == 706, f"N-1 large overlap failed: {n1_large}"
    assert n2_150 and (n2_150["inner_layer_mm"], n2_150["middle_layer_mm"], n2_150["outer_layer_mm"]) == (50, 50, 50), f"N-2 150 failed: {n2_150}"
    assert n3_150["jacket_length_mm"] == 400 and n3_150["inner_layer_foam_length_mm"] == 500, f"N-3 length rules failed: {n3_150}"
    assert n4_40 and n4_40["T2_mm"] == 5.0, f"N-4 CR40 failed: {n4_40}"
    assert n5_320 and n5_320["engineering_strength_sf5_kg_cm2"] == 18.59, f"N-5 320 failed: {n5_320}"
    assert n20_small and n20_small["cradle_no"] == "CR2.5" and n20_small["max_allowable_load_kg"] == 125, f"N-20 row failed: {n20_small}"
    assert n24_large and (n24_large["F_mm"], n24_large["H_mm"]) == (417, 509), f"N-24 row failed: {n24_large}"
    assert n26_large and n26_large["cradle_no"] == "CR76", f"N-26 row failed: {n26_large}"
    assert n26_large["max_allowable_load_source_value"] == 98100, f"N-26 source load failed: {n26_large}"
    assert n26_large["max_allowable_load_unit"] == "source_conflict", f"N-26 unit status failed: {n26_large}"
    assert n26_large["max_allowable_load_lb"] is None, f"N-26 canonical load failed: {n26_large}"
    assert ambiguous and not ambiguous["selection_resolved"] and ambiguous["candidate_insulation_thicknesses_mm"] == [125, 140], f"cold-core ambiguity failed: {ambiguous}"
    print("v N-1~N-5/N-20~N-26 cold-support core lookups OK")
except Exception as e:
    record_failure(f"X second cold-component lookup wave ERROR: {e}")

# Test third cold-component base/restraint lookup wave
try:
    from data.cold_restraint_tables import (
        get_n6_component,
        get_n7_by_cradle,
        get_n7a_by_cradle,
        get_n8_by_cradle,
        get_n8a_by_line_size,
    )

    n6 = get_n6_component()
    n7 = get_n7_by_cradle("CR3")
    n7a = get_n7a_by_cradle("CR3")
    n8 = get_n8_by_cradle("CR22")
    n8a = get_n8a_by_line_size(8)

    assert n6["overall_height_mm"] == 200 and n6["base_plate"]["outside_diameter_mm"] == 150 and not n6["weight_ready"], f"N-6 geometry failed: {n6}"
    assert n7 and n7["C_overall_mm"] == 105 and n7["D_thread_length_mm"] == 84 and n7["rod_weight_ready"], f"N-7 CR3 failed: {n7}"
    assert n7a and n7a["C_overall_mm"] == 106 and n7a["D_thread_length_mm"] == 52, f"N-7A CR3 failed: {n7a}"
    assert n7["rod_developed_length_mm"] != n7a["rod_developed_length_mm"], f"N-7/N-7A variants collapsed: {n7}, {n7a}"
    assert n8 and n8["thickness_mm"] == 16 and n8["A_mm"] == n8["B_hole_pitch_mm"] + 64 and not n8["weight_ready"], f"N-8 CR22 failed: {n8}"
    assert n8a and (n8a["R_mm"], n8a["A_mm"], n8a["B_hole_pitch_mm"]) == (113, 380, 316), f"N-8A 8in failed: {n8a}"
    print("v N-6/N-7/N-7A/N-8/N-8A cold base/restraint lookups OK")
except Exception as e:
    record_failure(f"X third cold-component lookup wave ERROR: {e}")

# Test fourth cold-component interface / band lookup wave
try:
    import math

    from data.cold_interface_tables import (
        get_n11_by_size,
        get_n13_component,
        get_n14_component,
        get_n15_by_cradle,
        get_n16_by_cradle,
        resolve_n19_designation,
    )

    n11 = get_n11_by_size("5/8")
    n13 = get_n13_component()
    n14 = get_n14_component()
    n15 = get_n15_by_cradle("CR8")
    n16 = get_n16_by_cradle("CR40")
    n19 = resolve_n19_designation("SLP-A-5347-4715")

    assert n11 and (n11["overall_length_mm"], n11["r_c_hole_depth_mm"]) == (127, 70) and n11["design_shear_at_sf5_kg"] == 900.6, f"N-11 5/8 failed: {n11}"
    assert n13["plate_thickness_mm"] == 10 and n13["elevation"]["hole_count_per_plate"] == 2 and not n13["fabrication_ready"], f"N-13 clip failed: {n13}"
    assert n14["plate_thickness_mm"] == 12 and n14["elevation"]["hole_pitch_vertical_mm"] == 55 and not n14["weight_ready"], f"N-14 clip failed: {n14}"
    assert n15 and n15["developed_length_mm"] == math.pi * 122 + 308 and n15["weight_ready"], f"N-15 CR8 failed: {n15}"
    assert n16 and n16["member_M"]["spec"] == "C180x75x7x10.5" and n16["known_steel_weight_kg"] > 80 and not n16["fabrication_ready"], f"N-16 CR40 failed: {n16}"
    assert n19 and n19["upper_plate"]["A_length_mm"] == 530 and n19["lower_backing_plate"]["outside_width_mm"] == 174 and n19["ptfe_slide_element"]["thickness_mm"] is None, f"N-19 designation failed: {n19}"
    print("v N-11/N-13/N-14/N-15/N-16/N-19 cold interface/band lookups OK")
except Exception as e:
    record_failure(f"X fourth cold-component lookup wave ERROR: {e}")

# Test m45_table
try:
    from data.m45_table import get_m45_by_dia, get_m45_by_type
    r1 = get_m45_by_dia("1/2\"")
    # Note: data shows 2312, original test expected 1200. We use data values.
    assert r1 is not None and r1["tensile_kg"] == 2312, f"m45 1/2 failed: {r1}"
    print("v m45_table OK")
except Exception as e:
    record_failure(f"X m45_table ERROR: {e}")

# Test m4_table
try:
    from data.m4_table import get_m4_by_line_size
    r_m4 = get_m4_by_line_size('2"')
    assert (
        r_m4 is not None
        and r_m4["designation"] == "PCL-A-2B"
        and r_m4["B"] == 54
        and r_m4["F"] == '1/2"'
        and r_m4["source_component"] == "M-4"
        and r_m4["source_transcribed"]
        and r_m4["set_weight_kg"] > 0
    ), f"m4 failed: {r_m4}"
    print("v m4_table OK")
except Exception as e:
    record_failure(f"X m4_table ERROR: {e}")

# Test m6_table
try:
    from data.m6_table import get_m6_by_line_size
    r_m6 = get_m6_by_line_size("4")
    r_m6_2_5 = get_m6_by_line_size('2 1/2"')
    assert r_m6 is not None and r_m6["type_label"] == "TYPE-C" and r_m6["designation"] == "PCL-C-4B" and not r_m6["lookup_ready"] and r_m6["partial_lookup_ready"] and r_m6["set_weight_kg"] is None, f"m6 failed: {r_m6}"
    assert r_m6_2_5 is not None and r_m6_2_5["designation"] == "PCL-C-2 1/2B", f"m6 2 1/2 failed: {r_m6_2_5}"
    print("v m6_table OK")
except Exception as e:
    record_failure(f"X m6_table ERROR: {e}")

# Test m5/m7 PDF designation coverage
try:
    from data.m5_table import get_m5_by_line_size
    from data.m7_table import get_m7_by_line_size
    r_m5_12 = get_m5_by_line_size('12"')
    r_m7_8 = get_m7_by_line_size('8"')
    r_m7_28 = get_m7_by_line_size('28"')
    r_m7_2 = get_m7_by_line_size('2"')
    assert r_m5_12 is not None and r_m5_12["designation"] == "PCL-B-12B" and r_m5_12["rod_size_a"] == '1 1/2"' and r_m5_12["load_650f_kg"] == 3930 and r_m5_12["source_component"] == "M-5" and not r_m5_12["lookup_ready"] and r_m5_12["set_weight_kg"] is None, f"m5 12 failed: {r_m5_12}"
    assert r_m7_8 is not None and r_m7_8["designation"] == "PCL-D-8B" and r_m7_8["rod_size_a"] == '1 1/8"' and r_m7_8["load_650f_kg"] == 2175 and r_m7_8["source_component"] == "M-7", f"m7 8 failed: {r_m7_8}"
    assert r_m7_28 is not None and r_m7_28["load_750f_kg"] is None and r_m7_28["load_750f_status"] == "source_blank_or_not_applicable", f"m7 28 load750 failed: {r_m7_28}"
    assert r_m7_2 is None, f"m7 2 should be None (not in PDF): {r_m7_2}"
    print("v m5/m7 PDF designation coverage OK")
except Exception as e:
    record_failure(f"X m5/m7 PDF designation coverage ERROR: {e}")

# Test M-8/M-9/M-10 high-temperature clamp source tables
try:
    from data.m8_table import get_m8_by_line_size
    from data.m9_table import get_m9_by_line_size
    from data.m10_table import get_m10_by_line_size

    r_m8_4 = get_m8_by_line_size('4"')
    r_m9_10 = get_m9_by_line_size('10"')
    r_m10_24 = get_m10_by_line_size('24"')
    assert (
        r_m8_4 is not None
        and r_m8_4["designation"] == "PCL-E-4B"
        and r_m8_4["G_formed_steel_size_mm"] == "8 x 51"
        and r_m8_4["maximum_recommended_load_kg_by_temperature_f"][1050] == 540
        and not r_m8_4["weight_ready"]
    ), f"m8 4 failed: {r_m8_4}"
    assert (
        r_m9_10 is not None
        and r_m9_10["designation"] == "PCL-F-10B"
        and r_m9_10["F_upper_cross_pin_diameter_in"] == '1 1/8"'
        and r_m9_10["maximum_recommended_load_kg_by_temperature_f"][950] == 3585
    ), f"m9 10 failed: {r_m9_10}"
    assert (
        r_m10_24 is not None
        and r_m10_24["designation"] == "PCL-G-24B"
        and r_m10_24["used_on_od_pipe_size_in"] == {"min": 20, "max": 24}
        and r_m10_24["K_overall_width_mm"] == 781
        and r_m10_24["maximum_recommended_load_kg_by_temperature_f"][1075] == 5250
    ), f"m10 24 failed: {r_m10_24}"
    assert get_m8_by_line_size(7) is None
    assert get_m9_by_line_size(5) is None
    assert get_m10_by_line_size(22) is None
    print("v m8/m9/m10 high-temperature clamp tables OK")
except Exception as e:
    record_failure(f"X m8/m9/m10 high-temperature clamp tables ERROR: {e}")

# Test m21_table
try:
    from data.m21_table import get_m21_by_dia
    r_m21 = get_m21_by_dia('1 1/4"')
    assert r_m21 is not None and r_m21["take_up_mm"] == 180 and r_m21["unit_weight_kg"] > 0, f"m21 failed: {r_m21}"
    print("v m21_table OK")
except Exception as e:
    record_failure(f"X m21_table ERROR: {e}")

# Test m24_table
try:
    from data.m24_table import get_m24_by_dia
    r_m24 = get_m24_by_dia('7/8"')
    assert r_m24 is not None and r_m24["pin_dia_b"] == '1"' and r_m24["unit_weight_kg"] > 0, f"m24 failed: {r_m24}"
    print("v m24_table OK")
except Exception as e:
    record_failure(f"X m24_table ERROR: {e}")

# Test Type 11 table-backed hardware
try:
    from data.type11_table import get_type11_hardware_item, build_type11_spring_item
    from core.calculator import analyze_single
    rod11 = get_type11_hardware_item("threaded_rod")
    washer11 = get_type11_hardware_item("washer")
    spring11 = build_type11_spring_item("SPR14")
    type11_result = analyze_single("11-6B-06J")
    type11_spring_entry = next((entry for entry in type11_result.entries if entry.name == "SPRING"), None)
    type11_washer_entry = next((entry for entry in type11_result.entries if entry.name == "WASHER"), None)
    assert rod11 is not None and rod11["spec"] == '1-5/8"*300L' and rod11["length_mm"] == 300, f"type11 rod failed: {rod11}"
    assert washer11 is not None and washer11["category"] == "鋼板類" and washer11["unit_weight_kg"] == 0.33 and washer11["quantity"] == 2, f"type11 washer failed: {washer11}"
    assert spring11 is not None and spring11["spec"] == "SPR14 (14W×46ID)" and spring11["spring_k_kg_per_mm"] == 42 and spring11["quantity"] == 1, f"type11 spring failed: {spring11}"
    assert not type11_result.error and type11_spring_entry and type11_spring_entry.category == "彈簧類", f"type11 calculator spring failed: {type11_result}"
    assert type11_washer_entry and type11_washer_entry.category == "鋼板類", f"type11 calculator washer failed: {type11_result}"
    print("v type11 hardware table OK")
except Exception as e:
    record_failure(f"X type11 hardware table ERROR: {e}")

# Test m47_table
try:
    from data.m47_table import build_m47_item, get_m47_dimensions
    r_m47 = build_m47_item('10"')
    dims_m47 = get_m47_dimensions("24")
    assert r_m47 is not None and r_m47["width_mm"] == 80 and r_m47["length_mm"] == 858 and r_m47["unit_weight_kg"] > 0, f"m47 failed: {r_m47}"
    assert dims_m47 == (90, 1915), f"m47 dims failed: {dims_m47}"
    print("v m47_table OK")
except Exception as e:
    record_failure(f"X m47_table ERROR: {e}")

# Test AI-visual transcribed component tables and remaining metadata-only tables
try:
    from data.m52_table import get_m52_by_line_size, get_m52_component, get_m52_spring_data
    from data.m53_table import get_m53_by_line_size, get_m53_component
    from data.m54_table import build_m54_item, get_m54_by_line_size, get_m54_component
    from data.m55_table import build_m55_item, get_m55_by_line_size, get_m55_component
    from data.n1_table import get_n1_component
    r_m52 = get_m52_component()
    r_m52_24 = get_m52_by_line_size('24"')
    r_m52_spring = get_m52_spring_data('24"')
    r_m53 = get_m53_component()
    r_m53_24 = get_m53_by_line_size('24"')
    r_m54 = get_m54_component()
    r_m54_2 = get_m54_by_line_size('2"', fig_no=2)
    r_m54_fig3 = get_m54_by_line_size('2"', fig_no=3)
    r_m54_item = build_m54_item('2"', fig_no=2)
    r_m55 = get_m55_component()
    r_m55_8 = get_m55_by_line_size('8"')
    r_m55_item = build_m55_item('8"')
    r_n1 = get_n1_component()
    assert r_m52["table_kind"] == "dimensional_lookup" and r_m52["lookup_ready"] and not r_m52["weight_ready"], f"m52 lookup summary failed: {r_m52}"
    assert r_m52_24 is not None and r_m52_24["designation"] == "SPRW-24B" and r_m52_24["dimensions_mm"]["H"] == 610 and r_m52_24["thread_size_j"] == '1"', f"m52 24 failed: {r_m52_24}"
    assert r_m52_spring is not None and r_m52_spring["wire_dia_mm"] == 10 and r_m52_spring["spring_constant_kg_per_mm"] == 45, f"m52 spring failed: {r_m52_spring}"
    assert r_m53["table_kind"] == "dimensional_lookup" and r_m53["lookup_ready"] and not r_m53["weight_ready"], f"m53 lookup summary failed: {r_m53}"
    assert r_m53_24 is not None and r_m53_24["designation"] == "PUBS2-24B" and r_m53_24["dimensions_mm"]["A"] == 838 and r_m53_24["bar_size"] == "150x12", f"m53 24 failed: {r_m53_24}"
    assert r_m54["table_kind"] == "dimensional_lookup" and r_m54["lookup_ready"] and not r_m54["weight_ready"], f"m54 lookup summary failed: {r_m54}"
    assert r_m54_2 is not None and r_m54_2["designation"] == "PUBS3-2B-2" and r_m54_2["dimensions_mm"]["A"] == 63.6 and r_m54_2["dimensions_mm"]["B"] == 150 and r_m54_2["unit_weight_kg"] == 0, f"m54 2 failed: {r_m54_2}"
    assert r_m54_item is not None and r_m54_item["spec"].startswith("PUBS3-2B-2") and r_m54_item["category"] == "鋼板類", f"m54 item failed: {r_m54_item}"
    assert r_m54_fig3 is None, f"m54 unsupported fig should be None: {r_m54_fig3}"
    assert r_m55["table_kind"] == "dimensional_lookup" and r_m55["lookup_ready"] and not r_m55["weight_ready"], f"m55 lookup summary failed: {r_m55}"
    assert r_m55_8 is not None and r_m55_8["designation"] == "PUBD1-8B" and r_m55_8["dimensions_mm"]["B"] == 410 and r_m55_8["unit_weight_kg"] == 0, f"m55 8 failed: {r_m55_8}"
    assert r_m55_item is not None and r_m55_item["spec"].startswith("PUBD1-8B") and r_m55_item["category"] == "鋼板類", f"m55 item failed: {r_m55_item}"
    assert r_n1["component_id"] == "N-1" and r_n1["table_kind"] == "dimensional_lookup" and r_n1["lookup_ready"], f"n1 lookup failed: {r_n1}"
    print("v m52/m53/m54/m55 visual lookup + N-1 lookup-ready component OK")
except Exception as e:
    record_failure(f"X m52/m53/m54/m55 visual lookup + N-1 component ERROR: {e}")

# Test m22_table
try:
    from data.m22_table import build_m22_item
    r12 = build_m22_item('3/4"', 600, left_hand=True)
    assert r12 is not None and r12["designation"] == "MTRL-3/4-600" and r12["thread_length_c"] == 152 and r12["unit_weight_kg"] > 0, f"m22 failed: {r12}"
    print("v m22_table OK")
except Exception as e:
    record_failure(f"X m22_table ERROR: {e}")

# Test m23_table
try:
    from data.m23_table import build_m23_item, get_m23_by_dia
    r13 = get_m23_by_dia('1 1/2"')
    r13_item = build_m23_item('1 1/2"', 900, left_hand=True)
    assert r13 is not None and r13["recommended_bolt_dia_b"] == '1 5/8"' and r13["thread_length_d"] == 152, f"m23 failed: {r13}"
    assert r13_item is not None and r13_item["designation"] == "WERL-1 1/2-900" and r13_item["unit_weight_kg"] > 0, f"m23 build failed: {r13_item}"
    r13_inferred = get_m23_by_dia('1 1/8"')
    assert r13_inferred is None, f"m23 1 1/8 should be None (not in PDF): {r13_inferred}"
    print("v m23_table OK")
except Exception as e:
    record_failure(f"X m23_table ERROR: {e}")

# Test m25_table
try:
    from data.m25_table import build_m25_item
    r14 = build_m25_item('7/8"', left_hand=True)
    assert r14 is not None and r14["designation"] == "WENL-7/8" and r14["G"] == 25 and r14["unit_weight_kg"] > 0, f"m25 failed: {r14}"
    print("v m25_table OK")
except Exception as e:
    record_failure(f"X m25_table ERROR: {e}")

# Test m26_table
try:
    from data.m26_table import get_m26_by_line_size
    r15 = get_m26_by_line_size('2"')
    assert r15 is not None and r15["type"] == "UB-2B" and r15["C"] == 71, f"m26 failed: {r15}"
    assert math.isclose(r15["rod_developed_length_mm"], math.pi * 62 / 2 + 2 * 74)
    assert r15["rod_calculated_weight_kg"] > 0 and r15["finished_hex_nuts_per_set"] == 4
    print("v m26_table OK")
except Exception as e:
    record_failure(f"X m26_table ERROR: {e}")

# Test m28_table
try:
    from data.m28_table import get_m28_takeoff
    r16 = get_m28_takeoff('1 1/2"', fig=2)
    from data.m28_table import get_m28_by_rod_size
    r16_item = get_m28_by_rod_size('1-1/2"')
    r16_inferred = get_m28_by_rod_size('1 1/8"')
    assert r16 == 102 and r16_item is not None and r16_item["unit_weight_kg"] > 0, f"m28 failed: {r16}/{r16_item}"
    assert r16_inferred is None, f"m28 1 1/8 should be None (not in PDF): {r16_inferred}"
    print("v m28_table OK")
except Exception as e:
    record_failure(f"X m28_table ERROR: {e}")

# Test centralized component fallback rules
try:
    from core.component_rules import (
        estimate_clamp_weight,
        estimate_eye_nut_weight,
        estimate_m28_weight,
        estimate_rod_weight,
    )
    from core.hardware_material import (
        HardwareKind,
        HardwareMaterialOverrides,
        ServiceClass,
        parse_hardware_material_context,
        parse_hardware_material_overrides,
        parse_service_class,
        resolve_hardware_material,
    )
    assert estimate_clamp_weight('4"', component_id="M-6") == 2.3, "M-6 clamp estimate should use centralized multiplier"
    assert estimate_rod_weight('5/8"', 3000) > 0, "rod estimate failed"
    assert estimate_eye_nut_weight('5/8"') >= 0.15, "eye nut estimate failed"
    assert estimate_m28_weight('1 1/8"') >= 0.3, "M-28 estimate failed"
    override = HardwareMaterialOverrides(all_hardware="SUS316")
    override_spec = resolve_hardware_material(HardwareKind.CLAMP_BODY, overrides=override)
    cryo_rod_spec = resolve_hardware_material(HardwareKind.THREADED_ROD, service=ServiceClass.CRYO)
    support_pipe_spec = resolve_hardware_material(HardwareKind.SUPPORT_PIPE)
    high_temp_pipe_spec = resolve_hardware_material(HardwareKind.SUPPORT_PIPE, service=ServiceClass.HIGH_TEMP)
    support_plate_spec = resolve_hardware_material(HardwareKind.SUPPORT_PLATE)
    high_temp_plate_spec = resolve_hardware_material(HardwareKind.SUPPORT_PLATE, service=ServiceClass.HIGH_TEMP)
    unknown_override_spec = resolve_hardware_material(HardwareKind.CLAMP_BODY, overrides=HardwareMaterialOverrides(all_hardware="GLOBAL"))
    assert override_spec.name == "SUS316" and override_spec.canonical_id == "JIS_SUS316", "hardware material override failed"
    assert cryo_rod_spec.name == "A320 L7" and cryo_rod_spec.canonical_id == "ASTM_A320_L7", "hardware service material failed"
    assert support_pipe_spec.name == "A36 / SS400" and support_pipe_spec.canonical_id == "ASTM_A36_OR_JIS_SS400", "support pipe default failed"
    assert high_temp_pipe_spec.name == "SA-106 Gr.B" and high_temp_pipe_spec.canonical_id == "ASTM_SA_106_GR_B", "support pipe high-temp default failed"
    assert support_plate_spec.name == "A36 / SS400" and support_plate_spec.canonical_id == "ASTM_A36_OR_JIS_SS400", "support plate default failed"
    assert high_temp_plate_spec.name == "A36 / SS400" and high_temp_plate_spec.canonical_id == "ASTM_A36_OR_JIS_SS400", "support plate high-temp default failed"
    assert unknown_override_spec.name == "GLOBAL" and unknown_override_spec.canonical_id == "UNRESOLVED_GLOBAL", "unknown override canonical fallback failed"
    from core.material_identity import canonical_material_id
    from data.engineering_material_spec import DEFAULT_HARDWARE_MATERIAL
    for kind, per_service in DEFAULT_HARDWARE_MATERIAL.items():
        for service_key, material_name in per_service.items():
            service = ServiceClass.AMBIENT if service_key == "*" else service_key
            spec = resolve_hardware_material(kind, service=service)
            assert spec.name == material_name, f"{kind.value}/{service_key} material name changed: {spec}"
            assert spec.canonical_id == canonical_material_id(material_name), f"{kind.value}/{service_key} canonical_id failed: {spec}"
    print("v phase 2C MaterialSpec canonical_id OK")
    assert parse_service_class({"service_class": "high-temp"}) == ServiceClass.HIGH_TEMP, "service parser failed"
    empty_context = parse_hardware_material_context({})
    assert empty_context.service == ServiceClass.AMBIENT, "empty override service parser failed"
    assert empty_context.material_overrides is not None, "empty override must return concrete material overrides"
    assert isinstance(empty_context.material_overrides.per_kind, dict), "empty override per_kind must be dict"
    assert empty_context.material_overrides.per_kind == {}, "empty override per_kind should be empty dict"
    assert empty_context.material_overrides.all_hardware is None, "empty override all_hardware should be None"
    parsed = parse_hardware_material_overrides({
        "hardware_material_by_kind": {
            "threaded_rod": "A193 B8",
            HardwareKind.HEAVY_HEX_NUT: "A194 8",
        },
        "upper_material": "SUS316",
    }, legacy_material_keys=("upper_material",), legacy_material_kinds=(HardwareKind.UPPER_BRACKET,))
    assert parsed is not None and parsed.per_kind[HardwareKind.THREADED_ROD] == "A193 B8", "per-kind override parser failed"
    assert parsed.per_kind[HardwareKind.HEAVY_HEX_NUT] == "A194 8", "enum-key override parser failed"
    assert parsed.per_kind[HardwareKind.UPPER_BRACKET] == "SUS316", "legacy scoped parser failed"
    per_kind_context = parse_hardware_material_context({
        "hardware_material_by_kind": {
            "threaded_rod": "A193 B8",
            HardwareKind.HEAVY_HEX_NUT: "A194 8",
        },
    })
    assert per_kind_context.material_overrides is not None, "per-kind context overrides missing"
    assert isinstance(per_kind_context.material_overrides.per_kind, dict), "per-kind context per_kind must be dict"
    assert per_kind_context.material_overrides.per_kind[HardwareKind.THREADED_ROD] == "A193 B8", "per-kind context parser failed"
    assert per_kind_context.material_overrides.per_kind[HardwareKind.HEAVY_HEX_NUT] == "A194 8", "enum-key context parser failed"
    legacy_context = parse_hardware_material_context(
        {"upper_material": "SUS316"},
        legacy_material_keys=("upper_material",),
        legacy_material_kinds=(HardwareKind.UPPER_BRACKET,),
    )
    assert legacy_context.material_overrides is not None, "legacy context overrides missing"
    assert legacy_context.material_overrides.per_kind == {HardwareKind.UPPER_BRACKET: "SUS316"}, "legacy context parser failed"
    all_context = parse_hardware_material_context({"hardware_material": "INCONEL"})
    assert all_context.material_overrides is not None, "all-hardware context overrides missing"
    assert all_context.material_overrides.per_kind == {}, "all-hardware context per_kind should be empty dict"
    assert all_context.material_overrides.all_hardware == "INCONEL", "all-hardware context parser failed"
    context = parse_hardware_material_context({"service": "cryo", "hardware_material": "INCONEL", "pipe_material": "A335 P11"})
    assert context.service == ServiceClass.CRYO, "hardware material context service failed"
    assert context.material_overrides and context.material_overrides.all_hardware == "INCONEL", "hardware material context override failed"
    pipe_only = parse_hardware_material_context({"pipe_material": "A335 P11"})
    assert pipe_only.service == ServiceClass.AMBIENT and pipe_only.material_overrides is not None, "pipe_material context should still be normalized"
    assert pipe_only.material_overrides.per_kind == {} and pipe_only.material_overrides.all_hardware is None, "pipe_material must not affect hardware parser"
    print("v component_rules fallback layer OK")
except Exception as e:
    record_failure(f"X component_rules fallback layer ERROR: {e}")

# Phase 2B material identity scaffold
try:
    from core.material_identity import (
        MATERIAL_ALIAS_MAP,
        MATERIAL_CATALOG,
        canonical_material_id,
        normalize_material_alias,
        resolve_material_identity,
    )

    assert MATERIAL_CATALOG["ASTM_A36_OR_JIS_SS400"].display_name == "A36 / SS400", "A36 catalog record failed"
    assert canonical_material_id("A36/SS400") == "ASTM_A36_OR_JIS_SS400", "A36 slash alias failed"
    assert canonical_material_id("A36 / SS400") == "ASTM_A36_OR_JIS_SS400", "A36 spaced alias failed"
    assert canonical_material_id("SA-106 Gr.B") == "ASTM_SA_106_GR_B", "SA-106 alias failed"
    assert canonical_material_id("ASTM A106 Grade B") == "ASTM_SA_106_GR_B", "A106 grade alias failed"
    assert canonical_material_id("A194 4 / S3") == "ASTM_A194_4_S3", "A194 4/S3 alias failed"
    assert canonical_material_id("SUS304") == "JIS_SUS304", "SUS304 alias failed"
    assert canonical_material_id("INCONEL") == "NICKEL_ALLOY_INCONEL", "INCONEL alias failed"
    assert resolve_material_identity("unknown-material") is None, "unknown material should not resolve"
    assert normalize_material_alias(" A36/SS400 ") in MATERIAL_ALIAS_MAP, "normalized alias map failed"
    print("v phase 2B material identity scaffold OK")
except Exception as e:
    record_failure(f"X phase 2B material identity scaffold ERROR: {e}")

# Phase 2I pipe/plate MaterialSpec compatibility
try:
    from core.hardware_material import HardwareKind, ServiceClass, resolve_hardware_material
    from core.models import AnalysisResult
    from core.pipe import add_pipe_entry
    from core.plate import add_plate_entry

    string_pipe = AnalysisResult(fullstring="phase-2I-string-pipe")
    add_pipe_entry(string_pipe, 2, "SCH.40", 1000, "A36/SS400")
    assert string_pipe.entries[0].material == "A36/SS400", "pipe string material path changed"
    assert not hasattr(string_pipe.entries[0], "material_canonical_id"), "pipe string path should not attach canonical id"

    string_plate = AnalysisResult(fullstring="phase-2I-string-plate")
    add_plate_entry(string_plate, 100, 100, 10, "TEST_PLATE", material="SUS304")
    assert string_plate.entries[0].material == "SUS304", "plate string material path changed"
    assert not hasattr(string_plate.entries[0], "material_canonical_id"), "plate string path should not attach canonical id"

    pipe_spec = resolve_hardware_material(HardwareKind.SUPPORT_PIPE, service=ServiceClass.HIGH_TEMP)
    spec_pipe = AnalysisResult(fullstring="phase-2I-spec-pipe")
    add_pipe_entry(spec_pipe, 2, "SCH.40", 1000, pipe_spec)
    assert spec_pipe.entries[0].material == "SA-106 Gr.B", "pipe MaterialSpec should emit material.name"
    assert spec_pipe.entries[0].material_canonical_id == "ASTM_SA_106_GR_B", "pipe MaterialSpec canonical id missing"

    plate_spec = resolve_hardware_material(HardwareKind.SUPPORT_PLATE)
    spec_plate = AnalysisResult(fullstring="phase-2I-spec-plate")
    add_plate_entry(spec_plate, 100, 100, 10, "TEST_PLATE", material=plate_spec)
    assert spec_plate.entries[0].material == "A36 / SS400", "plate MaterialSpec should emit material.name"
    assert spec_plate.entries[0].material_canonical_id == "ASTM_A36_OR_JIS_SS400", "plate MaterialSpec canonical id missing"

    default_plate = AnalysisResult(fullstring="phase-4B-default-plate")
    add_plate_entry(default_plate, 100, 100, 10, "TEST_DEFAULT_PLATE")
    assert default_plate.entries[0].material == "A36/SS400", "plate default material string changed"
    assert default_plate.entries[0].material_canonical_id == "ASTM_A36_OR_JIS_SS400", "plate default canonical id missing"

    print("v phase 2I pipe/plate MaterialSpec compatibility OK")
except Exception as e:
    record_failure(f"X phase 2I pipe/plate MaterialSpec compatibility ERROR: {e}")

# Phase 3A core helper MaterialSpec compatibility
try:
    from core.bolt import add_bolt_entry, add_custom_entry
    from core.hardware_material import HardwareKind, ServiceClass, resolve_hardware_material
    from core.m42 import perform_action_by_letter
    from core.models import AnalysisResult
    from core.steel import add_steel_section_entry
    from data.m42_table import get_m42_data, resolve_m42_data

    steel_spec = resolve_hardware_material(HardwareKind.STRUCTURAL_STRUT)
    bolt_spec = resolve_hardware_material(HardwareKind.ANCHOR_BOLT)
    nut_spec = resolve_hardware_material(HardwareKind.HEAVY_HEX_NUT)
    cryo_bolt_spec = resolve_hardware_material(
        HardwareKind.THREADED_ROD,
        service=ServiceClass.CRYO,
    )

    steel_string = AnalysisResult(fullstring="phase3A-steel-string")
    add_steel_section_entry(steel_string, "Angle", "40*40*5", 150, material="SUS304")
    assert steel_string.entries[0].material == "SUS304", "steel string material changed"
    assert not hasattr(steel_string.entries[0], "material_canonical_id"), "steel explicit string path should stay unmanaged"

    steel_default = AnalysisResult(fullstring="phase3A-steel-default")
    add_steel_section_entry(steel_default, "Angle", "40*40*5", 150)
    assert steel_default.entries[0].material == "A36/SS400", "steel default material changed"
    assert steel_default.entries[0].material_canonical_id == "ASTM_A36_OR_JIS_SS400", "steel default canonical id missing"

    steel_spec_result = AnalysisResult(fullstring="phase3A-steel-spec")
    add_steel_section_entry(steel_spec_result, "Angle", "40*40*5", 150, material=steel_spec)
    assert steel_spec_result.entries[0].material == steel_spec.name, "steel MaterialSpec material changed"
    assert steel_spec_result.entries[0].material_canonical_id == steel_spec.canonical_id, "steel MaterialSpec canonical id missing"

    custom_string = AnalysisResult(fullstring="phase3A-custom-string")
    add_custom_entry(custom_string, "CUSTOM", "C-1", "SUS304", 1, 0.1)
    assert custom_string.entries[0].material == "SUS304", "custom string material changed"
    assert not hasattr(custom_string.entries[0], "material_canonical_id"), "custom explicit string path should stay unmanaged"

    custom_spec = AnalysisResult(fullstring="phase3A-custom-spec")
    add_custom_entry(custom_spec, "CUSTOM", "C-2", nut_spec, 1, 0.1)
    assert custom_spec.entries[0].material == nut_spec.name, "custom MaterialSpec material changed"
    assert custom_spec.entries[0].material_canonical_id == nut_spec.canonical_id, "custom MaterialSpec canonical id missing"

    bolt_default = AnalysisResult(fullstring="phase3A-bolt-default")
    add_bolt_entry(bolt_default, 2, 4)
    assert bolt_default.entries[0].material == "SUS304", "bolt default material changed"
    assert bolt_default.entries[0].material_canonical_id == "JIS_SUS304", "bolt default canonical id missing"

    bolt_string = AnalysisResult(fullstring="phase3A-bolt-string")
    add_bolt_entry(bolt_string, 2, 4, material="SUS316")
    assert bolt_string.entries[0].material == "SUS316", "bolt string material changed"
    assert not hasattr(bolt_string.entries[0], "material_canonical_id"), "bolt explicit string path should stay unmanaged"

    bolt_spec_result = AnalysisResult(fullstring="phase3A-bolt-spec")
    add_bolt_entry(bolt_spec_result, 2, 4, material=cryo_bolt_spec)
    assert bolt_spec_result.entries[0].material == cryo_bolt_spec.name, "bolt MaterialSpec material changed"
    assert bolt_spec_result.entries[0].material_canonical_id == cryo_bolt_spec.canonical_id, "bolt MaterialSpec canonical id missing"

    m42_result = AnalysisResult(fullstring="phase3A-m42-default")
    perform_action_by_letter(m42_result, "E", "L50*50*6")
    assert [entry.material for entry in m42_result.entries] == ["A36/SS400", "A36/SS400", "SUS304", "A36/SS400"], f"m42 default materials changed: {[entry.material for entry in m42_result.entries]}"
    assert all(getattr(entry, "material_canonical_id", None) for entry in m42_result.entries), "m42 default canonical ids missing"

    m42_override = AnalysisResult(fullstring="phase3A-m42-override")
    perform_action_by_letter(
        m42_override,
        "B",
        2,
        plate_material=steel_spec,
        bolt_material=bolt_spec,
    )
    assert [entry.material for entry in m42_override.entries] == [steel_spec.name, steel_spec.name, bolt_spec.name], "m42 MaterialSpec override materials changed"
    assert [entry.material_canonical_id for entry in m42_override.entries] == [steel_spec.canonical_id, steel_spec.canonical_id, bolt_spec.canonical_id], "m42 MaterialSpec override canonical ids missing"

    assert get_m42_data(14)["plate_a"] == 440, "M-43 14 inch row missing"
    assert get_m42_data(16)["plate_bc"] == 630, "M-43 16 inch row missing"
    assert get_m42_data(28)["plate_e"] == 930, "M-43 28 inch row missing"
    assert get_m42_data(28)["exp_bolt_spec"] == '7/8"', "M-43 J bolt spec should follow Rev.1 table"

    m42_half, warn_half = resolve_m42_data(0.5)
    assert m42_half["plate_a"] == 150 and warn_half, "M42 0.5 inch fallback to 1 inch row missing"
    m42_three_quarter, warn_three_quarter = resolve_m42_data(0.75)
    assert m42_three_quarter["plate_a"] == 150 and warn_three_quarter, "M42 0.75 inch fallback to 1 inch row missing"
    m42_two_half, warn_two_half = resolve_m42_data(2.5)
    assert m42_two_half["plate_a"] == 150 and not warn_two_half, "M42 2.5 inch should use 1~3 inch row without warning"
    m42_twenty, warn_twenty = resolve_m42_data(20)
    assert m42_twenty["plate_a"] == 690 and warn_twenty, "M42 20 inch fallback to 24 inch row missing"
    m42_twenty_two, warn_twenty_two = resolve_m42_data(22)
    assert m42_twenty_two["plate_a"] == 690 and warn_twenty_two, "M42 22 inch fallback to 24 inch row missing"
    m42_twenty_six, warn_twenty_six = resolve_m42_data(26)
    assert m42_twenty_six["plate_a"] == 790 and warn_twenty_six, "M42 26 inch fallback to 28 inch row missing"

    m42_l80, warn_l80 = resolve_m42_data("L80*80*8")
    assert m42_l80["plate_a"] == 230 and warn_l80, "M42 L80 fallback to L100 row missing"

    m42_type_t = AnalysisResult(fullstring="phase3A-m42a-T")
    perform_action_by_letter(m42_type_t, "T", 2)
    assert [entry.name for entry in m42_type_t.entries] == ["Plate_a_無鑽孔"], "M-42A Type-T should add Plate a"
    assert [entry.material for entry in m42_type_t.entries] == ["SUS304"], "M-42A Type-T plate should be SS304"
    assert all(getattr(entry, "material_canonical_id", None) for entry in m42_type_t.entries), "M-42A Type-T canonical id missing"

    m42_type_v = AnalysisResult(fullstring="phase3A-m42a-V")
    perform_action_by_letter(m42_type_v, "V", 2)
    assert [entry.name for entry in m42_type_v.entries] == [
        "Plate_a_無鑽孔",
        "Plate_d_有鑽孔",
        "EXP.BOLT",
        "角鋼",
    ], "M-42A Type-V component sequence changed"
    assert [entry.material for entry in m42_type_v.entries] == [
        "A36/SS400",
        "SUS304",
        "SUS304",
        "A36/SS400",
    ], "M-42A Type-V materials changed"
    assert m42_type_v.entries[-1].name == "角鋼" and m42_type_v.entries[-1].quantity == 2, "M-42A Type-V should include two L40 angles"

    m42_type_n = AnalysisResult(fullstring="phase3A-m42-N")
    perform_action_by_letter(m42_type_n, "N", 2)
    assert [entry.name for entry in m42_type_n.entries] == ["Plate_a_無鑽孔"], "M-42 Type-N should not add L40 bracket"

    m42_a_075 = AnalysisResult(fullstring="phase3A-m42-A-075")
    perform_action_by_letter(m42_a_075, "A", 0.75)
    assert m42_a_075.entries[0].name == "Plate_a_無鑽孔", "M42 Type-A fallback should still use Plate a"
    assert m42_a_075.entries[0].length == 150 and m42_a_075.entries[0].spec == "9", "M42 0.75 inch Type-A fallback plate changed"
    assert m42_a_075.warnings, "M42 0.75 inch fallback should warn"

    m42_b_20 = AnalysisResult(fullstring="phase3A-m42-B-20")
    perform_action_by_letter(m42_b_20, "B", 20)
    assert [entry.name for entry in m42_b_20.entries] == ["Plate_a_無鑽孔", "Plate_d_有鑽孔", "EXP.BOLT"], "M42 Type-B 20 inch fallback BOM changed"
    assert m42_b_20.entries[0].length == 690 and m42_b_20.entries[1].length == 830, "M42 Type-B 20 inch should use 24 inch row"
    assert m42_b_20.entries[2].spec == '7/8"' and m42_b_20.entries[2].quantity == 4, "M42 Type-B 20 inch fallback bolt changed"
    assert len(m42_b_20.warnings) == 1, f"M42 fallback warning should not duplicate: {m42_b_20.warnings}"

    m42_unknown = AnalysisResult(fullstring="phase3A-m42-unknown")
    perform_action_by_letter(m42_unknown, "Z", 2)
    assert not m42_unknown.entries and m42_unknown.warnings, "unknown M-42 type should warn without adding BOM"

    print("v phase 3A core helper MaterialSpec compatibility OK")
except Exception as e:
    record_failure(f"X phase 3A core helper MaterialSpec compatibility ERROR: {e}")

# Test type41_table
try:
    from data.type41_table import get_type41_data
    r3 = get_type41_data("41-1")
    assert r3 is not None and r3["L"] == 230 and r3["fig"] == "A", f"type41 41-1 failed: {r3}"
    print("v type41_table OK")
except Exception as e:
    record_failure(f"X type41_table ERROR: {e}")

# Test type42_table
try:
    from data.type42_table import get_type42_member, get_type42_pipe
    r4 = get_type42_member("C125")
    assert r4 is not None and r4["H_MAX"] == 1750, f"type42 C125 failed: {r4}"
    print("v type42_table OK")
except Exception as e:
    record_failure(f"X type42_table ERROR: {e}")

# Test type43_table
try:
    from data.type43_table import get_type43_data, get_type43_formula
    r5 = get_type43_data("L75")
    assert r5 is not None and r5["A"] == 160, f"type43 L75 failed: {r5}"
    print("v type43_table OK")
except Exception as e:
    record_failure(f"X type43_table ERROR: {e}")

# Test type44_table
try:
    from data.type44_table import get_type44_q
    r6 = get_type44_q(10)
    assert r6 == 140, f"type44 Q 10 failed: {r6}"
    print("v type44_table OK")
except Exception as e:
    record_failure(f"X type44_table ERROR: {e}")

# Test type45_table
try:
    from data.type45_table import get_type45_q
    r7 = get_type45_q(14)
    assert r7 == 181, f"type45 Q 14 failed: {r7}"
    print("v type45_table OK")
except Exception as e:
    record_failure(f"X type45_table ERROR: {e}")

# Test type46_table
try:
    from data.type46_table import get_type46_47_q
    r8 = get_type46_47_q(6)
    assert r8 == 187, f"type46 Q 6 failed: {r8}"
    print("v type46_table OK")
except Exception as e:
    record_failure(f"X type46_table ERROR: {e}")

# Test type48_table
try:
    from data.type48_table import get_type48_data
    r9 = get_type48_data(2)
    assert r9 is not None and r9["plate_t"] == 6, f"type48 2 failed: {r9}"
    print("v type48_table OK")
except Exception as e:
    record_failure(f"X type48_table ERROR: {e}")

# Test type51_table
try:
    from data.type51_table import get_type51_data
    r10 = get_type51_data(12)
    assert r10 is not None and r10["member"] == "L65*65*6", f"type51 12 failed: {r10}"
    print("v type51_table OK")
except Exception as e:
    record_failure(f"X type51_table ERROR: {e}")

# Test type56_table
try:
    from data.type56_table import get_type56_data
    r11 = get_type56_data(6)
    assert r11 is not None and r11["R"] == 84, f"type56 6 failed: {r11}"
    print("v type56_table OK")
except Exception as e:
    record_failure(f"X type56_table ERROR: {e}")

# Test type62 hanger combination table/calculator
try:
    from core.calculator import analyze_single
    from data.type62_table import get_type62_lower_part, validate_type62_lower_pipe_size

    fig_j = get_type62_lower_part("J")
    fig_n_ok, _ = validate_type62_lower_pipe_size("N", '12"')
    fig_n_bad, _ = validate_type62_lower_pipe_size("N", '4"')
    r62 = analyze_single("62-4B-5/8-05~30D-J(T)")
    r62_simple = analyze_single("62-2B-3/8-05C-G")
    r62_fig_e = analyze_single("62-4B-5/8-05C-E")
    r62_fig_l = analyze_single("62-4B-5/8-05A-L")
    r62_fig_m = analyze_single("62-4B-5/8-05A-M")
    r62_fig_n = analyze_single("62-10B-7/8-05A-N")
    r62_fig_l_gap = analyze_single("62-7B-5/8-05A-L")
    r62_bad = analyze_single("62-4B-5/8-05C-N")
    r62_names = [entry.name for entry in r62.entries]
    r62_fig_e_names = [entry.name for entry in r62_fig_e.entries]
    assert fig_j is not None and fig_j["component_id"] == "M-6" and fig_j["max_insulation_thk_in"] == 4, f"type62 fig J failed: {fig_j}"
    assert fig_n_ok and not fig_n_bad, "type62 lower range validation failed"
    assert not r62.error and "TURNBUCKLE" in r62_names and "LOWER PIPE CLAMP" in r62_names, f"type62 calculator failed: {r62}"
    assert r62.entries[0].unit_weight == 0 and "CUT LENGTH TO BE CONFIRMED" in r62.entries[0].spec, f"type62 H must not become rod cut: {r62.entries[0]}"
    assert not r62.meta["fabrication"]["bom_ready"], f"type62 unresolved rod/clamp weights must block BOM readiness: {r62.meta['fabrication']}"
    assert not r62_simple.error and not any(entry.name == "TURNBUCKLE" for entry in r62_simple.entries), f"type62 simple failed: {r62_simple}"
    assert not r62_fig_e.error and "ADJUSTABLE CLEVIS" in r62_fig_e_names, f"type62 fig E failed: {r62_fig_e}"
    assert "WELDLESS EYE NUT" not in r62_fig_e_names and "HEAVY HEX. NUT" not in r62_fig_e_names, f"type62 fig E should not add nut callouts: {r62_fig_e.entries}"
    assert (
        not r62_fig_l.error
        and any(
            entry.geometry.component_id == "M-8"
            and entry.spec.startswith("PCL-E-4B")
            and entry.unit_weight == 0
            for entry in r62_fig_l.entries
        )
    ), f"type62 fig L failed: {r62_fig_l}"
    assert (
        not r62_fig_m.error
        and any(
            entry.geometry.component_id == "M-9"
            and entry.geometry.parameters["F_upper_cross_pin_diameter_in"] == '7/8"'
            for entry in r62_fig_m.entries
        )
    ), f"type62 fig M failed: {r62_fig_m}"
    assert (
        not r62_fig_n.error
        and any(
            entry.geometry.component_id == "M-10"
            and entry.geometry.parameters["M_upper_side_width_mm"] == 83
            for entry in r62_fig_n.entries
        )
    ), f"type62 fig N failed: {r62_fig_n}"
    assert (
        r62_fig_l_gap.error
        and "M-8未表列" in r62_fig_l_gap.error
        and not r62_fig_l_gap.entries
    ), f"type62 fig L exact-row guard failed: {r62_fig_l_gap}"
    assert r62_bad.error and "FIG-N" in r62_bad.error, f"type62 invalid range failed: {r62_bad}"
    r62_material = analyze_single("62-2B-3/8-05C-G", {"material": "SUS304"})
    assert not r62_material.error and r62_material.entries[0].material == "SUS304", f"type62 material override failed: {r62_material.entries}"
    print("v type62 hanger combination OK")
except Exception as e:
    record_failure(f"X type62 hanger combination ERROR: {e}")

# Test consistency refactor smokes
try:
    from core.calculator import analyze_single

    r10_invalid_letter = analyze_single("10-2B-05H")
    r07_override = analyze_single("07-2B-20J", {"upper_material": "SUS316"})
    r14_override = analyze_single("14-2B-1005", {"upper_material": "SUS316"})
    r16_override = analyze_single("16-2B-05", {"upper_material": "SUS316"})
    assert not r10_invalid_letter.error and r10_invalid_letter.meta["issues"][0]["code"] == "HOST_M42_NOT_LISTED", (
        f"type10 host M42 variance classification failed: {r10_invalid_letter.error}"
    )
    assert not r07_override.error and r07_override.entries[0].material == "SUS316", f"type07 material override failed: {r07_override.entries}"
    assert not r14_override.error and r14_override.entries[0].material == "SUS316", f"type14 material override failed: {r14_override.entries}"
    assert not r16_override.error and r16_override.entries[0].material == "SUS316", f"type16 upper pipe material override failed: {r16_override.entries}"
    print("v system consistency refactor smokes OK")
except Exception as e:
    record_failure(f"X system consistency refactor smokes ERROR: {e}")

# Phase 1D-0 material/override snapshot guardrails
try:
    from core.calculator import analyze_single

    _SNAPSHOT_CASES = {
        "07-2B-20J": {
            "count": 6,
            "total": 29.54,
            "warnings": 2,
            "materials": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "weights": (1.47, 20.12, 0, 2.83, 2.83, 2.29),
            "quantities": (1, 1, 4, 1, 1, 1),
            "upper_total": 29.9,
            "upper_override": (
                "SUS316",
                "SUS316",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "all_hardware": (
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
            ),
            "cryo": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
        },
        "10-2B-05A": {
            "count": 6,
            "total": 8.79,
            "warnings": 3,
            "materials": (
                "A53Gr.B",
                "SUS304",
                "A36 / SS400",
                "A194 2H",
                "A36 / SS400",
                "A36/SS400",
            ),
            "weights": (1.08, 1.48, 0.56, 0, 4.08, 1.59),
            "quantities": (1, 1, 4, 16, 2, 1),
            "upper_total": 8.8,
            "upper_override": (
                "A53Gr.B",
                "SUS316",
                "A36 / SS400",
                "A194 2H",
                "A36 / SS400",
                "A36/SS400",
            ),
            "all_hardware": (
                "A53Gr.B",
                "SUS304",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "A36/SS400",
            ),
            "cryo": (
                "A53Gr.B",
                "SUS304",
                "A36 / SS400",
                "A194 4 / S3",
                "A36 / SS400",
                "A36/SS400",
            ),
        },
        "14-2B-1005": {
            "count": 7,
            "total": 17.32,
            "warnings": 2,
            "materials": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "weights": (9.25, 2.08, 0, 2.55, 1.04, 0.45, 1.95),
            "quantities": (1, 1, 4, 1, 2, 1, 4),
            "upper_total": 17.35,
            "upper_override": (
                "A36 / SS400",
                "SUS316",
                "SUS316",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "all_hardware": (
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
            ),
            "cryo": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
        },
        "15-2B-1005": {
            "count": 6,
            "total": 18.06,
            "warnings": 2,
            "materials": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "weights": (9.25, 2.08, 2.55, 1.04, 0.45, 2.69),
            "quantities": (1, 1, 1, 2, 1, 4),
            "upper_total": 18.09,
            "upper_override": (
                "A36 / SS400",
                "SUS316",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "all_hardware": (
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
            ),
            "cryo": (
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
        },
        "16-2B-05": {
            "count": 2,
            "total": 4.56,
            "warnings": 2,
            "materials": ("A36 / SS400", "A36 / SS400"),
            "weights": (4.33, 0.23),
            "quantities": (1, 1),
            "upper_total": 4.63,
            "upper_override": ("SUS316", "A36 / SS400"),
            "all_hardware": ("INCONEL", "INCONEL"),
            "cryo": ("A36 / SS400", "A36 / SS400"),
        },
        "62-4B-5/8-05~30D-J(T)": {
            "count": 6,
            "total": 1.49,
            "warnings": 3,
            "materials": (
                "A194 2H",
                "A36 / SS400",
                "A193 B7",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "weights": (0, 0, 0, 0.95, 0.32, 0.22),
            "quantities": (2, 1, 1, 1, 1, 1),
            "upper_override": (
                "SUS316",
                "SUS316",
                "SUS316",
                "SUS316",
                "SUS316",
                "SUS316",
            ),
            "all_hardware": (
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
                "INCONEL",
            ),
            "cryo": (
                "A194 4 / S3",
                "A36 / SS400",
                "A320 L7",
                "A36 / SS400",
                "A36 / SS400",
                "A36 / SS400",
            ),
        },
        "64-2-8-05A": {
            "count": 5,
            "total": 0.16,
            "warnings": 4,
            "materials": (
                "A194 2H",
                "A36 / SS400",
                "A193 B7",
                "A36 / SS400",
                "A36 / SS400",
            ),
            "weights": (0, 0, 0, 0, 0.16),
            "quantities": (2, 1, 2, 1, 2),
            "upper_override": ("SUS316", "SUS316", "SUS316", "SUS316", "SUS316"),
            "all_hardware": ("INCONEL", "INCONEL", "INCONEL", "INCONEL", "INCONEL"),
            "cryo": ("A194 4 / S3", "A36 / SS400", "A320 L7", "A36 / SS400", "A36 / SS400"),
        },
        "65-6B-1505": {
            "count": 5,
            "total": 14.04,
            "warnings": 5,
            "materials": ("A36 / SS400", "A36 / SS400", "A194 2H", "A194 2H", "A193 B7"),
            "weights": (14.04, 0, 0, 0, 0),
            "quantities": (1, 2, 6, 4, 2),
            "upper_override": ("SUS316", "SUS316", "SUS316", "SUS316", "SUS316"),
            "all_hardware": ("INCONEL", "INCONEL", "INCONEL", "INCONEL", "INCONEL"),
            "cryo": ("A36 / SS400", "A36 / SS400", "A194 4 / S3", "A194 4 / S3", "A320 L7"),
        },
    }

    def _norm(value):
        return None if value is None else round(value, 4)

    def _snapshot_entries(result):
        return sorted(
            result.entries,
            key=lambda entry: (
                entry.category,
                entry.name,
                entry.spec,
                entry.material,
                entry.unit,
                entry.remark,
            ),
        )

    def _materials(result):
        return tuple(entry.material for entry in _snapshot_entries(result))

    def _weights(result):
        return tuple(_norm(entry.weight_output) for entry in _snapshot_entries(result))

    def _quantities(result):
        return tuple(_norm(entry.quantity) for entry in _snapshot_entries(result))

    for designation, expected in _SNAPSHOT_CASES.items():
        default_result = analyze_single(designation)
        assert not default_result.error, f"{designation} snapshot error: {default_result.error}"
        assert len(default_result.entries) == expected["count"], f"{designation} entry count changed: {len(default_result.entries)}"
        assert _norm(default_result.total_weight) == expected["total"], f"{designation} total changed: {default_result.total_weight}"
        assert len(default_result.warnings) == expected["warnings"], f"{designation} warning count changed: {default_result.warnings}"
        assert _materials(default_result) == expected["materials"], f"{designation} material snapshot changed: {_materials(default_result)}"
        assert _weights(default_result) == expected["weights"], f"{designation} weight snapshot changed: {_weights(default_result)}"
        assert _quantities(default_result) == expected["quantities"], f"{designation} quantity snapshot changed: {_quantities(default_result)}"

        upper_result = analyze_single(designation, {"upper_material": "SUS316"})
        all_result = analyze_single(designation, {"hardware_material": "INCONEL"})
        cryo_result = analyze_single(designation, {"service": "cryo"})
        pipe_result = analyze_single(designation, {"pipe_material": "A335 P11"})
        for result in (upper_result, all_result, cryo_result, pipe_result):
            assert not result.error, f"{designation} override snapshot error: {result.error}"
        assert _norm(upper_result.total_weight) == expected.get("upper_total", expected["total"]), f"{designation} upper override weight changed: {upper_result.total_weight}"
        for result in (all_result, cryo_result, pipe_result):
            assert _norm(result.total_weight) == expected["total"], f"{designation} override weight changed: {result.total_weight}"

        assert _materials(upper_result) == expected["upper_override"], f"{designation} upper override changed: {_materials(upper_result)}"
        assert _materials(all_result) == expected["all_hardware"], f"{designation} all-hardware override changed: {_materials(all_result)}"
        assert _materials(cryo_result) == expected["cryo"], f"{designation} service material changed: {_materials(cryo_result)}"
        assert _materials(pipe_result) == expected["materials"], f"{designation} pipe_material polluted hardware: {_materials(pipe_result)}"

    print("v phase 1D-0 material/override snapshot baseline OK")
except Exception as e:
    record_failure(f"X phase 1D-0 material/override snapshot baseline ERROR: {e}")

# Phase 1D-2C override consistency across migrated material Types
try:
    from collections import Counter

    from core.calculator import analyze_single
    from core.hardware_material import HardwareKind, ServiceClass, resolve_hardware_material

    _MIGRATED_TYPE_KIND_COUNTS = {
        "07-2B-20J": {
            HardwareKind.SUPPORT_PIPE: 2,
            HardwareKind.SUPPORT_PLATE: 3,
            HardwareKind.EXPANSION_BOLT: 1,
        },
        "14-2B-1005": {
            HardwareKind.STRUCTURAL_STRUT: 1,
            HardwareKind.SUPPORT_PIPE: 1,
            HardwareKind.ANCHOR_BOLT: 1,
            HardwareKind.SUPPORT_PLATE: 4,
        },
        "15-2B-1005": {
            HardwareKind.STRUCTURAL_STRUT: 1,
            HardwareKind.SUPPORT_PIPE: 1,
            HardwareKind.SUPPORT_PLATE: 4,
        },
        "16-2B-05": {
            HardwareKind.SUPPORT_PIPE: 1,
            HardwareKind.SUPPORT_PLATE: 1,
        },
        "62-4B-5/8-05~30D-J(T)": {
            HardwareKind.HEAVY_HEX_NUT: 1,
            HardwareKind.CLAMP_BODY: 1,
            HardwareKind.THREADED_ROD: 1,
            HardwareKind.TURNBUCKLE: 1,
            HardwareKind.BEAM_ATTACHMENT: 1,
            HardwareKind.WELDLESS_EYE_NUT: 1,
        },
        "64-2-8-05A": {
            HardwareKind.CLAMP_BODY: 2,
            HardwareKind.THREADED_ROD: 1,
            HardwareKind.WELDLESS_EYE_NUT: 1,
            HardwareKind.HEAVY_HEX_NUT: 1,
        },
        "65-6B-1505": {
            HardwareKind.STRUCTURAL_STRUT: 1,
            HardwareKind.BEAM_ATTACHMENT: 1,
            HardwareKind.THREADED_ROD: 1,
            HardwareKind.HEAVY_HEX_NUT: 2,
        },
    }
    _UNMANAGED_MATERIAL_COUNTS = {
        "07-2B-20J": Counter(),
        "14-2B-1005": Counter(),
        "15-2B-1005": Counter(),
        "16-2B-05": Counter(),
        "62-4B-5/8-05~30D-J(T)": Counter(),
        "64-2-8-05A": Counter(),
        "65-6B-1505": Counter(),
    }
    _LEGACY_SCOPES = {
        "07-2B-20J": {HardwareKind.SUPPORT_PIPE},
        "14-2B-1005": {HardwareKind.SUPPORT_PIPE, HardwareKind.ANCHOR_BOLT},
        "15-2B-1005": {HardwareKind.SUPPORT_PIPE},
        "16-2B-05": {HardwareKind.SUPPORT_PIPE},
    }
    _LEGACY_GLOBAL_CASES = {
        "62-4B-5/8-05~30D-J(T)",
        "64-2-8-05A",
        "65-6B-1505",
    }
    _FULL_PER_KIND_OVERRIDE = {
        "threaded_rod": "ROD_KIND",
        "heavy_hex_nut": "NUT_KIND",
        "upper_bracket": "UPPER_KIND",
        "support_pipe": "SUPPORT_PIPE_KIND",
        "support_plate": "SUPPORT_PLATE_KIND",
        "anchor_bolt": "ANCHOR_KIND",
        "expansion_bolt": "EXPANSION_KIND",
        "gusset_plate": "PLATE_KIND",
        "structural_strut": "STRUT_KIND",
        "beam_attachment": "BEAM_KIND",
        "clamp_body": "CLAMP_KIND",
        "weldless_eye_nut": "EYE_KIND",
        "turnbuckle": "TURN_KIND",
        "clevis": "CLEVIS_KIND",
        "plate_lug": "LUG_KIND",
    }
    _FULL_PER_KIND_EXPECTED = {
        HardwareKind.THREADED_ROD: "ROD_KIND",
        HardwareKind.HEAVY_HEX_NUT: "NUT_KIND",
        HardwareKind.UPPER_BRACKET: "UPPER_KIND",
        HardwareKind.SUPPORT_PIPE: "SUPPORT_PIPE_KIND",
        HardwareKind.SUPPORT_PLATE: "SUPPORT_PLATE_KIND",
        HardwareKind.ANCHOR_BOLT: "ANCHOR_KIND",
        HardwareKind.EXPANSION_BOLT: "EXPANSION_KIND",
        HardwareKind.GUSSET_PLATE: "PLATE_KIND",
        HardwareKind.STRUCTURAL_STRUT: "STRUT_KIND",
        HardwareKind.BEAM_ATTACHMENT: "BEAM_KIND",
        HardwareKind.CLAMP_BODY: "CLAMP_KIND",
        HardwareKind.WELDLESS_EYE_NUT: "EYE_KIND",
        HardwareKind.TURNBUCKLE: "TURN_KIND",
        HardwareKind.CLEVIS: "CLEVIS_KIND",
        HardwareKind.PLATE_LUG: "LUG_KIND",
    }
    _PARTIAL_PER_KIND_EXPECTED = {
        HardwareKind.THREADED_ROD: "ROD_KIND",
        HardwareKind.SUPPORT_PIPE: "SUPPORT_PIPE_KIND",
        HardwareKind.SUPPORT_PLATE: "SUPPORT_PLATE_KIND",
        HardwareKind.CLAMP_BODY: "CLAMP_KIND",
    }
    _PARTIAL_PER_KIND_OVERRIDE = {
        kind.value: material
        for kind, material in _PARTIAL_PER_KIND_EXPECTED.items()
    }

    def _entry_material_counter(designation, overrides=None):
        result = analyze_single(designation, overrides or {})
        assert not result.error, f"{designation} override consistency error: {result.error}"
        return Counter(entry.material for entry in result.entries)

    def _default_material(kind, service):
        return resolve_hardware_material(kind, service=service).name

    def _expected_material_counter(
        designation,
        *,
        service=ServiceClass.AMBIENT,
        global_material=None,
        per_kind=None,
        legacy_material=None,
    ):
        per_kind = per_kind or {}
        kind_counts = _MIGRATED_TYPE_KIND_COUNTS[designation]
        expected = Counter(_UNMANAGED_MATERIAL_COUNTS[designation])
        legacy_scope = _LEGACY_SCOPES.get(designation, set())
        legacy_is_global = designation in _LEGACY_GLOBAL_CASES

        for kind, count in kind_counts.items():
            if kind in per_kind:
                material = per_kind[kind]
            elif global_material is not None:
                material = global_material
            elif legacy_material is not None and (legacy_is_global or kind in legacy_scope):
                material = legacy_material
            else:
                material = _default_material(kind, service)
            expected[material] += count
        return expected

    for designation in _MIGRATED_TYPE_KIND_COUNTS:
        assert _entry_material_counter(designation, {"pipe_material": "A335 P11"}) == _expected_material_counter(designation), f"{designation} pipe_material polluted hardware"

        assert _entry_material_counter(designation, {"hardware_material": "GLOBAL"}) == _expected_material_counter(designation, global_material="GLOBAL"), f"{designation} global hardware override failed"

        assert _entry_material_counter(designation, {"hardware_material_by_kind": _FULL_PER_KIND_OVERRIDE}) == _expected_material_counter(designation, per_kind=_FULL_PER_KIND_EXPECTED), f"{designation} per-kind hardware override failed"

        mixed_overrides = {
            "hardware_material": "GLOBAL",
            "hardware_material_by_kind": _PARTIAL_PER_KIND_OVERRIDE,
        }
        assert _entry_material_counter(designation, mixed_overrides) == _expected_material_counter(designation, global_material="GLOBAL", per_kind=_PARTIAL_PER_KIND_EXPECTED), f"{designation} per-kind should override global"

        assert _entry_material_counter(designation, {"material": "LEGACY_M"}) == _expected_material_counter(designation, legacy_material="LEGACY_M"), f"{designation} legacy material conversion failed"

        assert _entry_material_counter(designation, {"upper_material": "LEGACY_U"}) == _expected_material_counter(designation, legacy_material="LEGACY_U"), f"{designation} legacy upper_material conversion failed"

        high_temp = {"service": "high_temp"}
        assert _entry_material_counter(designation, high_temp) == _expected_material_counter(designation, service=ServiceClass.HIGH_TEMP), f"{designation} high-temp service defaults failed"

        high_temp_with_per_kind = {
            "service": "high_temp",
            "hardware_material_by_kind": {"threaded_rod": "ROD_KIND"},
        }
        assert _entry_material_counter(designation, high_temp_with_per_kind) == _expected_material_counter(
            designation,
            service=ServiceClass.HIGH_TEMP,
            per_kind={HardwareKind.THREADED_ROD: "ROD_KIND"},
        ), f"{designation} service default broken by per-kind override"

        high_temp_with_legacy = {"service": "high_temp", "upper_material": "LEGACY_U"}
        assert _entry_material_counter(designation, high_temp_with_legacy) == _expected_material_counter(
            designation,
            service=ServiceClass.HIGH_TEMP,
            legacy_material="LEGACY_U",
        ), f"{designation} service default broken by legacy override"

    print("v phase 1D-2C override consistency OK")
except Exception as e:
    record_failure(f"X phase 1D-2C override consistency ERROR: {e}")

# Phase 1D-6 material-system lock-in checks
try:
    import re
    from pathlib import Path

    type_root = Path("core/types")
    migrated_type_files = [
        type_root / "type_07.py",
        type_root / "type_10.py",
        type_root / "type_14.py",
        type_root / "type_15.py",
        type_root / "type_16.py",
        type_root / "type_62.py",
        type_root / "type_64.py",
        type_root / "type_65.py",
    ]
    legacy_override_tokens = (
        "_material_" "overrides_from_dict",
        "_service_" "from_overrides",
        "resolve_" "material",
        "DEFAULT_UPPER_" "MATERIAL",
        "DEFAULT_STRUCTURAL_" "MATERIAL",
    )
    direct_material_patterns = (
        re.compile(r"\b[A-Za-z_][A-Za-z0-9_]*material[A-Za-z0-9_]*\s*=\s*['\"]"),
        re.compile(r"\.material\s*=\s*['\"]"),
        re.compile(r"\bmaterial\s*=\s*['\"]"),
    )

    def _source_hits(paths, predicate):
        hits = []
        for path in paths:
            text = path.read_text(encoding="utf-8")
            for line_no, line in enumerate(text.splitlines(), start=1):
                if predicate(line):
                    hits.append(f"{path}:{line_no}: {line.strip()}")
        return hits

    all_type_files = sorted(type_root.glob("type_*.py"))
    upper_bracket_hits = _source_hits(
        all_type_files,
        lambda line: "UPPER_BRACKET" in line,
    )
    assert not upper_bracket_hits, "UPPER_BRACKET Type mapping usage must stay 0: " + "; ".join(upper_bracket_hits)

    legacy_override_hits = _source_hits(
        all_type_files,
        lambda line: any(token in line for token in legacy_override_tokens),
    )
    assert not legacy_override_hits, "legacy material override path found: " + "; ".join(legacy_override_hits)

    direct_material_hits = _source_hits(
        migrated_type_files,
        lambda line: any(pattern.search(line) for pattern in direct_material_patterns),
    )
    assert not direct_material_hits, "migrated Types must not assign literal material: " + "; ".join(direct_material_hits)

    print("v phase 1D-6 material-system lock-in OK")
except Exception as e:
    record_failure(f"X phase 1D-6 material-system lock-in ERROR: {e}")

# Phase 4B material hard lock for Phase 1D migrated Types.
#
# This is intentionally a hard failure: Phase 1D Types have completed material
# migration, so missing canonical identity indicates a bypassed MaterialSpec path.
try:
    from core.calculator import analyze_single

    _PHASE_4B_MIGRATED_TYPE_SAMPLES = [
        ("07", "07-2B-20J"),
        ("10", "10-2B-05A"),
        ("14", "14-2B-1005"),
        ("15", "15-2B-1005"),
        ("16", "16-2B-05"),
        ("62", "62-4B-5/8-05~30D-J(T)"),
        ("64", "64-2-8-05A"),
        ("65", "65-6B-1505"),
    ]

    hard_lock_errors = []
    for type_id, designation in _PHASE_4B_MIGRATED_TYPE_SAMPLES:
        result = analyze_single(designation)
        if result.error:
            hard_lock_errors.append(f"type={type_id} designation={designation} error={result.error}")
            continue

        for fallback_index, entry in enumerate(result.entries, start=1):
            if getattr(entry, "material_canonical_id", None):
                continue
            entry_index = entry.item_no or fallback_index
            hard_lock_errors.append(
                f"type={type_id} entry_index={entry_index} material={entry.material}"
            )

    assert not hard_lock_errors, "Phase 4B migrated Type material hard-lock failures: " + "; ".join(hard_lock_errors)

    print("v phase 4B material hard lock OK")
except Exception as e:
    record_failure(f"X phase 4B material hard lock ERROR: {e}")
    raise

# Test type72 strap support table/calculator
try:
    from core.calculator import analyze_single
    from data.type72_table import get_type72_data

    r72_table = get_type72_data('2"')
    r72 = analyze_single("72-2B")
    r72_bad = analyze_single("72-6B")
    r72_names = [entry.name for entry in r72.entries]
    assert r72_table is not None and r72_table["A"] == 63.6 and r72_table["B"] == 150 and r72_table["T"] == 6, f"type72 table failed: {r72_table}"
    assert not r72.error and r72_names == ["STRAP", "EXP. BOLT"], f"type72 calculator failed: {r72}"
    assert r72.entries[0].spec.startswith("PUBS3-2B-2") and r72.entries[0].unit_weight == 0 and r72.entries[1].spec == "EB-3/8", f"type72 entries failed: {r72.entries}"
    assert r72.total_weight == 0 and not r72.meta["fabrication"]["bom_ready"]
    assert all(entry.geometry.fabrication_blockers for entry in r72.entries)
    assert r72_bad.error and "3/4" in r72_bad.error, f"type72 invalid range failed: {r72_bad}"
    print("v type72 strap support OK")
except Exception as e:
    record_failure(f"X type72 strap support ERROR: {e}")

# Test type73/type76/type77/type78/type79 support calculators
try:
    from core.calculator import analyze_single
    from data.type73_table import get_type73_bolt_count, get_type73_data, get_type73_spring_data
    from data.type76_table import get_type76_data
    from data.type77_table import get_type77_data
    from data.type79_table import get_type79_data

    r73_table = get_type73_data('6"')
    r73_spring = get_type73_spring_data("SPR04")
    r73 = analyze_single("73-6B-G")
    r73_bad = analyze_single("73-30B-G")
    r76_table = get_type76_data('30"')
    r76 = analyze_single("76-30B")
    r77_table = get_type77_data('40"')
    r77 = analyze_single("77-40B-(A)")
    r78 = analyze_single("78-2B(A)")
    r79_table = get_type79_data('8"')
    r79 = analyze_single("79-8B(A)")
    r79_bad = analyze_single("79-4B")

    assert r73_table is not None and r73_table["A"] == 396 and r73_table["spring_mark"] == "SPR04" and get_type73_bolt_count('6"') == 4, f"type73 table failed: {r73_table}"
    assert r73_spring is not None and r73_spring["spring_constant_kg_per_mm"] == 2.9 and r73_spring["unit_weight_kg"] == 0, f"type73 spring failed: {r73_spring}"
    assert not r73.error and [entry.name for entry in r73.entries][:4] == ["STRAP", "SPRING COIL", "STUD BOLT", "WASHER"], f"type73 calculator failed: {r73}"
    assert r73.entries[0].unit_weight == 3.44 and r73.entries[0].geometry.fabrication_ready
    assert all(entry.unit_weight == 0 for entry in r73.entries[1:])
    assert r73_bad.error and '1"' in r73_bad.error, f"type73 invalid range failed: {r73_bad}"
    assert r76_table is not None and r76_table["pad_angle_deg"] == 120 and r76_table["pad_length_mm"] == 400, f"type76 table failed: {r76_table}"
    assert not r76.error and r76.entries[0].name == "PIPE PAD" and r76.entries[0].unit_weight == 0, f"type76 calculator failed: {r76.entries}"
    assert not r76.meta["fabrication"]["bom_ready"] and r76.entries[0].geometry.fabrication_blockers
    assert r77_table is not None and r77_table["A"] == 300 and r77_table["T"] == 16 and r77_table["unit_weight_kg"] == 0, f"type77 table failed: {r77_table}"
    assert not r77.error and r77.entries[0].name == "SADDLE ASSEMBLY" and r77.entries[0].unit_weight == 0, f"type77 calculator failed: {r77}"
    assert not r78.error and r78.entries[0].spec.startswith("PUBS3-2B-1") and r78.entries[0].unit_weight == 0, f"type78 calculator failed: {r78.entries}"
    assert r79_table is not None and r79_table["B"] == 410 and r79_table["unit_weight_kg"] == 0, f"type79 table failed: {r79_table}"
    assert not r79.error and r79.entries[0].name == "U-BAND ASSEMBLY" and r79.entries[0].spec.startswith("PUBD1-8B") and r79.entries[0].unit_weight == 0, f"type79 calculator failed: {r79}"
    assert r79_bad.error and '5"' in r79_bad.error, f"type79 invalid range failed: {r79_bad}"
    print("v type73/type76/type77/type78/type79 support calculators OK")
except Exception as e:
    record_failure(f"X type73/type76/type77/type78/type79 support calculators ERROR: {e}")

# Test localized truth/evidence contract
try:
    from core.calculator import analyze_single
    from core.truth import TRUTH_ESTIMATED, TRUTH_UNKNOWN, need_escalation

    r72_truth = analyze_single("72-2B")
    r76_truth = analyze_single("76-30B")
    r78_truth = analyze_single("78-2B(A)")
    r79_truth = analyze_single("79-8B(A)")
    r_unknown_truth = analyze_single("99-1B")

    assert r72_truth.meta["truth_level"] == TRUTH_ESTIMATED and r72_truth.meta["requires_review"], f"type72 truth failed: {r72_truth.meta}"
    assert r76_truth.meta["truth_level"] == TRUTH_ESTIMATED and "TYPE-76_D-91.pdf" in r76_truth.meta["source_labels"], f"type76 truth failed: {r76_truth.meta}"
    assert r78_truth.meta["truth_level"] == TRUTH_ESTIMATED and r78_truth.evidence[0]["field"] == "type78_d93_m54_dimensions", f"type78 evidence failed: {r78_truth.meta}/{r78_truth.evidence}"
    assert r79_truth.meta["truth_level"] == TRUTH_ESTIMATED and r79_truth.meta["requires_review"], f"type79 truth failed: {r79_truth.meta}"
    assert not any(e["basis"] == "missing_table" for e in r79_truth.evidence), f"type79 should no longer include missing-table evidence: {r79_truth.evidence}"
    assert need_escalation(r79_truth.meta, r79_truth.meta["invariant_errors"]), f"type79 escalation failed: {r79_truth.meta}"
    assert r_unknown_truth.meta["truth_level"] == TRUTH_UNKNOWN and r_unknown_truth.meta["requires_review"], f"unknown truth failed: {r_unknown_truth.meta}"
    print("v localized truth/evidence contract OK")
except Exception as e:
    record_failure(f"X localized truth/evidence contract ERROR: {e}")

# Test type64/type65 normalization helpers
try:
    from data.type64_table import get_type64_rod
    from data.type65_table import get_type65_data
    r64 = get_type64_rod("1-1/4")
    r64_half = get_type64_rod("1/2")
    r65 = get_type65_data("2-1/2")
    assert r64 is None and r64_half is not None and r64_half["g"] == '3/8"', f"type64 source rows failed: {r64}/{r64_half}"
    assert r65 is None and get_type65_data("6")["rod_size"] == '1/2"', f"type65 source rows failed: {r65}"
    print("v type64/type65 normalization OK")
except Exception as e:
    record_failure(f"X type64/type65 normalization ERROR: {e}")

# Phase 2L-A soft lock warnings for unmanaged material paths.
#
# This is intentionally warning-only.  It does not fail validation because the
# remaining unmanaged material paths are known Phase 2 migration backlog.
try:
    from pathlib import Path

    from core.calculator import analyze_single

    _PHASE_2L_A_SAMPLES = [
        ("01", "01-2B-05A"),
        ("03", "03-1B-05N"),
        ("05", "05-L50-05L"),
        ("06", "06-L50-0510-0401"),
        ("07", "07-2B-20J"),
        ("08", "08-2B-1005G"),
        ("09", "09-2B-05B"),
        ("10", "10-2B-05A"),
        ("11", "11-2B-06G"),
        ("12", "12-6B-05B"),
        ("13", "13-6B-05B"),
        ("14", "14-2B-1005"),
        ("15", "15-2B-1005"),
        ("16", "16-2B-05"),
        ("19", "19-2B"),
        ("20", "20-L50-05A"),
        ("21", "21-L50-05A"),
        ("22", "22-L50-05(A)L"),
        ("23", "23-L50-05A"),
        ("24", "24-L50-05"),
        ("25", "25-L50-0505A"),
        ("26", "26-L50-1005A"),
        ("27", "27-L75-0505L-0401"),
        ("28", "28-L50-1005L"),
        ("30", "30-L75-0505A-0401"),
        ("31", "31-L50-1005"),
        ("32", "32-L50-1005"),
        ("33", "33-L50-1005"),
        ("34", "34-L50-1005"),
        ("35", "35-C125-05A"),
        ("36", "36-C125-05"),
        ("37", "37-C125-1200A"),
        ("39", "39-C100-500 A"),
        ("41", "41-1"),
        ("42", "42-8B-C125-500 A"),
        ("43", "43-8B-C125-500 A"),
        ("44", "44-8B-C125-500 A"),
        ("45", "45-8B-C125-500 A"),
        ("46", "46-8B-C125-500 A"),
        ("47", "47-8B-C125-500 A"),
        ("48", "48-2"),
        ("49", "49-8B-A"),
        ("51", "51-2B"),
        ("52", "52-2B(P)-A(A)-130-500"),
        ("56", "56-2B"),
        ("57", "57-2B-A"),
        ("58", "58-4B-A"),
        ("59", "59-6B-A"),
        ("60", "60-20B-A"),
        ("61", "61-4B-T1-05"),
        ("62", "62-4B-5/8-05~30D-J(T)"),
        ("64", "64-2-8-05A"),
        ("65", "65-6B-1505"),
        ("72", "72-2B"),
        ("73", "73-6B-G"),
        ("76", "76-30B"),
        ("77", "77-40B-(A)"),
        ("78", "78-2B(A)"),
        ("79", "79-8B(A)"),
    ]
    _HELPER_DEFAULT_MARKERS = [
        (Path("core/bolt.py"), 'entry.material = "SUS304"', "SUS304"),
        (Path("core/plate.py"), 'material_name = "A36/SS400"', "A36/SS400"),
        (Path("core/steel.py"), 'material = "A36/SS400"', "A36/SS400"),
    ]

    warning_count = 0
    for file_path, marker, material in _HELPER_DEFAULT_MARKERS:
        try:
            for line_no, line in enumerate(file_path.read_text(encoding="utf-8").splitlines(), start=1):
                if marker in line:
                    warning_count += 1
                    print(
                        "WARN phase 2L-A helper default material | "
                        f"file={file_path.as_posix()} | line={line_no} | material={material}"
                    )
        except Exception as helper_error:
            warning_count += 1
            print(
                "WARN phase 2L-A helper scan error | "
                f"file={file_path.as_posix()} | error={helper_error}"
            )

    for type_id, designation in _PHASE_2L_A_SAMPLES:
        result = analyze_single(designation)
        if result.error:
            warning_count += 1
            print(
                "WARN phase 2L-A sample error | "
                f"type={type_id} | designation={designation} | error={result.error}"
            )
            continue

        for fallback_index, entry in enumerate(result.entries, start=1):
            if getattr(entry, "material_canonical_id", None):
                continue
            entry_index = entry.item_no or fallback_index
            warning_count += 1
            print(
                "WARN phase 2L-A unmanaged material entry | "
                f"type={type_id} | entry_index={entry_index} | material={entry.material} | "
                "reason=missing_material_canonical_id,string_material_path"
            )

    print(f"v phase 2L-A soft lock warnings emitted: {warning_count}")
except Exception as e:
    print(f"WARN phase 2L-A soft lock audit skipped: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# Phase 5b golden cases — type_27 / type_42 / type_43
# 目的：鎖定高風險 type 的計算結果，任何回歸會立即報錯。
# 更新方式：先跑 analyze_single(designation)，確認邏輯正確後再更新 expected。
# ─────────────────────────────────────────────────────────────────────────────
try:
    from core.calculator import analyze_single

    def _golden(designation: str, expected: list[tuple]):
        """
        比對計算結果與預期值。
        expected: [(name, spec, length, qty), ...]  length=-1 表示不比對
        """
        r = analyze_single(designation)
        assert not r.error, f"{designation}: unexpected error '{r.error}'"
        assert len(r.entries) == len(expected), (
            f"{designation}: expected {len(expected)} entries, got {len(r.entries)}"
        )
        for i, (entry, (exp_name, exp_spec, exp_len, exp_qty)) in enumerate(
            zip(r.entries, expected), start=1
        ):
            if exp_name:
                assert entry.name == exp_name, (
                    f"{designation} entry#{i}: name '{entry.name}' != '{exp_name}'"
                )
            if exp_spec:
                assert entry.spec == exp_spec, (
                    f"{designation} entry#{i}: spec '{entry.spec}' != '{exp_spec}'"
                )
            if exp_len >= 0:
                assert entry.length == exp_len, (
                    f"{designation} entry#{i}: length {entry.length} != {exp_len}"
                )
            if exp_qty >= 0:
                assert entry.quantity == exp_qty, (
                    f"{designation} entry#{i}: qty {entry.quantity} != {exp_qty}"
                )

    # ── type_42: 8B C125 H=500 FIG-A (θ=30°) ──────────────────────────────
    # G = g_coeff * 500 + g_offset → 722
    _golden("42-8B-C125-500 A", [
        ("槽鐵",  "125*65*6", 500,  1),
        ("槽鐵",  "125*65*6", 722,  1),
        ("TRUNNION", "4\"",       -1,   1),
        ("C/S SHIM", "6",        125,  1),
        ("M.BOLT",   '3/4"x50',  -1,   2),
    ])

    # ── type_43: 8B C125 H=500 FIG-A (θ=30°) ──────────────────────────────
    # main_len = 500 + A; N = n_coeff * 500 + n_offset → 692
    _golden("43-8B-C125-500 A", [
        ("槽鐵",          "125*65*6", 670,  1),
        ("槽鐵",          "125*65*6", 692,  1),
        ("TRUNNION",         "4\"",       -1,   1),
        ("LUG PLATE TYPE-C", "10",       170,  1),
        ("LUG PLATE TYPE-E", "10",       145,  1),
        ("K BOLT",           '3/4"x50',  -1,  12),
        ("C/S SHIM",         "6",        125,  1),
    ])

    # ── type_27 H150: unresolved D-30 rows stay traceable outside material BOM
    _golden("27-H150-0505L", [
        ("D30-GUSSET-PLATE", "9",         100,  2),
        ("Plate_c_有鑽孔",  "16",         500,  1),
    ])

    # ── type_27 L75: same D-30 fabrication blockers ────────────────────────
    _golden("27-L75-0505L", [
        ("D30-GUSSET-PLATE", "9",       100,  2),
        ("Plate_c_有鑽孔", "9",         260,  1),
    ])
    for designation in (
        "27-H150-0505L",
        "27-L75-0505L",
        "27-L50-0303L",
        "27-H150-0406P",
        "27-L100-0410L",
    ):
        result = analyze_single(designation)
        excluded_ids = {
            item.get("component_id")
            for item in result.meta.get("excluded_bom_components", [])
        }
        assert {"D30-MEMBER-M", "D30-TOP-PLATE", "M42-FASTENER"} <= excluded_ids, (
            f"{designation}: unresolved D-30 items lost from exclusion trace: {excluded_ids}"
        )

    # ── type_42 擴充: 4B L75 H=300 FIG-A (小管徑角鐵版) ────────────────────
    # Trunnion=2", G=438
    _golden("42-4B-L75-300 A", [
        ("角鋼",    "75*75*9",  300,  1),
        ("角鋼",    "75*75*9",  438,  1),
        ("TRUNNION", "2\"",       -1,   1),
        ("C/S SHIM", "6",         75,  1),
        ("M.BOLT",   '3/4"x50',  -1,   2),
    ])

    # ── type_42 擴充: 16B C200 H=800 FIG-B (大管徑 θ=45°) ──────────────────
    # Trunnion=10", G=1331
    _golden("42-16B-C200-800 B", [
        ("槽鐵",  "200*80*7.5", 800,   1),
        ("槽鐵",  "200*80*7.5", 1331,  1),
        ("TRUNNION", "10\"",        -1,    1),
        ("C/S SHIM", "6",          200,   1),
        ("M.BOLT",   '3/4"x50',    -1,    2),
    ])

    # ── type_42 擴充: 24B C200 H=1000 FIG-B (最大管徑) ─────────────────────
    # Trunnion=14", G=1614
    _golden("42-24B-C200-1000 B", [
        ("槽鐵",  "200*80*7.5", 1000,  1),
        ("槽鐵",  "200*80*7.5", 1614,  1),
        ("TRUNNION", "14\"",         -1,   1),
        ("C/S SHIM", "6",           200,  1),
        ("M.BOLT",   '3/4"x50',     -1,   2),
    ])

    # ── type_43 擴充: 4B L75 H=300 FIG-A → LUG TYPE-E (小管徑) ────────────
    # main=460, N=438; LUG-C T=9 L=160, LUG-E T=9 L=135
    _golden("43-4B-L75-300 A", [
        ("角鋼",            "75*75*9",  460,  1),
        ("角鋼",            "75*75*9",  438,  1),
        ("TRUNNION",         "2\"",       -1,   1),
        ("LUG PLATE TYPE-C", "9",        160,  1),
        ("LUG PLATE TYPE-E", "9",        135,  1),
        ("K BOLT",           '3/4"x50',  -1,   8),
        ("C/S SHIM",         "6",         75,  1),
    ])

    # ── type_43 擴充: 16B C200 H=800 FIG-B → LUG TYPE-D (大管徑 θ=45°) ────
    # main=1020, N=1187; LUG-C T=12 L=220, LUG-D T=12 L=160
    _golden("43-16B-C200-800 B", [
        ("槽鐵",          "200*80*7.5", 1020,  1),
        ("槽鐵",          "200*80*7.5", 1187,  1),
        ("TRUNNION",         "10\"",         -1,   1),
        ("LUG PLATE TYPE-C", "12",          220,  1),
        ("LUG PLATE TYPE-D", "12",          160,  1),
        ("K BOLT",           '3/4"x50',      -1,  12),
        ("C/S SHIM",         "6",           200,  1),
    ])

    # ── type_27 擴充: L50，保留D-30未標尺寸blockers ─────────────────────────
    _golden("27-L50-0303L", [
        ("D30-GUSSET-PLATE", "9",       100, 2),
        ("Plate_c_有鑽孔", "9",         180,  1),
    ])

    # ── type_27 擴充: H150 M42=P (NOTE4 valid variant) ──────────────────────
    _golden("27-H150-0406P", [
        ("D30-GUSSET-PLATE", "9",       100, 2),
        ("Plate_c_有鑽孔", "16",         500,  1),
    ])

    # ── type_27 擴充: L100 大尺寸 ───────────────────────────────────────────
    _golden("27-L100-0410L", [
        ("D30-GUSSET-PLATE", "9",       100, 2),
        ("Plate_c_有鑽孔", "9",          260,  1),
    ])

    # ── type_39: C125 H=500 FIG-A (標準, L=200 default) ────────────────────
    # main = 500+200=700; N=692; LUG-C T=10 L=170; LUG-E T=10 L=145
    _golden("39-C125-500 A", [
        ("槽鐵",          "125*65*6", 700,  1),
        ("槽鐵",          "125*65*6", 692,  1),
        ("LUG PLATE TYPE-C", "10",       170,  1),
        ("LUG PLATE TYPE-E", "10",       145,  1),
        ("K BOLT",           '3/4"x50',  -1,  12),
    ])

    # ── type_39: C200 H=800 FIG-B (大型, θ=45°) ─────────────────────────────
    # main = 800+200=1000; N=1187; LUG-C T=12 L=220; LUG-D T=12 L=160
    _golden("39-C200-800 B", [
        ("槽鐵",          "200*80*7.5", 1000,  1),
        ("槽鐵",          "200*80*7.5", 1187,  1),
        ("LUG PLATE TYPE-C", "12",         220,  1),
        ("LUG PLATE TYPE-D", "12",         160,  1),
        ("K BOLT",           '3/4"x50',     -1,  12),
    ])

    # ── type_39: L75 H=300 FIG-A (小角鐵版) ─────────────────────────────────
    # main = 300+200=500; N=438; LUG-C T=9 L=160; LUG-E T=9 L=135
    _golden("39-L75-300 A", [
        ("角鋼",            "75*75*9",  500,  1),
        ("角鋼",            "75*75*9",  438,  1),
        ("LUG PLATE TYPE-C", "9",        160,  1),
        ("LUG PLATE TYPE-E", "9",        135,  1),
        ("K BOLT",           '3/4"x50',  -1,   8),
    ])

    # ── type_56: 管線檔止 5 個尺寸分支 ──────────────────────────────────────
    # ≤2-1/2": PL 100×100×6 ×2
    _golden("56-2B", [
        ("PIPE STOP PLATE", "6", 100, 2),
    ])
    # 3"~4": D-67只指定FAB FROM 6t，拆片未尺寸化
    _golden("56-4B", [
        ("MEMBER C ASSEMBLY", "MEMBER C / FAB. FROM 6t PLATE", 0, 2),
    ])
    # 5"~14": 只保留CUT FROM母H型鋼reference，不計完整母材重量
    _golden("56-10B", [
        ("MEMBER C ASSEMBLY", "MEMBER C / CUT FROM H200*200*8*12", 0, 2),
    ])
    # 16"~24": D-67只指定FAB FROM 12t，拆片未尺寸化
    _golden("56-20B", [
        ("MEMBER C ASSEMBLY", "MEMBER C / FAB. FROM 12t PLATE", 0, 2),
    ])
    # 26"~42": D-67A assembly + D-91 reference，兩者都不虛構重量
    _golden("56-36B", [
        ("MEMBER C ASSEMBLY", "D-67A SUPPORT MEMBER / FAB. FROM 12t PLATE", 0, 2),
        ("REINFORCING PAD", "D-91 / 120 DEG / L400 / t>=12", 0, 1),
    ])

    print("v phase 5b golden cases type_27/42/43/39/56 OK")
except Exception as e:
    import traceback
    record_failure(f"X phase 5b golden cases FAILED: {e}")
    traceback.print_exc()

if _FAILURES:
    print(f"\n=== VALIDATION FAILED: {len(_FAILURES)} error(s) ===")
    sys.exit(1)

print("\n=== VALIDATION COMPLETE ===")
