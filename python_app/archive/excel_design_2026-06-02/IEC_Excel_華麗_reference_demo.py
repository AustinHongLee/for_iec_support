"""華麗版 dashboard mockup — proves the openpyxl ceiling. Real numbers from analysis_result.xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, GradientFill, Border, Side, Alignment
from openpyxl.chart import DoughnutChart, BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import IconSetRule, DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter

INK="1F3864"; INK2="2E5395"; ROYAL="3A5BA0"; ACCENT="BF8F00"; GOLD2="E8B923"
OK="70AD47"; OKF="E2EFDA"; BAD="C00000"; BADF="FCE4D6"; WARN="ED7D31"; WARNF="FFF2CC"
CANVAS="F2F4F8"; ZEBRA="F7F9FC"; MUTE="595959"; CARD="FFFFFF"; CARDB="D0D7E2"
CJK="Microsoft JhengHei"

wb=openpyxl.Workbook(); ws=wb.active; ws.title="專案摘要 (華麗版)"
ws.sheet_view.showGridLines=False
ws.sheet_properties.tabColor=INK
widths=[2.5,15,9,5,15,9,5,15,9,5,15,9,5,2.5]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

def mrg(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
def C(r,c): return ws.cell(row=r,column=c)

# ============ R1-R3 GRADIENT MASTHEAD ============
mrg(1,1,2,14)
m=C(1,1); m.value="IEC 管架支撐 ·  專案材料統計總覽"
m.fill=GradientFill(degree=30, stop=[INK, INK2, ROYAL])
m.font=Font(name=CJK,size=22,bold=True,color="FFFFFF")
m.alignment=Alignment(horizontal="left",vertical="center",indent=2)
ws.row_dimensions[1].height=30; ws.row_dimensions[2].height=22
# right-aligned meta sits in same band via a second overlay row isn't possible on merge;
# put meta on R3 sub-band
mrg(3,1,3,14)
s=C(3,1); s.value="製表 2026-06-01     全案總重 5,360.00 kg     支撐 1,045 組     材料 74 項     資料狀態 87 項待確認"
s.fill=PatternFill("solid",fgColor=CANVAS)
s.font=Font(name=CJK,size=10,italic=True,color=MUTE)
s.alignment=Alignment(horizontal="right",vertical="center",indent=2)
ws.row_dimensions[3].height=20
# thin amber accent rule
mrg(4,1,4,14); a=C(4,1); a.fill=PatternFill("solid",fgColor=ACCENT); ws.row_dimensions[4].height=4

def section(r,text):
    mrg(r,2,r,13); c=C(r,2); c.value=f"▌  {text}"
    c.font=Font(name=CJK,size=12,bold=True,color=INK)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    thick=Side(style="thick",color=ACCENT)
    for col in range(2,14): C(r,col).border=Border(bottom=thick)
    ws.row_dimensions[r].height=24

# ============ KPI CARDS (accent-bordered, glyph) ============
section(6,"關鍵指標")
def kpi(r,c,glyph,label,value,unit,accent_hex,note,big_color=INK):
    thick=Side(style="thick",color=accent_hex); thin=Side(style="thin",color=CARDB)
    # left accent bar
    C(r,c).border=Border(left=thick,top=thin); C(r+1,c).border=Border(left=thick)
    C(r+2,c).border=Border(left=thick,bottom=thin)
    for rr in range(r,r+3):
        for cc in range(c,c+3):
            cell=C(rr,cc); cell.fill=PatternFill("solid",fgColor=CARD)
            b=cell.border
            top=Side(style="thin",color=CARDB) if rr==r else b.top
            bot=Side(style="thin",color=CARDB) if rr==r+2 else b.bottom
            rgt=Side(style="thin",color=CARDB) if cc==c+2 else b.right
            lft=Side(style="thick",color=accent_hex) if cc==c else b.left
            cell.border=Border(left=lft,right=rgt,top=top,bottom=bot)
    # label + glyph
    mrg(r,c,r,c+2); lc=C(r,c); lc.value=f"{glyph}  {label}"
    lc.font=Font(name=CJK,size=10,color=MUTE); lc.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    # value
    mrg(r+1,c,r+1,c+1); vc=C(r+1,c); vc.value=value
    vc.font=Font(name="Calibri",size=24,bold=True,color=big_color)
    vc.alignment=Alignment(horizontal="right",vertical="center"); vc.number_format="#,##0.##"
    uc=C(r+1,c+2); uc.value=unit; uc.font=Font(name=CJK,size=10,color=MUTE)
    uc.alignment=Alignment(horizontal="left",vertical="bottom",indent=1)
    # note
    mrg(r+2,c,r+2,c+2); nc=C(r+2,c); nc.value=note
    nc.font=Font(name=CJK,size=9,italic=True,color="808080"); nc.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[r].height=18; ws.row_dimensions[r+1].height=34; ws.row_dimensions[r+2].height=16

kpi(7,2,"◆","全案總重",5360.00,"kg",ACCENT,"全案累計總重",big_color=ACCENT)
kpi(7,5,"●","支撐組數",1045,"組",INK2,"本批設計支撐總數")
kpi(7,8,"▲","平均使用率",0.575,"",OK,"原料切割利用率",big_color=OK)
C(8,8).number_format="0.0%"
kpi(7,11,"⚠","需確認項目",87,"項",BAD,"請見支撐統計明細",big_color=BAD)

# ============ CHARTS: doughnut + bar (native) ============
section(11,"材料重量分佈 ＆ 重型支撐 Top 5")
# chart data (place off-screen to the right, cols 16+)
mat=[("角鋼",994.1),("H型鋼",890.7),("槽鐵 (A)",773.5),("槽鐵 (B)",495.8),
     ("角鋼 (C)",341.4),("其他",5360-994.1-890.7-773.5-495.8-341.4)]
C(30,16).value="材料"; C(30,17).value="總重"
for i,(n,v) in enumerate(mat,1):
    C(30+i,16).value=n; C(30+i,17).value=round(v,1)
dough=DoughnutChart(); dough.title="材料重量分佈 (kg)"; dough.holeSize=55
labels=Reference(ws,min_col=16,min_row=31,max_row=30+len(mat))
data=Reference(ws,min_col=17,min_row=30,max_row=30+len(mat))
dough.add_data(data,titles_from_data=True); dough.set_categories(labels)
dough.dataLabels=DataLabelList(); dough.dataLabels.showPercent=True
dough.height=6.6; dough.width=10.5
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
palette=[INK,INK2,ROYAL,ACCENT,GOLD2,"BFBFBF"]
sset=dough.series[0]
from openpyxl.chart.marker import DataPoint as DP
for idx,col in enumerate(palette):
    dp=DP(idx=idx); dp.graphicalProperties.solidFill=col; sset.data_points.append(dp)
ws.add_chart(dough,"B12")

# Top5 bar
sup=[("15-8B-1532",210.98),("H150-1800",133.91),("2-C150-161",85.68),
     ("2-C150-150",138.04),("C150-19C",133.28)]
C(40,16).value="型號"; C(40,17).value="總重"
for i,(n,v) in enumerate(sup,1):
    C(40+i,16).value=n; C(40+i,17).value=v
bar=BarChart(); bar.type="bar"; bar.title="重型支撐 Top 5 (kg)"
bd=Reference(ws,min_col=17,min_row=40,max_row=40+len(sup))
bl=Reference(ws,min_col=16,min_row=41,max_row=40+len(sup))
bar.add_data(bd,titles_from_data=True); bar.set_categories(bl)
bar.dataLabels=DataLabelList(); bar.dataLabels.showVal=True
bar.legend=None; bar.height=6.6; bar.width=11
bar.series[0].graphicalProperties.solidFill=ACCENT
ws.add_chart(bar,"H12")

# ============ icon-set + databar status mini table ============
section(26,"下料使用率健康度 (Icon Set + Data Bar 範例)")
hdr=["材料","原料#","使用率","餘料(mm)","狀態"]
for c,h in enumerate(hdr,2):
    cell=C(27,c); cell.value=h; cell.fill=PatternFill("solid",fgColor=INK2)
    cell.font=Font(name=CJK,size=10,bold=True,color="FFFFFF")
    cell.alignment=Alignment(horizontal="center",vertical="center")
    cell.border=Border(bottom=Side(style="medium",color=INK))
sample=[("角鋼 L65x6","#1",0.94,42),("H型鋼 H150","#3",0.71,820),
        ("槽鐵 C150","#5",0.45,3210),("角鋼 L50x4","#2",0.22,5400)]
r=28
for name,bid,util,rem in sample:
    C(r,2).value=name; C(r,3).value=bid; C(r,4).value=util; C(r,5).value=rem
    C(r,4).number_format="0.0%"; C(r,5).number_format="#,##0"
    C(r,4).alignment=Alignment(horizontal="center",vertical="center")
    C(r,3).alignment=Alignment(horizontal="center",vertical="center")
    C(r,5).alignment=Alignment(horizontal="right",vertical="center",indent=1)
    C(r,2).alignment=Alignment(horizontal="left",vertical="center",indent=1)
    st="健康" if util>=0.7 else "偏低" if util>=0.4 else "浪費"
    stf=OKF if util>=0.7 else WARNF if util>=0.4 else BADF
    sc=C(r,6); sc.value=st; sc.fill=PatternFill("solid",fgColor=stf)
    sc.alignment=Alignment(horizontal="center",vertical="center"); sc.font=Font(name=CJK,size=10,bold=True)
    if r%2==0:
        for c in range(2,6): C(r,c).fill=PatternFill("solid",fgColor=ZEBRA)
    for c in range(2,7): C(r,c).border=Border(bottom=Side(style="thin",color="D9D9D9"))
    ws.row_dimensions[r].height=20
    r+=1
# icon set on 使用率, data bar on 餘料
ws.conditional_formatting.add(f"D28:D{r-1}", IconSetRule('3TrafficLights1','percent',[0,40,70],showValue=True))
ws.conditional_formatting.add(f"E28:E{r-1}", DataBarRule(start_type='min',end_type='max',color=ROYAL,showValue=True))

ws.print_area=f"A1:N{r}"
ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_margins=openpyxl.worksheet.page.PageMargins(left=0.2,right=0.2,top=0.3,bottom=0.3,header=0.1,footer=0.1)
wb.save("haoli_demo.xlsx")
print("saved haoli_demo.xlsx")
