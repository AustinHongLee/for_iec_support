"""
PyQt6 主視窗
IEC 管架支撐分析工具

三面板佈局:
  左: 輸入清單 (QListWidget, 可勾選/點選)
  中: 結果表格
  右: Side Panel (選中項目的設定, 可單筆覆寫)
"""
import csv
import json
import os
import re
from dataclasses import replace
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTextEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QFileDialog, QLabel, QSplitter, QGroupBox, QMessageBox,
    QComboBox, QHeaderView, QStatusBar, QTabWidget, QSpinBox,
    QDoubleSpinBox, QLineEdit, QFormLayout, QDialog,
    QListWidget, QListWidgetItem, QRadioButton, QButtonGroup,
    QFrame, QScrollArea, QTextBrowser, QInputDialog,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import (
    QCursor, QFont, QColor, QIcon, QImage, QPixmap, QShortcut, QKeySequence,
)

try:
    import fitz  # PyMuPDF
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

from core.calculator import (
    analyze_single, get_supported_types,
    set_analysis_setting, get_analysis_setting,
)
from core.models import AnalysisResult
from core.parser import get_type_code, get_part, get_lookup_value
from core.project_aggregation import ProjectInputRow, analyze_project_rows
from core.config_loader import load_config, get_type_table_as_dict
from ui.type_manager import TypeManagerWidget, load_catalog
from ui.ontology_browser import OntologyBrowserWidget
from ui.material_cutting_page import MaterialCuttingPage

# PDF/資源路徑
_UI_DIR = os.path.dirname(os.path.abspath(__file__))
_APP_DIR = os.path.dirname(_UI_DIR)
_PDF_DIR = os.path.join(_APP_DIR, "assets", "Type")

# 結果表格群組背景色 (header_row_color, body_row_color)
_RESULT_GROUP_COLORS = [
    ("#E2EDF8", "#FAFCFE"),   # blue-gray
    ("#E8EEF5", "#FFFFFF"),   # cool gray
    ("#E5F0F4", "#FBFDFE"),   # steel blue
    ("#EDF1F6", "#FFFFFF"),   # neutral gray
]

_RESULT_HEADERS = [
    "流水號.sort", "數量", "單位", "型號", "Type", "項次", "品名", "規格", "材質",
    "長度(mm)", "寬度(mm)", "單件數量", "總數量",
    "單組重(kg)", "總重(kg)", "屬性", "物件類別", "製造方式",
    "計算說明", "零件ID", "庫存ID",
]

_RESULT_COLUMN_WIDTHS = [
    92, 54, 48, 120, 48, 48, 150, 230, 95,
    78, 78, 74, 74, 86, 86, 82, 112, 112, 300, 260, 96,
]

_PROJECT_ROW_ALIASES = {
    "serial": ("流水號.sort", "流水號", "serial", "serial_no", "seq", "sort", "序號", "編號"),
    "designation": ("型號", "designation", "support_designation", "support_no", "支撐編碼", "編碼", "支撐型號", "model"),
    "quantity": ("數量", "quantity", "qty", "count", "組數", "支數"),
    "unit": ("單位", "unit", "uom"),
    "enabled": ("enabled", "啟用"),
    "overrides": ("overrides_json", "overrides"),
    "description": ("description", "desc", "描述", "中文說明", "說明", "品名"),
    "item_code": ("item_code", "item code", "料號", "code"),
}

_PROJECT_XLSX_FIELD_LABELS = {
    "serial": "流水號.sort",
    "quantity": "數量",
    "unit": "單位",
    "designation": "型號",
    "description": "說明備援",
    "item_code": "料號備援",
}


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("IEC 管架支撐分析工具 (Python Edition)")
        self.setMinimumSize(1300, 720)
        self._project_rows = []    # list[ProjectInputRow]
        self._results = []
        self._project_result = None
        self._selected_index = -1
        self._apply_stylesheet()
        self._init_ui()

    # ══════════════════════════════════════════
    #  全域樣式
    # ══════════════════════════════════════════
    def _apply_stylesheet(self):
        self.setStyleSheet("""
            QMainWindow, QDialog {
                background-color: #F5F6FA;
            }
            QTabWidget::pane {
                border: 1px solid #C8CDD5;
                background: #FFFFFF;
                border-radius: 0 4px 4px 4px;
            }
            QTabBar::tab {
                padding: 7px 18px;
                background: #E8ECF1;
                color: #555;
                border: 1px solid #C8CDD5;
                border-bottom: none;
                border-radius: 4px 4px 0 0;
                font-size: 12px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #FFFFFF;
                color: #1565C0;
                font-weight: bold;
                border-bottom: 2px solid #FFFFFF;
            }
            QTabBar::tab:hover:!selected {
                background: #DDE3EC;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                color: #333;
                border: 1px solid #C8CDD5;
                border-radius: 6px;
                margin-top: 18px;
                padding: 12px 8px 8px 8px;
                background: #FFFFFF;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #1565C0;
                font-size: 12px;
            }
            QPushButton {
                padding: 5px 12px;
                border: 1px solid #B8C0CC;
                border-radius: 4px;
                background-color: #F0F2F5;
                color: #333;
                font-size: 12px;
                min-height: 24px;
            }
            QPushButton:hover {
                background-color: #E2E8F0;
                border-color: #8898B0;
            }
            QPushButton:pressed {
                background-color: #D0DAEA;
            }
            QPushButton:disabled {
                color: #AAA;
                background-color: #EAEAEA;
                border-color: #D0D0D0;
            }
            QLineEdit, QComboBox {
                border: 1px solid #C0C8D4;
                border-radius: 4px;
                padding: 4px 8px;
                background: #FFFFFF;
                color: #222;
                selection-background-color: #BBDEFB;
                font-size: 12px;
            }
            QLineEdit:focus, QComboBox:focus {
                border-color: #1565C0;
            }
            QListWidget {
                border: 1px solid #C0C8D4;
                border-radius: 4px;
                background: #FFFFFF;
                outline: none;
            }
            QListWidget::item {
                padding: 3px 6px;
                border-bottom: 1px solid #EEF0F3;
            }
            QListWidget::item:selected {
                background: #BBDEFB;
                color: #0D47A1;
                border-radius: 2px;
            }
            QListWidget::item:hover:!selected {
                background: #E8F0F8;
            }
            QTableWidget {
                border: 1px solid #C0C8D4;
                border-radius: 4px;
                gridline-color: #E4E8EE;
                background: #FFFFFF;
                alternate-background-color: #F8FAFB;
                selection-background-color: #BBDEFB;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #EEF2F8;
                color: #2C3E60;
                font-weight: bold;
                font-size: 11px;
                border: none;
                border-right: 1px solid #D2D8E2;
                border-bottom: 1px solid #BBDEFB;
                padding: 5px 8px;
            }
            QScrollBar:vertical {
                width: 10px;
                background: #F0F2F5;
            }
            QScrollBar::handle:vertical {
                background: #C0CBD8;
                border-radius: 5px;
                min-height: 20px;
            }
            QStatusBar {
                color: #666;
                font-size: 12px;
                background: #F0F2F5;
                border-top: 1px solid #D0D5DC;
            }
            QSplitter::handle {
                background: #D0D5DC;
                width: 2px;
            }
        """)

    # ══════════════════════════════════════════
    #  UI 初始化
    # ══════════════════════════════════════════
    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(4, 4, 4, 4)

        # ── 主 Tab ──
        self.main_tabs = QTabWidget()
        self.main_tabs.setFont(QFont("Microsoft JhengHei UI", 10))

        # Tab 1: 分析頁面
        analysis_page = self._build_analysis_page()
        self.main_tabs.addTab(analysis_page, "📊 重量分析")

        # Tab 2: 材料合計 / 下料
        self.material_cutting_page = MaterialCuttingPage()
        self.material_cutting_page.btn_generate.clicked.connect(self._on_generate_material)
        self.main_tabs.addTab(self.material_cutting_page, "📦 材料合計 / 下料")

        # Tab 3: Type 管理器
        self.type_manager = TypeManagerWidget()
        self.main_tabs.addTab(self.type_manager, "📋 Type 總覽")

        # Tab 4: 支撐架構
        self.ontology_browser = OntologyBrowserWidget()
        self.main_tabs.addTab(self.ontology_browser, "🌳 支撐架構")

        main_layout.addWidget(self.main_tabs)
        self.statusBar().showMessage("就緒 — 新增支撐編碼後按「開始分析」")
        self._install_shortcuts()

    def _install_shortcuts(self):
        """Keyboard shortcuts for high-volume input work."""
        self._shortcuts = [
            QShortcut(
                QKeySequence("Delete"),
                self.item_list,
                activated=self._on_delete_item,
            ),
            QShortcut(
                QKeySequence("F2"),
                self.item_list,
                activated=self._on_edit_current_item,
            ),
            QShortcut(
                QKeySequence("Ctrl+Return"),
                self,
                activated=self._on_batch_paste,
            ),
            QShortcut(
                QKeySequence("Ctrl+Enter"),
                self,
                activated=self._on_batch_paste,
            ),
            QShortcut(
                QKeySequence("Ctrl+F"),
                self,
                activated=self._focus_result_filter,
            ),
            QShortcut(
                QKeySequence("Escape"),
                self.result_filter_input,
                activated=self.result_filter_input.clear,
            ),
        ]

    def _build_analysis_page(self) -> QWidget:
        """建構分析頁面 (原始三面板)"""
        page = QWidget()
        page_layout = QVBoxLayout(page)

        # ── 頂部工具列 ──
        toolbar = QHBoxLayout()
        supported = get_supported_types()
        info_label = QLabel(f"已支援 Type: {', '.join(supported)}")
        info_label.setStyleSheet("color: #555; font-size: 12px;")
        toolbar.addWidget(info_label)
        toolbar.addStretch()

        toolbar.addWidget(QLabel("全域上段管材質:"))
        self.material_combo = QComboBox()
        self.material_combo.addItems([
            "SUS304", "SUS316", "A53Gr.B", "A106Gr.B",
            "A335-P11", "A335-P22", "A312-TP304", "A312-TP316",
        ])
        self.material_combo.setEditable(True)
        self.material_combo.setCurrentText("SUS304")
        self.material_combo.currentTextChanged.connect(self._on_material_changed)
        toolbar.addWidget(self.material_combo)

        self.btn_config = QPushButton("⚙ Type 資料管理")
        self.btn_config.clicked.connect(self._on_open_config)
        toolbar.addWidget(self.btn_config)

        page_layout.addLayout(toolbar)

        # ── 三面板 ──
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # [左] 輸入清單
        left_panel = self._build_left_panel()
        splitter.addWidget(left_panel)

        # [中] 結果表格
        center_panel = self._build_center_panel()
        splitter.addWidget(center_panel)

        # [右] Side Panel
        self.side_panel = SidePanel()
        self.side_panel.overrideChanged.connect(self._on_override_changed)
        splitter.addWidget(self.side_panel)

        splitter.setSizes([260, 680, 300])
        page_layout.addWidget(splitter)

        return page

    def _build_left_panel(self):
        panel = QGroupBox("輸入清單")
        layout = QVBoxLayout(panel)

        # 新增列
        add_row = QHBoxLayout()
        self.add_input = QLineEdit()
        self.add_input.setPlaceholderText("型號，或 流水號 數量 單位 型號")
        self.add_input.setFont(QFont("Consolas", 11))
        self.add_input.returnPressed.connect(self._on_add_item)
        add_row.addWidget(self.add_input)
        btn_add = QPushButton("+")
        btn_add.setFixedWidth(32)
        btn_add.clicked.connect(self._on_add_item)
        add_row.addWidget(btn_add)
        layout.addLayout(add_row)

        # 清單
        self.item_list = QListWidget()
        self.item_list.currentRowChanged.connect(self._on_item_selected)
        self.item_list.itemChanged.connect(self._on_item_check_changed)
        self.item_list.setFont(QFont("Consolas", 10))
        layout.addWidget(self.item_list)

        # 按鈕列
        btn_row1 = QHBoxLayout()
        btn_batch = QPushButton("批次貼上...")
        btn_batch.clicked.connect(self._on_batch_paste)
        btn_row1.addWidget(btn_batch)
        btn_qty = QPushButton("流水號/組數...")
        btn_qty.clicked.connect(self._on_set_quantities)
        btn_row1.addWidget(btn_qty)
        layout.addLayout(btn_row1)

        btn_row2 = QHBoxLayout()
        btn_load = QPushButton("從檔案載入...")
        btn_load.clicked.connect(self._on_load_file)
        btn_row2.addWidget(btn_load)
        btn_save_list = QPushButton("儲存清單...")
        btn_save_list.clicked.connect(self._on_save_file)
        btn_row2.addWidget(btn_save_list)
        layout.addLayout(btn_row2)

        btn_row3 = QHBoxLayout()
        btn_del = QPushButton("刪除選中")
        btn_del.clicked.connect(self._on_delete_item)
        btn_row3.addWidget(btn_del)
        btn_clear = QPushButton("全部清除")
        btn_clear.clicked.connect(self._on_clear_all)
        btn_row3.addWidget(btn_clear)
        layout.addLayout(btn_row3)

        # 分析按鈕
        self.btn_analyze = QPushButton("▶ 開始分析")
        self.btn_analyze.setStyleSheet(
            "QPushButton { background-color: #1976D2; color: white; "
            "border: 1px solid #1565C0; padding: 8px; "
            "font-size: 14px; font-weight: bold; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1565C0; }"
            "QPushButton:pressed { background-color: #0D47A1; }"
            "QPushButton:disabled { background-color: #B0BEC5; "
            "border-color: #90A4AE; color: #ECEFF1; }"
        )
        self.btn_analyze.clicked.connect(self._on_analyze)
        layout.addWidget(self.btn_analyze)

        return panel

    def _build_center_panel(self):
        panel = QGroupBox("分析結果")
        layout = QVBoxLayout(panel)

        layout.addWidget(self._build_result_summary_bar())
        layout.addLayout(self._build_result_filter_row())

        self.result_table = QTableWidget()
        self._set_project_result_headers()
        self._apply_result_table_column_layout()
        self.result_table.setAlternatingRowColors(False)
        self.result_table.setWordWrap(False)
        self.result_table.setTextElideMode(Qt.TextElideMode.ElideRight)
        self.result_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        self.result_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        self.result_table.verticalHeader().setDefaultSectionSize(22)
        self.result_table.verticalHeader().setVisible(False)
        layout.addWidget(self.result_table)

        # 匯出列
        export_row = QHBoxLayout()
        self.export_format = QComboBox()
        self.export_format.addItems(["Excel (.xlsx)", "CSV (.csv)", "PDF (.pdf)"])
        export_row.addWidget(QLabel("匯出格式:"))
        export_row.addWidget(self.export_format)
        self.btn_export = QPushButton("匯出結果")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._on_export)
        export_row.addWidget(self.btn_export)
        export_row.addStretch()
        layout.addLayout(export_row)

        return panel

    def _build_result_filter_row(self):
        row = QHBoxLayout()
        row.setSpacing(6)

        label = QLabel("搜尋:")
        label.setStyleSheet("color: #607080; font-size: 12px;")
        row.addWidget(label)

        self.result_filter_input = QLineEdit()
        self.result_filter_input.setPlaceholderText("流水號 / 型號 / 品名 / 規格 / 材質 / 零件ID / 庫存ID / 說明")
        self.result_filter_input.setClearButtonEnabled(True)
        self.result_filter_input.textChanged.connect(self._apply_result_filter)
        row.addWidget(self.result_filter_input, 1)

        self.result_filter_count_label = QLabel("顯示 0 列")
        self.result_filter_count_label.setMinimumWidth(86)
        self.result_filter_count_label.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        self.result_filter_count_label.setStyleSheet(
            "color: #607080; font-size: 12px;"
        )
        row.addWidget(self.result_filter_count_label)
        return row

    def _build_result_summary_bar(self):
        bar = QFrame()
        bar.setObjectName("resultSummaryBar")
        bar.setStyleSheet(
            "QFrame#resultSummaryBar {"
            "background: #F8FBFE;"
            "border: 1px solid #D6E2EF;"
            "border-radius: 4px;"
            "}"
            "QLabel { background: transparent; }"
        )

        row = QHBoxLayout(bar)
        row.setContentsMargins(10, 6, 10, 6)
        row.setSpacing(12)

        self.total_weight_label = self._make_summary_value_label(
            "-- kg", "#0D47A1", 13, True
        )
        self.summary_success_label = self._make_summary_value_label("--", "#1B5E20")
        self.summary_error_label = self._make_summary_value_label("--", "#546E7A")
        self.summary_support_label = self._make_summary_value_label("--", "#263238")

        row.addLayout(self._make_summary_metric("總重量", self.total_weight_label))
        row.addWidget(self._make_summary_separator())
        row.addLayout(self._make_summary_metric("成功項目", self.summary_success_label))
        row.addWidget(self._make_summary_separator())
        row.addLayout(self._make_summary_metric("錯誤項目", self.summary_error_label))
        row.addWidget(self._make_summary_separator())
        row.addLayout(self._make_summary_metric("支撐組數", self.summary_support_label))
        row.addStretch()
        return bar

    def _make_summary_metric(self, title: str, value_label: QLabel):
        metric = QHBoxLayout()
        metric.setContentsMargins(0, 0, 0, 0)
        metric.setSpacing(5)

        title_label = QLabel(title)
        title_label.setFont(QFont("Microsoft JhengHei UI", 12))
        title_label.setStyleSheet(
            "color: #607080; font-weight: normal; background: transparent;"
        )
        metric.addWidget(title_label)
        metric.addWidget(value_label)
        return metric

    def _make_summary_value_label(
        self,
        text: str,
        color: str,
        size: int = 12,
        bold: bool = True,
    ):
        label = QLabel(text)
        weight = QFont.Weight.Bold if bold else QFont.Weight.Normal
        label.setFont(QFont("Microsoft JhengHei UI", size, weight))
        label.setStyleSheet(f"color: {color}; background: transparent;")
        label.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
        return label

    def _make_summary_separator(self):
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setFrameShadow(QFrame.Shadow.Plain)
        sep.setStyleSheet("color: #D6E2EF; background: #D6E2EF;")
        sep.setFixedHeight(20)
        return sep

    def _set_project_result_headers(self):
        """Result table columns: human BOM + machine IDs without overloading notes."""
        self.result_table.setColumnCount(len(_RESULT_HEADERS))
        self.result_table.setHorizontalHeaderLabels(_RESULT_HEADERS)

    def _apply_result_table_column_layout(self):
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(False)
        for col, width in enumerate(_RESULT_COLUMN_WIDTHS):
            self.result_table.setColumnWidth(col, width)

    def _result_item(self, value) -> QTableWidgetItem:
        text = "" if value is None else str(value)
        item = QTableWidgetItem(text)
        if text:
            item.setToolTip(text)
        return item

    # ══════════════════════════════════════════
    #  清單操作
    # ══════════════════════════════════════════
    def _on_add_item(self):
        text = self.add_input.text().strip()
        if not text:
            return
        try:
            rows = self._read_project_rows_text([text])
        except ValueError as exc:
            QMessageBox.warning(self, "輸入格式錯誤", str(exc))
            return
        if not rows:
            return
        row = rows[0]
        self._add_item_to_list(
            row.designation,
            quantity=row.quantity,
            enabled=row.enabled,
            overrides=row.overrides or None,
            serial=row.serial,
            unit=row.unit,
        )
        self.add_input.clear()
        self.add_input.setFocus()

    def _add_item_to_list(
        self,
        text: str,
        invalidate: bool = True,
        *,
        quantity: int = 1,
        enabled: bool = True,
        overrides: dict | None = None,
        serial: str = "",
        unit: str = "組",
    ):
        idx = len(self._project_rows)
        self._project_rows.append(
            ProjectInputRow(
                designation=text,
                quantity=quantity,
                enabled=enabled,
                overrides=overrides or None,
                serial=str(serial or "").strip(),
                unit=str(unit or "組").strip(),
            )
        )
        item_widget = QListWidgetItem(text)
        item_widget.setCheckState(
            Qt.CheckState.Checked if enabled else Qt.CheckState.Unchecked
        )
        self._update_item_display(idx, item_widget)
        self.item_list.addItem(item_widget)
        if invalidate:
            self._invalidate_analysis_outputs("輸入清單已變更，請重新分析")

    def _update_item_display(self, idx: int, item_widget: QListWidgetItem = None):
        """更新清單顯示文字 (有覆寫時加標記)"""
        if item_widget is None:
            item_widget = self.item_list.item(idx)
        if item_widget is None:
            return
        row = self._project_rows[idx]
        text = f"{row.serial} | {row.designation}" if row.serial else row.designation
        overrides = row.overrides or {}
        tags = []
        if row.quantity != 1 or row.serial:
            tags.append(f"{row.quantity}{row.unit or '組'}")
        if overrides.get("connection"):
            tags.append("Tee" if overrides["connection"] == "tee" else "Elbow")
        if overrides.get("upper_material"):
            tags.append(overrides["upper_material"])
        if any(overrides.get(k) for k in ("pipe_size", "schedule", "l_value")):
            tags.append("自訂值")

        if tags:
            item_widget.setText(f"{text}  ◆ [{', '.join(tags)}]")
            item_widget.setForeground(QColor("#1565C0"))
        else:
            item_widget.setText(text)
            item_widget.setForeground(QColor("black"))
        item_widget.setToolTip(
            f"流水號.sort: {row.serial or '-'}\n"
            f"型號: {row.designation}\n"
            f"數量: {row.quantity}\n"
            f"單位: {row.unit or '組'}"
        )

    def _refresh_item_list_display(self):
        """Refresh all list labels and checkbox states from project rows."""
        self.item_list.blockSignals(True)
        for idx, row in enumerate(self._project_rows):
            item_widget = self.item_list.item(idx)
            if item_widget is None:
                continue
            item_widget.setCheckState(
                Qt.CheckState.Checked if row.enabled else Qt.CheckState.Unchecked
            )
            self._update_item_display(idx, item_widget)
        self.item_list.blockSignals(False)

    def _on_batch_paste(self):
        """批次貼上多筆"""
        dlg = QDialog(self)
        dlg.setWindowTitle("批次貼上")
        dlg.setMinimumSize(400, 300)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("每行一筆支撐編碼；也可貼 流水號.sort,數量,單位,型號:"))
        text_edit = QTextEdit()
        text_edit.setFont(QFont("Consolas", 11))
        text_edit.setPlaceholderText(
            "01-2B-05A\n"
            "01-3B-08B,2\n"
            "12,1,組,57-1B-A"
        )
        lay.addWidget(text_edit)
        count_label = QLabel("已輸入 0 筆")
        count_label.setStyleSheet("color: #666; font-size: 12px;")
        lay.addWidget(count_label)
        btn = QPushButton("加入清單")
        btn.clicked.connect(dlg.accept)
        lay.addWidget(btn)

        def parse_lines() -> list[str]:
            return [
                line.strip()
                for line in text_edit.toPlainText().splitlines()
                if line.strip()
            ]

        def row_key(row: ProjectInputRow) -> str:
            return f"{row.serial}|{row.designation}".casefold() if row.serial else row.designation.casefold()

        def duplicate_count(rows: list[ProjectInputRow]) -> int:
            seen = {row_key(row) for row in self._project_rows}
            pasted = set()
            duplicates = 0
            for row in rows:
                key = row_key(row)
                if key in seen or key in pasted:
                    duplicates += 1
                pasted.add(key)
            return duplicates

        def update_count():
            lines = parse_lines()
            try:
                rows = self._read_project_rows_text(lines)
            except ValueError:
                count_label.setText(f"已輸入 {len(lines)} 筆；格式待確認")
                count_label.setStyleSheet("color: #C62828; font-size: 12px;")
                return
            dupes = duplicate_count(rows)
            if dupes:
                count_label.setText(f"已輸入 {len(rows)} 筆；其中 {dupes} 筆重複")
                count_label.setStyleSheet("color: #E65100; font-size: 12px;")
            else:
                count_label.setText(f"已輸入 {len(rows)} 筆")
                count_label.setStyleSheet("color: #666; font-size: 12px;")

        text_edit.textChanged.connect(update_count)
        if dlg.exec():
            lines = parse_lines()
            if not lines:
                return
            try:
                rows = self._read_project_rows_text(lines)
            except ValueError as exc:
                QMessageBox.warning(self, "批次格式錯誤", str(exc))
                return
            for row in rows:
                self._add_item_to_list(
                    row.designation,
                    invalidate=False,
                    quantity=row.quantity,
                    enabled=row.enabled,
                    overrides=row.overrides or None,
                    serial=row.serial,
                    unit=row.unit,
                )
            self._invalidate_analysis_outputs("批次貼上已加入清單，請重新分析")
            self.statusBar().showMessage(f"已加入 {len(rows)} 筆支撐編碼")

    def _on_load_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "載入支撐清單", "",
            "Excel/CSV/Text (*.xlsx *.xlsm *.csv *.txt);;Excel (*.xlsx *.xlsm);;CSV (*.csv);;文字檔 (*.txt);;所有檔案 (*)"
        )
        if not filepath:
            return

        try:
            rows = self._read_project_rows_file(filepath)
        except Exception as exc:
            QMessageBox.critical(self, "載入失敗", str(exc))
            return

        if not rows:
            QMessageBox.information(self, "載入支撐清單", "檔案內沒有可載入的支撐編碼。")
            return

        replace_existing = False
        if self._project_rows:
            choice = QMessageBox.question(
                self,
                "載入支撐清單",
                "要取代目前清單嗎？\n\n選「否」會將檔案內容追加到目前清單後方。",
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Yes,
            )
            if choice == QMessageBox.StandardButton.Cancel:
                return
            replace_existing = choice == QMessageBox.StandardButton.Yes

        if replace_existing:
            self._project_rows.clear()
            self.item_list.clear()

        for row in rows:
            self._add_item_to_list(
                row.designation,
                invalidate=False,
                quantity=row.quantity,
                enabled=row.enabled,
                overrides=row.overrides or None,
                serial=row.serial,
                unit=row.unit,
            )
        self._invalidate_analysis_outputs("載入檔案已更新清單，請重新分析")
        action = "取代並載入" if replace_existing else "追加載入"
        self.statusBar().showMessage(f"已{action} {len(rows)} 筆: {filepath}")

    def _on_save_file(self):
        if not self._project_rows:
            QMessageBox.warning(self, "提示", "目前沒有支撐清單可儲存")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "儲存支撐清單",
            "support_list.csv",
            "CSV (*.csv);;所有檔案 (*)",
        )
        if not filepath:
            return
        if not os.path.splitext(filepath)[1]:
            filepath += ".csv"

        try:
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["流水號.sort", "數量", "單位", "型號", "enabled", "overrides_json"],
                )
                writer.writeheader()
                for row in self._project_rows:
                    writer.writerow({
                        "流水號.sort": row.serial,
                        "數量": row.quantity,
                        "單位": row.unit or "組",
                        "型號": row.designation,
                        "enabled": "1" if row.enabled else "0",
                        "overrides_json": json.dumps(
                            row.overrides or {},
                            ensure_ascii=False,
                            sort_keys=True,
                        ),
                    })
        except Exception as exc:
            QMessageBox.critical(self, "儲存失敗", str(exc))
            return

        self.statusBar().showMessage(f"已儲存 {len(self._project_rows)} 筆清單: {filepath}")

    def _read_project_rows_file(self, filepath: str) -> list[ProjectInputRow]:
        ext = os.path.splitext(filepath)[1].lower()
        if ext in {".xlsx", ".xlsm"}:
            return self._read_project_rows_xlsx(filepath)

        with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
            text = f.read()
        if not text.strip():
            return []

        lines = [line for line in text.splitlines() if line.strip()]
        first = lines[0].lower()
        if "," in lines[0] and (
            "designation" in first
            or "型號" in first
            or "流水號" in first
            or "quantity" in first
            or "數量" in first
        ):
            return self._read_project_rows_csv(lines)
        return self._read_project_rows_text(lines)

    def _read_project_rows_csv(self, lines: list[str]) -> list[ProjectInputRow]:
        rows: list[ProjectInputRow] = []
        reader = csv.DictReader(lines)
        for idx, raw in enumerate(reader, start=2):
            designation = self._project_field_value(raw, _PROJECT_ROW_ALIASES["designation"])
            if not designation:
                designation = (
                    self._extract_designation_from_text(
                        self._project_field_value(raw, _PROJECT_ROW_ALIASES["description"])
                    )
                    or self._extract_designation_from_text(
                        self._project_field_value(raw, _PROJECT_ROW_ALIASES["item_code"])
                    )
                )
            if not designation:
                continue
            quantity_text = self._project_field_value(raw, _PROJECT_ROW_ALIASES["quantity"]) or "1"
            enabled_text = self._project_field_value(raw, _PROJECT_ROW_ALIASES["enabled"]) or "1"
            overrides_text = self._project_field_value(raw, _PROJECT_ROW_ALIASES["overrides"])
            rows.append(
                ProjectInputRow(
                    designation=designation,
                    quantity=self._parse_list_quantity(quantity_text, idx),
                    enabled=self._parse_list_enabled(enabled_text),
                    overrides=self._parse_list_overrides(overrides_text, idx),
                    serial=self._project_field_value(raw, _PROJECT_ROW_ALIASES["serial"]),
                    unit=self._project_field_value(raw, _PROJECT_ROW_ALIASES["unit"]) or "組",
                )
            )
        return rows

    def _read_project_rows_xlsx(self, filepath: str) -> list[ProjectInputRow]:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            layout = self._detect_project_xlsx_layout(wb)
            if layout is None:
                raise ValueError("找不到可用的 MTO 表頭；請確認 Excel 內有流水號/數量/單位/型號等欄位。")

            ws = layout["worksheet"]
            header_row = layout["header_row"]
            headers = layout["headers"]
            mapping = dict(layout["mapping"])
            mapping = self._confirm_project_xlsx_mapping(ws.title, header_row, headers, mapping)

            if not self._has_project_designation_source(mapping):
                raise ValueError("xlsx 匯入至少需要指定「型號」欄，或指定可抽型號的說明/料號備援欄。")
            if mapping.get("quantity") is None:
                raise ValueError("xlsx 匯入至少需要指定「數量」欄。")

            rows: list[ProjectInputRow] = []
            for row_idx, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                designation = self._project_mapped_value(values, mapping, "designation")
                if not designation:
                    designation = (
                        self._extract_designation_from_text(
                            self._project_mapped_value(values, mapping, "description")
                        )
                        or self._extract_designation_from_text(
                            self._project_mapped_value(values, mapping, "item_code")
                        )
                    )
                if not designation:
                    continue

                quantity_text = self._project_mapped_value(values, mapping, "quantity") or "1"
                rows.append(
                    ProjectInputRow(
                        designation=designation,
                        quantity=self._parse_list_quantity(quantity_text, row_idx),
                        enabled=True,
                        serial=self._project_mapped_value(values, mapping, "serial"),
                        unit=self._project_mapped_value(values, mapping, "unit") or "組",
                    )
                )
            return rows
        finally:
            wb.close()

    def _read_project_rows_text(self, lines: list[str]) -> list[ProjectInputRow]:
        rows: list[ProjectInputRow] = []
        for idx, line in enumerate(lines, start=1):
            parts = self._split_project_row_parts(line)
            if not parts or not parts[0]:
                continue
            if self._looks_like_project_header(parts):
                continue
            rows.append(self._project_row_from_text_parts(parts, idx))
        return rows

    def _detect_project_xlsx_layout(self, wb):
        best = None
        for ws in wb.worksheets:
            scan_rows = list(
                ws.iter_rows(
                    min_row=1,
                    max_row=min(ws.max_row, 30),
                    values_only=True,
                )
            )
            for row_offset, row_values in enumerate(scan_rows, start=1):
                headers = ["" if cell is None else str(cell).strip() for cell in row_values]
                if sum(1 for header in headers if header) < 2:
                    continue
                sample_rows = scan_rows[row_offset:row_offset + 12]
                mapping, score = self._infer_project_column_mapping(headers, sample_rows)
                score += self._project_sheet_name_bonus(ws.title)
                score += 60 if mapping.get("designation") is not None and mapping.get("quantity") is not None else 0
                candidate = {
                    "worksheet": ws,
                    "header_row": row_offset,
                    "headers": headers,
                    "mapping": mapping,
                    "score": score,
                }
                if best is None or candidate["score"] > best["score"]:
                    best = candidate
        return best

    @staticmethod
    def _project_sheet_name_bonus(title: str) -> int:
        normalized = re.sub(r"\s+", "", str(title or "").strip().lower())
        if normalized == "supportmto":
            return 30
        bonus = 0
        if "support" in normalized:
            bonus += 8
        if "mto" in normalized:
            bonus += 8
        if "材料" in normalized or "支撐" in normalized:
            bonus += 4
        return bonus

    def _infer_project_column_mapping(self, headers: list[str], sample_rows: list[tuple]) -> tuple[dict[str, int | None], float]:
        field_order = ("designation", "quantity", "serial", "unit", "description", "item_code")
        thresholds = {
            "designation": 35,
            "quantity": 35,
            "serial": 45,
            "unit": 35,
            "description": 35,
            "item_code": 35,
        }
        scores = []
        for field in field_order:
            for col_idx, header in enumerate(headers):
                column_values = [
                    row[col_idx] if col_idx < len(row) else None
                    for row in sample_rows
                ]
                score = self._project_column_score(field, header, column_values)
                if score >= thresholds[field]:
                    scores.append((score, field, col_idx))

        mapping: dict[str, int | None] = {field: None for field in field_order}
        used_cols: set[int] = set()
        total_score = 0.0
        for score, field, col_idx in sorted(scores, reverse=True):
            if mapping[field] is not None or col_idx in used_cols:
                continue
            mapping[field] = col_idx
            used_cols.add(col_idx)
            total_score += score
        return mapping, total_score

    def _project_column_score(self, field: str, header, values: list) -> float:
        header_text = str(header or "").strip()
        normalized = self._normalize_project_header(header_text)
        if not normalized:
            return 0.0

        score = 0.0
        aliases = [self._normalize_project_header(alias) for alias in _PROJECT_ROW_ALIASES.get(field, ())]
        if normalized in aliases:
            score += 100
        elif any(alias and alias in normalized for alias in aliases):
            score += 65

        keyword_scores = {
            "serial": (("流水", "序號", "編號", "serial", "serialno", "seq", "sort", "rowno", "lineno"), 42),
            "quantity": (("數量", "組數", "qty", "quantity", "count"), 42),
            "unit": (("單位", "unit", "uom"), 42),
            "designation": (("型號", "designation", "model", "supportno", "supportdesignation", "支撐編碼"), 42),
            "description": (("description", "desc", "描述", "說明", "中文說明", "品名"), 42),
            "item_code": (("itemcode", "料號", "code"), 42),
        }
        keywords, keyword_score = keyword_scores.get(field, ((), 0))
        if any(self._normalize_project_header(keyword) in normalized for keyword in keywords):
            score += keyword_score

        nonempty = [str(value).strip() for value in values if value is not None and str(value).strip()]
        if not nonempty:
            return score

        if field == "designation":
            hit_ratio = sum(1 for value in nonempty if self._looks_like_designation(value)) / len(nonempty)
            score += hit_ratio * 70
        elif field == "quantity":
            numeric_ratio = sum(1 for value in nonempty if self._looks_like_list_quantity(value)) / len(nonempty)
            score += numeric_ratio * 45
        elif field == "unit":
            unit_ratio = sum(1 for value in nonempty if self._looks_like_project_unit(value)) / len(nonempty)
            score += unit_ratio * 70
        elif field == "serial":
            unique_ratio = len(set(nonempty)) / len(nonempty)
            score += min(unique_ratio, 1.0) * 15
        elif field == "description":
            extracted_ratio = sum(1 for value in nonempty if self._extract_designation_from_text(value) != value) / len(nonempty)
            score += extracted_ratio * 35
        return score

    def _project_mapped_value(self, values, mapping: dict, field: str) -> str:
        col_idx = mapping.get(field)
        if col_idx is None or col_idx >= len(values):
            return ""
        value = values[col_idx]
        return "" if value is None else str(value).strip()

    @staticmethod
    def _has_project_designation_source(mapping: dict) -> bool:
        return any(mapping.get(field) is not None for field in ("designation", "description", "item_code"))

    def _confirm_project_xlsx_mapping(
        self,
        sheet_name: str,
        header_row: int,
        headers: list[str],
        mapping: dict[str, int | None],
    ) -> dict[str, int | None]:
        if not self._can_show_xlsx_mapping_dialog():
            return mapping

        dlg = QDialog(self)
        dlg.setWindowTitle("xlsx 匯入欄位對應")
        dlg.setMinimumSize(560, 360)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel(f"Sheet: {sheet_name}    表頭列: {header_row}"))
        lay.addWidget(QLabel("請確認欄位對應；型號與數量為必填，其餘可不使用。"))

        form = QFormLayout()
        combos: dict[str, QComboBox] = {}
        fields = ("serial", "quantity", "unit", "designation", "description", "item_code")
        for field in fields:
            combo = QComboBox()
            combo.addItem("不使用", None)
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                combo.addItem(f"{self._xlsx_column_label(col_idx)}  {header}", col_idx)
            current_col = mapping.get(field)
            if current_col is not None:
                for item_idx in range(combo.count()):
                    if combo.itemData(item_idx) == current_col:
                        combo.setCurrentIndex(item_idx)
                        break
            combos[field] = combo
            label = _PROJECT_XLSX_FIELD_LABELS.get(field, field)
            if field in {"designation", "quantity"}:
                label += " *"
            form.addRow(label, combo)
        lay.addLayout(form)

        button_row = QHBoxLayout()
        btn_ok = QPushButton("套用匯入")
        btn_cancel = QPushButton("取消")
        button_row.addStretch()
        button_row.addWidget(btn_ok)
        button_row.addWidget(btn_cancel)
        lay.addLayout(button_row)

        def accept_mapping():
            selected = {field: combos[field].currentData() for field in fields}
            if not self._has_project_designation_source(selected) or selected.get("quantity") is None:
                QMessageBox.warning(dlg, "欄位不足", "請至少指定「數量」欄，以及「型號」或可抽型號的備援欄。")
                return
            dlg.accept()

        btn_ok.clicked.connect(accept_mapping)
        btn_cancel.clicked.connect(dlg.reject)

        if not dlg.exec():
            raise ValueError("已取消 xlsx 匯入")

        return {field: combos[field].currentData() for field in fields}

    def _can_show_xlsx_mapping_dialog(self) -> bool:
        try:
            return QApplication.instance() is not None and "item_list" in object.__getattribute__(self, "__dict__")
        except RuntimeError:
            return False

    @staticmethod
    def _xlsx_column_label(col_idx: int) -> str:
        from openpyxl.utils import get_column_letter

        return get_column_letter(col_idx + 1)

    def _split_project_row_parts(self, line: str) -> list[str]:
        text = str(line or "").strip().replace("，", ",")
        if not text:
            return []
        if "\t" in text:
            return [part.strip() for part in text.split("\t") if part.strip()]
        if "," in text:
            return [part.strip() for part in next(csv.reader([text])) if part.strip()]
        if "|" in text:
            return [part.strip() for part in text.split("|") if part.strip()]
        return [part.strip() for part in text.split() if part.strip()]

    def _project_row_from_text_parts(self, parts: list[str], row_number: int) -> ProjectInputRow:
        if len(parts) >= 4 and self._looks_like_list_quantity(parts[1]):
            if self._looks_like_project_unit(parts[2]):
                serial, quantity_text, unit, designation = parts[0], parts[1], parts[2], parts[3]
            else:
                serial, quantity_text, designation, unit = parts[0], parts[1], parts[2], parts[3]
            return ProjectInputRow(
                designation=designation,
                quantity=self._parse_list_quantity(quantity_text, row_number),
                serial=serial,
                unit=unit or "組",
            )

        if len(parts) >= 3 and self._looks_like_list_quantity(parts[1]):
            return ProjectInputRow(
                designation=parts[2],
                quantity=self._parse_list_quantity(parts[1], row_number),
                serial=parts[0],
                unit=parts[3] if len(parts) >= 4 and parts[3] else "組",
            )

        if len(parts) >= 3 and self._looks_like_list_quantity(parts[2]):
            return ProjectInputRow(
                designation=parts[1],
                quantity=self._parse_list_quantity(parts[2], row_number),
                serial=parts[0],
                unit=parts[3] if len(parts) >= 4 and parts[3] else "組",
            )

        quantity = self._parse_list_quantity(parts[1], row_number) if len(parts) >= 2 and parts[1] else 1
        return ProjectInputRow(designation=parts[0], quantity=quantity)

    def _parse_list_quantity(self, text: str, row_number: int) -> int:
        try:
            numeric = float(str(text).strip().replace(",", ""))
            if not numeric.is_integer():
                raise ValueError
            value = int(numeric)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {row_number} 列組數不是整數: {text!r}") from exc
        if value <= 0:
            raise ValueError(f"第 {row_number} 列組數必須大於 0")
        return value

    def _looks_like_list_quantity(self, text: str) -> bool:
        try:
            self._parse_list_quantity(text, 0)
        except ValueError:
            return False
        return True

    def _looks_like_project_header(self, parts: list[str]) -> bool:
        normalized = {self._normalize_project_header(part) for part in parts}
        has_designation = any(
            self._normalize_project_header(alias) in normalized
            for alias in _PROJECT_ROW_ALIASES["designation"]
        )
        has_quantity = any(
            self._normalize_project_header(alias) in normalized
            for alias in _PROJECT_ROW_ALIASES["quantity"]
        )
        has_serial = any(
            self._normalize_project_header(alias) in normalized
            for alias in _PROJECT_ROW_ALIASES["serial"]
        )
        return has_designation and (has_quantity or has_serial)

    @staticmethod
    def _looks_like_project_unit(text: str) -> bool:
        return str(text or "").strip().lower() in {"組", "set", "sets", "kg", "ea", "pc", "m"}

    @staticmethod
    def _looks_like_designation(text: str) -> bool:
        return bool(re.fullmatch(r"\d{2}[A-Z]?(?:-[A-Z0-9./()]+)+", str(text or "").strip().upper()))

    @staticmethod
    def _normalize_project_header(text) -> str:
        return re.sub(r"[\s._-]+", "", str(text or "").strip().lower())

    def _project_field_value(self, raw: dict, aliases: tuple[str, ...]) -> str:
        normalized = {
            self._normalize_project_header(key): value
            for key, value in raw.items()
            if key is not None
        }
        for alias in aliases:
            value = normalized.get(self._normalize_project_header(alias))
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    @staticmethod
    def _extract_designation_from_text(text: str) -> str:
        text = str(text or "").strip()
        if not text:
            return ""
        matches = re.findall(r"\b\d{2}[A-Z]?(?:-[A-Z0-9./()]+)+\b", text.upper())
        return matches[-1] if matches else text

    @staticmethod
    def _parse_list_enabled(text: str) -> bool:
        return str(text).strip().lower() not in {"0", "false", "no", "n", "否", "停用"}

    @staticmethod
    def _parse_list_overrides(text: str, row_number: int) -> dict | None:
        text = str(text or "").strip()
        if not text:
            return None
        try:
            overrides = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"第 {row_number} 列 overrides_json 不是有效 JSON") from exc
        if not isinstance(overrides, dict):
            raise ValueError(f"第 {row_number} 列 overrides_json 必須是 JSON object")
        return overrides or None

    def _on_set_quantities(self):
        """Edit project source trace and quantities in one grid."""
        if not self._project_rows:
            QMessageBox.warning(self, "提示", "請先新增支撐編碼")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("設定流水號與組數")
        dlg.setMinimumSize(760, 520)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("可直接修改流水號.sort、型號、數量、單位；變更型號後需重新分析。"))

        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["流水號.sort", "型號", "數量", "單位"])
        table.setRowCount(len(self._project_rows))
        table.setAlternatingRowColors(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.setColumnWidth(0, 120)
        table.setColumnWidth(1, 260)
        table.setColumnWidth(2, 90)
        table.setColumnWidth(3, 80)

        for idx, row in enumerate(self._project_rows):
            serial_item = QTableWidgetItem(row.serial)
            desig_item = QTableWidgetItem(row.designation)
            unit_item = QTableWidgetItem(row.unit or "組")
            table.setItem(idx, 0, serial_item)
            table.setItem(idx, 1, desig_item)
            table.setItem(idx, 2, QTableWidgetItem(str(row.quantity)))
            table.setItem(idx, 3, unit_item)

        lay.addWidget(table)

        lay.addWidget(QLabel("批次貼入校正值：可貼單欄數量、流水號+數量，或 流水號 數量 單位 型號。"))
        qty_text = QTextEdit()
        qty_text.setPlaceholderText(
            "2\n"
            "A-001 2\n"
            "A-002 1 組 51-1B\n"
            "A-003,57-1B-A,4"
        )
        qty_text.setMaximumHeight(110)
        lay.addWidget(qty_text)

        btn_row = QHBoxLayout()
        btn_apply = QPushButton("套用批次校正")
        btn_fill_serial = QPushButton("補空流水號")
        btn_ok = QPushButton("確定")
        btn_cancel = QPushButton("取消")
        btn_row.addWidget(btn_apply)
        btn_row.addWidget(btn_fill_serial)
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

        def parse_quantity(text: str, row_number: int) -> int:
            return self._parse_list_quantity(text, row_number)

        def ensure_item(row_idx: int, col_idx: int) -> QTableWidgetItem:
            item = table.item(row_idx, col_idx)
            if item is None:
                item = QTableWidgetItem("")
                table.setItem(row_idx, col_idx, item)
            return item

        def parse_correction_line(text: str, row_number: int) -> dict:
            parts = self._split_project_row_parts(text)
            if not parts:
                return {}
            if len(parts) == 1:
                return {"quantity": self._parse_list_quantity(parts[0], row_number)}
            if len(parts) == 2 and self._looks_like_list_quantity(parts[1]):
                if self._looks_like_designation(parts[0]):
                    return {
                        "designation": parts[0],
                        "quantity": self._parse_list_quantity(parts[1], row_number),
                    }
                return {
                    "serial": parts[0],
                    "quantity": self._parse_list_quantity(parts[1], row_number),
                }
            parsed = self._project_row_from_text_parts(parts, row_number)
            return {
                "serial": parsed.serial,
                "designation": parsed.designation,
                "quantity": parsed.quantity,
                "unit": parsed.unit or "組",
            }

        def apply_batch_corrections():
            lines = [line.strip() for line in qty_text.toPlainText().splitlines() if line.strip()]
            if not lines:
                return
            if len(lines) > len(self._project_rows):
                QMessageBox.warning(
                    dlg,
                    "校正列數過多",
                    f"貼入 {len(lines)} 列，但目前只有 {len(self._project_rows)} 筆項目。",
                )
                return
            try:
                corrections = [
                    parse_correction_line(line, idx + 1)
                    for idx, line in enumerate(lines)
                ]
            except ValueError as exc:
                QMessageBox.warning(dlg, "校正格式錯誤", str(exc))
                return
            for idx, correction in enumerate(corrections):
                if "serial" in correction:
                    ensure_item(idx, 0).setText(str(correction["serial"]))
                if "designation" in correction:
                    ensure_item(idx, 1).setText(str(correction["designation"]))
                if "quantity" in correction:
                    ensure_item(idx, 2).setText(str(correction["quantity"]))
                if "unit" in correction:
                    ensure_item(idx, 3).setText(str(correction["unit"] or "組"))

        def fill_empty_serials():
            for idx in range(table.rowCount()):
                item = ensure_item(idx, 0)
                if not item.text().strip():
                    item.setText(str(idx + 1))

        btn_apply.clicked.connect(apply_batch_corrections)
        btn_fill_serial.clicked.connect(fill_empty_serials)
        btn_ok.clicked.connect(dlg.accept)
        btn_cancel.clicked.connect(dlg.reject)

        if not dlg.exec():
            return

        updated_rows = []
        try:
            for idx, row in enumerate(self._project_rows):
                serial_item = table.item(idx, 0)
                desig_item = table.item(idx, 1)
                qty_item = table.item(idx, 2)
                unit_item = table.item(idx, 3)
                designation = (desig_item.text() if desig_item else "").strip()
                if not designation:
                    raise ValueError(f"第 {idx + 1} 列型號不可空白")
                quantity = parse_quantity(qty_item.text() if qty_item else "", idx + 1)
                overrides = row.overrides
                if get_type_code(designation) != get_type_code(row.designation):
                    overrides = None
                updated_rows.append(
                    replace(
                        row,
                        serial=(serial_item.text() if serial_item else "").strip(),
                        designation=designation,
                        quantity=quantity,
                        unit=(unit_item.text() if unit_item and unit_item.text().strip() else "組").strip(),
                        overrides=overrides,
                    )
                )
        except ValueError as exc:
            QMessageBox.warning(self, "校正格式錯誤", str(exc))
            return

        self._project_rows = updated_rows
        self._refresh_item_list_display()
        if 0 <= self._selected_index < len(self._project_rows):
            selected = self._project_rows[self._selected_index]
            self.side_panel.show_item(
                self._selected_index,
                selected.designation,
                selected.overrides or {},
            )
        self._clear_analysis_outputs()
        total_supports = sum(row.quantity for row in self._project_rows if row.enabled)
        self.statusBar().showMessage(f"已更新流水號/組數，啟用項目合計 {total_supports} 組")

    def _focus_result_filter(self):
        self.result_filter_input.setFocus()
        self.result_filter_input.selectAll()

    def _set_result_row_group(self, row: int, group_key: str):
        item = self.result_table.item(row, 0)
        if item is None:
            item = QTableWidgetItem("")
            self.result_table.setItem(row, 0, item)
        item.setData(Qt.ItemDataRole.UserRole, group_key)

    def _result_row_group(self, row: int) -> str:
        item = self.result_table.item(row, 0)
        if item is not None:
            group_key = item.data(Qt.ItemDataRole.UserRole)
            if group_key:
                return str(group_key)
        return f"row:{row}"

    def _result_row_text(self, row: int) -> str:
        values = []
        for col in range(self.result_table.columnCount()):
            item = self.result_table.item(row, col)
            if item and item.text():
                values.append(item.text())
        return " ".join(values).casefold()

    def _apply_result_filter(self):
        if not hasattr(self, "result_table"):
            return
        query = self.result_filter_input.text().strip().casefold()
        terms = [term for term in query.split() if term]
        row_count = self.result_table.rowCount()

        if not terms:
            for row in range(row_count):
                self.result_table.setRowHidden(row, False)
            self.result_filter_count_label.setText(f"顯示 {row_count} 列")
            return

        row_groups = [self._result_row_group(row) for row in range(row_count)]
        matched_groups = set()
        for row, group_key in enumerate(row_groups):
            row_text = self._result_row_text(row)
            if all(term in row_text for term in terms):
                matched_groups.add(group_key)

        visible_count = 0
        for row, group_key in enumerate(row_groups):
            is_visible = group_key in matched_groups
            self.result_table.setRowHidden(row, not is_visible)
            if is_visible:
                visible_count += 1

        self.result_filter_count_label.setText(f"顯示 {visible_count}/{row_count} 列")

    def _set_result_summary(
        self,
        *,
        total_weight: float | None = None,
        total_precision: int = 3,
        success_count: int | None = None,
        error_count: int | None = None,
        support_count: int | None = None,
        reset: bool = False,
    ):
        if reset or total_weight is not None:
            text = "-- kg" if total_weight is None else f"{total_weight:.{total_precision}f} kg"
            self.total_weight_label.setText(text)

        if reset or success_count is not None:
            text = "--" if success_count is None else str(success_count)
            self.summary_success_label.setText(text)

        if reset or error_count is not None:
            if error_count is None:
                text = "--"
                color = "#546E7A"
            else:
                text = str(error_count)
                color = "#C62828" if error_count else "#546E7A"
            self.summary_error_label.setText(text)
            self.summary_error_label.setStyleSheet(
                f"color: {color}; background: transparent;"
            )

        if reset or support_count is not None:
            text = "--" if support_count is None else f"{support_count} 組"
            self.summary_support_label.setText(text)

    def _clear_analysis_outputs(self):
        """Clear stale analysis/material outputs after project inputs change."""
        self._results.clear()
        self._project_result = None
        self.result_table.setRowCount(0)
        self._apply_result_filter()
        self.btn_export.setEnabled(False)
        self._set_result_summary(reset=True)
        self.material_cutting_page.set_results_ready(False)
        self.material_cutting_page.clear_outputs()
        self.side_panel.mark_result_stale()

    def _invalidate_analysis_outputs(self, message: str = ""):
        """Invalidate analysis/material outputs after input rows change."""
        had_outputs = (
            bool(self._results)
            or self._project_result is not None
            or self.result_table.rowCount() > 0
        )
        self._clear_analysis_outputs()
        if message:
            suffix = "，已清除舊結果" if had_outputs else ""
            self.statusBar().showMessage(f"{message}{suffix}")

    def _on_item_check_changed(self, item: QListWidgetItem):
        row = self.item_list.row(item)
        if row < 0 or row >= len(self._project_rows):
            return
        enabled = item.checkState() == Qt.CheckState.Checked
        if self._project_rows[row].enabled == enabled:
            return
        self._project_rows[row] = replace(self._project_rows[row], enabled=enabled)
        self._invalidate_analysis_outputs("啟用項目已變更，請重新分析")

    def _on_edit_current_item(self):
        row = self.item_list.currentRow()
        if row < 0 or row >= len(self._project_rows):
            return
        old_designation = self._project_rows[row].designation
        new_designation, accepted = QInputDialog.getText(
            self,
            "編輯支撐編碼",
            "支撐編碼:",
            text=old_designation,
        )
        if not accepted:
            return
        new_designation = new_designation.strip()
        if not new_designation or new_designation == old_designation:
            return
        overrides = self._project_rows[row].overrides
        if get_type_code(new_designation) != get_type_code(old_designation):
            overrides = None
        self._project_rows[row] = replace(
            self._project_rows[row],
            designation=new_designation,
            overrides=overrides,
        )
        self._update_item_display(row)
        self.side_panel.show_item(
            row,
            new_designation,
            self._project_rows[row].overrides or {},
        )
        self._invalidate_analysis_outputs("支撐編碼已編輯，請重新分析")

    def _on_delete_item(self):
        row = self.item_list.currentRow()
        if row < 0:
            return
        deleted_designation = self._project_rows[row].designation
        self.item_list.blockSignals(True)
        self.item_list.takeItem(row)
        self.item_list.blockSignals(False)
        self._project_rows.pop(row)
        self._selected_index = -1
        self._refresh_item_list_display()
        self._invalidate_analysis_outputs("輸入清單已變更，請重新分析")
        if self._project_rows:
            self.item_list.setCurrentRow(min(row, len(self._project_rows) - 1))
        else:
            self.side_panel.clear_panel()
        self.statusBar().showMessage(f"已刪除 {deleted_designation}，請重新分析")

    def _on_clear_all(self):
        if not self._project_rows and not self._results:
            return
        reply = QMessageBox.question(
            self,
            "確認全部清除",
            "確定要清除所有支撐編碼與目前分析結果嗎？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.item_list.clear()
        self._project_rows.clear()
        self._clear_analysis_outputs()
        self._selected_index = -1
        self.side_panel.clear_panel()
        self.statusBar().showMessage("已清除")

    def _on_item_selected(self, row):
        """清單項目被點選 → 更新 Side Panel"""
        if row < 0 or row >= len(self._project_rows):
            self._selected_index = -1
            self.side_panel.clear_panel()
            return
        # 同步 checkbox 狀態
        item_widget = self.item_list.item(row)
        self._project_rows[row] = replace(
            self._project_rows[row],
            enabled=item_widget.checkState() == Qt.CheckState.Checked,
        )
        self._selected_index = row
        project_row = self._project_rows[row]
        self.side_panel.show_item(
            row, project_row.designation, project_row.overrides or {}
        )
        # 若已有分析結果，一併顯示
        if 0 <= row < len(self._results):
            self.side_panel.update_result(self._results[row])

    def _on_override_changed(self, idx: int, overrides: dict):
        """Side Panel 發出覆寫變更"""
        # 移除空值
        clean = {k: v for k, v in overrides.items() if v}
        if 0 <= idx < len(self._project_rows):
            old_clean = self._project_rows[idx].overrides or {}
            if old_clean == clean:
                return
            self._project_rows[idx] = replace(
                self._project_rows[idx],
                overrides=clean or None,
            )
            self._update_item_display(idx)
            self._invalidate_analysis_outputs("覆寫設定已變更，請重新分析")

    # ══════════════════════════════════════════
    #  分析
    # ══════════════════════════════════════════
    def _set_analyze_busy(self, busy: bool):
        self.btn_analyze.setEnabled(not busy)
        self.btn_analyze.setText("分析中..." if busy else "▶ 開始分析")

    def _on_analyze(self):
        if not self._project_rows:
            QMessageBox.warning(self, "提示", "請先新增支撐編碼")
            return

        self._set_analyze_busy(True)
        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        self.statusBar().showMessage("分析中...")
        QApplication.processEvents()
        error_message = ""
        try:
            # 同步 checkbox 狀態
            for i in range(self.item_list.count()):
                w = self.item_list.item(i)
                self._project_rows[i] = replace(
                    self._project_rows[i],
                    enabled=w.checkState() == Qt.CheckState.Checked,
                )

            self._project_result = analyze_project_rows(self._project_rows)
            self._results = [row.scaled_result for row in self._project_result.rows]
            self._display_results()
            self.btn_export.setEnabled(True)

            error_count = sum(1 for r in self._results if r.error)
            success_count = len(self._results) - error_count
            self._set_result_summary(
                success_count=success_count,
                error_count=error_count,
                support_count=self._project_result.total_support_count,
            )
            self.statusBar().showMessage(
                f"分析完成: {len(self._results)} 筆 / "
                f"{self._project_result.total_support_count} 組 "
                f"(成功 {success_count}, 錯誤 {error_count})"
            )

            # 更新 side panel 的計算結果
            if 0 <= self._selected_index < len(self._results):
                self.side_panel.update_result(self._results[self._selected_index])

            # 啟用材料合計 Tab
            self.material_cutting_page.set_results_ready(True)
        except Exception as exc:
            self.statusBar().showMessage("分析失敗")
            error_message = str(exc)
        finally:
            QApplication.restoreOverrideCursor()
            self._set_analyze_busy(False)
        if error_message:
            QMessageBox.critical(self, "分析錯誤", error_message)

    def _display_results(self):
        self.result_table.setRowCount(0)
        total_weight = 0.0

        if self._project_result is not None:
            self._display_project_results()
            return

        for result_index, result in enumerate(self._results):
            group_key = f"legacy:{result_index}"
            if result.error:
                row = self.result_table.rowCount()
                self.result_table.insertRow(row)
                desc = self._result_item(result.fullstring)
                desc.setForeground(QColor("red"))
                self.result_table.setItem(row, 3, desc)
                err = self._result_item(f"錯誤: {result.error}")
                err.setForeground(QColor("red"))
                self.result_table.setItem(row, 6, err)
                self._set_result_row_group(row, group_key)
                continue

            for entry in result.entries:
                row = self.result_table.rowCount()
                self.result_table.insertRow(row)
                values = [
                    "",
                    "1",
                    "組",
                    result.fullstring if entry.item_no == 1 else "",
                    get_type_code(result.fullstring) if entry.item_no == 1 else "",
                    str(entry.item_no),
                    entry.name,
                    entry.display_spec,
                    entry.material,
                    str(entry.length) if entry.length else "",
                    str(entry.width) if entry.width else "",
                    str(entry.quantity),
                    str(entry.quantity),
                    f"{entry.weight_output:.3f}",
                    f"{entry.weight_output:.3f}",
                    entry.category,
                    entry.item_class,
                    entry.manufacturing_type,
                    entry.display_remark,
                    entry.part_key,
                    entry.stock_id,
                ]
                for col, value in enumerate(values):
                    self.result_table.setItem(row, col, self._result_item(value))
                self._set_result_row_group(row, group_key)
                total_weight += entry.weight_output

        self._set_result_summary(total_weight=total_weight, total_precision=2)
        self._apply_result_filter()

    def _display_project_results(self):
        """Display project results in a flat layout with visual grouping."""
        # 數字欄 (右對齊): 長度/寬度、數量、重量
        RIGHT_ALIGN = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        CENTER = Qt.AlignmentFlag.AlignCenter
        LEFT = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        # col_idx → alignment
        COL_ALIGN = {
            0: LEFT, 1: CENTER, 2: CENTER, 3: LEFT, 4: CENTER, 5: CENTER,
            6: LEFT, 7: LEFT, 8: CENTER, 9: RIGHT_ALIGN, 10: RIGHT_ALIGN,
            11: CENTER, 12: CENTER, 13: RIGHT_ALIGN, 14: RIGHT_ALIGN,
            15: CENTER, 16: CENTER, 17: CENTER, 18: LEFT, 19: LEFT, 20: LEFT,
        }

        total_weight = 0.0
        g_idx = 0  # 群組色輪 index

        for result_index, row_result in enumerate(self._project_result.rows):
            input_row = row_result.input_row
            single_result = row_result.single_result
            scaled_result = row_result.scaled_result
            group_key = f"project:{result_index}"
            hdr_color, body_color = _RESULT_GROUP_COLORS[g_idx % len(_RESULT_GROUP_COLORS)]
            g_idx += 1

            if single_result.error:
                row = self.result_table.rowCount()
                self.result_table.insertRow(row)
                self.result_table.setRowHeight(row, 24)
                err_bg = QColor("#FDE8E8")
                err_fg = QColor("#C62828")
                for col in range(self.result_table.columnCount()):
                    cell = QTableWidgetItem()
                    cell.setBackground(err_bg)
                    self.result_table.setItem(row, col, cell)
                desc = self.result_table.item(row, 3)
                desc.setText(input_row.designation)
                desc.setForeground(err_fg)
                desc.setFont(QFont("Microsoft JhengHei UI", 10, QFont.Weight.Bold))
                self.result_table.item(row, 0).setText(input_row.serial)
                self.result_table.item(row, 1).setText(str(input_row.quantity))
                self.result_table.item(row, 2).setText(input_row.unit or "組")
                self.result_table.item(row, 4).setText(get_type_code(input_row.designation))
                err_cell = self.result_table.item(row, 6)
                err_cell.setText(f"⚠ {single_result.error}")
                err_cell.setForeground(err_fg)
                self._set_result_row_group(row, group_key)
                continue

            is_first = True
            group_weight = 0.0
            group_start_row = self.result_table.rowCount()

            for single_entry, scaled_entry in zip(single_result.entries, scaled_result.entries):
                row = self.result_table.rowCount()
                self.result_table.insertRow(row)
                bg = QColor(hdr_color if is_first else body_color)

                values = [
                    input_row.serial if is_first else "",                # 0 流水號.sort
                    str(input_row.quantity) if is_first else "",         # 1 數量
                    (input_row.unit or "組") if is_first else "",        # 2 單位
                    input_row.designation if is_first else "",           # 3 型號
                    get_type_code(input_row.designation) if is_first else "",  # 4 Type
                    str(single_entry.item_no),                            # 5 項次
                    single_entry.name,                                    # 6 品名
                    single_entry.display_spec,                            # 7 規格
                    single_entry.material,                                # 8 材質
                    str(single_entry.length) if single_entry.length else "",  # 9 長度
                    str(single_entry.width)  if single_entry.width  else "",  # 10 寬度
                    str(single_entry.quantity),                           # 11 單件數量
                    str(scaled_entry.quantity),                           # 12 總數量
                    f"{single_entry.weight_output:.3f}",                  # 13 單組重
                    f"{scaled_entry.weight_output:.3f}",                  # 14 總重
                    single_entry.category,                                # 15 屬性
                    single_entry.item_class,                              # 16 物件類別
                    single_entry.manufacturing_type,                      # 17 製造方式
                    single_entry.display_remark,                          # 18 計算說明
                    single_entry.part_key,                                # 19 零件ID
                    single_entry.stock_id,                                # 20 庫存ID
                ]
                for col, val in enumerate(values):
                    item = self._result_item(val)
                    item.setBackground(bg)
                    item.setTextAlignment(COL_ALIGN.get(col, LEFT))
                    if col == 3 and is_first:
                        item.setFont(QFont("Microsoft JhengHei UI", 10, QFont.Weight.Bold))
                        item.setForeground(QColor("#1A3A6B"))
                    elif col == 14:
                        item.setFont(QFont("Microsoft JhengHei UI", 10, QFont.Weight.Bold))
                    self.result_table.setItem(row, col, item)
                self._set_result_row_group(row, group_key)

                group_weight += scaled_entry.weight_output
                total_weight += scaled_entry.weight_output
                is_first = False

            if group_start_row < self.result_table.rowCount():
                self.result_table.setRowHeight(group_start_row, 24)

            # ── 群組小計列 ─────────────────────────────────────
            sub_row = self.result_table.rowCount()
            self.result_table.insertRow(sub_row)
            sub_bg = QColor(hdr_color).darker(104)
            source_label = f"{input_row.serial} | {input_row.designation}" if input_row.serial else input_row.designation
            sub_label = QTableWidgetItem(
                f"小計  {source_label} ({input_row.quantity}{input_row.unit or '組'})"
            )
            sub_label.setBackground(sub_bg)
            sub_label.setForeground(QColor("#1A3A6B"))
            sub_label.setFont(QFont("Microsoft JhengHei UI", 10, QFont.Weight.Bold))
            sub_label.setTextAlignment(LEFT)
            self.result_table.setItem(sub_row, 0, sub_label)
            for col in range(1, self.result_table.columnCount()):
                filler = QTableWidgetItem("")
                filler.setBackground(sub_bg)
                self.result_table.setItem(sub_row, col, filler)
            sub_wt = self.result_table.item(sub_row, 14)
            sub_wt.setText(f"{group_weight:.3f}")
            sub_wt.setTextAlignment(RIGHT_ALIGN)
            sub_wt.setForeground(QColor("#1A3A6B"))
            sub_wt.setFont(QFont("Microsoft JhengHei UI", 10, QFont.Weight.Bold))
            self.result_table.setRowHeight(sub_row, 22)
            self._set_result_row_group(sub_row, group_key)

        self._set_result_summary(total_weight=total_weight, total_precision=3)
        self._apply_result_filter()

    # ══════════════════════════════════════════
    #  匯出 / 設定
    # ══════════════════════════════════════════
    def _set_export_busy(self, busy: bool):
        self.btn_export.setEnabled(False if busy else bool(self._results))
        self.btn_export.setText("匯出中..." if busy else "匯出結果")

    def _on_export(self):
        if not self._results:
            return
        fmt = self.export_format.currentText()
        if "xlsx" in fmt:
            filt, ext = "Excel (*.xlsx)", ".xlsx"
        elif "csv" in fmt:
            filt, ext = "CSV (*.csv)", ".csv"
        else:
            filt, ext = "PDF (*.pdf)", ".pdf"

        filepath, _ = QFileDialog.getSaveFileName(
            self, "匯出結果", f"analysis_result{ext}", filt
        )
        if not filepath:
            return
        self._set_export_busy(True)
        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        self.statusBar().showMessage("匯出中...")
        QApplication.processEvents()
        success_message = ""
        warning_message = ""
        error_message = ""
        try:
            if ext == ".xlsx":
                from export.excel_export import (
                    export_project_workbook,
                    export_to_excel,
                )
                if self._project_result is not None:
                    export_project_workbook(self._project_result, filepath)
                else:
                    export_to_excel(self._results, filepath)
            elif ext == ".csv":
                from export.csv_export import export_to_csv
                export_to_csv(self._results, filepath)
            else:
                from export.pdf_export import export_to_pdf
                export_to_pdf(self._results, filepath)
            success_message = f"已匯出至:\n{filepath}"
            self.statusBar().showMessage(f"已匯出: {filepath}")
        except ImportError as e:
            self.statusBar().showMessage("匯出失敗: 缺少套件")
            warning_message = (
                f"匯出失敗，請安裝必要套件:\n{e}\n\n"
                "pip install openpyxl reportlab"
            )
        except Exception as e:
            self.statusBar().showMessage("匯出失敗")
            error_message = str(e)
        finally:
            QApplication.restoreOverrideCursor()
            self._set_export_busy(False)
        if success_message:
            QMessageBox.information(self, "匯出成功", success_message)
        elif warning_message:
            QMessageBox.warning(self, "缺少套件", warning_message)
        elif error_message:
            QMessageBox.critical(self, "匯出錯誤", error_message)

    def _on_generate_material(self):
        """材料合計 Tab 按下產生按鈕"""
        if not self._results:
            QMessageBox.warning(self, "提示", "請先在重量分析頁完成分析")
            return
        if self._project_result is not None:
            self.material_cutting_page.generate_project(self._project_result)
        else:
            self.material_cutting_page.generate(self._results)
        self.main_tabs.setCurrentWidget(self.material_cutting_page)

    def _on_material_changed(self, text):
        set_analysis_setting("upper_material", text)
        self.statusBar().showMessage(f"全域上段管材質: {text}")

    def _on_open_config(self):
        dialog = ConfigDialog(self)
        dialog.exec()


# ══════════════════════════════════════════════════
#  Side Panel — 單筆項目覆寫設定
# ══════════════════════════════════════════════════
class PdfPreviewScrollArea(QScrollArea):
    zoomRequested = pyqtSignal(float)

    def wheelEvent(self, event):
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            delta_y = event.angleDelta().y()
            if delta_y:
                self.zoomRequested.emit(0.15 if delta_y > 0 else -0.15)
                event.accept()
                return
        super().wheelEvent(event)


class SidePanel(QGroupBox):
    """右側面板：上半 PDF 圖面預覽（可縮放/滑動），下半計算明細與覆寫設定"""

    overrideChanged = pyqtSignal(int, dict)
    _catalog_cache: list = []   # 類別層級快取，避免重複讀檔

    # ── 常用按鈕樣式 ─────────────────────────────────────────
    _ZOOM_BTN = (
        "QPushButton { font-size: 13px; font-weight: bold; background: #EEEEEE; "
        "border: 1px solid #CCC; color: #333; padding: 0; }"
        "QPushButton:hover { background: #BDBDBD; }"
    )
    _ZOOM_FIT_BTN = (
        "QPushButton { font-size: 10px; background: #EEEEEE; "
        "border: 1px solid #CCC; color: #333; padding: 0 6px; }"
        "QPushButton:hover { background: #BDBDBD; }"
    )

    def __init__(self):
        super().__init__("項目設定")
        self._idx = -1
        self._overrides = {}
        self._building = False
        self._preview_pixmap = None
        self._zoom_level = 1.0
        self._current_type_code = ""
        # form widget refs (reset each show_item call)
        self._rb_elbow = None
        self._rb_tee = None
        self._mat_combo = None
        self._pipe_edit = None
        self._sch_edit = None
        self._l_edit = None
        self._result_browser = None
        self._btn_inventor = None
        self._current_result = None
        self._current_designation = ""
        self.setMinimumWidth(280)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(4, 16, 4, 4)
        outer.setSpacing(0)

        # placeholder（未選中時顯示）
        self._placeholder = QLabel("← 點選左側項目\n   以檢視詳情")
        self._placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setStyleSheet("color: #999; font-size: 12px;")
        outer.addWidget(self._placeholder)

        # 主分割器（垂直）
        self._splitter = QSplitter(Qt.Orientation.Vertical)
        self._splitter.setHandleWidth(5)
        self._splitter.setStyleSheet(
            "QSplitter::handle { background: #DDE3EC; }"
        )
        self._splitter.setVisible(False)
        outer.addWidget(self._splitter)

        self._build_pdf_panel()
        self._build_detail_panel()
        self._splitter.setSizes([400, 320])

    # ══════════════════════════════════════════
    #  PDF 預覽面板（上半）
    # ══════════════════════════════════════════
    def _build_pdf_panel(self):
        pane = QWidget()
        pane.setStyleSheet("background: white;")
        vbox = QVBoxLayout(pane)
        vbox.setContentsMargins(2, 2, 2, 2)
        vbox.setSpacing(3)

        # 縮放控制列
        zrow = QHBoxLayout()
        zrow.setSpacing(4)
        lbl = QLabel("圖面預覽")
        lbl.setFont(QFont("Microsoft JhengHei UI", 9, QFont.Weight.Bold))
        lbl.setStyleSheet("color: #555;")
        zrow.addWidget(lbl)
        zrow.addStretch()

        self._btn_zoom_out = QPushButton("－")
        self._btn_zoom_out.setFixedSize(26, 24)
        self._btn_zoom_out.setStyleSheet(self._ZOOM_BTN)

        self._lbl_zoom_pct = QLabel("—")
        self._lbl_zoom_pct.setFont(QFont("Consolas", 9))
        self._lbl_zoom_pct.setStyleSheet("color: #555;")
        self._lbl_zoom_pct.setFixedWidth(44)
        self._lbl_zoom_pct.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self._btn_zoom_in = QPushButton("＋")
        self._btn_zoom_in.setFixedSize(26, 24)
        self._btn_zoom_in.setStyleSheet(self._ZOOM_BTN)

        self._btn_zoom_fit = QPushButton("適合寬度")
        self._btn_zoom_fit.setFixedHeight(24)
        self._btn_zoom_fit.setStyleSheet(self._ZOOM_FIT_BTN)

        zrow.addWidget(self._btn_zoom_out)
        zrow.addWidget(self._lbl_zoom_pct)
        zrow.addWidget(self._btn_zoom_in)
        zrow.addWidget(self._btn_zoom_fit)
        vbox.addLayout(zrow)

        # 可捲動的圖片區
        self._pdf_scroll = PdfPreviewScrollArea()
        self._pdf_scroll.setWidgetResizable(False)
        self._pdf_scroll.setStyleSheet(
            "QScrollArea { background: white; border: 1px solid #E0E0E0; }"
            "QScrollBar:vertical { width: 8px; } QScrollBar:horizontal { height: 8px; }"
        )
        self._lbl_pdf = QLabel("選擇項目後顯示圖面")
        self._lbl_pdf.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lbl_pdf.setMinimumSize(100, 80)
        self._lbl_pdf.setStyleSheet(
            "background: white; color: #AAA; font-size: 11px; padding: 12px;"
        )
        self._pdf_scroll.setWidget(self._lbl_pdf)
        vbox.addWidget(self._pdf_scroll)

        self._btn_zoom_in.clicked.connect(lambda: self._zoom_preview(0.15))
        self._btn_zoom_out.clicked.connect(lambda: self._zoom_preview(-0.15))
        self._btn_zoom_fit.clicked.connect(self._zoom_fit)
        self._pdf_scroll.zoomRequested.connect(self._zoom_preview)

        self._splitter.addWidget(pane)

    # ══════════════════════════════════════════
    #  計算明細 + 覆寫設定面板（下半）
    # ══════════════════════════════════════════
    def _build_detail_panel(self):
        pane = QWidget()
        pane.setStyleSheet("background: white;")
        vbox = QVBoxLayout(pane)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(0)

        self._detail_scroll = QScrollArea()
        self._detail_scroll.setWidgetResizable(True)
        self._detail_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._detail_scroll.setStyleSheet(
            "QScrollArea { background: white; } QScrollBar:vertical { width: 8px; }"
        )
        self._detail_content = QWidget()
        self._detail_content.setStyleSheet("background: white;")
        self._detail_layout = QVBoxLayout(self._detail_content)
        self._detail_layout.setContentsMargins(8, 8, 8, 8)
        self._detail_layout.setSpacing(8)
        self._detail_layout.addStretch()

        self._detail_scroll.setWidget(self._detail_content)
        vbox.addWidget(self._detail_scroll)
        self._splitter.addWidget(pane)

        self._detail_widgets: list = []

    # ══════════════════════════════════════════
    #  PDF 渲染 / 縮放
    # ══════════════════════════════════════════
    def _load_pdf_for_type(self, type_code: str):
        """根據 type_code 找 PDF 並渲染第一頁"""
        self._preview_pixmap = None

        cat_entry = self._get_catalog_entry(type_code)
        pdf_file = cat_entry.get("pdf_file", "") or f"{type_code.zfill(2)}.pdf"

        pdf_path = os.path.join(_PDF_DIR, pdf_file)
        if _FITZ_AVAILABLE and os.path.exists(pdf_path):
            try:
                doc = fitz.open(pdf_path)
                page = doc[0]
                mat = fitz.Matrix(2.5, 2.5)
                pix = page.get_pixmap(matrix=mat)
                img = QImage(
                    pix.samples, pix.width, pix.height,
                    pix.stride, QImage.Format.Format_RGB888,
                )
                self._preview_pixmap = QPixmap.fromImage(img)
                doc.close()
                self._zoom_fit()
                return
            except Exception:
                pass

        # fallback
        self._lbl_pdf.clear()
        self._lbl_pdf.setMinimumSize(100, 80)
        self._lbl_pdf.setText(
            f"尚無此 Type 的 PDF\n({pdf_file})"
            if _FITZ_AVAILABLE else
            "PyMuPDF 未安裝，無法顯示 PDF 預覽"
        )
        self._lbl_pdf.setStyleSheet(
            "background: #FAFAFA; color: #AAA; font-size: 11px; padding: 12px;"
        )
        self._lbl_zoom_pct.setText("—")

    def _apply_zoom(self):
        if not self._preview_pixmap:
            return
        new_w = int(self._preview_pixmap.width() * self._zoom_level)
        scaled = self._preview_pixmap.scaledToWidth(
            max(60, new_w), Qt.TransformationMode.SmoothTransformation
        )
        self._lbl_pdf.setPixmap(scaled)
        self._lbl_pdf.resize(scaled.size())
        self._lbl_pdf.setStyleSheet("background: white;")
        self._lbl_zoom_pct.setText(f"{int(self._zoom_level * 100)}%")

    def _zoom_preview(self, delta: float):
        self._zoom_level = max(0.2, min(3.0, self._zoom_level + delta))
        self._apply_zoom()

    def _zoom_fit(self):
        if not self._preview_pixmap:
            return
        avail = self._pdf_scroll.viewport().width() - 4
        if avail <= 0:
            avail = 270
        self._zoom_level = max(0.2, min(3.0, avail / max(1, self._preview_pixmap.width())))
        self._apply_zoom()

    # ══════════════════════════════════════════
    #  型錄查詢
    # ══════════════════════════════════════════
    @classmethod
    def _get_catalog_entry(cls, type_code: str) -> dict:
        if not cls._catalog_cache:
            try:
                cls._catalog_cache = load_catalog()
            except Exception:
                return {}
        tc = type_code.lstrip("0") or "0"
        for entry in cls._catalog_cache:
            tid = entry.get("type_id", "")
            if tid == type_code or tid.lstrip("0") == tc:
                return entry
        return {}

    # ══════════════════════════════════════════
    #  下半動態內容管理
    # ══════════════════════════════════════════
    def _clear_detail(self):
        for w in self._detail_widgets:
            self._detail_layout.removeWidget(w)
            w.deleteLater()
        self._detail_widgets.clear()
        # 重置 form widget 參照
        self._rb_elbow = None
        self._rb_tee = None
        self._mat_combo = None
        self._pipe_edit = None
        self._sch_edit = None
        self._l_edit = None
        self._result_browser = None
        self._btn_inventor = None

    def _add_dw(self, w: QWidget):
        """插入在 stretch 之前"""
        self._detail_layout.insertWidget(self._detail_layout.count() - 1, w)
        self._detail_widgets.append(w)

    def _add_sep(self):
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFrameShadow(QFrame.Shadow.Sunken)
        self._add_dw(sep)

    def _section_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "font-weight: bold; font-size: 11px; color: #1565C0; "
            "padding: 4px 0 2px 0;"
        )
        return lbl

    # ══════════════════════════════════════════
    #  Public API
    # ══════════════════════════════════════════
    def clear_panel(self):
        self._idx = -1
        self._current_type_code = ""
        self._preview_pixmap = None
        self._lbl_pdf.clear()
        self._lbl_pdf.setText("選擇項目後顯示圖面")
        self._lbl_pdf.setStyleSheet(
            "background: white; color: #AAA; font-size: 11px; padding: 12px;"
        )
        self._lbl_zoom_pct.setText("—")
        self._clear_detail()
        self._splitter.setVisible(False)
        self._placeholder.setVisible(True)

    def mark_result_stale(self):
        """Clear only the calculated-detail area when inputs changed."""
        self._current_result = None
        if self._btn_inventor:
            self._btn_inventor.setEnabled(False)
        if self._result_browser is not None:
            self._result_browser.setHtml(
                '<p style="color:#AAA; font-size:10pt;">'
                '（輸入已變更，請重新執行分析）</p>'
            )

    def show_item(self, idx: int, item_text: str, current_overrides: dict):
        self._building = True
        self._idx = idx
        self._overrides = dict(current_overrides)
        self._current_designation = item_text
        self._current_result = None
        self._clear_detail()
        self._placeholder.setVisible(False)
        self._splitter.setVisible(True)

        type_code = get_type_code(item_text)
        self._current_type_code = type_code

        # ── 載入 PDF 預覽 ──
        self._load_pdf_for_type(type_code)

        # ── 標題 ──
        title = QLabel(f"📌 {item_text}")
        title.setWordWrap(True)
        title.setStyleSheet("font-weight: bold; font-size: 13px; padding: 4px 0;")
        self._add_dw(title)

        # ── Type 資訊 ──
        cat = self._get_catalog_entry(type_code)
        config = load_config(type_code.replace("T", ""))
        type_name = (
            cat.get("name_zh")
            or (config.get("name") if config else "")
            or f"Type {type_code}"
        )
        info_lbl = QLabel(f"Type {type_code}  ·  {type_name}")
        info_lbl.setStyleSheet("color: #666; font-size: 11px; padding-bottom: 2px;")
        self._add_dw(info_lbl)

        self._add_sep()

        # ── 覆寫設定 ──
        self._add_dw(self._section_label("覆寫設定"))
        if type_code in ("01", "01T"):
            self._build_type01_form(item_text, type_code, current_overrides, config)
        else:
            self._build_generic_form(current_overrides)

        # ── 還原按鈕 ──
        btn_reset = QPushButton("↩ 還原為 Type 預設")
        btn_reset.clicked.connect(self._on_reset)
        self._add_dw(btn_reset)

        self._add_sep()

        # ── 計算邏輯 ──
        self._add_dw(self._section_label("計算邏輯"))
        calc_logic = cat.get("calc_logic", "")
        if not calc_logic and config:
            calc_logic = config.get("calc_logic", "")
        calc_browser = QTextBrowser()
        calc_browser.setFont(QFont("Consolas", 10))
        calc_browser.setStyleSheet(
            "QTextBrowser { border: 1px solid #E0E0E0; background: #FAFAFA; "
            "color: #212121; padding: 8px; border-radius: 4px; }"
        )
        calc_browser.setMinimumHeight(70)
        calc_browser.setMaximumHeight(220)
        if calc_logic:
            calc_browser.setPlainText(calc_logic)
        else:
            calc_browser.setHtml(
                '<p style="color:#AAA; font-size:10pt;">（尚未填寫計算邏輯）</p>'
            )
        self._add_dw(calc_browser)

        self._add_sep()

        # ── 計算結果（等候 update_result 填入）──
        self._add_dw(self._section_label("計算結果"))
        self._result_browser = QTextBrowser()
        self._result_browser.setFont(QFont("Microsoft JhengHei UI", 10))
        self._result_browser.setStyleSheet(
            "QTextBrowser { border: 1px solid #E0E0E0; background: #F8FFF8; "
            "color: #222; padding: 8px; border-radius: 4px; }"
        )
        self._result_browser.setMinimumHeight(60)
        self._result_browser.setHtml(
            '<p style="color:#AAA; font-size:10pt;">'
            '（尚未計算，請按「▶ 開始分析」）</p>'
        )
        self._add_dw(self._result_browser)

        # ── Inventor 匯出（僅 Pipe Shoe 家族）────────────────────────────────
        from core.pipe_shoe_engine import PIPE_SHOE_TYPE_IDS
        if type_code in PIPE_SHOE_TYPE_IDS:
            self._add_sep()
            self._btn_inventor = QPushButton("📐 匯出 Inventor 參數 (CSV)")
            self._btn_inventor.setEnabled(False)   # 計算完成後才啟用
            self._btn_inventor.setToolTip(
                "執行計算後可匯出 Inventor iLogic 可讀的 CSV 參數檔案"
            )
            self._btn_inventor.setStyleSheet(
                "QPushButton { background: #1565C0; color: white; font-size: 10pt; "
                "padding: 5px 10px; border-radius: 4px; border: none; }"
                "QPushButton:hover  { background: #1976D2; }"
                "QPushButton:disabled { background: #B0BEC5; color: #ECEFF1; }"
            )
            self._btn_inventor.clicked.connect(self._on_export_inventor)
            self._add_dw(self._btn_inventor)

        self._building = False

    def update_result(self, result):
        """分析完成後由 MainWindow 呼叫，填入計算結果"""
        self._current_result = result
        if self._result_browser is None:
            return
        if result is None or result.error:
            msg = result.error if result else "無結果"
            self._result_browser.setHtml(
                f'<p style="color:#D32F2F; font-size:10pt;">錯誤: {msg}</p>'
            )
            if self._btn_inventor:
                self._btn_inventor.setEnabled(False)
            return

        html = (
            '<style>'
            'table { border-collapse: collapse; width: 100%; font-size: 10pt; }'
            'th { background: #E3F0FF; color: #1565C0; padding: 4px 6px;'
            '     border-bottom: 2px solid #90C2F0; text-align: left; }'
            'td { padding: 3px 6px; border-bottom: 1px solid #EEF0F3; vertical-align: top; }'
            'tr:nth-child(even) td { background: #F8FAFB; }'
            '.r { text-align: right; }'
            '.c { text-align: center; }'
            '</style>'
        )
        html += (
            f'<p style="font-weight:bold; color:#333; margin-bottom:6px;">'
            f'總重量: <span style="color:#1565C0;">{result.total_weight:.2f} kg</span></p>'
        )
        html += (
            '<table><tr>'
            '<th>#</th><th>品名</th><th>公式 / 備註</th>'
            '<th>材質</th><th class="r">L(mm)</th>'
            '<th class="c">數</th><th class="r">重(kg)</th>'
            '</tr>'
        )
        for e in result.entries:
            formula = (
                (e.geometry.formula if e.geometry and e.geometry.formula else "")
                or e.remark or "—"
            )
            html += (
                f'<tr>'
                f'<td class="c">{e.item_no}</td>'
                f'<td>{e.name}</td>'
                f'<td style="font-family:Consolas; font-size:9pt;">{formula}</td>'
                f'<td>{e.material}</td>'
                f'<td class="r">{e.length:.0f}</td>'
                f'<td class="c">{e.quantity}</td>'
                f'<td class="r">{e.weight_output:.3f}</td>'
                f'</tr>'
            )
        html += '</table>'
        if result.warnings:
            html += (
                '<p style="color:#E65100; font-size:9pt; margin-top:6px;">⚠ '
                + '<br>'.join(result.warnings)
                + '</p>'
            )
        self._result_browser.setHtml(html)

        # 計算成功後啟用 Inventor 匯出按鈕
        if self._btn_inventor:
            self._btn_inventor.setEnabled(True)

    # ══════════════════════════════════════════
    #  Inventor 參數匯出
    # ══════════════════════════════════════════
    def _on_export_inventor(self):
        """匯出 Pipe Shoe 計算結果為 Inventor iLogic CSV 參數檔案。"""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from core.pipe_shoe_engine import PIPE_SHOE_TYPE_IDS
        from export.inventor_params import (
            export_ilogic_snippet,
            export_to_csv,
            extract_params,
        )

        designation = self._current_designation
        type_code = self._current_type_code

        if type_code not in PIPE_SHOE_TYPE_IDS:
            return

        default_stem = designation.replace("/", "_").replace("-", "_")
        csv_path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出 Inventor 參數 CSV",
            f"{default_stem}_inventor.csv",
            "CSV 檔案 (*.csv)",
        )
        if not csv_path:
            return

        try:
            params = extract_params(designation, type_code)
            if params is None:
                QMessageBox.warning(self, "匯出失敗", "無法取得計算參數，請先執行分析。")
                return

            export_to_csv(params, csv_path)

            vb_path = os.path.splitext(csv_path)[0] + "_iLogic_LoadParams.vb"
            export_ilogic_snippet(vb_path, csv_path)

            warn_html = (
                "<br><br><b>⚠ 計算警告：</b><br>" + "<br>".join(params["warnings"])
                if params["warnings"] else ""
            )
            QMessageBox.information(
                self,
                "匯出完成",
                f"<b>CSV 參數檔案已儲存：</b><br>{csv_path}"
                f"<br><br><b>iLogic 讀取範本：</b><br>{vb_path}"
                f"<br><br>在 Inventor 中開啟 iLogic Rule Editor，"
                f"貼入 .vb 檔的程式碼後執行即可套用參數。"
                f"{warn_html}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "匯出錯誤", str(exc))

    # ══════════════════════════════════════════
    #  覆寫設定表單
    # ══════════════════════════════════════════
    def _build_type01_form(self, item_text, type_code, overrides, config):
        # 接入方式
        conn_group = QGroupBox("接入方式")
        conn_lay = QHBoxLayout(conn_group)
        self._rb_elbow = QRadioButton("Elbow (彎頭)")
        self._rb_tee = QRadioButton("Tee (三通)")
        effective_conn = overrides.get("connection")
        if effective_conn == "tee" or (not effective_conn and type_code == "01T"):
            self._rb_tee.setChecked(True)
        else:
            self._rb_elbow.setChecked(True)
        self._rb_elbow.toggled.connect(self._on_field_changed)
        self._rb_tee.toggled.connect(self._on_field_changed)
        conn_lay.addWidget(self._rb_elbow)
        conn_lay.addWidget(self._rb_tee)
        self._add_dw(conn_group)

        # 材質
        mat_group = QGroupBox("上段管材質")
        mat_lay = QHBoxLayout(mat_group)
        self._mat_combo = QComboBox()
        self._mat_combo.addItems([
            "", "SUS304", "SUS316", "A53Gr.B", "A106Gr.B",
            "A335-P11", "A335-P22", "A312-TP304", "A312-TP316",
        ])
        self._mat_combo.setEditable(True)
        self._mat_combo.setCurrentText(overrides.get("upper_material", ""))
        self._mat_combo.currentTextChanged.connect(self._on_field_changed)
        hint = QLabel("空白=跟隨全域")
        hint.setStyleSheet("color: #999; font-size: 10px;")
        mat_lay.addWidget(self._mat_combo)
        mat_lay.addWidget(hint)
        self._add_dw(mat_group)

        # 查表值覆寫
        part2 = get_part(item_text, 2)
        try:
            line_size = int(get_lookup_value(part2))
        except (ValueError, TypeError):
            line_size = 0
        cfg_table = get_type_table_as_dict("01") or {}
        row_data = cfg_table.get(line_size, {})

        table_group = QGroupBox("查表值 (留空=用 Config 預設)")
        form = QFormLayout(table_group)

        self._pipe_edit = QLineEdit(overrides.get("pipe_size", ""))
        self._pipe_edit.setPlaceholderText(f"預設: {row_data.get('pipe_size', '?')}")
        self._pipe_edit.textChanged.connect(self._on_field_changed)
        form.addRow("支撐管徑:", self._pipe_edit)

        self._sch_edit = QLineEdit(overrides.get("schedule", ""))
        self._sch_edit.setPlaceholderText(f"預設: {row_data.get('schedule', '?')}")
        self._sch_edit.textChanged.connect(self._on_field_changed)
        form.addRow("Schedule:", self._sch_edit)

        self._l_edit = QLineEdit(
            str(overrides["l_value"]) if overrides.get("l_value") else ""
        )
        self._l_edit.setPlaceholderText(f"預設: {row_data.get('L', '?')}")
        self._l_edit.textChanged.connect(self._on_field_changed)
        form.addRow("L 值 (mm):", self._l_edit)

        part3 = get_part(item_text, 3) or ""
        if part3:
            letter = part3[-1] if part3[-1].isalpha() else ""
            h_code = part3[:-1] if letter else part3
            try:
                h_mm = int(h_code) * 100
            except ValueError:
                h_mm = 0
            form.addRow("H 高度:", QLabel(f"{h_mm} mm"))
            form.addRow("M42 底板:", QLabel(f"代碼 {letter}" if letter else "無"))

        self._add_dw(table_group)

    def _build_generic_form(self, overrides):
        mat_group = QGroupBox("上段管材質")
        mat_lay = QHBoxLayout(mat_group)
        self._mat_combo = QComboBox()
        self._mat_combo.addItems([
            "", "SUS304", "SUS316", "A53Gr.B", "A106Gr.B",
            "A335-P11", "A335-P22", "A312-TP304", "A312-TP316",
        ])
        self._mat_combo.setEditable(True)
        self._mat_combo.setCurrentText(overrides.get("upper_material", ""))
        self._mat_combo.currentTextChanged.connect(self._on_field_changed)
        hint = QLabel("空白=跟隨全域")
        hint.setStyleSheet("color: #999; font-size: 10px;")
        mat_lay.addWidget(self._mat_combo)
        mat_lay.addWidget(hint)
        self._add_dw(mat_group)

    # ══════════════════════════════════════════
    #  欄位變更 / 還原
    # ══════════════════════════════════════════
    def _on_field_changed(self):
        if self._building or self._idx < 0:
            return
        overrides = {}
        if self._rb_tee is not None:
            overrides["connection"] = "tee" if self._rb_tee.isChecked() else "elbow"
        if self._mat_combo is not None:
            mat = self._mat_combo.currentText().strip()
            if mat:
                overrides["upper_material"] = mat
        if self._pipe_edit is not None:
            v = self._pipe_edit.text().strip()
            if v:
                overrides["pipe_size"] = v
        if self._sch_edit is not None:
            v = self._sch_edit.text().strip()
            if v:
                overrides["schedule"] = v
        if self._l_edit is not None:
            v = self._l_edit.text().strip()
            if v:
                try:
                    overrides["l_value"] = int(v)
                except ValueError:
                    pass
        self._overrides = overrides
        self.overrideChanged.emit(self._idx, overrides)

    def _on_reset(self):
        if self._idx < 0:
            return
        self._overrides = {}
        self.overrideChanged.emit(self._idx, {})
        parent = self.parent()
        if parent and hasattr(parent, "parent"):
            mw = parent.parent()
            if hasattr(mw, "_project_rows") and self._idx < len(mw._project_rows):
                row = mw._project_rows[self._idx]
                self.show_item(self._idx, row.designation, {})


# ══════════════════════════════════════════════════
#  Config 管理對話框
# ══════════════════════════════════════════════════
class ConfigDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Type 資料管理")
        self.setMinimumSize(800, 500)
        self._current_config = None
        self._current_type_id = None
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("選擇 Type:"))
        self.type_combo = QComboBox()
        self._load_config_list()
        self.type_combo.currentTextChanged.connect(self._on_type_selected)
        top_row.addWidget(self.type_combo)
        top_row.addStretch()
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color: #666;")
        top_row.addWidget(self.info_label)
        layout.addLayout(top_row)

        self.notes_label = QLabel("")
        self.notes_label.setWordWrap(True)
        self.notes_label.setStyleSheet(
            "background: #FFF9C4; padding: 8px; border-radius: 4px; font-size: 11px;"
        )
        layout.addWidget(self.notes_label)

        self.config_table = QTableWidget()
        self.config_table.setAlternatingRowColors(True)
        layout.addWidget(self.config_table)

        btn_row = QHBoxLayout()
        self.btn_save = QPushButton("💾 儲存變更")
        self.btn_save.clicked.connect(self._on_save)
        self.btn_save.setEnabled(False)
        btn_row.addWidget(self.btn_save)
        self.btn_add_row = QPushButton("+ 新增列")
        self.btn_add_row.clicked.connect(lambda: self.config_table.insertRow(
            self.config_table.rowCount()))
        btn_row.addWidget(self.btn_add_row)
        self.btn_del_row = QPushButton("- 刪除列")
        self.btn_del_row.clicked.connect(lambda: self.config_table.removeRow(
            self.config_table.currentRow()) if self.config_table.currentRow() >= 0 else None)
        btn_row.addWidget(self.btn_del_row)
        btn_row.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        if self.type_combo.count() > 0:
            self._on_type_selected(self.type_combo.currentText())

    def _load_config_list(self):
        from core.config_loader import list_configs
        for c in list_configs():
            self.type_combo.addItem(
                f"Type {c['type_id']} - {c['name']}", c["type_id"]
            )

    def _on_type_selected(self, text):
        type_id = self.type_combo.currentData()
        if not type_id:
            return
        self._current_config = load_config(type_id)
        self._current_type_id = type_id
        if not self._current_config:
            return
        cfg = self._current_config
        self.info_label.setText(
            f"圖號: {cfg.get('drawing_no', '?')} | "
            f"適用: {cfg.get('applicable_range', '?')} | "
            f"v{cfg.get('version', '?')}"
        )
        notes = cfg.get("notes", [])
        self.notes_label.setText("📌 " + "\n📌 ".join(notes) if notes else "")
        self._display_table(cfg.get("table", []))
        self.btn_save.setEnabled(True)
        self.config_table.cellChanged.connect(
            lambda: self.btn_save.setStyleSheet(
                "QPushButton { background-color: #FF9800; color: white; }"
            )
        )

    def _display_table(self, table):
        if not table:
            self.config_table.setRowCount(0)
            return
        headers = list(table[0].keys())
        self.config_table.blockSignals(True)
        self.config_table.setColumnCount(len(headers))
        self.config_table.setHorizontalHeaderLabels(headers)
        self.config_table.setRowCount(len(table))
        for r, row in enumerate(table):
            for c, key in enumerate(headers):
                self.config_table.setItem(r, c, QTableWidgetItem(str(row.get(key, ""))))
        self.config_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.config_table.blockSignals(False)

    def _on_save(self):
        if not self._current_config:
            return
        table = self._current_config.get("table", [])
        if not table:
            return
        headers = list(table[0].keys())
        new_table = []
        for r in range(self.config_table.rowCount()):
            row_data = {}
            for c, key in enumerate(headers):
                item = self.config_table.item(r, c)
                val = item.text() if item else ""
                try:
                    val = float(val) if "." in val else int(val)
                except ValueError:
                    pass
                row_data[key] = val
            new_table.append(row_data)
        self._current_config["table"] = new_table

        from core.config_loader import save_config
        save_config(self._current_type_id, self._current_config,
                    "GUI 手動修改資料表")
        self.btn_save.setStyleSheet("")
        QMessageBox.information(self, "已儲存",
                                f"Type {self._current_type_id} 設定已儲存")
