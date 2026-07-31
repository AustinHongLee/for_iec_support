import openpyxl

from core.project_aggregation import ProjectInputRow, analyze_project_rows
from export.excel_export import (
    FULL_WORKBOOK_SHEETS,
    WORKBOOK_PACKAGE_PROFILES,
    export_project_workbook_package,
)
from export.excel.workbook import build_project_workbook


def _sample_project():
    return analyze_project_rows(
        [
            ProjectInputRow("51-1.1/2B", 10, drawing_line_number="DL-001", serial="S-001"),
            ProjectInputRow("54-10B-A-150-250", 3, drawing_line_number="DL-002", serial="S-002"),
        ]
    )


def _internal_link_sheet(target: str) -> str | None:
    if not target.startswith("#'") or "'!" not in target:
        return None
    return target[2:].split("'!", 1)[0]


def test_project_workbook_package_exports_role_focused_files(tmp_path):
    exported = export_project_workbook_package(_sample_project(), tmp_path)

    assert tuple(exported) == tuple(WORKBOOK_PACKAGE_PROFILES)
    assert set(path.name for path in exported.values()) == {
        "完整活頁簿.xlsx",
        "長官業主包.xlsx",
        "工程明細包.xlsx",
        "採購材料包.xlsx",
        "下料製造包.xlsx",
    }

    for label, expected_sheets in WORKBOOK_PACKAGE_PROFILES.items():
        path = exported[label]
        assert path.exists(), label
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            assert tuple(wb.sheetnames) == expected_sheets
        finally:
            wb.close()


def test_complete_workbook_package_keeps_full_sheet_order(tmp_path):
    exported = export_project_workbook_package(_sample_project(), tmp_path)
    wb = openpyxl.load_workbook(exported["完整活頁簿"], read_only=True, data_only=True)
    try:
        assert tuple(wb.sheetnames) == FULL_WORKBOOK_SHEETS
    finally:
        wb.close()


def test_package_internal_links_only_target_included_sheets(tmp_path):
    exported = export_project_workbook_package(_sample_project(), tmp_path)

    for label, path in exported.items():
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            sheet_names = set(wb.sheetnames)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if not cell.hyperlink:
                            continue
                        sheet = _internal_link_sheet(cell.hyperlink.target)
                        if sheet:
                            assert sheet in sheet_names, f"{label}: {ws.title}!{cell.coordinate} -> {sheet}"
        finally:
            wb.close()


def test_claim_summary_drills_into_reconciled_weight_sources():
    project = analyze_project_rows(
        [
                ProjectInputRow("10-6B-14A", 2, drawing_line_number="DL-010", serial="10"),
            ProjectInputRow("15-8B-1040", 1, drawing_line_number="DL-015", serial="15"),
        ]
    )
    workbook = build_project_workbook(project, ("長官-支撐分類", "查核-支撐明細"))
    summary = workbook["長官-支撐分類"]
    evidence = workbook["查核-支撐明細"]
    contract = "CS(熱鍍鋅)管支撐(Pipe Support)製裝>15Kg"

    summary_row = next(
        row for row in range(1, summary.max_row + 1)
        if summary.cell(row, 2).value == contract
    )
    evidence_rows = [
        row for row in range(4, evidence.max_row + 1)
        if evidence.cell(row, 1).value == contract
    ]
    expected_total = round(sum(row.scaled_result.total_weight for row in project.rows), 3)

    assert summary.cell(summary_row, 4).value == expected_total
    assert summary.cell(summary_row, 8).value == "查看 2 筆來源"
    assert summary.cell(summary_row, 8).hyperlink.target == f"#'查核-支撐明細'!A{evidence_rows[0]}"
    assert evidence_rows == list(range(evidence_rows[0], evidence_rows[-1] + 1))
    assert {evidence.cell(row, 2).value for row in evidence_rows} == {expected_total}
    assert round(sum(evidence.cell(row, 11).value for row in evidence_rows), 3) == expected_total
    for row, project_row in zip(evidence_rows, project.rows):
        assert evidence.cell(row, 8).value == round(project_row.single_result.total_weight, 3)
        assert evidence.cell(row, 11).value == round(project_row.scaled_result.total_weight, 3)
        assert "kg/組 ×" in evidence.cell(row, 10).value
