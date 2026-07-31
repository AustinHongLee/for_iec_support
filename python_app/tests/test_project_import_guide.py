import os
import sys

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook, load_workbook
from PyQt6.QtWidgets import QApplication, QLabel

from core.project_aggregation import ProjectInputRow
from core.project_import import (
    PROJECT_IMPORT_TEMPLATE_HEADERS,
    read_project_rows_xlsx,
    write_project_import_template,
)
from ui.project_import_dialog import (
    ProjectImportGuideDialog,
    ProjectImportPreviewDialog,
    summarize_project_rows,
)


_APP = QApplication.instance() or QApplication(sys.argv)


def test_blank_import_template_documents_required_and_optional_columns(tmp_path):
    path = tmp_path / "project_import_template.xlsx"
    write_project_import_template(str(path))

    wb = load_workbook(path)
    try:
        ws = wb["支撐清單"]
        headers = tuple(cell.value for cell in ws[1])
        assert headers == PROJECT_IMPORT_TEMPLATE_HEADERS
        assert ws.max_row == 1
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == "A1:H1"
        assert len(ws.data_validations.dataValidation) == 3
        assert ws["C1"].comment is not None
        assert ws["E1"].comment is not None
        assert ws["H1"].comment is not None
        assert "必要欄" in wb["填寫說明"]["A2"].value
    finally:
        wb.close()


def test_import_guide_explains_sources_required_fields_and_penetration_holes():
    dialog = ProjectImportGuideDialog()
    try:
        text = " ".join(label.text() for label in dialog.findChildren(QLabel))
        assert "Support MTO Excel" in text
        assert "型號" in text
        assert "數量" in text
        assert "PENETRATION HOLE" in text
        assert "Drawing line number" in text
    finally:
        dialog.close()


def test_import_preview_calls_out_missing_traceability():
    rows = [
        ProjectInputRow(
            designation="01-2B-05A",
            quantity=1,
            drawing_line_number="DL-001",
            serial="1",
        ),
        ProjectInputRow(designation="51-1.1/2B", quantity=2),
    ]
    summary, warnings = summarize_project_rows(rows)
    assert "Drawing：1/2" in summary
    assert "流水號：1/2" in summary
    assert any("無法依原始圖面追溯" in warning for warning in warnings)

    dialog = ProjectImportPreviewDialog("support.xlsx", rows, existing_count=5)
    try:
        assert dialog.replace_existing
        assert "1 筆沒有 Drawing" in dialog.warning_label.text()
        dialog.append_radio.setChecked(True)
        assert not dialog.replace_existing
    finally:
        dialog.close()


def test_import_preview_flags_penetration_hole_without_size():
    rows = [ProjectInputRow(designation="PENETRATION HOLE", quantity=1)]
    _, warnings = summarize_project_rows(rows)
    assert any("缺少管徑" in warning for warning in warnings)


def test_import_preview_reports_skipped_and_defaulted_source_rows():
    rows = [ProjectInputRow(designation="01-2B-05A", quantity=1)]
    _, warnings = summarize_project_rows(
        rows,
        import_report={
            "skipped_missing_designation": 2,
            "quantity_defaulted": 1,
            "unit_defaulted": 1,
        },
    )
    text = " ".join(warnings)
    assert "2 列找不到型號" in text
    assert "1 筆數量空白" in text
    assert "1 筆單位空白" in text


def test_xlsx_import_report_exposes_skipped_and_defaulted_rows(tmp_path):
    path = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Support MTO"
    ws.append(["型號", "數量", "單位", "Drawing line number"])
    ws.append(["01-2B-05A", "", "", "DL-001"])
    ws.append(["", 3, "組", "DL-002"])
    ws.append(["57-1B-A", 1.5, "組", "DL-003"])
    wb.save(path)

    report = {}
    rows = read_project_rows_xlsx(str(path), report=report)

    assert len(rows) == 1
    assert rows[0].quantity == 1
    assert report["source_rows"] == 3
    assert report["skipped_missing_designation"] == 1
    assert report["skipped_invalid_quantity"] == 1
    assert report["quantity_defaulted"] == 1
    assert report["unit_defaulted"] == 1
    problems = report["problems"]
    assert any(
        problem["row"] == 3
        and problem["field"] == "型號"
        and "此列不會匯入" in problem["issue"]
        and "DL-002" in problem["raw"]
        for problem in problems
    )
    assert any(
        problem["row"] == 4
        and problem["field"] == "數量"
        and "1.5" in problem["issue"]
        and "大於 0 的整數" in problem["resolution"]
        for problem in problems
    )


def test_import_preview_shows_and_copies_source_row_problem_details():
    rows = [ProjectInputRow(designation="01-2B-05A", quantity=1)]
    report = {
        "source_rows": 2,
        "skipped_missing_designation": 1,
        "problems": [
            {
                "row": 7,
                "severity": "error",
                "field": "型號",
                "issue": "找不到可辨識的支撐型號；此列不會匯入",
                "raw": "Drawing=DL-007；數量=2",
                "resolution": "在型號欄填入例如 57-1B-A。",
            }
        ],
    }
    dialog = ProjectImportPreviewDialog(
        "support.xlsx", rows, import_report=report
    )
    try:
        assert "原檔資料列：2" in dialog.summary_label.text()
        assert "不匯入：1" in dialog.summary_label.text()
        assert dialog.problem_table.rowCount() == 1
        assert dialog.problem_table.item(0, 0).text() == "7"
        assert "型號" in dialog.problem_table.item(0, 2).text()
        assert "DL-007" in dialog.problem_table.item(0, 3).text()
        assert "57-1B-A" in dialog.problem_table.item(0, 4).text()

        dialog.copy_problems_button.click()
        copied = _APP.clipboard().text()
        assert "原檔列\t程度\t欄位" in copied
        assert "DL-007" in copied
        assert dialog.copy_problems_button.text() == "已複製"
    finally:
        dialog.close()
