import os
import tempfile

import openpyxl

from core.models import AnalysisResult
from core.project_aggregation import (
    ProjectAnalysisResult,
    ProjectInputRow,
    ProjectRowResult,
    analyze_project_rows,
)
from export.excel.confidence_summary import (
    format_confidence_counts,
    project_confidence_counts,
    review_required_count,
    worst_confidence_level,
)
from export.excel_export import export_project_workbook


def _row(level: str, *, error: str = "", requires_review: bool = False) -> ProjectRowResult:
    single = AnalysisResult(
        fullstring=f"{level}-sample",
        meta={"truth_level": level, "requires_review": requires_review},
    )
    single.error = error
    scaled = AnalysisResult(fullstring=single.fullstring, meta=single.meta.copy())
    return ProjectRowResult(
        input_row=ProjectInputRow(single.fullstring, 1),
        single_result=single,
        scaled_result=scaled,
    )


def test_project_confidence_counts_use_existing_meta_and_errors():
    project = ProjectAnalysisResult(
        rows=[
            _row("精確"),
            _row("推導"),
            _row("估算", requires_review=True),
            _row("未知", requires_review=True),
            _row("精確", error="broken"),
        ]
    )

    counts = project_confidence_counts(project)

    assert counts == {"精確": 1, "推導": 1, "估算": 1, "未知": 1, "錯誤": 1}
    assert worst_confidence_level(counts) == "錯誤"
    assert review_required_count(project) == 3
    assert format_confidence_counts(counts) == "精確 1 / 推導 1 / 估算 1 / 未知 1 / 錯誤 1"


def test_project_workbook_keeps_manager_claim_focused_and_confidence_in_engineering_sheets():
    project = analyze_project_rows(
        [
            ProjectInputRow("51-1.1/2B", 10, drawing_line_number="DL-001", serial="S-001"),
            ProjectInputRow("72-2B", 2, drawing_line_number="DL-002", serial="S-002"),
        ]
    )
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        export_project_workbook(project, path)
        wb = openpyxl.load_workbook(path, data_only=True)

        ws_manager = wb["長官-摘要"]
        manager_values = [cell.value for row in ws_manager.iter_rows() for cell in row]
        assert "資料可信度" not in manager_values
        assert "合約名稱怎麼來的？" in manager_values
        assert "哪些資料不能直接說明？" in manager_values

        ws_summary = wb["專案摘要"]
        headers = [ws_summary.cell(row=15, column=col).value for col in range(1, 10)]
        assert "資料狀態" in headers
        status_col = headers.index("資料狀態") + 1
        status_values = [
            ws_summary.cell(row=row, column=status_col).value
            for row in range(16, ws_summary.max_row + 1)
        ]
        assert any(isinstance(value, str) and ("未知" in value or "估算" in value) for value in status_values)
        assert any(
            ws_summary.cell(row=row, column=status_col).fill.fgColor.rgb not in (None, "00000000", "FFFFFFFF")
            for row in range(16, ws_summary.max_row + 1)
            if ws_summary.cell(row=row, column=status_col).value
        )

        ws_reference = wb["計算標準與假設"]
        reference_values = [
            [cell.value for cell in row]
            for row in ws_reference.iter_rows()
        ]
        header_row = next(row for row in reference_values if "資料版本" in row)
        version_col = header_row.index("資料版本")
        assert any(
            row[version_col] not in (None, "")
            for row in reference_values[reference_values.index(header_row) + 1 :]
            if len(row) > version_col
        )
    finally:
        try:
            os.remove(path)
        except OSError:
            pass
