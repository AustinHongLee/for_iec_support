import openpyxl

from core.calculator import analyze_single
from core.project_aggregation import ProjectInputRow, analyze_project_rows
from core.project_import import read_project_rows_xlsx
from export.excel.workbook import build_project_workbook


def test_penetration_hole_builds_fb50x6_from_pipe_size_and_insulation():
    result = analyze_single(
        "PENETRATION HOLE",
        {"nominal_size": "1", "insulation": ""},
    )

    assert not result.error
    assert result.meta["type_id"] == "PENETRATION HOLE"
    assert len(result.entries) == 1
    entry = result.entries[0]
    assert entry.name == "FB50×6 開孔補強"
    assert entry.spec == "FB50×6"
    assert entry.length == 360
    assert entry.weight_per_unit == 2.355
    assert entry.weight_output == 0.848
    assert "OPEN-1\"" in entry.display_remark
    assert "開孔=90mm" in entry.display_remark


def test_penetration_hole_matches_rev01_insulated_example():
    result = analyze_single(
        "PENETRATION HOLE",
        {"nominal_size": "4", "insulation": "H50"},
    )

    assert not result.error
    entry = result.entries[0]
    assert entry.length == 1080
    assert entry.weight_output == 2.543
    assert "OPEN-4\"-H50" in entry.display_remark
    assert "開孔=270mm" in entry.display_remark


def test_penetration_hole_requires_nominal_size():
    result = analyze_single("PENETRATION HOLE")

    assert "nominal_size" in result.error


def test_support_mto_import_attaches_this_project_opening_rule(tmp_path):
    source = tmp_path / "support_mto.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUPPORT MTO"
    ws.append(["line_group", "流水號", "description", "nominal_size", "quantity", "unit", "保溫厚度"])
    ws.append(["DRAW-009", "9", "PENETRATION HOLE", "1", 1, "組", ""])
    ws.append(["DRAW-010", "10", "PENETRATION HOLE", "4", 2, "組", "H50"])
    wb.save(source)

    rows = read_project_rows_xlsx(str(source))

    assert [row.designation for row in rows] == ["PENETRATION HOLE", "PENETRATION HOLE"]
    assert [row.display_designation for row in rows] == ['OPEN-1"', 'OPEN-4"-H50']
    assert rows[0].overrides == {"nominal_size": "1", "insulation": ""}
    assert rows[1].overrides == {"nominal_size": "4", "insulation": "H50"}


def test_penetration_hole_uses_mto_code_in_weight_analysis_output():
    project = analyze_project_rows(
        [
            ProjectInputRow(
                "PENETRATION HOLE",
                1,
                overrides={"nominal_size": "1", "insulation": ""},
                display_designation='OPEN-1"',
                serial="9",
            )
        ]
    )
    wb = build_project_workbook(
        project,
        ("重量明細表", "單組重量明細", "查核-支撐明細", "重量分析"),
    )

    weight_detail = wb["重量分析"]
    basis = wb["重量明細表"]
    unit_weight = wb["單組重量明細"]
    leader_detail = wb["查核-支撐明細"]
    assert weight_detail["A4"].value == 'OPEN-1"'
    assert weight_detail["D4"].value == "FB50×6 開孔補強"
    assert basis["A4"].value == 'OPEN-1"'
    assert basis["O4"].value == 0.848
    assert unit_weight["A4"].value == 'OPEN-1"'
    assert unit_weight["B3"].value == "單組重量(kg)"
    assert unit_weight["B4"].value == 0.848
    assert unit_weight.max_column == 2
    assert leader_detail["F4"].value == 'OPEN-1"'
    assert leader_detail["H4"].value == 0.848
    assert "單組 0.848 kg ≤ 15 kg" in leader_detail["J4"].value
    assert not any(
        cell.value == "PENETRATION HOLE"
        for row in leader_detail.iter_rows()
        for cell in row
    )


def test_unit_weight_sheet_does_not_multiply_project_groups():
    project = analyze_project_rows(
        [
            ProjectInputRow(
                "PENETRATION HOLE",
                2,
                overrides={"nominal_size": "1", "insulation": ""},
                display_designation='OPEN-1"',
            )
        ]
    )
    wb = build_project_workbook(project, ("重量明細表", "單組重量明細"))

    assert wb["重量明細表"]["O4"].value == 1.696
    assert wb["單組重量明細"]["B4"].value == 0.848
    assert wb["單組重量明細"].max_column == 2
