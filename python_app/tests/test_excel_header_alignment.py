import os
import tempfile

import openpyxl
import pytest

from core.project_aggregation import ProjectInputRow, analyze_project_rows
from export.excel.headers import LEADER_DETAIL_HEADERS, PROJECT_HEADERS, _CALC_BASIS_HEADERS
from export.excel_export import export_project_to_excel, export_project_workbook


def _header_map(ws, row: int) -> dict[str, int]:
    values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
    return {header: index + 1 for index, header in enumerate(values) if header}


def _exported_workbook(project, exporter):
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        exporter(project, path)
        return openpyxl.load_workbook(path, data_only=True)
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


def _sample_project():
    return analyze_project_rows(
        [
            ProjectInputRow(
                "51-1.1/2B",
                10,
                drawing_line_number="DL-001",
                serial="S-001",
            )
        ]
    )


def _multi_type_project():
    return analyze_project_rows(
        [
            ProjectInputRow(
                "51-1.1/2B",
                10,
                drawing_line_number="DL-001",
                serial="S-001",
            ),
            ProjectInputRow(
                "54-10B-A-150-250",
                3,
                drawing_line_number="DL-002",
                serial="S-002",
            ),
        ]
    )


def _row_values_by_header(ws, row: int, headers: list[str]) -> dict:
    columns = _header_map(ws, row=3)
    return {header: ws.cell(row=row, column=columns[header]).value for header in headers}


def _assert_row_values(actual: dict, expected: dict) -> None:
    assert set(actual) == set(expected)
    for key, expected_value in expected.items():
        if isinstance(expected_value, float):
            assert actual[key] == pytest.approx(expected_value, abs=0.000001), key
        else:
            assert actual[key] == expected_value, key


def test_flat_project_export_headers_align_with_values():
    wb = _exported_workbook(_sample_project(), export_project_to_excel)
    ws = wb["Project_Weight_Analysis"]

    assert [ws.cell(row=1, column=col).value for col in range(1, len(PROJECT_HEADERS) + 1)] == PROJECT_HEADERS
    columns = _header_map(ws, 1)
    assert ws.cell(row=2, column=columns["型號"]).value == "51-1.1/2B"
    assert ws.cell(row=2, column=columns["來源圖號"]).value == "DL-001"
    assert ws.cell(row=2, column=columns["流水號"]).value == "S-001"
    assert ws.cell(row=2, column=columns["輸入數量"]).value == 10
    assert ws.cell(row=2, column=columns["輸入單位"]).value == "組"


def test_workbook_project_weight_headers_align_with_values():
    wb = _exported_workbook(_sample_project(), export_project_workbook)
    ws = wb["重量分析"]

    assert [ws.cell(row=3, column=col).value for col in range(1, len(PROJECT_HEADERS) + 1)] == PROJECT_HEADERS
    columns = _header_map(ws, 3)
    assert ws.cell(row=4, column=columns["型號"]).value == "51-1.1/2B"
    assert ws.cell(row=4, column=columns["來源圖號"]).value == "DL-001"
    assert ws.cell(row=4, column=columns["流水號"]).value == "S-001"
    assert ws.cell(row=4, column=columns["輸入數量"]).value == 10
    assert ws.cell(row=4, column=columns["輸入單位"]).value == "組"


def test_calculation_basis_headers_align_with_subtotal_trace_values():
    wb = _exported_workbook(_sample_project(), export_project_workbook)
    ws = wb["重量明細表"]

    assert [ws.cell(row=3, column=col).value for col in range(1, len(_CALC_BASIS_HEADERS) + 1)] == _CALC_BASIS_HEADERS
    columns = _header_map(ws, 3)
    subtotal_row = next(
        row for row in range(4, ws.max_row + 1)
        if ws.cell(row=row, column=columns["列型"]).value == "小計"
    )
    assert ws.cell(row=subtotal_row, column=columns["型號"]).value == "小計 51-1.1/2B"
    assert ws.cell(row=subtotal_row, column=columns["來源圖號"]).value == "DL-001"
    assert ws.cell(row=subtotal_row, column=columns["流水號"]).value == "S-001"
    assert ws.cell(row=subtotal_row, column=columns["輸入數量"]).value == 10
    assert ws.cell(row=subtotal_row, column=columns["輸入單位"]).value == "組"


def test_leader_detail_headers_align_with_trace_values():
    wb = _exported_workbook(_sample_project(), export_project_workbook)
    ws = wb["查核-支撐明細"]

    assert [ws.cell(row=3, column=col).value for col in range(1, len(LEADER_DETAIL_HEADERS) + 1)] == LEADER_DETAIL_HEADERS
    columns = _header_map(ws, 3)
    detail_row = next(
        row for row in range(4, ws.max_row + 1)
        if ws.cell(row=row, column=columns["型號"]).value == "51-1.1/2B"
    )
    assert ws.cell(row=detail_row, column=columns["來源圖號"]).value == "DL-001"
    assert ws.cell(row=detail_row, column=columns["流水號"]).value == "S-001"
    assert ws.cell(row=detail_row, column=columns["數量"]).value == 10
    assert ws.cell(row=detail_row, column=columns["單位"]).value == "組"


def test_calculation_basis_multitype_rows_align_every_exported_column():
    wb = _exported_workbook(_multi_type_project(), export_project_workbook)
    ws = wb["重量明細表"]

    assert [ws.cell(row=3, column=col).value for col in range(1, len(_CALC_BASIS_HEADERS) + 1)] == _CALC_BASIS_HEADERS
    columns = _header_map(ws, 3)
    type54_detail_row = next(
        row for row in range(4, ws.max_row + 1)
        if (
            ws.cell(row=row, column=columns["型號"]).value == "54-10B-A-150-250"
            and ws.cell(row=row, column=columns["項次"]).value == 1
        )
    )
    type54_subtotal_row = next(
        row for row in range(4, ws.max_row + 1)
        if ws.cell(row=row, column=columns["型號"]).value == "小計 54-10B-A-150-250"
    )

    _assert_row_values(
        _row_values_by_header(ws, type54_detail_row, _CALC_BASIS_HEADERS),
        {
            "型號": "54-10B-A-150-250",
            "型號類別": "54",
            "項次": 1,
            "品名": "角鋼",
            "規格": "40*40*5",
            "材質": "A36/SS400",
            "長度(mm)": 150,
            "寬度(mm)": None,
            "屬性": "型鋼類",
            "單件數量": 2,
            "組數": 3,
            "總數量": 6,
            "單件重(kg)": 0.44,
            "單組小計(kg)": 0.88,
            "總重(kg)": 2.64,
            "重量計算式": "2件 × 150mm ÷ 1,000 × 2.960 kg/m = 0.880 kg",
            "物件類別": "primary_structure",
            "製造方式": "raw_cut",
            "列型": "明細",
            "來源圖號": None,
            "流水號": None,
            "輸入數量": None,
            "輸入單位": None,
        },
    )
    _assert_row_values(
        _row_values_by_header(ws, type54_subtotal_row, _CALC_BASIS_HEADERS),
        {
            "型號": "小計 54-10B-A-150-250",
            "型號類別": "54",
            "項次": None,
            "品名": None,
            "規格": None,
            "材質": None,
            "長度(mm)": None,
            "寬度(mm)": None,
            "屬性": None,
            "單件數量": None,
            "組數": 3,
            "總數量": None,
            "單件重(kg)": None,
            "單組小計(kg)": None,
            "總重(kg)": 49.35,
            "重量計算式": None,
            "物件類別": None,
            "製造方式": None,
            "列型": "小計",
            "來源圖號": "DL-002",
            "流水號": "S-002",
            "輸入數量": 3,
            "輸入單位": "組",
        },
    )


def test_leader_detail_pipe_shoe_row_aligns_every_exported_column():
    wb = _exported_workbook(_multi_type_project(), export_project_workbook)
    ws = wb["查核-支撐明細"]

    assert [ws.cell(row=3, column=col).value for col in range(1, len(LEADER_DETAIL_HEADERS) + 1)] == LEADER_DETAIL_HEADERS
    columns = _header_map(ws, 3)
    pipe_shoe_row = next(
        row for row in range(4, ws.max_row + 1)
        if ws.cell(row=row, column=columns["型號"]).value == "54-10B-A-150-250"
    )

    _assert_row_values(
        _row_values_by_header(ws, pipe_shoe_row, LEADER_DETAIL_HEADERS),
        {
            "狀態": "命中",
            "類別": "Pipe Shoe",
            "統計項目": 'PIPE SHOE 5"~10" 熱浸鍍鋅',
            "來源圖號": "DL-002",
            "流水號": "S-002",
            "數量": 3,
            "單位": "組",
            "型號": "54-10B-A-150-250",
            "型號類別": "54",
            "管徑(吋)": 10,
            "計入數量": 3,
            "計入單位": "組",
            "命中明細": 'Type 54 Pipe Shoe，10"',
            "材質判定": "整組材質 -> HDG/CS",
            "統計條件": 'Type 52/53/54/55/66/67/80/85，管徑 5"~10"，整組不含 SUS304',
            "備註": None,
        },
    )
