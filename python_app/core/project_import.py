"""Project input import helpers that do not depend on the GUI layer."""

from __future__ import annotations

import re
from collections.abc import Callable

from .project_aggregation import ProjectInputRow
from .source_profiles import normalize_source_profile, source_profile_choices

ProjectXlsxMapping = dict[str, int | None]
ProjectXlsxMappingConfirmer = Callable[
    [str, int, list[str], ProjectXlsxMapping],
    ProjectXlsxMapping,
]

PROJECT_ROW_ALIASES = {
    "drawing_line_number": (
        "Drawing line number", "drawing_line_number", "drawing line", "drawing", "line_group",
        "line group", "line no", "line number", "dwg", "dwg no", "圖名", "圖號",
    ),
    "serial": ("流水號.sort", "流水號", "serial", "serial_no", "seq", "sort", "序號", "編號"),
    "designation": ("型號", "designation", "support_designation", "support_no", "支撐編碼", "編碼", "支撐型號", "model"),
    "quantity": ("數量", "quantity", "qty", "count", "組數", "支數"),
    "unit": ("單位", "unit", "uom"),
    "enabled": ("enabled", "啟用"),
    "source_profile": (
        "source_profile", "圖面來源", "圖面來源覆寫", "來源設定"
    ),
    "overrides": ("overrides_json", "overrides"),
    "description": ("description", "desc", "描述", "中文說明", "說明", "品名"),
    "item_code": ("item_code", "item code", "料號", "code"),
    "nominal_size": ("nominal_size", "nominal size", "管徑", "管徑(吋)", "size"),
    "insulation": ("保溫厚度", "insulation", "insulation thickness", "insul thk"),
}

PROJECT_XLSX_FIELDS = (
    "designation", "quantity", "drawing_line_number", "serial", "unit", "description", "item_code",
    "nominal_size", "insulation", "source_profile",
)

PROJECT_IMPORT_TEMPLATE_HEADERS = (
    "Drawing line number",
    "流水號.sort",
    "數量",
    "單位",
    "型號",
    "管徑",
    "保溫厚度",
    "圖面來源覆寫",
)


def append_import_problem(
    report: dict | None,
    *,
    row_number: int,
    severity: str,
    field: str,
    issue: str,
    raw: str,
    resolution: str,
) -> None:
    """Append one source-row problem in the shared preview-report format."""
    if report is None:
        return
    report.setdefault("problems", []).append(
        {
            "row": int(row_number),
            "severity": str(severity),
            "field": str(field),
            "issue": str(issue),
            "raw": str(raw),
            "resolution": str(resolution),
        }
    )


def format_import_raw(values, headers=None) -> str:
    """Return a compact, readable snapshot of one original source row."""
    headers = list(headers or [])
    parts = []
    if isinstance(values, dict):
        items = values.items()
    else:
        items = enumerate(values)
    for key, value in items:
        if value is None or not str(value).strip():
            continue
        if isinstance(key, int):
            label = headers[key] if key < len(headers) and headers[key] else f"欄{key + 1}"
        else:
            label = str(key or "欄位")
        text = str(value).strip().replace("\r", " ").replace("\n", " ")
        if len(text) > 80:
            text = text[:77] + "…"
        parts.append(f"{label}={text}")
    result = "；".join(parts) or "（空白列）"
    return result if len(result) <= 500 else result[:497] + "…"


def write_project_import_template(filepath: str) -> str:
    """Create a blank, documented workbook accepted by the project importer."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "支撐清單"
    ws.append(PROJECT_IMPORT_TEMPLATE_HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:H1"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    for column, header in enumerate(PROJECT_IMPORT_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(1, column)
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = required_fill if header in {"數量", "型號"} else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    comments = {
        "Drawing line number": "選填。原始圖面或管線群組，用於追溯來源。",
        "流水號.sort": "選填。專案內排序或識別號。",
        "數量": "必填。正整數。",
        "單位": "選填。空白時使用「組」。",
        "型號": "必填。例如 57-1B-A、01-2B-05A；開孔列填 PENETRATION HOLE。",
        "管徑": "僅 PENETRATION HOLE 必填，例如 4。",
        "保溫厚度": "僅 PENETRATION HOLE 使用；無保溫可留空。",
        "圖面來源覆寫": (
            "選填。一般留空跟隨專案；混用例外才填 "
            + "、".join(profile_id for profile_id, _ in source_profile_choices())
            + "。"
        ),
    }
    for column, header in enumerate(PROJECT_IMPORT_TEMPLATE_HEADERS, start=1):
        ws.cell(1, column).comment = Comment(comments[header], "IEC Support Tool")

    widths = (28, 14, 10, 10, 24, 12, 14, 24)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, column).column_letter].width = width

    quantity_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
    )
    quantity_validation.error = "數量必須是大於 0 的整數"
    quantity_validation.errorTitle = "數量格式錯誤"
    quantity_validation.prompt = "輸入大於 0 的整數"
    quantity_validation.promptTitle = "必要欄位"
    quantity_validation.showErrorMessage = True
    quantity_validation.showInputMessage = True
    ws.add_data_validation(quantity_validation)
    quantity_validation.add("C2:C5000")

    unit_validation = DataValidation(
        type="list",
        formula1='"組,set,kg,ea,pc,m"',
        allow_blank=True,
    )
    ws.add_data_validation(unit_validation)
    unit_validation.add("D2:D5000")

    source_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(
            profile_id for profile_id, _ in source_profile_choices()
        ) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(source_validation)
    source_validation.add("H2:H5000")

    guide = wb.create_sheet("填寫說明")
    guide.column_dimensions["A"].width = 95
    instructions = (
        "使用方式：在「支撐清單」分頁由第 2 列開始填寫，一列代表一筆支撐。",
        "必要欄：型號、數量。黃色表頭代表必要欄位。",
        "建議欄：Drawing line number、流水號.sort、單位；缺少時仍可計算，但不利於追溯與核對。",
        "一般範例：Drawing=43--1-1.2-S11UG-N4-60371，流水號=43，數量=1，單位=組，型號=23-L75-07C-10。",
        "開孔範例：型號=PENETRATION HOLE，並填寫管徑與保溫厚度。",
        "也可以直接匯入原始 Support MTO Excel，程式會先要求確認欄位對應。",
    )
    for row, text in enumerate(instructions, start=1):
        cell = guide.cell(row, 1, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row == 1:
            cell.font = Font(bold=True, color="1F4E78")
        guide.row_dimensions[row].height = 30

    wb.save(filepath)
    return filepath


def read_project_rows_xlsx(
    filepath: str,
    mapping_confirmer: ProjectXlsxMappingConfirmer | None = None,
    report: dict | None = None,
) -> list[ProjectInputRow]:
    """Read project rows from xlsx/xlsm without importing PyQt UI classes."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        layout = detect_project_xlsx_layout(wb)
        if layout is None:
            raise ValueError("找不到可用的 MTO 表頭；請確認 Excel 內有流水號/數量/單位/型號等欄位。")

        ws = layout["worksheet"]
        header_row = layout["header_row"]
        headers = layout["headers"]
        mapping = dict(layout["mapping"])
        if mapping_confirmer is not None:
            mapping = mapping_confirmer(ws.title, header_row, headers, mapping)

        if not has_project_designation_source(mapping):
            raise ValueError("xlsx 匯入至少需要指定「型號」欄，或指定可抽型號的說明/料號備援欄。")
        if mapping.get("quantity") is None:
            raise ValueError("xlsx 匯入至少需要指定「數量」欄。")

        if report is not None:
            report.update(
                {
                    "sheet": ws.title,
                    "header_row": header_row,
                    "mapped_fields": [
                        field for field, column in mapping.items() if column is not None
                    ],
                    "source_rows": 0,
                    "skipped_missing_designation": 0,
                    "skipped_invalid_quantity": 0,
                    "quantity_defaulted": 0,
                    "unit_defaulted": 0,
                    "problems": [],
                }
            )

        rows: list[ProjectInputRow] = []
        for row_idx, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            if report is not None:
                report["source_rows"] += 1
            raw = format_import_raw(values, headers)
            designation = project_mapped_value(values, mapping, "designation")
            if not designation:
                designation = (
                    extract_designation_from_text(project_mapped_value(values, mapping, "description"))
                    or extract_designation_from_text(project_mapped_value(values, mapping, "item_code"))
                )
            if not designation:
                if report is not None:
                    report["skipped_missing_designation"] += 1
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="error",
                    field="型號",
                    issue="找不到可辨識的支撐型號；此列不會匯入",
                    raw=raw,
                    resolution="在型號欄填入例如 57-1B-A；若資料藏在說明欄，請重新確認欄位對應。",
                )
                continue

            quantity_value = project_mapped_value(values, mapping, "quantity")
            unit_value = project_mapped_value(values, mapping, "unit")
            if report is not None and not quantity_value:
                report["quantity_defaulted"] += 1
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="warning",
                    field="數量",
                    issue="數量空白；本次暫按 1 組匯入",
                    raw=raw,
                    resolution="回原檔補上大於 0 的整數；若確實只有一組，請明確填 1。",
                )
            if report is not None and not unit_value:
                report["unit_defaulted"] += 1
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="info",
                    field="單位",
                    issue="單位空白；本次使用「組」",
                    raw=raw,
                    resolution="建議回原檔填入組、set、ea 等實際單位。",
                )
            quantity_text = quantity_value or "1"
            try:
                quantity = parse_list_quantity(quantity_text, row_idx)
            except ValueError:
                if report is None:
                    raise
                if report is not None:
                    report["skipped_invalid_quantity"] += 1
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="error",
                    field="數量",
                    issue=f"數量 {quantity_text!r} 不是大於 0 的整數；此列不會匯入",
                    raw=raw,
                    resolution="改成 1、2、3 等大於 0 的整數。",
                )
                continue

            drawing_line_number = project_mapped_value(
                values, mapping, "drawing_line_number"
            )
            serial = project_mapped_value(values, mapping, "serial")
            if not drawing_line_number:
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="warning",
                    field="Drawing",
                    issue="缺少 Drawing line number；仍可計算但無法依圖面追溯",
                    raw=raw,
                    resolution="填入原始圖面或管線群組編號。",
                )
            if not serial:
                append_import_problem(
                    report,
                    row_number=row_idx,
                    severity="warning",
                    field="流水號",
                    issue="缺少流水號；仍可計算但逐筆核對較困難",
                    raw=raw,
                    resolution="填入原始 MTO 的流水號或專案排序編號。",
                )
            overrides = None
            display_designation = ""
            source_profile = project_mapped_value(
                values, mapping, "source_profile"
            )
            if source_profile:
                source_profile = normalize_source_profile(source_profile)
            if str(designation).strip().upper() == "PENETRATION HOLE":
                from .penetration_hole import build_item_code

                nominal_size = project_mapped_value(values, mapping, "nominal_size")
                insulation = project_mapped_value(values, mapping, "insulation")
                overrides = {"nominal_size": nominal_size, "insulation": insulation}
                display_designation = build_item_code(nominal_size, insulation)
                if not nominal_size:
                    append_import_problem(
                        report,
                        row_number=row_idx,
                        severity="error",
                        field="管徑",
                        issue="PENETRATION HOLE 缺少管徑；型號尚不完整",
                        raw=raw,
                        resolution="在管徑欄填入例如 4、6 或 8。",
                    )

            rows.append(
                ProjectInputRow(
                    designation=designation,
                    quantity=quantity,
                    enabled=True,
                    drawing_line_number=drawing_line_number,
                    serial=serial,
                    unit=normalize_project_unit_value(unit_value),
                    overrides=overrides,
                    display_designation=display_designation,
                    source_profile=source_profile,
                )
            )
        return rows
    finally:
        wb.close()


def detect_project_xlsx_layout(wb):
    best = None
    for ws in wb.worksheets:
        scan_rows = list(
            ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row, 30),
                values_only=True,
            )
        )
        for row_offset, row_values in enumerate(scan_rows, start=1):
            headers = ["" if cell is None else str(cell).strip() for cell in row_values]
            if sum(1 for header in headers if header) < 2:
                continue
            sample_rows = scan_rows[row_offset:row_offset + 12]
            mapping, score = infer_project_column_mapping(headers, sample_rows)
            score += project_sheet_name_bonus(ws.title)
            score += 60 if mapping.get("designation") is not None and mapping.get("quantity") is not None else 0
            candidate = {
                "worksheet": ws,
                "header_row": row_offset,
                "headers": headers,
                "mapping": mapping,
                "score": score,
            }
            if best is None or candidate["score"] > best["score"]:
                best = candidate
    return best


def project_sheet_name_bonus(title: str) -> int:
    normalized = re.sub(r"\s+", "", str(title or "").strip().lower())
    if normalized == "supportmto":
        return 30
    bonus = 0
    if "support" in normalized:
        bonus += 8
    if "mto" in normalized:
        bonus += 8
    if "材料" in normalized or "支撐" in normalized:
        bonus += 4
    return bonus


def infer_project_column_mapping(headers: list[str], sample_rows: list[tuple]) -> tuple[ProjectXlsxMapping, float]:
    thresholds = {
        "designation": 35,
        "quantity": 35,
        "drawing_line_number": 35,
        "serial": 45,
        "unit": 35,
        "description": 35,
        "item_code": 35,
        "nominal_size": 35,
        "insulation": 35,
        "source_profile": 35,
    }
    scores = []
    for field in PROJECT_XLSX_FIELDS:
        for col_idx, header in enumerate(headers):
            column_values = [
                row[col_idx] if col_idx < len(row) else None
                for row in sample_rows
            ]
            score = project_column_score(field, header, column_values)
            if score >= thresholds[field]:
                scores.append((score, field, col_idx))

    mapping: ProjectXlsxMapping = {field: None for field in PROJECT_XLSX_FIELDS}
    used_cols: set[int] = set()
    total_score = 0.0
    for score, field, col_idx in sorted(scores, reverse=True):
        if mapping[field] is not None or col_idx in used_cols:
            continue
        mapping[field] = col_idx
        used_cols.add(col_idx)
        total_score += score
    return mapping, total_score


def project_column_score(field: str, header, values: list) -> float:
    header_text = str(header or "").strip()
    normalized = normalize_project_header(header_text)
    if not normalized:
        return 0.0

    score = 0.0
    aliases = [normalize_project_header(alias) for alias in PROJECT_ROW_ALIASES.get(field, ())]
    if normalized in aliases:
        score += 100
    elif any(alias and alias in normalized for alias in aliases):
        score += 65

    keyword_scores = {
        "serial": (("流水", "序號", "編號", "serial", "serialno", "seq", "sort", "rowno", "lineno"), 42),
        "drawing_line_number": (
            ("drawinglinenumber", "drawingline", "linegroup", "line_group", "dwg", "圖名", "圖號"), 42
        ),
        "quantity": (("數量", "組數", "qty", "quantity", "count"), 42),
        "unit": (("單位", "unit", "uom"), 42),
        "designation": (("型號", "designation", "model", "supportno", "supportdesignation", "支撐編碼"), 42),
        "description": (("description", "desc", "描述", "說明", "中文說明", "品名"), 42),
        "item_code": (("itemcode", "料號", "code"), 42),
        "nominal_size": (("nominalsize", "管徑", "size"), 42),
        "insulation": (("保溫厚度", "insulation", "insulthk"), 42),
    }
    keywords, keyword_score = keyword_scores.get(field, ((), 0))
    if any(normalize_project_header(keyword) in normalized for keyword in keywords):
        score += keyword_score

    nonempty = [str(value).strip() for value in values if value is not None and str(value).strip()]
    if not nonempty:
        return score

    if field == "designation":
        hit_ratio = sum(1 for value in nonempty if looks_like_designation(value)) / len(nonempty)
        score += hit_ratio * 70
    elif field == "quantity":
        numeric_ratio = sum(1 for value in nonempty if looks_like_list_quantity(value)) / len(nonempty)
        score += numeric_ratio * 45
    elif field == "unit":
        unit_ratio = sum(1 for value in nonempty if looks_like_project_unit(value)) / len(nonempty)
        score += unit_ratio * 70
    elif field == "serial":
        unique_ratio = len(set(nonempty)) / len(nonempty)
        score += min(unique_ratio, 1.0) * 15
    elif field == "drawing_line_number":
        drawing_ratio = sum(1 for value in nonempty if looks_like_drawing_line_number(value)) / len(nonempty)
        score += drawing_ratio * 35
    elif field == "description":
        extracted_ratio = sum(1 for value in nonempty if extract_designation_from_text(value) != value) / len(nonempty)
        score += extracted_ratio * 35
    return score


def project_mapped_value(values, mapping: dict, field: str) -> str:
    col_idx = mapping.get(field)
    if col_idx is None or col_idx >= len(values):
        return ""
    value = values[col_idx]
    return "" if value is None else str(value).strip()


def has_project_designation_source(mapping: dict) -> bool:
    return any(mapping.get(field) is not None for field in ("designation", "description", "item_code"))


def parse_list_quantity(text: str, row_number: int) -> int:
    try:
        numeric = float(str(text).strip().replace(",", ""))
        if not numeric.is_integer():
            raise ValueError
        value = int(numeric)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {row_number} 列組數不是整數: {text!r}") from exc
    if value <= 0:
        raise ValueError(f"第 {row_number} 列組數必須大於 0")
    return value


def looks_like_list_quantity(text: str) -> bool:
    try:
        parse_list_quantity(text, 0)
    except ValueError:
        return False
    return True


def normalize_project_unit_value(text: str) -> str:
    value = str(text or "").strip()
    normalized = value.replace(" ", "").lower()
    if normalized in {"", "ве", "вe", "be"}:
        return "組"
    return value


def looks_like_project_unit(text: str) -> bool:
    return normalize_project_unit_value(text).lower() in {"組", "set", "sets", "kg", "ea", "pc", "m"}


def looks_like_designation(text: str) -> bool:
    return bool(re.fullmatch(r"\d{2}[A-Z]?(?:-[A-Z0-9./()]+)+", str(text or "").strip().upper()))


def looks_like_drawing_line_number(text: str) -> bool:
    value = str(text or "").strip()
    return bool(re.search(r"[A-Z]", value, re.IGNORECASE) and ("--" in value or value.count("-") >= 2))


def normalize_project_header(text) -> str:
    return re.sub(r"[\s._-]+", "", str(text or "").strip().lower())


def extract_designation_from_text(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    matches = re.findall(r"\b\d{2}[A-Z]?(?:-[A-Z0-9./()]+)+\b", text.upper())
    return matches[-1] if matches else text
