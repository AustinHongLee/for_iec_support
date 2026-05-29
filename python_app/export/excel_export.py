"""
Excel 匯出模組 — 2026-05-29 視覺風格更新（現代深靛 + 琥珀強調）
"""
from dataclasses import dataclass
from typing import List
from core.models import AnalysisResult
from core.project_aggregation import ProjectAnalysisResult
from core.material_summary import aggregate_project, MaterialSummary
from core.cutting_optimizer import CuttingPlan, optimize_from_summary
from core.parser import get_lookup_value, get_part

# 表頭定義 (對應 VBA 的 headers)
HEADERS = [
    "材料描述欄", "項次", "品名", "尺寸/規格", "長度", "寬度",
    "材質", "數量", "每米重", "單重", "總重小計", "單位",
    "係數", "長度小計", "數量小計", "總重合計", "屬性",
    "物件類別", "製造方式", "零件ID", "庫存ID", "計算說明"
]

PROJECT_HEADERS = [
    "型號", "項次", "品名", "規格", "材質",
    "長度(mm)", "寬度(mm)",
    "單件數量", "組數", "總數量", "單組重(kg)", "總重(kg)", "屬性",
    "物件類別", "製造方式", "零件ID", "庫存ID", "計算說明",
]

SUMMARY_HEADERS = [
    "品名", "規格", "材質", "屬性", "物件類別", "製造方式",
    "需求總長(mm)", "需求件數/數量", "總重(kg)", "原料長度(mm)",
    "建議採購量", "單位", "來源編碼",
]

CUTTING_HEADERS = [
    "材料", "原料 #", "切割段", "需求長(mm)", "含損耗(mm)",
    "累計(mm)", "餘料(mm)", "使用率", "用於",
]

LEADER_STAT_HEADERS = ["類別", "統計項目", "統計條件", "單位", "數量", "命中/確認筆數"]
LEADER_GROUP_DETAIL_HEADERS = [
    "狀態", "型號", "組數", "管徑(in)", "計入數量", "單位", "命中明細", "材質判定", "備註",
]
LEADER_DETAIL_HEADERS = [
    "狀態", "類別", "統計項目", "型號", "組數", "管徑(in)",
    "計入數量", "單位", "命中明細", "材質判定", "統計條件", "備註",
]

VISUAL_SLOT_COUNT = 30


@dataclass(frozen=True)
class LeaderStatRow:
    item: str
    label: str
    unit: str
    key: str
    criteria: str


@dataclass(frozen=True)
class LeaderHitDetail:
    stat_key: str
    status: str
    category: str
    label: str
    designation: str
    project_qty: int
    pipe_size: float | None
    amount: float
    unit: str
    matched_detail: str
    material_basis: str
    criteria: str
    note: str = ""


def _styles():
    """
    現代深靛 + 琥珀強調 配色（2026-05 更新）

    palette 設計目標：
      - 主色 1F3864 深靛：標題、表頭、合計列 → 沉穩、長官感
      - 次色 2E5395 中靛：副區塊強調
      - 強調 BF8F00 琥珀：KPI 大數字、重點數據
      - 背景 F2F4F8 極淺灰藍：副標題列、KPI 卡片底
      - zebra F7F9FC：隔行底色，極輕微對比，避免疲勞
      - 狀態色（OK 70AD47 / Warn ED7D31 / Bad C00000 / Info 4472C4）一致 PowerBI 系
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    border_thin = Side(style="thin", color="BFBFBF")
    border_medium = Side(style="medium", color="1F3864")

    return {
        # === Fills ===
        "title_fill": PatternFill("solid", fgColor="1F3864"),       # 主標題列
        "subtitle_fill": PatternFill("solid", fgColor="F2F4F8"),    # 副標題列（R2）
        "section_fill": PatternFill("solid", fgColor="D9E1F2"),     # 區塊標題
        "header_fill": PatternFill("solid", fgColor="2E5395"),      # 表頭
        "subheader_fill": PatternFill("solid", fgColor="DEE3EE"),   # 次級表頭
        "kpi_label_fill": PatternFill("solid", fgColor="F2F4F8"),   # KPI 卡片 label
        "kpi_value_fill": PatternFill("solid", fgColor="FFFFFF"),   # KPI 卡片 value
        "zebra_fill": PatternFill("solid", fgColor="F7F9FC"),       # 隔行
        "subtotal_fill": PatternFill("solid", fgColor="FFF2CC"),    # 小計列（暖色凸顯）
        "grand_total_fill": PatternFill("solid", fgColor="1F3864"), # 全案合計
        # 狀態色
        "ok_fill": PatternFill("solid", fgColor="E2EFDA"),
        "warn_fill": PatternFill("solid", fgColor="FFF2CC"),
        "bad_fill": PatternFill("solid", fgColor="FCE4D6"),
        "info_fill": PatternFill("solid", fgColor="DDEBF7"),
        # 下料圖示
        "used_fill": PatternFill("solid", fgColor="4472C4"),
        "remnant_fill": PatternFill("solid", fgColor="A9D18E"),

        # === Borders ===
        "border": Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin),
        "border_top": Border(top=border_medium),
        "border_bottom": Border(bottom=border_medium),

        # === Fonts ===
        "title_font": Font(name="Calibri", bold=True, color="FFFFFF", size=18),
        "subtitle_font": Font(name="Calibri", color="595959", size=10, italic=True),
        "header_font": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "section_font": Font(name="Calibri", bold=True, color="1F3864", size=12),
        "kpi_label_font": Font(name="Calibri", color="595959", size=10),
        "kpi_value_font": Font(name="Calibri", bold=True, color="1F3864", size=22),
        "kpi_accent_font": Font(name="Calibri", bold=True, color="BF8F00", size=22),
        "kpi_unit_font": Font(name="Calibri", color="595959", size=10),
        "kpi_note_font": Font(name="Calibri", color="808080", size=9, italic=True),
        "grand_total_font": Font(name="Calibri", bold=True, color="FFFFFF", size=12),
        "bold_font": Font(bold=True),

        # === Alignments ===
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center", indent=1),
        "right": Alignment(horizontal="right", vertical="center"),
        "wrap": Alignment(vertical="center", wrap_text=True),
        "wrap_top": Alignment(vertical="top", wrap_text=True),
    }


# === 通用視覺輔助 ====================================================

def _setup_sheet(ws, title: str, merge_to: str = "K1", subtitle: str = ""):
    """
    設定 sheet 的標題列（R1）與可選副標題列（R2）。

    Args:
        ws: worksheet
        title: 主標題文字
        merge_to: 合併儲存格的終點（如 "K1"）
        subtitle: 可選副標題（會放在 R2，使用副標題樣式）
    """
    styles = _styles()
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{merge_to}")
    cell = ws["A1"]
    cell.value = title
    cell.font = styles["title_font"]
    cell.fill = styles["title_fill"]
    cell.alignment = styles["center"]
    ws.row_dimensions[1].height = 40

    if subtitle:
        merge_end_col = merge_to.rstrip("0123456789")
        ws.merge_cells(f"A2:{merge_end_col}2")
        sub = ws["A2"]
        sub.value = subtitle
        sub.font = styles["subtitle_font"]
        sub.fill = styles["subtitle_fill"]
        sub.alignment = Alignment_for_subtitle()
        ws.row_dimensions[2].height = 22


def Alignment_for_subtitle():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="center", vertical="center")


def _apply_zebra(ws, start_row: int, end_row: int, max_col: int):
    """資料區隔行底色（偶數列 zebra）。需在資料寫入後呼叫。"""
    styles = _styles()
    zebra = styles["zebra_fill"]
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = zebra


def _section_header(ws, row: int, text: str, span_cols: int = 1):
    """區塊小標題列（▌ XXX），淺藍底深靛字。"""
    styles = _styles()
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row=row, column=1, value=f"▌ {text}")
    cell.font = styles["section_font"]
    cell.fill = styles["section_fill"]
    from openpyxl.styles import Alignment
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _kpi_card(ws, row: int, col: int, label: str, value, unit: str = "",
              note: str = "", accent: bool = False, value_format: str = None):
    """
    繪製 KPI 卡片（3 列：label / value+unit / note）

    Args:
        row: 起始列
        col: 起始欄
        label: 上方標題
        value: 中央大數字
        unit: 數字右側單位
        note: 下方說明（小灰字）
        accent: True 時數字用琥珀色（強調）
        value_format: openpyxl number format（如 '#,##0.00'）
    """
    styles = _styles()
    from openpyxl.styles import Alignment, Border, Side

    # 卡片邊框
    side = Side(style="thin", color="D0D7E2")
    border = Border(left=side, right=side, top=side, bottom=side)

    # Row 1: Label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    lcell = ws.cell(row=row, column=col, value=label)
    lcell.font = styles["kpi_label_font"]
    lcell.fill = styles["kpi_label_fill"]
    lcell.alignment = Alignment(horizontal="center", vertical="center")
    lcell.border = border
    ws.row_dimensions[row].height = 18

    # Row 2: Value + Unit
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    vcell = ws.cell(row=row + 1, column=col, value=value)
    vcell.font = styles["kpi_accent_font"] if accent else styles["kpi_value_font"]
    vcell.fill = styles["kpi_value_fill"]
    vcell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    vcell.border = border
    if value_format:
        vcell.number_format = value_format

    ucell = ws.cell(row=row + 1, column=col + 2, value=unit)
    ucell.font = styles["kpi_unit_font"]
    ucell.fill = styles["kpi_value_fill"]
    ucell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ucell.border = border
    ws.row_dimensions[row + 1].height = 34

    # Row 3: Note
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)
    ncell = ws.cell(row=row + 2, column=col, value=note)
    ncell.font = styles["kpi_note_font"]
    ncell.fill = styles["kpi_value_fill"]
    ncell.alignment = Alignment(horizontal="center", vertical="center")
    ncell.border = border
    ws.row_dimensions[row + 2].height = 16


def _add_data_bar(ws, cell_range: str, color: str = "BF8F00"):
    """加 Excel 內建 data bar（條形圖視覺化）。"""
    try:
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(
            start_type='min', start_value=None,
            end_type='max', end_value=None,
            color=color, showValue=True,
            minLength=0, maxLength=100,
        )
        ws.conditional_formatting.add(cell_range, rule)
    except Exception:
        pass


def _format_number_block(ws, row_start: int, row_end: int, col_formats: dict):
    """批次套用 number_format。col_formats: {col_index: format_string}"""
    for r in range(row_start, row_end + 1):
        for col, fmt in col_formats.items():
            ws.cell(row=r, column=col).number_format = fmt


def _write_headers(ws, row: int, headers: list[str]):
    styles = _styles()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]


def _apply_table_style(ws, min_row: int, max_row: int, max_col: int):
    styles = _styles()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = styles["border"]
            if cell.row > min_row:
                cell.alignment = styles["wrap"]


def _set_widths(ws, widths: list[float]):
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _format_number_columns(ws, row_start: int, row_end: int, columns: list[int], fmt: str):
    for row in range(row_start, row_end + 1):
        for col in columns:
            ws.cell(row=row, column=col).number_format = fmt


def _format_sheet(ws, headers):
    from openpyxl.styles import Font, Alignment, PatternFill

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)


def export_to_excel(results: List[AnalysisResult], filepath: str):
    """匯出分析結果至 Excel"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weight_Analysis"

    _format_sheet(ws, HEADERS)

    # 資料
    row = 2
    for result in results:
        # 寫入原始字串到 A 欄
        ws.cell(row=row, column=1, value=result.fullstring)

        if result.error:
            ws.cell(row=row, column=2, value="Error")
            ws.cell(row=row, column=3, value=result.error)
            row += 1
            continue

        for entry in result.entries:
            ws.cell(row=row, column=1, value=result.fullstring if entry.item_no == 1 else "")
            ws.cell(row=row, column=2, value=entry.item_no)
            ws.cell(row=row, column=3, value=entry.name)
            ws.cell(row=row, column=4, value=entry.display_spec)
            ws.cell(row=row, column=5, value=entry.length)
            ws.cell(row=row, column=6, value=entry.width if entry.width else "")
            ws.cell(row=row, column=7, value=entry.material)
            ws.cell(row=row, column=8, value=entry.quantity)
            ws.cell(row=row, column=9, value=entry.weight_per_unit if entry.weight_per_unit else "")
            ws.cell(row=row, column=10, value=entry.unit_weight)
            ws.cell(row=row, column=11, value=entry.total_weight)
            ws.cell(row=row, column=12, value=entry.unit)
            ws.cell(row=row, column=13, value=entry.factor)
            ws.cell(row=row, column=14, value=entry.length_subtotal if entry.length_subtotal else "")
            ws.cell(row=row, column=15, value=entry.qty_subtotal if entry.qty_subtotal else "")
            ws.cell(row=row, column=16, value=entry.weight_output)
            ws.cell(row=row, column=17, value=entry.category)
            ws.cell(row=row, column=18, value=entry.item_class)
            ws.cell(row=row, column=19, value=entry.manufacturing_type)
            ws.cell(row=row, column=20, value=entry.part_key)
            ws.cell(row=row, column=21, value=entry.stock_id)
            ws.cell(row=row, column=22, value=entry.display_remark)
            row += 1

    _format_sheet(ws, HEADERS)

    wb.save(filepath)


def export_project_to_excel(project: ProjectAnalysisResult, filepath: str):
    """匯出 project-aware 分析結果（平坦表格，每列填滿型號/組數）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project_Weight_Analysis"

    _format_sheet(ws, PROJECT_HEADERS)

    row = 2
    for row_result in project.rows:
        input_row = row_result.input_row
        single_result = row_result.single_result
        scaled_result = row_result.scaled_result

        if single_result.error:
            ws.cell(row=row, column=1, value=input_row.designation)
            ws.cell(row=row, column=2, value="Error")
            ws.cell(row=row, column=3, value=single_result.error)
            ws.cell(row=row, column=9, value=input_row.quantity)
            row += 1
            continue

        for single_entry, scaled_entry in zip(single_result.entries, scaled_result.entries):
            ws.cell(row=row, column=1, value=input_row.designation)   # 每列填滿
            ws.cell(row=row, column=2, value=single_entry.item_no)
            ws.cell(row=row, column=3, value=single_entry.name)
            ws.cell(row=row, column=4, value=single_entry.display_spec)
            ws.cell(row=row, column=5, value=single_entry.material)
            ws.cell(row=row, column=6, value=single_entry.length)
            ws.cell(row=row, column=7, value=single_entry.width if single_entry.width else "")
            ws.cell(row=row, column=8, value=single_entry.quantity)
            ws.cell(row=row, column=9, value=input_row.quantity)
            ws.cell(row=row, column=10, value=scaled_entry.quantity)
            ws.cell(row=row, column=11, value=single_entry.weight_output)
            ws.cell(row=row, column=12, value=scaled_entry.weight_output)
            ws.cell(row=row, column=13, value=single_entry.category)
            ws.cell(row=row, column=14, value=single_entry.item_class)
            ws.cell(row=row, column=15, value=single_entry.manufacturing_type)
            ws.cell(row=row, column=16, value=single_entry.part_key)
            ws.cell(row=row, column=17, value=single_entry.stock_id)
            ws.cell(row=row, column=18, value=single_entry.display_remark)
            row += 1

    _format_sheet(ws, PROJECT_HEADERS)

    wb.save(filepath)


def export_project_workbook(project: ProjectAnalysisResult, filepath: str):
    """
    匯出專案級整合 workbook。

    Sheets:
      1. 專案摘要
      2. 重量明細表  (Flat pivot-friendly table)
      3. 計算標準與假設  (Static manager/client reference page)
      4. 支撐分類統計
      5. 支撐統計明細
      6. 重量分析
      7. 材料合計
      8. 下料明細
      9. 下料圖示
    """
    import openpyxl

    summary = aggregate_project(project)
    cutting_plans = _build_cutting_plans(summary)

    wb = openpyxl.Workbook()
    _write_project_summary_sheet(wb.active, project, summary, cutting_plans)
    _write_calculation_basis_sheet(wb.create_sheet("重量明細表"), project)
    _write_calc_reference_sheet(wb.create_sheet("計算標準與假設"), project)
    _write_leader_procurement_sheet(wb.create_sheet("支撐分類統計"), project)
    _write_leader_detail_sheet(wb.create_sheet("支撐統計明細"), project)
    _write_project_weight_sheet(wb.create_sheet("重量分析"), project)
    _write_material_summary_sheet(wb.create_sheet("材料合計"), summary)
    _write_cutting_detail_sheet(wb.create_sheet("下料明細"), cutting_plans)
    _write_cutting_visual_sheet(wb.create_sheet("下料圖示"), cutting_plans)

    wb.save(filepath)


def _build_cutting_plans(summary: MaterialSummary) -> list[CuttingPlan]:
    plans: list[CuttingPlan] = []
    for ln in summary.get_linear_lines():
        plan = optimize_from_summary(ln)
        if plan and plan.total_pieces > 0:
            plans.append(plan)
    return plans


def _write_project_summary_sheet(ws, project: ProjectAnalysisResult, summary: MaterialSummary, plans: list[CuttingPlan]):
    """專案摘要 — 長官第一眼看的儀表板頁。

    版面：
      R1     主標題列（深靛底白字 18pt）
      R2     副標題列（製表日期 + 系統版本 + 總重）
      R3     空
      R4     ▌ 關鍵指標
      R5~R7  KPI 卡片第一列 × 4 個（支撐組數 / 材料種類 / 專案總重 / 平均單組重）
      R8     空
      R9~R11 KPI 卡片第二列 × 4 個（下料材料 / 建議原料根數 / 下料段數 / 平均使用率）
      R12    空
      R13    ▌ 重型支撐 Top 5
      R14    表頭
      R15~19 Top 5 資料 + 條形視覺
      R20    空
      R21    ▌ 材料重量分佈（Top 8 by 總重）
      R22    表頭
      R23~30 資料 + 條形視覺
      R31    空
      R32    ▌ Workbook 索引（各分頁用途）
      R33    表頭
      R34~42 各 sheet 一列
      R43    空
      R44    ▌ 注意事項
      R45~47 注意事項條列
    """
    import datetime as _dt
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    styles = _styles()
    ws.title = "專案摘要"

    # === R1 主標題 + R2 副標題 ====================================
    today_str = _dt.date.today().strftime("%Y-%m-%d")
    subtitle = (
        f"製表日期 {today_str}    "
        f"全案總重 {summary.total_weight:,.2f} kg    "
        f"支撐 {project.total_support_count} 組    "
        f"材料 {len(summary.lines)} 項"
    )
    _setup_sheet(ws, "專案材料統計總覽", "L1", subtitle=subtitle)

    # 欄寬：12 個欄（3 欄一組 × 4 組 KPI）
    _set_widths(ws, [16, 9, 5, 16, 9, 5, 16, 9, 5, 16, 9, 5])

    # === KPI 計算 ================================================
    total_bars = sum(plan.total_bars for plan in plans)
    total_cut_pieces = sum(plan.total_pieces for plan in plans)
    avg_util = (
        sum(plan.avg_utilization for plan in plans) / len(plans)
        if plans else 0.0
    )
    successful_rows = [r for r in project.rows if not r.single_result.error]
    avg_unit_weight = (
        sum(r.single_result.total_weight for r in successful_rows) / len(successful_rows)
        if successful_rows else 0.0
    )

    # === R4 區塊標題 + R5~R7 第一列 KPI ===========================
    _section_header(ws, 4, "關鍵指標", span_cols=12)

    row_kpi_1 = 5
    _kpi_card(ws, row_kpi_1, 1, "支撐總組數",
              project.total_support_count, "組",
              note="本批設計含支撐總數", accent=True, value_format="#,##0")
    _kpi_card(ws, row_kpi_1, 4, "材料種類",
              len(summary.lines), "項",
              note="合計表獨立材料品項", value_format="#,##0")
    _kpi_card(ws, row_kpi_1, 7, "專案總重",
              round(summary.total_weight, 2), "kg",
              note="全案累計總重", accent=True, value_format="#,##0.00")
    _kpi_card(ws, row_kpi_1, 10, "平均單組重",
              round(avg_unit_weight, 2), "kg/組",
              note=f"成功項 {len(successful_rows)} 組之均值", value_format="#,##0.00")

    # === R9~R11 第二列 KPI =======================================
    row_kpi_2 = 9
    _kpi_card(ws, row_kpi_2, 1, "下料材料",
              len(plans), "種",
              note="需切割的線性材料種類", value_format="#,##0")
    _kpi_card(ws, row_kpi_2, 4, "建議原料根數",
              total_bars, "根",
              note="依下料規劃所需原料數", accent=True, value_format="#,##0")
    _kpi_card(ws, row_kpi_2, 7, "下料段數",
              total_cut_pieces, "段",
              note="所有原料切割段累計", value_format="#,##0")
    _kpi_card(ws, row_kpi_2, 10, "平均使用率",
              round(avg_util, 1) / 100 if avg_util else 0, "",
              note="原料切割平均利用率", accent=True, value_format="0.0%")

    # === R13~R19 Top 5 重型支撐 ===================================
    _section_header(ws, 13, "重型支撐 Top 5（依單組重）", span_cols=12)

    top5 = sorted(
        successful_rows,
        key=lambda r: r.single_result.total_weight,
        reverse=True,
    )[:5]

    top5_headers = ["排名", "型號", "組數", "單組重 (kg)", "累計重 (kg)", "佔比"]
    header_row = 14
    for col, h in enumerate(top5_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.row_dimensions[header_row].height = 24
    # 將後面欄合併為條形圖視覺欄
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=12)
    bar_h = ws.cell(row=header_row, column=7, value="視覺比例")
    bar_h.fill = styles["header_fill"]
    bar_h.font = styles["header_font"]
    bar_h.alignment = styles["center"]
    bar_h.border = styles["border"]

    total_proj = summary.total_weight if summary.total_weight > 0 else 1.0
    for idx, r in enumerate(top5, start=1):
        rr = header_row + idx
        unit_w = r.single_result.total_weight
        total_w = r.scaled_result.total_weight
        pct = total_w / total_proj
        values = [
            idx,
            r.input_row.designation,
            r.input_row.quantity,
            round(unit_w, 2),
            round(total_w, 2),
            pct,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=rr, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col in (1, 2) else styles["right"]
        ws.cell(rr, 4).number_format = "#,##0.00"
        ws.cell(rr, 5).number_format = "#,##0.00"
        ws.cell(rr, 6).number_format = "0.0%"
        ws.row_dimensions[rr].height = 20
        # 視覺欄：用值寫入隱藏值，套 data bar
        ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=12)
        bcell = ws.cell(row=rr, column=7, value=round(total_w, 2))
        bcell.border = styles["border"]
        bcell.number_format = "#,##0.00"
        bcell.alignment = styles["right"]

    if top5:
        _add_data_bar(ws, f"G{header_row + 1}:G{header_row + len(top5)}", color="BF8F00")

    # === R21~R30 材料重量分佈 Top 8 ===============================
    _section_header(ws, 21, "材料重量分佈 Top 8（依總重）", span_cols=12)

    top_mats = sorted(summary.lines, key=lambda ln: ln.total_weight, reverse=True)[:8]
    mat_headers = ["#", "品名", "規格", "材質", "總重 (kg)", "佔比"]
    mat_header_row = 22
    for col, h in enumerate(mat_headers, 1):
        cell = ws.cell(row=mat_header_row, column=col, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.row_dimensions[mat_header_row].height = 24
    ws.merge_cells(start_row=mat_header_row, start_column=7, end_row=mat_header_row, end_column=12)
    mb_h = ws.cell(row=mat_header_row, column=7, value="視覺比例")
    mb_h.fill = styles["header_fill"]
    mb_h.font = styles["header_font"]
    mb_h.alignment = styles["center"]
    mb_h.border = styles["border"]

    for idx, ln in enumerate(top_mats, start=1):
        rr = mat_header_row + idx
        pct = ln.total_weight / total_proj
        values = [
            idx,
            ln.name,
            ln.spec,
            ln.material,
            round(ln.total_weight, 2),
            pct,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=rr, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col in (1, 2, 3, 4) else styles["right"]
        ws.cell(rr, 5).number_format = "#,##0.00"
        ws.cell(rr, 6).number_format = "0.0%"
        ws.row_dimensions[rr].height = 20
        ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=12)
        bcell = ws.cell(row=rr, column=7, value=round(ln.total_weight, 2))
        bcell.border = styles["border"]
        bcell.number_format = "#,##0.00"
        bcell.alignment = styles["right"]

    if top_mats:
        _add_data_bar(ws, f"G{mat_header_row + 1}:G{mat_header_row + len(top_mats)}", color="4472C4")

    # === R32~R42 Workbook 索引 ====================================
    _section_header(ws, 32, "Workbook 索引（各分頁用途）", span_cols=12)

    idx_headers = ["#", "分頁名稱", "主要用途", "資料量"]
    idx_header_row = 33
    # 合併寬欄位
    for col, h in enumerate(idx_headers, 1):
        cell = ws.cell(row=idx_header_row, column=col, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    # 第 3 欄合併到 11、第 4 欄落 12
    ws.merge_cells(start_row=idx_header_row, start_column=3, end_row=idx_header_row, end_column=11)
    ws.cell(idx_header_row, 3).value = "主要用途"
    ws.cell(idx_header_row, 3).fill = styles["header_fill"]
    ws.cell(idx_header_row, 3).font = styles["header_font"]
    ws.cell(idx_header_row, 3).alignment = styles["center"]
    ws.cell(idx_header_row, 12).value = "資料量"
    ws.cell(idx_header_row, 12).fill = styles["header_fill"]
    ws.cell(idx_header_row, 12).font = styles["header_font"]
    ws.cell(idx_header_row, 12).alignment = styles["center"]
    ws.cell(idx_header_row, 12).border = styles["border"]
    ws.row_dimensions[idx_header_row].height = 24

    workbook_index = [
        ("專案摘要", "長官第一眼總覽：KPI、Top 5、材料分佈", "本頁"),
        ("重量明細表", "單件 × 組數 × 總重 平表（樞紐分析用）", f"{len(project.rows)} 支撐"),
        ("計算標準與假設", "材料計算所依引用標準與資料狀態圖例", "靜態說明"),
        ("支撐分類統計", "U-Bolt/Pipe Shoe/管支撐製裝採購數量彙總", "依規則"),
        ("支撐統計明細", "命中、需確認、未納入的逐筆查核表", "依規則"),
        ("重量分析", "單件與總量並列、按 entry 展開", f"{len(project.rows)} 列"),
        ("材料合計", "依材質聚合的採購清單", f"{len(summary.lines)} 項"),
        ("下料明細", "每根原料的切割順序與餘料", f"{len(plans)} 種材料"),
        ("下料圖示", "原料使用率視覺化條塊", f"{sum(p.total_bars for p in plans)} 根原料"),
    ]
    for idx, (name, purpose, count) in enumerate(workbook_index, start=1):
        rr = idx_header_row + idx
        ws.cell(rr, 1, idx).alignment = styles["center"]
        ws.cell(rr, 2, name).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(rr, 2).font = styles["bold_font"]
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=11)
        ws.cell(rr, 3, purpose).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(rr, 12, count).alignment = styles["center"]
        for c in (1, 2, 3, 12):
            ws.cell(rr, c).border = styles["border"]
        ws.row_dimensions[rr].height = 20
        # zebra
        if idx % 2 == 0:
            for c in (1, 2, 3, 12):
                ws.cell(rr, c).fill = styles["zebra_fill"]

    # === R44~ 注意事項 ============================================
    note_section_row = idx_header_row + len(workbook_index) + 2
    _section_header(ws, note_section_row, "注意事項", span_cols=12)
    notes = [
        "本表所有材料與重量以各 Type calculator 與 component table 計算為準。",
        "下料圖示為現場規劃輔助；實際餘料與鋸口條件仍需現場確認。",
        "若 KPI 與其他分頁加總有微小差異，係四捨五入造成；以「材料合計」分頁為基準。",
        "支撐分類統計如有未命中項，請至「支撐統計明細」查核並回饋規則維護人員。",
    ]
    for i, note in enumerate(notes):
        rr = note_section_row + 1 + i
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
        cell = ws.cell(rr, 1, value=f"・{note}")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.font = Font(name="Calibri", size=10, color="595959")
        ws.row_dimensions[rr].height = 18


def _leader_stat_template() -> list[LeaderStatRow]:
    return [
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band <= 6" 熱浸鍍鋅', "組", "uband_hdg_le6", '明細含 U-Bolt/Band，管徑 <= 6"，材質非 SUS304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band >= 8" 熱浸鍍鋅', "組", "uband_hdg_ge8", '明細含 U-Bolt/Band，管徑 >= 8"，材質非 SUS304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band <= 6" (SUS304)', "組", "uband_304_le6", '明細含 U-Bolt/Band，管徑 <= 6"，材質含 304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band >= 8" (SUS304)', "組", "uband_304_ge8", '明細含 U-Bolt/Band，管徑 >= 8"，材質含 304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE <= 4" 熱浸鍍鋅', "組", "shoe_hdg_le4", 'Type 52/53/54/55/66/67/80/85，管徑 <= 4"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE 5"~10" 熱浸鍍鋅', "組", "shoe_hdg_5_10", 'Type 52/53/54/55/66/67/80/85，管徑 5"~10"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE 12"~24" 熱浸鍍鋅', "組", "shoe_hdg_12_24", 'Type 52/53/54/55/66/67/80/85，管徑 12"~24"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE >= 26" 熱浸鍍鋅', "組", "shoe_hdg_ge26", 'Type 52/53/54/55/66/67/80/85，管徑 >= 26"，整組不含 SUS304'),
        LeaderStatRow("Cold Support", "保冷支撐座（長春帶料）", "組", "cold_support", "Type 代碼尾碼為 C 的保冷支撐"),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE <= 4" (SUS304)', "組", "shoe_304_le4", 'Type 52/53/54/55/66/67/80/85，管徑 <= 4"，整組任一明細材質含 304'),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE 5"~10" (SUS304)', "組", "shoe_304_5_10", 'Type 52/53/54/55/66/67/80/85，管徑 5"~10"，整組任一明細材質含 304'),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE 12"~24" (SUS304)', "組", "shoe_304_12_24", 'Type 52/53/54/55/66/67/80/85，管徑 12"~24"，整組任一明細材質含 304'),
        LeaderStatRow("CS Support", "CS 管支撐製裝 <= 15 kg/組", "組", "cs_support_le15", "整組不含 SUS304，單組總重 <= 15 kg，按支撐組數統計"),
        LeaderStatRow("CS Support", "CS 管支撐製裝 > 15 kg/組", "KG", "cs_support_gt15", "整組不含 SUS304，單組總重 > 15 kg，按總重量統計"),
        LeaderStatRow("SUS304 Support", "SUS304 管支撐製裝 <= 15 kg/組", "組", "ss_support_le15", "整組含 SUS304，單組總重 <= 15 kg，按支撐組數統計"),
        LeaderStatRow("SUS304 Support", "SUS304 管支撐製裝 > 15 kg/組", "KG", "ss_support_gt15", "整組含 SUS304，單組總重 > 15 kg，按總重量統計"),
    ]


def _parse_designation_type(designation: str) -> str:
    return (get_part(designation, 1) or "").strip()


def _parse_designation_pipe_size(designation: str) -> float | None:
    token = (get_part(designation, 2) or "").strip()
    if not token:
        return None
    token = token.split("(")[0].replace('"', "").strip()
    if token.upper().endswith("B"):
        token = token[:-1]
    try:
        return float(get_lookup_value(token))
    except (TypeError, ValueError):
        return None


def _is_304_material(material: str) -> bool:
    return "304" in str(material or "").upper().replace(" ", "")


def _is_ubolt_or_band_entry(name: str) -> bool:
    upper = str(name or "").upper().replace(" ", "")
    return "U-BOLT" in upper or "UBOLT" in upper or "U-BAND" in upper or "UBAND" in upper


def _support_has_304_material(row_result) -> bool:
    return any(_is_304_material(entry.material) for entry in row_result.scaled_result.entries)


def _is_cold_support_type(type_id: str) -> bool:
    return type_id.endswith("C") and type_id[:-1].isdigit()


def _leader_size_bucket(size: float | None, buckets: tuple[tuple[str, float, float], ...]) -> str | None:
    if size is None:
        return None
    for key, lower, upper in buckets:
        if lower <= size <= upper:
            return key
    return None


def _leader_procurement_stats(
    project: ProjectAnalysisResult,
) -> tuple[dict[str, float], dict[str, list[str]], list[LeaderHitDetail]]:
    template = _leader_stat_template()
    template_by_key = {row.key: row for row in template}
    stats = {row.key: 0.0 for row in template}
    sources = {row.key: [] for row in template}
    details: list[LeaderHitDetail] = []
    pipe_shoe_types = {"52", "53", "54", "55", "66", "67", "80", "85"}

    def material_label(is_304: bool) -> str:
        return "SUS304" if is_304 else "HDG/CS"

    def pipe_size_label(size: float | None) -> str:
        return "" if size is None else f'{size:g}"'

    def add_detail(
        *,
        status: str,
        key: str,
        designation: str,
        project_qty: int,
        pipe_size: float | None,
        amount: float,
        unit: str,
        matched_detail: str,
        material_basis: str,
        note: str = "",
    ) -> None:
        stat_row = template_by_key.get(key)
        if stat_row is None:
            return
        details.append(
            LeaderHitDetail(
                stat_key=stat_row.key,
                status=status,
                category=stat_row.item,
                label=stat_row.label,
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                amount=amount,
                unit=unit,
                matched_detail=matched_detail,
                material_basis=material_basis,
                criteria=stat_row.criteria,
                note=note,
            )
        )

    def add_stat(key: str, amount: float, source: str, **detail_kwargs) -> None:
        stats[key] += amount
        if source and source not in sources[key]:
            sources[key].append(source)
        add_detail(status="命中", key=key, amount=amount, **detail_kwargs)

    def add_issue(
        *,
        key: str,
        designation: str,
        project_qty: int,
        pipe_size: float | None,
        matched_detail: str,
        material_basis: str,
        note: str,
    ) -> None:
        add_detail(
            status="需確認",
            key=key,
            designation=designation,
            project_qty=project_qty,
            pipe_size=pipe_size,
            amount=0.0,
            unit="",
            matched_detail=matched_detail,
            material_basis=material_basis,
            note=note,
        )

    def add_unmatched(
        *,
        designation: str,
        project_qty: int,
        pipe_size: float | None,
        row_result,
    ) -> None:
        has_304 = _support_has_304_material(row_result)
        entry_names = "、".join(entry.name for entry in row_result.scaled_result.entries[:5])
        if len(row_result.scaled_result.entries) > 5:
            entry_names += "..."
        details.append(
            LeaderHitDetail(
                stat_key="unmatched",
                status="未納入",
                category="未納入支撐分類統計",
                label="未命中摘要規則",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                amount=0.0,
                unit="",
                matched_detail=entry_names or "無材料明細",
                material_basis="整組含 SUS304" if has_304 else "整組不含 SUS304",
                criteria="目前摘要統計 U-Bolt/Band、Pipe Shoe、Cold Support、CS/SUS304 管支撐製裝",
                note=(
                    "整組含 SUS304，但未符合目前摘要規則，請確認是否需要新增採購/製裝分類。"
                    if has_304
                    else "未符合目前支撐分類統計規則，請確認是否需要新增採購/製裝分類。"
                ),
            )
        )

    for row_result in project.rows:
        designation = row_result.input_row.designation
        project_qty = row_result.input_row.quantity
        type_id = _parse_designation_type(designation)
        pipe_size = _parse_designation_pipe_size(designation)
        detail_count_before = len(details)

        if row_result.single_result.error:
            add_issue(
                key="cs_support_le15",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                matched_detail="分析失敗",
                material_basis="",
                note=row_result.single_result.error,
            )
            continue

        for entry in row_result.scaled_result.entries:
            if not _is_ubolt_or_band_entry(entry.name):
                continue
            entry_is_304 = _is_304_material(entry.material)
            material_key = "304" if entry_is_304 else "hdg"
            bucket = _leader_size_bucket(pipe_size, (
                ("le6", 0.0, 6.0),
                ("ge8", 8.0, 999.0),
            ))
            if bucket:
                key = f"uband_{material_key}_{bucket}"
                add_stat(
                    key,
                    entry.quantity,
                    f"{designation} ×{project_qty}: {entry.name} {entry.quantity:g}{entry.unit}",
                    designation=designation,
                    project_qty=project_qty,
                    pipe_size=pipe_size,
                    unit=entry.unit,
                    matched_detail=(
                        f"項次{entry.item_no} {entry.name} {entry.display_spec}"
                        f" ×{entry.quantity:g}"
                    ),
                    material_basis=f"{entry.material} -> {material_label(entry_is_304)}",
                )
            else:
                add_issue(
                    key=f"uband_{material_key}_le6",
                    designation=designation,
                    project_qty=project_qty,
                    pipe_size=pipe_size,
                    matched_detail=f"項次{entry.item_no} {entry.name} {entry.display_spec}",
                    material_basis=f"{entry.material} -> {material_label(entry_is_304)}",
                    note="U-Bolt/Band 命中，但管徑無法落入 <=6 或 >=8 統計區間。",
                )

        if type_id in pipe_shoe_types:
            support_is_304 = _support_has_304_material(row_result)
            material_key = "304" if support_is_304 else "hdg"
            bucket = _leader_size_bucket(pipe_size, (
                ("le4", 0.0, 4.0),
                ("5_10", 5.0, 10.0),
                ("12_24", 12.0, 24.0),
                ("ge26", 26.0, 999.0),
            ))
            if bucket and f"shoe_{material_key}_{bucket}" in stats:
                key = f"shoe_{material_key}_{bucket}"
                add_stat(
                    key,
                    project_qty,
                    f"{designation}: {project_qty}組",
                    designation=designation,
                    project_qty=project_qty,
                    pipe_size=pipe_size,
                    unit="組",
                    matched_detail=f"Type {type_id} Pipe Shoe，{pipe_size_label(pipe_size)}",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                )
            elif bucket:
                add_issue(
                    key="shoe_hdg_ge26",
                    designation=designation,
                    project_qty=project_qty,
                    pipe_size=pipe_size,
                    matched_detail=f"Type {type_id} Pipe Shoe，{pipe_size_label(pipe_size)}",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                    note=f"Pipe Shoe 命中 {bucket} 區間，但摘要表尚無對應統計列。",
                )
            else:
                add_issue(
                    key="shoe_hdg_le4",
                    designation=designation,
                    project_qty=project_qty,
                    pipe_size=pipe_size,
                    matched_detail=f"Type {type_id} Pipe Shoe",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                    note="Pipe Shoe Type 命中，但管徑無法落入統計區間。",
                )

        if _is_cold_support_type(type_id):
            add_stat(
                "cold_support",
                project_qty,
                f"{designation}: {project_qty}組",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                unit="組",
                matched_detail=f"Type {type_id} 保冷支撐",
                material_basis="Type 代碼尾碼 C",
            )

        support_is_304 = _support_has_304_material(row_result)
        material_prefix = "ss" if support_is_304 else "cs"
        material_basis = "整組含 SUS304" if support_is_304 else "整組不含 SUS304"
        single_weight = row_result.single_result.total_weight
        scaled_weight = row_result.scaled_result.total_weight
        if single_weight <= 15:
            add_stat(
                f"{material_prefix}_support_le15",
                project_qty,
                f"{designation}: {project_qty}組，單組 {single_weight:.2f}kg",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                unit="組",
                matched_detail=f"單組總重 {single_weight:.2f}kg <= 15kg",
                material_basis=material_basis,
            )
        else:
            add_stat(
                f"{material_prefix}_support_gt15",
                scaled_weight,
                f"{designation}: {scaled_weight:.2f}kg，單組 {single_weight:.2f}kg",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                unit="KG",
                matched_detail=f"單組總重 {single_weight:.2f}kg > 15kg",
                material_basis=material_basis,
            )

        if len(details) == detail_count_before:
            add_unmatched(
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                row_result=row_result,
            )

    return stats, sources, details


def _write_leader_procurement_sheet(ws, project: ProjectAnalysisResult):
    styles = _styles()
    rows = _leader_stat_template()
    stats, _, details = _leader_procurement_stats(project)

    details_by_key: dict[str, list[LeaderHitDetail]] = {stat_row.key: [] for stat_row in rows}
    for detail in details:
        if detail.stat_key in details_by_key:
            details_by_key[detail.stat_key].append(detail)

    def stat_value(stat_row: LeaderStatRow):
        value = stats.get(stat_row.key, 0.0)
        return int(value) if stat_row.unit == "組" else round(value, 2)

    def write_detail_header(row: int) -> None:
        for col, header in enumerate(LEADER_GROUP_DETAIL_HEADERS, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = styles["subheader_fill"]
            cell.font = styles["bold_font"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]

    def write_detail_row(row: int, detail: LeaderHitDetail) -> None:
        values = [
            detail.status,
            detail.designation,
            detail.project_qty,
            "" if detail.pipe_size is None else detail.pipe_size,
            round(detail.amount, 3) if detail.unit == "KG" else int(detail.amount),
            detail.unit,
            detail.matched_detail,
            detail.material_basis,
            detail.note,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            cell.alignment = styles["wrap"]
            if col in (3, 4, 5):
                cell.alignment = styles["right"]
            if col == 1:
                if detail.status == "需確認":
                    cell.fill = styles["bad_fill"]
                elif detail.status == "未納入":
                    cell.fill = styles["warn_fill"]
                else:
                    cell.fill = styles["ok_fill"]
                cell.alignment = styles["center"]
        ws.cell(row=row, column=4).number_format = '0.##'
        ws.cell(row=row, column=5).number_format = "0.000" if detail.unit == "KG" else "0"

    _setup_sheet(ws, "支撐分類統計", "I1")
    ws.cell(
        row=2,
        column=1,
        value=(
            "業主/長官摘要：僅列本批有數量或需確認的支撐分類統計；"
            "每個統計項目下方列出命中型號與判定依據，完整查核請見「支撐統計明細」。"
        ),
    )
    ws.cell(row=2, column=1).font = styles["section_font"]
    ws.merge_cells("A2:I2")

    active_rows = [
        stat_row for stat_row in rows
        if stat_value(stat_row) != 0 or details_by_key.get(stat_row.key)
    ]

    row = 4
    if not active_rows:
        ws.cell(row=row, column=1, value="本批無需列示之支撐分類統計項目")
        ws.cell(row=row, column=1).font = styles["bold_font"]
        ws.cell(row=row, column=1).fill = styles["section_fill"]
        ws.cell(row=row, column=1).border = styles["border"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

    for stat_row in active_rows:
        stat_details = details_by_key.get(stat_row.key, [])
        summary_values = [
            stat_row.item,
            stat_row.label,
            stat_row.criteria,
            stat_row.unit,
            stat_value(stat_row),
            len(stat_details),
        ]
        for col, header in enumerate(LEADER_STAT_HEADERS, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = styles["header_fill"]
            cell.font = styles["header_font"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        note_cell = ws.cell(row=row, column=7, value="命中型號依據")
        note_cell.fill = styles["header_fill"]
        note_cell.font = styles["header_font"]
        note_cell.alignment = styles["center"]
        note_cell.border = styles["border"]

        row += 1
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["section_fill"]
            cell.font = styles["bold_font"]
            cell.border = styles["border"]
            cell.alignment = styles["wrap"]
        for col, value in enumerate(summary_values, 1):
            ws.cell(row=row, column=col, value=value)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        ws.cell(row=row, column=7, value=f"{len(stat_details)} 筆命中/確認")
        ws.cell(row=row, column=4).alignment = styles["center"]
        ws.cell(row=row, column=5).alignment = styles["right"]
        ws.cell(row=row, column=6).alignment = styles["right"]
        ws.cell(row=row, column=5).number_format = "0.00" if stat_row.unit == "KG" else "0"

        row += 1
        write_detail_header(row)
        row += 1
        for detail in stat_details:
            write_detail_row(row, detail)
            row += 1
        row += 1

    last_row = max(row - 1, 4)
    ws.freeze_panes = "A4"
    _set_widths(ws, [12, 24, 8, 10, 12, 8, 38, 24, 42])


def _write_leader_detail_sheet(ws, project: ProjectAnalysisResult):
    styles = _styles()
    _, _, details = _leader_procurement_stats(project)

    _setup_sheet(ws, "支撐統計明細（製表者查核）", "L1")
    ws.cell(
        row=2,
        column=1,
        value="本表逐筆列出支撐分類統計的命中、需確認與未納入來源；未納入者代表目前分類規則未統計該支撐。",
    )
    ws.cell(row=2, column=1).font = styles["section_font"]
    ws.merge_cells("A2:L2")
    _write_headers(ws, 3, LEADER_DETAIL_HEADERS)

    row = 4
    if not details:
        ws.cell(row=row, column=1, value="無命中資料")
        ws.cell(row=row, column=1).border = styles["border"]
        row += 1
    else:
        issue_fill = _styles()["bad_fill"]
        hit_fill = _styles()["ok_fill"]
        unmatched_fill = _styles()["warn_fill"]
        for detail in details:
            values = [
                detail.status,
                detail.category,
                detail.label,
                detail.designation,
                detail.project_qty,
                "" if detail.pipe_size is None else detail.pipe_size,
                round(detail.amount, 3) if detail.unit == "KG" else int(detail.amount),
                detail.unit,
                detail.matched_detail,
                detail.material_basis,
                detail.criteria,
                detail.note,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = styles["border"]
                cell.alignment = styles["wrap"]
                if col in (5, 6, 7):
                    cell.alignment = styles["right"]
                if col == 1:
                    if detail.status == "需確認":
                        cell.fill = issue_fill
                    elif detail.status == "未納入":
                        cell.fill = unmatched_fill
                    else:
                        cell.fill = hit_fill
                    cell.alignment = styles["center"]
            ws.cell(row=row, column=6).number_format = '0.##'
            ws.cell(row=row, column=7).number_format = "0.000" if detail.unit == "KG" else "0"
            row += 1

    last_row = max(row - 1, 3)
    _apply_table_style(ws, 3, last_row, len(LEADER_DETAIL_HEADERS))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{last_row}"
    _set_widths(ws, [10, 16, 34, 22, 8, 10, 12, 8, 36, 22, 52, 42])


def _write_project_weight_sheet(ws, project: ProjectAnalysisResult):
    _setup_sheet(ws, "重量分析明細", "R1")
    _write_headers(ws, 3, PROJECT_HEADERS)

    row = 4
    for row_result in project.rows:
        input_row = row_result.input_row
        single_result = row_result.single_result
        scaled_result = row_result.scaled_result

        if single_result.error:
            ws.cell(row=row, column=1, value=input_row.designation)
            ws.cell(row=row, column=2, value="Error")
            ws.cell(row=row, column=3, value=single_result.error)
            ws.cell(row=row, column=9, value=input_row.quantity)
            row += 1
            continue

        for single_entry, scaled_entry in zip(single_result.entries, scaled_result.entries):
            values = [
                input_row.designation,                                  # 型號 - 每列填滿
                single_entry.item_no,
                single_entry.name,
                single_entry.display_spec,
                single_entry.material,
                single_entry.length,
                single_entry.width if single_entry.width else "",
                single_entry.quantity,                                   # 單件數量
                input_row.quantity,                                      # 組數
                scaled_entry.quantity,                                  # 總數量
                single_entry.weight_output,                             # 單組重(kg)
                scaled_entry.weight_output,                             # 總重(kg)
                single_entry.category,
                single_entry.item_class,
                single_entry.manufacturing_type,
                single_entry.part_key,
                single_entry.stock_id,
                single_entry.display_remark,
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    last_row = max(row - 1, 3)
    _apply_table_style(ws, 3, last_row, len(PROJECT_HEADERS))
    _format_number_columns(ws, 4, last_row, [6, 7, 11, 12], "0.00")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:R{last_row}"
    _set_widths(ws, [20, 8, 16, 22, 14, 12, 12, 10, 8, 10, 14, 14, 10, 14, 14, 28, 12, 34])


def _write_material_summary_sheet(ws, summary: MaterialSummary):
    styles = _styles()
    _setup_sheet(ws, "材料合計與採購清單", "M1")
    _write_headers(ws, 3, SUMMARY_HEADERS)

    row = 4
    for ln in summary.lines:
        length_value = round(ln.total_length_mm, 1) if ln.aggregate_type == "linear" else ""
        qty_value = ln.piece_count if ln.aggregate_type == "linear" else ln.total_qty
        stock_length = round(ln.stock_length, 0) if ln.stock_length else ""
        sources = ", ".join(ln.source_fullstrings[:8])
        if len(ln.source_fullstrings) > 8:
            sources += f" ...+{len(ln.source_fullstrings) - 8}"
        values = [
            ln.name,
            ln.spec,
            ln.material,
            ln.category,
            ln.item_class,
            ln.manufacturing_type,
            length_value,
            qty_value,
            round(ln.total_weight, 2),
            stock_length,
            ln.purchase_qty,
            ln.purchase_unit,
            sources,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=8, value="合計總重").font = styles["bold_font"]
    ws.cell(row=total_row, column=9, value=round(summary.total_weight, 2)).font = styles["bold_font"]

    last_row = max(row - 1, 3)
    _apply_table_style(ws, 3, last_row, len(SUMMARY_HEADERS))
    _format_number_columns(ws, 4, last_row, [7, 9, 10], "0.00")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{last_row}"
    _set_widths(ws, [16, 22, 16, 10, 14, 14, 14, 12, 12, 12, 12, 8, 46])


def _write_cutting_detail_sheet(ws, plans: list[CuttingPlan]):
    styles = _styles()
    _setup_sheet(ws, "下料明細", "I1")

    row = 3
    if not plans:
        ws.cell(row=row, column=1, value="無線性材料需要下料。")
        return

    for plan in plans:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value=f"{plan.name}  {plan.spec}  ({plan.material})")
        cell.fill = styles["title_fill"]
        cell.font = styles["header_font"]
        row += 1

        ws.cell(row=row, column=1, value="需求段數")
        ws.cell(row=row, column=2, value=plan.total_pieces)
        ws.cell(row=row, column=3, value="需求總長(mm)")
        ws.cell(row=row, column=4, value=round(plan.total_demand_length, 1))
        ws.cell(row=row, column=5, value="原料根數")
        ws.cell(row=row, column=6, value=plan.total_bars)
        ws.cell(row=row, column=7, value="平均使用率")
        ws.cell(row=row, column=8, value=f"{plan.avg_utilization:.1f}%")
        row += 1

        _write_headers(ws, row, CUTTING_HEADERS)
        header_row = row
        row += 1

        for bar_idx, bar in enumerate(plan.bars, start=1):
            cumulative = 0.0
            for piece_idx, piece in enumerate(bar.pieces, start=1):
                cumulative += piece.cut_length
                values = [
                    f"#{bar_idx}" if piece_idx == 1 else "",
                    f"段 {piece_idx}",
                    round(piece.demand_length, 1),
                    round(piece.cut_length, 1),
                    round(cumulative, 1),
                    "",
                    "",
                    "",
                    piece.source,
                ]
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

            note = "廢料" if bar.remnant < 100 else "短料" if bar.remnant < 300 else ""
            fill = styles["bad_fill"] if note == "廢料" else styles["warn_fill"] if note == "短料" else styles["ok_fill"]
            values = ["", "餘料", "", "", "", round(bar.remnant, 1), f"{bar.utilization:.1f}%", note, ""]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill
            row += 1

        _apply_table_style(ws, header_row, row - 1, len(CUTTING_HEADERS))
        row += 1

    ws.freeze_panes = "A3"
    _set_widths(ws, [10, 10, 13, 13, 13, 12, 10, 10, 28])


def _write_cutting_visual_sheet(ws, plans: list[CuttingPlan]):
    styles = _styles()
    _setup_sheet(ws, "下料圖示", "AH1")
    ws.cell(row=2, column=1, value="每列代表一根原料；藍色=使用段，綠/黃/紅=餘料狀態。")
    ws.cell(row=2, column=1).font = styles["section_font"]

    headers = ["材料", "原料 #", "使用率", "餘料(mm)", "下料配置"] + ["" for _ in range(VISUAL_SLOT_COUNT - 1)] + ["用於"]
    _write_headers(ws, 4, headers)

    row = 5
    if not plans:
        ws.cell(row=row, column=1, value="無線性材料需要下料。")
        return

    for plan in plans:
        for bar_idx, bar in enumerate(plan.bars, start=1):
            used_ratio = 0 if bar.effective_length <= 0 else max(0, min(1, bar.used_length / bar.effective_length))
            used_slots = max(1, min(VISUAL_SLOT_COUNT, round(used_ratio * VISUAL_SLOT_COUNT))) if bar.pieces else 0
            remnant_fill = styles["bad_fill"] if bar.remnant < 100 else styles["warn_fill"] if bar.remnant < 300 else styles["remnant_fill"]

            ws.cell(row=row, column=1, value=f"{plan.name} {plan.spec} ({plan.material})")
            ws.cell(row=row, column=2, value=f"#{bar_idx}")
            ws.cell(row=row, column=3, value=f"{bar.utilization:.1f}%")
            ws.cell(row=row, column=4, value=round(bar.remnant, 1))

            for slot in range(VISUAL_SLOT_COUNT):
                cell = ws.cell(row=row, column=5 + slot)
                if slot < used_slots:
                    cell.fill = styles["used_fill"]
                else:
                    cell.fill = remnant_fill
                cell.border = styles["border"]

            pieces = " | ".join(f"{piece.demand_length:.0f}({piece.source})" for piece in bar.pieces)
            ws.cell(row=row, column=5 + VISUAL_SLOT_COUNT, value=pieces)
            row += 1

    from openpyxl.utils import get_column_letter
    for col in range(5, 5 + VISUAL_SLOT_COUNT):
        ws.column_dimensions[get_column_letter(col)].width = 2.2
    _set_widths(ws, [28, 10, 10, 12])
    ws.column_dimensions[get_column_letter(5 + VISUAL_SLOT_COUNT)].width = 54
    ws.freeze_panes = "A5"



# ══════════════════════════════════════════════════════════════════════
#  重量明細表 Sheet  (Pivot-friendly flat table)
# ══════════════════════════════════════════════════════════════════════

_CALC_BASIS_HEADERS = [
    "型號", "項次", "品名", "規格", "材質",
    "長度(mm)", "寬度(mm)", "屬性",
    "單件數量", "組數", "總數量",
    "單件重(kg)", "單組小計(kg)", "總重(kg)",
    "重量計算式", "物件類別", "製造方式",
]

_CONFIDENCE_FILL = {
    "精確": "C6EFCE",
    "推導": "BDD7EE",
    "估算": "FFEB9C",
    "未知": "FFC7CE",
}

_STANDARDS_TABLE = [
    ("管道重量",    "ASME B36.10M / JIS G3454",     "以管徑及 Schedule 查表取 kg/m，再乘以長度(m)"),
    ("不鏽鋼管道",  "ASME B36.19M / JIS G3459",     "SUS 系列管道；查表值已含不鏽鋼密度修正"),
    ("角鋼/槽鋼",   "JIS G3192 / CNS 2948",          "H 型鋼、角鋼、槽鋼查表取 kg/m，再乘以長度(m)"),
    ("鋼板",        "密度公式計算",                   "W = L × W × t(mm) × ρ ÷ 10⁶  [ρ = 7,850 kg/m³ (CS) / 7,930 kg/m³ (SUS)]"),
    ("膨脹螺栓",    "HILTI/RAWL 產品目錄",            "以管徑查標準五金組合重量（保守估算，不含埋入砂漿）"),
    ("自訂五金",    "工程判斷 / 廠商目錄",            "由計算程式依規格查表，如有疑義請參閱備註欄"),
    ("係數說明",    "factor = 1.0 (一般)",            "熱浸鍍鋅塗裝：+6%；特殊接頭：依廠商資料"),
    ("重量加總",    "單件數量 × 組數 = 總數量",      "單組小計 × 組數 = 總重，各項次加總為全案總重"),
]


def _weight_formula_str(entry) -> str:
    """從 AnalysisEntry 欄位重建人類可讀的計算式。"""
    qty = entry.quantity
    unit_weight = entry.unit_weight
    line_total = entry.weight_output

    if entry.unit == "M" and entry.weight_per_unit and entry.weight_per_unit > 0:
        wpm = entry.weight_per_unit
        return (
            f"{qty}件 × {entry.length:.0f}mm ÷ 1,000"
            f" × {wpm:.3f} kg/m"
            f" = {line_total:.3f} kg"
        )

    if entry.unit == "PC" and entry.width and entry.width > 0:
        try:
            t = float(entry.spec)
            density = 7.93 if "304" in entry.material else 7.82 if entry.material == "AS" else 7.85
            geometry = getattr(entry, "geometry", None)
            net_area = getattr(geometry, "net_area_mm2", 0) if geometry else 0
            formula = getattr(geometry, "formula", "") if geometry else ""
            if net_area and net_area != entry.length * entry.width:
                area_label = formula or f"{entry.length:.0f}×{entry.width:.0f}"
                return (
                    f"{qty}件 × [{area_label} = {net_area:.0f} mm2]"
                    f" × t{t:.0f} × {density:.2f} ÷ 1,000,000"
                    f" = {entry.weight_output:.3f} kg"
                )
            return (
                f"{qty}件 × {entry.length:.0f}×{entry.width:.0f}×{t:.0f}mm"
                f" × {density:.2f} t/m³"
                f" = {entry.weight_output:.3f} kg"
            )
        except (ValueError, TypeError):
            return (
                f"{qty}件 × {entry.length:.0f}×{entry.width:.0f}mm"
                f" × t×ρ = {line_total:.3f} kg"
            )

    if entry.unit in ("SET", "EA", "KG"):
        return (
            f"{qty} {entry.unit} × {unit_weight:.3f} kg/{entry.unit}"
            f" = {line_total:.3f} kg"
        )

    return f"{qty} × {unit_weight:.3f} kg = {line_total:.3f} kg"


def _confidence_label(meta: dict) -> str:
    return meta.get("truth_level", "未知")


def _write_calculation_basis_sheet(ws, project: ProjectAnalysisResult):
    """重量明細表 — Flat 表，含小計與全案合計，適合審查與樞紐分析。"""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    styles = _styles()
    qty_fill = PatternFill("solid", fgColor="EAF2F8")
    subtotal_fill = styles["subtotal_fill"]
    alt_fill = styles["zebra_fill"]
    error_fill = PatternFill("solid", fgColor="FCE4D6")
    total_fill = styles["grand_total_fill"]
    total_font = styles["grand_total_font"]

    n_cols = len(_CALC_BASIS_HEADERS)
    last_col_letter = get_column_letter(n_cols)

    subtitle = (
        f"支撐總組數 {project.total_support_count} 組    "
        f"成功項目 {len(project.rows) - len(project.errors)}    "
        f"錯誤項目 {len(project.errors)}    "
        f"全案總重 {project.total_weight:,.3f} kg"
    )
    _setup_sheet(ws, "IEC 管架支撐 - 重量明細表", f"{last_col_letter}1", subtitle=subtitle)

    HEADER_ROW = 3
    _write_headers(ws, HEADER_ROW, _CALC_BASIS_HEADERS)
    ws.row_dimensions[HEADER_ROW].height = 28

    data_row = HEADER_ROW + 1

    for row_result in project.rows:
        inp = row_result.input_row
        single = row_result.single_result
        scaled = row_result.scaled_result

        if single.error:
            vals = [inp.designation, "錯誤", single.error] + [""] * (n_cols - 3)
            vals[9] = inp.quantity
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.fill = error_fill
                cell.border = styles["border"]
                cell.alignment = styles["wrap"]
            data_row += 1
            continue

        group_start_row = data_row
        for s_entry, sc_entry in zip(single.entries, scaled.entries):
            formula_str = _weight_formula_str(s_entry)
            single_unit_w = round(s_entry.unit_weight, 3)
            single_group_w = round(s_entry.weight_output, 3)
            total_w = round(sc_entry.weight_output, 3)

            vals = [
                inp.designation, s_entry.item_no, s_entry.name,
                s_entry.display_spec, s_entry.material,
                s_entry.length if s_entry.length else "",
                s_entry.width if s_entry.width else "",
                getattr(s_entry, "category", ""),
                s_entry.quantity, inp.quantity, sc_entry.quantity,
                single_unit_w, single_group_w, total_w,
                formula_str,
                getattr(s_entry, "item_class", ""),
                getattr(s_entry, "manufacturing_type", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = styles["border"]
                if data_row % 2 == 0:
                    cell.fill = alt_fill
                if col in (9, 10, 11):
                    cell.fill = qty_fill
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="right" if col in (6, 7, 9, 10, 11, 12, 13, 14) else "left",
                    wrap_text=(col == 15),
                    indent=1 if col in (1, 3, 4, 5, 8) else 0,
                )
                if col in (9, 10, 11):
                    cell.number_format = "#,##0"
                elif col in (6, 7):
                    cell.number_format = "#,##0"
                elif col in (12, 13, 14):
                    cell.number_format = "#,##0.000"
            ws.row_dimensions[data_row].height = 16
            data_row += 1

        if data_row > group_start_row:
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=data_row, column=col)
                cell.fill = subtotal_fill
                cell.border = styles["border"]
                cell.font = styles["bold_font"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=data_row, column=1, value=f"小計 {inp.designation}")
            ws.cell(row=data_row, column=10, value=inp.quantity)
            ws.cell(row=data_row, column=14, value=round(scaled.total_weight, 3))
            ws.cell(row=data_row, column=14).number_format = "#,##0.000"
            ws.row_dimensions[data_row].height = 18
            data_row += 1

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=data_row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=data_row, column=1, value="全案合計")
    ws.cell(row=data_row, column=10, value=project.total_support_count)
    ws.cell(row=data_row, column=14, value=round(project.total_weight, 3))
    ws.cell(row=data_row, column=14).number_format = "#,##0.000"
    ws.row_dimensions[data_row].height = 24
    last_data_row = data_row

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    if last_data_row >= HEADER_ROW + 1:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{last_data_row}"

    col_widths = [20, 7, 18, 22, 14, 10, 10, 10, 10, 8, 10, 12, 13, 13, 50, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_calc_reference_sheet(ws, project: ProjectAnalysisResult):
    """計算標準與假設 — 給長官或客戶看的靜態說明頁。"""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    styles = _styles()

    _setup_sheet(ws, "計算標準與假設", "F1",
                 subtitle="本頁列示重量計算依據與資料狀態判讀依據")

    row = 4
    _section_header(ws, row, "計算標準與假設", span_cols=6)
    row += 1

    std_headers = ["計算項目", "引用標準 / 依據", "計算方式說明"]
    for col, h in enumerate(std_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = styles["subheader_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.cell(row, 3).value = "計算方式說明"
    ws.cell(row, 3).fill = styles["subheader_fill"]
    ws.cell(row, 3).font = styles["bold_font"]
    ws.cell(row, 3).alignment = styles["center"]
    ws.cell(row, 3).border = styles["border"]
    ws.row_dimensions[row].height = 24
    row += 1

    for item, standard, desc in _STANDARDS_TABLE:
        for col, val in enumerate([item, standard, desc], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3),
                                       horizontal="left", indent=1)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    _section_header(ws, row, "資料狀態圖例", span_cols=6)
    row += 1
    legends = [
        ("精確 — 直接查表", "E2EFDA"),
        ("推導 — 公式計算", "DDEBF7"),
        ("估算 — 工程假設", "FFF2CC"),
        ("未知 — 需複核",   "FCE4D6"),
    ]
    for col_off, (label, color) in enumerate(legends, 1):
        cell = ws.cell(row=row, column=col_off + 1, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.border = styles["border"]
        cell.alignment = styles["center"]
        cell.font = styles["bold_font"]
    ws.row_dimensions[row].height = 22
    row += 2

    _section_header(ws, row, "各支撐重量彙整", span_cols=6)
    row += 1

    summ_headers = ["型號", "組數", "資料狀態", "單組重(kg)", "合計重(kg)", "備註"]
    for col, h in enumerate(summ_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = styles["subheader_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.row_dimensions[row].height = 24
    row += 1

    grand_total = 0.0

    for row_result in project.rows:
        inp = row_result.input_row
        single = row_result.single_result
        scaled = row_result.scaled_result

        meta = single.meta or {}
        confidence = _confidence_label(meta)
        conf_color = _CONFIDENCE_FILL.get(confidence, "FCE4D6")

        if single.error:
            for col, val in enumerate([inp.designation, inp.quantity, "錯誤", "", "", single.error], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
                cell.border = styles["border"]
                cell.alignment = styles["center"]
        else:
            single_total = round(single.total_weight, 3)
            scaled_total = round(scaled.total_weight, 3)
            grand_total += scaled_total
            vals = [inp.designation, inp.quantity, confidence, single_total, scaled_total, ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = styles["border"]
                cell.alignment = styles["center"]
                if col in (4, 5):
                    cell.number_format = "#,##0.000"
                if col == 3:
                    cell.fill = PatternFill("solid", fgColor=conf_color)
                    cell.font = styles["bold_font"]
        row += 1

    grand_fill = styles["grand_total_fill"]
    grand_font = styles["grand_total_font"]
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = grand_fill
        cell.font = grand_font
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1, value="■ 全案合計總重")
    ws.cell(row=row, column=5, value=round(grand_total, 3))
    ws.cell(row=row, column=5).number_format = "#,##0.000"
    ws.row_dimensions[row].height = 24

    col_widths = [22, 8, 12, 14, 14, 36]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
