"""Shared openpyxl styling helpers for Excel exports."""

COLORS = {
    "ink": "1F3864",
    "ink2": "2E5395",
    "royal": "3A5BA0",
    "teal": "2E7D8A",
    "grey": "808080",
    "nav_sub": "DEE3EE",
    "section": "D9E1F2",
    "canvas": "F2F4F8",
    "zebra": "F7F9FC",
    "accent": "BF8F00",
    "gold2": "E8B923",
    "subtotal": "FFF2CC",
    "card_border": "D0D7E2",
    "grid": "BFBFBF",
    "text_mute": "595959",
    "text_note": "808080",
    "ok_fill": "E2EFDA",
    "ok_mark": "70AD47",
    "warn_fill": "FFF2CC",
    "warn_mark": "ED7D31",
    "bad_fill": "FCE4D6",
    "bad_mark": "C00000",
    "info_fill": "DDEBF7",
    "info_mark": "4472C4",
    "conf_exact": "C6EFCE",
    "conf_derive": "BDD7EE",
    "conf_estimate": "FFEB9C",
    "conf_unknown": "FFC7CE",
    "bar_used": "4472C4",
    "bar_remnant": "A9D18E",
}

NUMFMT = {
    "WEIGHT_KG": "#,##0.00",
    "WEIGHT_KG3": "#,##0.000",
    "LEN_MM": "#,##0",
    "LEN_MM1": "#,##0.0",
    "QTY_INT": "#,##0",
    "PCT": "0.0%",
    "PIPE_IN": "0.##",
    "MONEY": "#,##0",
}

FONT_CJK = "Microsoft JhengHei"

ROW_H = {
    "title": 40,
    "subtitle": 22,
    "section": 24,
    "header": 24,
    "header_tall": 28,
    "data": 18,
    "kpi_label": 18,
    "kpi_value": 34,
    "kpi_note": 16,
    "total": 24,
}

GRAD = {
    "title": ("ink", "ink2", "royal"),
    "plan": ("ink", "ink2"),
}

GLYPH = {
    "總重": "◆",
    "數量": "●",
    "良好": "▲",
    "警示": "⚠",
    "重點": "★",
    "資訊": "◇",
    "方塊": "■",
}

TAB_COLORS = {
    "manager": COLORS["ink"],
    "procurement": COLORS["teal"],
    "engineering": COLORS["grey"],
}


def _styles():
    """
    現代深靛 + 琥珀強調 配色（2026-05 更新）

    palette 設計目標：
      - 主色 1F3864 深靛：標題、表頭、合計列 → 沉穩、長官感
      - 次色 2E5395 中靛：副區塊強調
      - 強調 BF8F00 琥珀：KPI 大數字、重點數據
      - 背景 F2F4F8 極淺灰藍：副標題列、KPI 卡片底
      - zebra F7F9FC：隔行底色，極輕微對比，避免疲勞
      - 狀態色（OK 70AD47 / Warn ED7D31 / Bad C00000 / Info 4472C4）一致 PowerBI 系
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    border_thin = Side(style="thin", color=COLORS["grid"])
    border_medium = Side(style="medium", color=COLORS["ink"])
    header_border = Border(
        left=border_thin,
        right=border_thin,
        top=border_thin,
        bottom=border_medium,
    )

    return {
        # === Fills ===
        "title_fill": PatternFill("solid", fgColor=COLORS["ink"]),       # 主標題列
        "subtitle_fill": PatternFill("solid", fgColor=COLORS["canvas"]), # 副標題列（R2）
        "section_fill": PatternFill("solid", fgColor=COLORS["section"]), # 區塊標題
        "header_fill": PatternFill("solid", fgColor=COLORS["ink2"]),     # 表頭
        "subheader_fill": PatternFill("solid", fgColor=COLORS["nav_sub"]), # 次級表頭
        "kpi_label_fill": PatternFill("solid", fgColor=COLORS["canvas"]), # KPI 卡片 label
        "kpi_value_fill": PatternFill("solid", fgColor="FFFFFF"),   # KPI 卡片 value
        "zebra_fill": PatternFill("solid", fgColor=COLORS["zebra"]), # 隔行
        "subtotal_fill": PatternFill("solid", fgColor=COLORS["subtotal"]), # 小計列
        "grand_total_fill": PatternFill("solid", fgColor=COLORS["ink"]), # 全案合計
        # 狀態色
        "ok_fill": PatternFill("solid", fgColor=COLORS["ok_fill"]),
        "warn_fill": PatternFill("solid", fgColor=COLORS["warn_fill"]),
        "bad_fill": PatternFill("solid", fgColor=COLORS["bad_fill"]),
        "info_fill": PatternFill("solid", fgColor=COLORS["info_fill"]),
        # 下料圖示
        "used_fill": PatternFill("solid", fgColor=COLORS["bar_used"]),
        "remnant_fill": PatternFill("solid", fgColor=COLORS["bar_remnant"]),

        # === Borders ===
        "border": Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin),
        "header_border": header_border,
        "border_top": Border(top=border_medium),
        "border_bottom": Border(bottom=border_medium),
        "border_top_full": Border(left=border_thin, right=border_thin, top=border_medium, bottom=border_thin),

        # === Fonts ===
        "title_font": Font(name=FONT_CJK, bold=True, color="FFFFFF", size=18),
        "subtitle_font": Font(name="Calibri", color=COLORS["text_mute"], size=10, italic=True),
        "header_font": Font(name=FONT_CJK, bold=True, color="FFFFFF", size=11),
        "section_font": Font(name=FONT_CJK, bold=True, color=COLORS["ink"], size=12),
        "kpi_label_font": Font(name="Calibri", color=COLORS["text_mute"], size=10),
        "kpi_value_font": Font(name="Calibri", bold=True, color=COLORS["ink"], size=22),
        "kpi_accent_font": Font(name="Calibri", bold=True, color=COLORS["accent"], size=22),
        "kpi_bad_font": Font(name="Calibri", bold=True, color=COLORS["bad_mark"], size=22),
        "kpi_warn_font": Font(name="Calibri", bold=True, color=COLORS["warn_mark"], size=22),
        "kpi_unit_font": Font(name="Calibri", color=COLORS["text_mute"], size=10),
        "kpi_note_font": Font(name="Calibri", color=COLORS["text_note"], size=9, italic=True),
        "grand_total_font": Font(name=FONT_CJK, bold=True, color="FFFFFF", size=12),
        "bold_font": Font(name=FONT_CJK, bold=True),
        "data_font": Font(name=FONT_CJK, size=10),
        "number_font": Font(name="Calibri", size=10),

        # === Alignments ===
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center", indent=1),
        "right": Alignment(horizontal="right", vertical="center"),
        "wrap": Alignment(vertical="center", wrap_text=True),
        "wrap_top": Alignment(vertical="top", wrap_text=True),
    }


# === 通用視覺輔助 ====================================================

def _setup_sheet(
    ws,
    title: str,
    merge_to: str = "K1",
    subtitle: str = "",
    freeze_title: bool = False,
    audience: str = "",
):
    """
    設定 sheet 的標題列（R1）與可選副標題列（R2）。

    Args:
        ws: worksheet
        title: 主標題文字
        merge_to: 合併儲存格的終點（如 "K1"）
        subtitle: 可選副標題（會放在 R2，使用副標題樣式）
    """
    styles = _styles()
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{merge_to}")
    cell = ws["A1"]
    cell.value = title
    cell.font = styles["title_font"]
    cell.fill = make_gradient_fill(GRAD["title"])
    cell.alignment = styles["center"]
    ws.row_dimensions[1].height = ROW_H["title"]

    if subtitle:
        merge_end_col = merge_to.rstrip("0123456789")
        ws.merge_cells(f"A2:{merge_end_col}2")
        sub = ws["A2"]
        sub.value = f"{subtitle}    適用：{audience}" if audience else subtitle
        sub.font = styles["subtitle_font"]
        sub.fill = styles["subtitle_fill"]
        sub.alignment = Alignment_for_subtitle()
        ws.row_dimensions[2].height = ROW_H["subtitle"]

    if freeze_title:
        ws.freeze_panes = "A3"


def make_gradient_fill(stops: tuple[str, ...] | list[str], degree: int = 30):
    """Return an openpyxl gradient fill from color tokens or literal hex values."""
    from openpyxl.styles import GradientFill

    colors = [COLORS.get(stop, stop) for stop in stops]
    return GradientFill(degree=degree, stop=colors)


def set_tab_color(ws, group: str) -> None:
    color = TAB_COLORS.get(group, COLORS.get(group, group))
    if color:
        ws.sheet_properties.tabColor = color


def make_title_band(ws, title: str, merge_to: str = "K1", subtitle: str = "", **kwargs):
    """Compatibility wrapper for rich title bands."""
    _setup_sheet(ws, title, merge_to=merge_to, subtitle=subtitle, **kwargs)


def Alignment_for_subtitle():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="center", vertical="center")


def _apply_zebra(ws, start_row: int, end_row: int, max_col: int):
    """資料區隔行底色（偶數列 zebra）。需在資料寫入後呼叫。"""
    styles = _styles()
    zebra = styles["zebra_fill"]
    protected = {
        f"00{COLORS['ok_fill']}",
        f"00{COLORS['warn_fill']}",
        f"00{COLORS['bad_fill']}",
        f"00{COLORS['subtotal']}",
        f"00{COLORS['ink']}",
    }
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb in (None, "00000000", "FFFFFFFF") and fill_rgb not in protected:
                    cell.fill = zebra


def _section_header(ws, row: int, text: str, span_cols: int = 1):
    """區塊小標題列（▌ XXX），淺藍底深靛字。"""
    styles = _styles()
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row=row, column=1, value=f"▌ {text}")
    cell.font = styles["section_font"]
    cell.fill = styles["section_fill"]
    from openpyxl.styles import Alignment, Border, Side
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thick = Side(style="thick", color=COLORS["accent"])
    thin = Side(style="thin", color=COLORS["grid"])
    for col in range(1, span_cols + 1):
        ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thick)
    ws.row_dimensions[row].height = ROW_H["section"]


def _kpi_card(ws, row: int, col: int, label: str, value, unit: str = "",
              note: str = "", accent: bool = False, value_format: str = None,
              tone: str = "neutral"):
    """
    繪製 KPI 卡片（3 列：label / value+unit / note）

    Args:
        row: 起始列
        col: 起始欄
        label: 上方標題
        value: 中央大數字
        unit: 數字右側單位
        note: 下方說明（小灰字）
        accent: True 時數字用琥珀色（強調）
        value_format: openpyxl number format（如 '#,##0.00'）
    """
    styles = _styles()
    from openpyxl.styles import Alignment, Border, Side

    # 卡片邊框
    side = Side(style="thin", color=COLORS["card_border"])
    accent_side = Side(
        style="thick",
        color=COLORS["bad_mark"] if tone == "bad" else COLORS["warn_mark"] if tone == "warn" else COLORS["accent"] if accent else COLORS["ink2"],
    )
    border = Border(left=accent_side, right=side, top=side, bottom=side)
    tone_fill = {
        "bad": styles["bad_fill"],
        "warn": styles["warn_fill"],
        "neutral": styles["kpi_value_fill"],
    }.get(tone, styles["kpi_value_fill"])

    # Row 1: Label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    lcell = ws.cell(row=row, column=col, value=label)
    lcell.font = styles["kpi_label_font"]
    lcell.fill = styles["kpi_label_fill"]
    lcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lcell.border = border
    ws.row_dimensions[row].height = ROW_H["kpi_label"]

    # Row 2: Value + Unit
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    vcell = ws.cell(row=row + 1, column=col, value=value)
    if tone == "bad":
        vcell.font = styles["kpi_bad_font"]
    elif tone == "warn":
        vcell.font = styles["kpi_warn_font"]
    else:
        vcell.font = styles["kpi_accent_font"] if accent else styles["kpi_value_font"]
    vcell.fill = tone_fill
    vcell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    vcell.border = border
    if value_format:
        vcell.number_format = value_format

    ucell = ws.cell(row=row + 1, column=col + 2, value=unit)
    ucell.font = styles["kpi_unit_font"]
    ucell.fill = tone_fill
    ucell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ucell.border = border
    ws.row_dimensions[row + 1].height = ROW_H["kpi_value"]

    # Row 3: Note
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)
    ncell = ws.cell(row=row + 2, column=col, value=note)
    ncell.font = styles["kpi_note_font"]
    ncell.fill = styles["kpi_value_fill"]
    ncell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ncell.border = border
    ws.row_dimensions[row + 2].height = ROW_H["kpi_note"]


def _add_data_bar(ws, cell_range: str, color: str = COLORS["accent"]):
    """加 Excel 內建 data bar（條形圖視覺化）。"""
    try:
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(
            start_type='min', start_value=None,
            end_type='max', end_value=None,
            color=color, showValue=True,
            minLength=0, maxLength=100,
        )
        ws.conditional_formatting.add(cell_range, rule)
    except Exception:
        pass


def _format_number_block(ws, row_start: int, row_end: int, col_formats: dict):
    """批次套用 number_format。col_formats: {col_index: format_string}"""
    for r in range(row_start, row_end + 1):
        for col, fmt in col_formats.items():
            ws.cell(row=r, column=col).number_format = fmt


def _write_headers(ws, row: int, headers: list[str]):
    styles = _styles()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["header_border"]
    ws.row_dimensions[row].height = ROW_H["header"]


def _apply_table_style(ws, min_row: int, max_row: int, max_col: int):
    styles = _styles()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = styles["border"]
            if cell.row > min_row:
                cell.alignment = styles["wrap"]


def _set_widths(ws, widths: list[float]):
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _format_number_columns(ws, row_start: int, row_end: int, columns: list[int], fmt: str):
    for row in range(row_start, row_end + 1):
        for col in columns:
            ws.cell(row=row, column=col).number_format = fmt


def _format_sheet(ws, headers):
    from openpyxl.styles import Font, Alignment, PatternFill

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)


def apply_status_fill(cell, status: str, *, set_font: bool = False) -> bool:
    """Apply the workbook action-status color ramp to a cell."""
    from openpyxl.styles import Font

    styles = _styles()
    status_text = str(status or "")
    mapping = {
        "命中": ("ok_fill", "ok_mark"),
        "正常": ("ok_fill", "ok_mark"),
        "需確認": ("bad_fill", "bad_mark"),
        "錯誤": ("bad_fill", "bad_mark"),
        "廢料": ("bad_fill", "bad_mark"),
        "未納入": ("warn_fill", "warn_mark"),
        "短料": ("warn_fill", "warn_mark"),
    }
    match = mapping.get(status_text)
    if not match:
        return False
    fill_key, mark_key = match
    cell.fill = styles[fill_key]
    if set_font:
        cell.font = Font(name=FONT_CJK, bold=True, color=COLORS[mark_key])
    return True


def apply_confidence_fill(cell, level: str) -> bool:
    """Apply the data-confidence color ramp to a cell."""
    from openpyxl.styles import PatternFill

    mapping = {
        "精確": COLORS["conf_exact"],
        "推導": COLORS["conf_derive"],
        "估算": COLORS["conf_estimate"],
        "未知": COLORS["conf_unknown"],
        "錯誤": COLORS["bad_fill"],
    }
    color = mapping.get(str(level or ""))
    if not color:
        return False
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = _styles()["bold_font"]
    return True


def set_print_layout(
    ws,
    *,
    orientation: str = "landscape",
    fit_width: int = 1,
    fit_height: int | None = None,
    title_rows: str | None = "3:3",
    area: str | None = None,
    footer_title: str | None = None,
):
    """Set a consistent A4 print layout for exported report sheets."""
    import datetime as _dt
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0 if fit_height is None else fit_height
    if title_rows:
        ws.print_title_rows = title_rows
    if area:
        ws.print_area = area
    ws.oddHeader.center.text = f"製表日期 {_dt.date.today():%Y-%m-%d}"
    ws.oddFooter.left.text = footer_title or ws.title
    ws.oddFooter.center.text = ws.title
    ws.oddFooter.right.text = "Page &P / &N"


def freeze_and_filter(
    ws,
    header_row: int,
    last_row: int,
    last_col_letter: str,
    *,
    autofilter: bool = True,
):
    ws.freeze_panes = f"A{header_row + 1}"
    if autofilter and last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"


def apply_report_table(
    ws,
    header_row: int,
    headers: list[str],
    first_data_row: int,
    last_data_row: int,
    *,
    col_formats: dict[int, str] | None = None,
    widths: list[float] | None = None,
    zebra: bool = True,
    freeze: bool = True,
    autofilter: bool = True,
):
    """Apply the standard flat report table styling."""
    from openpyxl.utils import get_column_letter

    _write_headers(ws, header_row, headers)
    max_col = len(headers)
    table_last = max(last_data_row, header_row)
    _apply_table_style(ws, header_row, table_last, max_col)
    if zebra and last_data_row >= first_data_row:
        _apply_zebra(ws, first_data_row, last_data_row, max_col)
    if col_formats and last_data_row >= first_data_row:
        _format_number_block(ws, first_data_row, last_data_row, col_formats)
    if widths:
        _set_widths(ws, widths)
    if freeze:
        freeze_and_filter(
            ws,
            header_row,
            last_data_row,
            get_column_letter(max_col),
            autofilter=autofilter,
        )


def write_grand_total_band(
    ws,
    row: int,
    last_col: int,
    label: str,
    value_col: int,
    value,
    *,
    fmt: str = NUMFMT["WEIGHT_KG3"],
    label_col: int = 1,
):
    """Write a consistent deep-ink grand-total band."""
    styles = _styles()
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["grand_total_fill"]
        cell.font = styles["grand_total_font"]
        cell.border = styles["border_top_full"]
        cell.alignment = styles["center"]
    ws.cell(row=row, column=label_col, value=label)
    value_cell = ws.cell(row=row, column=value_col, value=value)
    value_cell.number_format = fmt
    value_cell.alignment = styles["right"]
    ws.row_dimensions[row].height = ROW_H["total"]


def add_color_scale(ws, cell_range: str, kind: str):
    """Add a semantic color scale to a numeric range."""
    try:
        from openpyxl.formatting.rule import ColorScaleRule

        if kind == "util":
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color=COLORS["bad_mark"],
                mid_type="num", mid_value=0.5, mid_color=COLORS["warn_mark"],
                end_type="num", end_value=1, end_color=COLORS["ok_mark"],
            )
        elif kind == "remnant":
            rule = ColorScaleRule(
                start_type="min", start_color=COLORS["bar_remnant"],
                end_type="max", end_color=COLORS["bad_mark"],
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color=COLORS["accent"],
            )
        ws.conditional_formatting.add(cell_range, rule)
    except Exception:
        pass


def write_kpi_strip(ws, row: int, specs: list[dict]):
    """Write KPI cards in 3-column slots from a list of spec dictionaries."""
    for index, spec in enumerate(specs):
        _kpi_card(ws, row, 1 + index * 3, **spec)


def write_kpi_card_v2(
    ws,
    row: int,
    col: int,
    glyph: str,
    label: str,
    value,
    unit: str,
    accent: str,
    note: str,
    *,
    big_color: str | None = None,
    value_format: str | None = None,
):
    """Write a bolder KPI card with a glyph and accent strip."""
    token_color = COLORS.get(accent, accent)
    _kpi_card(
        ws,
        row,
        col,
        f"{glyph}  {label}",
        value,
        unit,
        note=note,
        accent=token_color == COLORS["accent"],
        value_format=value_format,
    )
    from openpyxl.styles import Font, Side, Border

    value_cell = ws.cell(row=row + 1, column=col)
    if big_color:
        value_cell.font = Font(name="Calibri", bold=True, color=COLORS.get(big_color, big_color), size=24)
    thick = Side(style="thick", color=token_color)
    for r in range(row, row + 3):
        cell = ws.cell(row=r, column=col)
        old = cell.border
        cell.border = Border(left=thick, right=old.right, top=old.top, bottom=old.bottom)


def add_doughnut_chart(ws, labels_ref, data_ref, anchor: str, palette: list[str], title: str):
    """Add a native Excel doughnut chart and return it."""
    try:
        from openpyxl.chart import DoughnutChart
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.marker import DataPoint

        chart = DoughnutChart()
        chart.title = title
        chart.holeSize = 55
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.height = 7
        chart.width = 10.5
        if chart.series:
            for idx, color in enumerate(palette):
                point = DataPoint(idx=idx)
                point.graphicalProperties.solidFill = COLORS.get(color, color)
                chart.series[0].data_points.append(point)
        ws.add_chart(chart, anchor)
        return chart
    except Exception:
        return None


def add_bar_chart(ws, labels_ref, data_ref, anchor: str, color: str, title: str, horizontal: bool = True):
    """Add a native Excel bar chart and return it."""
    try:
        from openpyxl.chart import BarChart
        from openpyxl.chart.label import DataLabelList

        chart = BarChart()
        chart.type = "bar" if horizontal else "col"
        chart.title = title
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.legend = None
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.height = 7
        chart.width = 11
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = COLORS.get(color, color)
        ws.add_chart(chart, anchor)
        return chart
    except Exception:
        return None


def apply_icon_set(ws, cell_range: str, kind: str = "3TrafficLights1", thresholds: list[int] | None = None):
    """Add a semantic icon set to a numeric range."""
    try:
        from openpyxl.formatting.rule import IconSetRule

        ws.conditional_formatting.add(
            cell_range,
            IconSetRule(kind, "percent", thresholds or [0, 40, 70], showValue=True),
        )
    except Exception:
        pass
