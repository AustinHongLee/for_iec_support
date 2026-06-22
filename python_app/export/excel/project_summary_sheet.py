"""Renderer for the project summary sheet."""

from core.cutting_optimizer import CuttingPlan
from core.material_summary import MaterialSummary
from core.parser import get_type_code
from core.project_aggregation import ProjectAnalysisResult

from .confidence_summary import (
    confidence_level_for_row,
    format_confidence_counts,
    worst_confidence_level,
)
from .navigation import workbook_navigation
from .styles import (
    GLYPH,
    NUMFMT,
    _add_data_bar,
    _section_header,
    _set_widths,
    _setup_sheet,
    _styles,
    apply_confidence_fill,
    set_print_layout,
    write_kpi_card_v2,
)


def _type_sort_key(type_id: str) -> tuple[int, str]:
    raw = str(type_id or "").rstrip("C")
    return (int(raw) if raw.isdigit() else 9999, str(type_id or ""))


def _project_type_summary(project: ProjectAnalysisResult) -> list[dict]:
    stats: dict[str, dict] = {}
    for row_result in project.rows:
        designation = row_result.input_row.designation
        type_id = get_type_code(designation) or "未解析"
        stat = stats.setdefault(
            type_id,
            {
                "type_id": type_id,
                "support_count": 0,
                "row_count": 0,
                "total_weight": 0.0,
                "status_counts": {
                    "精確": 0,
                    "推導": 0,
                    "估算": 0,
                    "未知": 0,
                    "錯誤": 0,
                },
                "designations": [],
            },
        )
        stat["support_count"] += row_result.input_row.quantity
        stat["row_count"] += 1
        stat["status_counts"][confidence_level_for_row(row_result)] += 1
        if not row_result.single_result.error:
            stat["total_weight"] += row_result.scaled_result.total_weight
        if designation not in stat["designations"]:
            stat["designations"].append(designation)
    return [stats[key] for key in sorted(stats, key=_type_sort_key)]


def _write_table_header(ws, row: int, headers: list[str], *, span_cols: int) -> None:
    styles = _styles()
    for col in range(1, span_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    ws.row_dimensions[row].height = 24


def _write_index_row(ws, row: int, sheet: str, purpose: str, print_note: str) -> None:
    from openpyxl.styles import Alignment, Font

    styles = _styles()
    ws.cell(row=row, column=1, value=sheet)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=purpose)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    ws.cell(row=row, column=7, value=print_note)
    for col in (1, 2, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    name_cell = ws.cell(row=row, column=1)
    name_cell.font = Font(name="Microsoft JhengHei", bold=True, color="1F3864", underline="single")
    name_cell.hyperlink = f"#'{sheet}'!A1"
    name_cell.style = "Hyperlink"
    ws.row_dimensions[row].height = 28


def _package_navigation(total_bars: int, available_sheets: tuple[str, ...] | None) -> list:
    items = workbook_navigation(total_bars)
    if available_sheets is None:
        return items
    return [item for item in items if item.sheet in available_sheets]


def _write_project_summary_sheet(
    ws,
    project: ProjectAnalysisResult,
    summary: MaterialSummary,
    plans: list[CuttingPlan],
    *,
    available_sheets: tuple[str, ...] | None = None,
):
    """專案摘要 — A4 直式、以專案 Type 與查閱路徑為主。"""
    import datetime as _dt
    from openpyxl.styles import Alignment, Font

    styles = _styles()
    ws.title = "專案摘要"

    type_rows = _project_type_summary(project)
    type_count = len(type_rows)
    total_rows = len(project.rows)
    total_weight = summary.total_weight if summary.total_weight > 0 else 0.0
    successful_rows = [r for r in project.rows if not r.single_result.error]
    total_bars = sum(plan.total_bars for plan in plans)

    try:
        from .leader_sheets import _leader_procurement_stats

        _, _, leader_details = _leader_procurement_stats(project)
        confirm_count = sum(1 for detail in leader_details if detail.status == "需確認")
    except Exception:
        confirm_count = 0

    today_str = _dt.date.today().strftime("%Y-%m-%d")
    subtitle = (
        f"製表日期 {today_str}    "
        f"支撐 {project.total_support_count:,} 組    "
        f"使用 Type {type_count:,} 種    "
        f"全案總重 {summary.total_weight:,.2f} kg"
    )
    _setup_sheet(ws, "專案材料統計總覽", "I1", subtitle=subtitle, audience="主管 / 工程 / 採購", freeze_title=True)
    _set_widths(ws, [14, 11, 10, 13, 10, 20, 13, 13, 31])

    # A4 portrait first page summary.
    _section_header(ws, 4, "專案概況", span_cols=9)
    write_kpi_card_v2(
        ws, 5, 1, GLYPH["數量"], "支撐總組數",
        project.total_support_count, "組", "ink2",
        "此專案納入統計的支撐數", big_color="ink", value_format=NUMFMT["QTY_INT"],
    )
    write_kpi_card_v2(
        ws, 5, 4, GLYPH["資訊"], "使用 Type 數",
        type_count, "種", "royal",
        "由型號第一段彙整", big_color="royal", value_format=NUMFMT["QTY_INT"],
    )
    write_kpi_card_v2(
        ws, 5, 7, GLYPH["總重"], "全案總重",
        round(summary.total_weight, 2), "kg", "accent",
        "以材料合計頁為準", big_color="accent", value_format=NUMFMT["WEIGHT_KG"],
    )
    write_kpi_card_v2(
        ws, 9, 1, GLYPH["方塊"], "型號列數",
        total_rows, "列", "info_mark",
        "輸入型號列數，不等於支撐組數", big_color="info_mark", value_format=NUMFMT["QTY_INT"],
    )
    write_kpi_card_v2(
        ws, 9, 4, GLYPH["警示"], "錯誤項目",
        len(project.errors), "項", "bad_mark" if project.errors else "ink2",
        "計算失敗需先處理", big_color="bad_mark" if project.errors else "ink", value_format=NUMFMT["QTY_INT"],
    )
    write_kpi_card_v2(
        ws, 9, 7, GLYPH["警示"], "需確認分類",
        confirm_count, "筆", "warn_mark" if confirm_count else "ink2",
        "支撐分類統計規則需人工確認", big_color="warn_mark" if confirm_count else "ink", value_format=NUMFMT["QTY_INT"],
    )

    row = 14
    _section_header(ws, row, "使用 Type 統計", span_cols=9)
    row += 1
    _write_table_header(ws, row, ["型號類別", "支撐組數", "型號列數", "總重(kg)", "佔比", "資料狀態", "代表型號"], span_cols=9)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    ws.cell(row=row, column=7).alignment = styles["center"]
    row += 1

    total_weight_for_pct = total_weight or 1.0
    for stat in type_rows:
        status_counts = stat["status_counts"]
        ws.cell(row=row, column=1, value=f"Type {stat['type_id']}")
        ws.cell(row=row, column=2, value=stat["support_count"])
        ws.cell(row=row, column=3, value=stat["row_count"])
        ws.cell(row=row, column=4, value=round(stat["total_weight"], 2))
        ws.cell(row=row, column=5, value=stat["total_weight"] / total_weight_for_pct)
        ws.cell(row=row, column=6, value=format_confidence_counts(status_counts))
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        examples = "、".join(stat["designations"][:3])
        if len(stat["designations"]) > 3:
            examples += f" 等 {len(stat['designations'])} 種"
        ws.cell(row=row, column=7, value=examples)
        for col in (1, 2, 3, 4, 5, 6, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.alignment = Alignment(
                horizontal="right" if col in (2, 3, 4, 5) else "left",
                vertical="center",
                wrap_text=(col in (6, 7)),
                indent=1 if col in (1, 6, 7) else 0,
            )
        apply_confidence_fill(ws.cell(row=row, column=6), worst_confidence_level(status_counts))
        ws.cell(row=row, column=2).number_format = NUMFMT["QTY_INT"]
        ws.cell(row=row, column=3).number_format = NUMFMT["QTY_INT"]
        ws.cell(row=row, column=4).number_format = NUMFMT["WEIGHT_KG"]
        ws.cell(row=row, column=5).number_format = NUMFMT["PCT"]
        if (row - 16) % 2 == 1:
            for col in (1, 2, 3, 4, 5, 7):
                ws.cell(row=row, column=col).fill = styles["zebra_fill"]
        ws.row_dimensions[row].height = 24
        row += 1

    if type_rows:
        _add_data_bar(ws, f"D16:D{row - 1}", color="4472C4")

    row += 2
    _section_header(ws, row, "分頁導覽", span_cols=9)
    row += 1
    _write_table_header(ws, row, ["分頁", "適合查看", "", "", "", "", "列印建議"], span_cols=9)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    row += 1
    for nav_item in _package_navigation(total_bars, available_sheets):
        _write_index_row(ws, row, nav_item.sheet, nav_item.purpose, nav_item.print_note)
        row += 1

    row += 2
    _section_header(ws, row, "輔助觀察（非主要判斷）", span_cols=9)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    note = ws.cell(
        row=row,
        column=1,
        value="重量排行與材料分佈保留在材料合計、重量明細等分頁；摘要頁不再把它們當成主要結論。",
    )
    note.font = Font(name="Microsoft JhengHei", size=10, italic=True, color="595959")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 24
    row += 2

    _section_header(ws, row, "報表備註", span_cols=9)
    notes = [
        "材料與重量計算以各 Type calculator、component table 與材料合計頁為準。",
        "Type 統計依型號第一段彙整；例如 15-... 歸入 Type 15。",
    ]
    if available_sheets is None or "查核-支撐明細" in available_sheets:
        notes.append("若分類顯示需確認，請先看「查核-支撐明細」，再決定是否更新分類規則。")
    else:
        notes.append("若分類顯示需確認，請開啟完整活頁簿或採購材料包查看「查核-支撐明細」。")
    for note in notes:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value=f"・{note}")
        cell.font = Font(name="Microsoft JhengHei", size=10, color="595959")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 20

    set_print_layout(ws, orientation="portrait", title_rows=None, area=f"A1:I{row}", footer_title="專案摘要")
