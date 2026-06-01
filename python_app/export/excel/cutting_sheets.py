"""Renderers for cutting detail and visual sheets."""

from core.cutting_optimizer import CuttingPlan

from .headers import CUTTING_HEADERS, VISUAL_SLOT_COUNT
from .styles import (
    COLORS,
    NUMFMT,
    add_color_scale,
    apply_status_fill,
    set_print_layout,
    _apply_table_style,
    _set_widths,
    _setup_sheet,
    _styles,
    _write_headers,
)


def _write_cutting_detail_sheet(ws, plans: list[CuttingPlan]):
    from openpyxl.styles import Alignment

    styles = _styles()
    _setup_sheet(ws, "下料明細", "I1")

    row = 3
    if not plans:
        ws.cell(row=row, column=1, value="無線性材料需要下料。")
        set_print_layout(ws, title_rows=None, area="A1:I3", footer_title="下料明細")
        return

    for plan in plans:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value=f"{plan.name}  {plan.spec}  ({plan.material})")
        cell.fill = styles["title_fill"]
        cell.font = styles["header_font"]
        row += 1

        summary_row = row
        ws.cell(row=row, column=1, value="需求段數")
        ws.cell(row=row, column=2, value=plan.total_pieces)
        ws.cell(row=row, column=3, value="需求總長(mm)")
        ws.cell(row=row, column=4, value=round(plan.total_demand_length, 1))
        ws.cell(row=row, column=5, value="原料根數")
        ws.cell(row=row, column=6, value=plan.total_bars)
        ws.cell(row=row, column=7, value="平均使用率")
        ws.cell(row=row, column=8, value=plan.avg_utilization / 100 if plan.avg_utilization else 0)
        for col in range(1, 9):
            cell = ws.cell(row=summary_row, column=col)
            cell.fill = styles["subtitle_fill"]
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col % 2 else styles["right"]
            if col % 2 == 1:
                cell.font = styles["bold_font"]
        ws.cell(row=summary_row, column=2).number_format = NUMFMT["QTY_INT"]
        ws.cell(row=summary_row, column=4).number_format = NUMFMT["LEN_MM1"]
        ws.cell(row=summary_row, column=6).number_format = NUMFMT["QTY_INT"]
        ws.cell(row=summary_row, column=8).number_format = NUMFMT["PCT"]
        row += 1

        _write_headers(ws, row, CUTTING_HEADERS)
        header_row = row
        row += 1

        for bar_idx, bar in enumerate(plan.bars, start=1):
            cumulative = 0.0
            for piece_idx, piece in enumerate(bar.pieces, start=1):
                cumulative += piece.cut_length
                values = [
                    f"#{bar_idx}" if piece_idx == 1 else "",
                    f"段 {piece_idx}",
                    round(piece.demand_length, 1),
                    round(piece.cut_length, 1),
                    round(cumulative, 1),
                    "",
                    "",
                    "",
                    piece.source,
                ]
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

            status = "廢料" if bar.remnant < 100 else "短料" if bar.remnant < 300 else "正常"
            note = "" if status == "正常" else status
            values = ["", "餘料", "", "", "", round(bar.remnant, 1), bar.utilization / 100, note, ""]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                apply_status_fill(cell, status)
            row += 1

        _apply_table_style(ws, header_row, row - 1, len(CUTTING_HEADERS))
        for r in range(header_row + 1, row):
            for col in (3, 4, 5, 6):
                ws.cell(row=r, column=col).number_format = NUMFMT["LEN_MM1"]
                ws.cell(row=r, column=col).alignment = styles["right"]
            ws.cell(row=r, column=7).number_format = NUMFMT["PCT"]
            ws.cell(row=r, column=7).alignment = styles["right"]
        if row - 1 >= header_row + 1:
            add_color_scale(ws, f"G{header_row + 1}:G{row - 1}", "util")
        row += 1

    ws.freeze_panes = "A3"
    _set_widths(ws, [10, 10, 13, 13, 13, 12, 10, 10, 28])
    set_print_layout(ws, title_rows=None, area=f"A1:I{row - 1}", footer_title="下料明細")


def _write_cutting_visual_sheet(ws, plans: list[CuttingPlan]):
    from openpyxl.styles import Alignment, Border, Font, Side
    styles = _styles()
    _setup_sheet(ws, "下料圖示", "AH1")
    ws.cell(row=2, column=1, value="每列代表一根原料；藍色=使用段，綠/黃/紅=餘料狀態。")
    ws.cell(row=2, column=1).font = styles["section_font"]

    scale_points = [(5, "0%"), (5 + VISUAL_SLOT_COUNT // 2, "50%"), (4 + VISUAL_SLOT_COUNT, "100%")]
    for col, label in scale_points:
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name="Calibri", size=9, color=COLORS["text_note"], italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["材料", "原料 #", "使用率", "餘料(mm)", "下料配置"] + ["" for _ in range(VISUAL_SLOT_COUNT - 1)] + ["用於"]
    _write_headers(ws, 4, headers)

    row = 5
    if not plans:
        ws.cell(row=row, column=1, value="無線性材料需要下料。")
        set_print_layout(ws, title_rows=None, area="A1:AI5", footer_title="下料圖示")
        return

    slot_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )

    for plan in plans:
        for bar_idx, bar in enumerate(plan.bars, start=1):
            used_ratio = 0 if bar.effective_length <= 0 else max(0, min(1, bar.used_length / bar.effective_length))
            used_slots = max(1, min(VISUAL_SLOT_COUNT, round(used_ratio * VISUAL_SLOT_COUNT))) if bar.pieces else 0
            remnant_fill = styles["bad_fill"] if bar.remnant < 100 else styles["warn_fill"] if bar.remnant < 300 else styles["remnant_fill"]

            ws.cell(row=row, column=1, value=f"{plan.name} {plan.spec} ({plan.material})")
            ws.cell(row=row, column=2, value=f"#{bar_idx}")
            ws.cell(row=row, column=3, value=bar.utilization / 100)
            ws.cell(row=row, column=4, value=round(bar.remnant, 1))
            ws.cell(row=row, column=3).number_format = NUMFMT["PCT"]
            ws.cell(row=row, column=4).number_format = NUMFMT["LEN_MM1"]

            for slot in range(VISUAL_SLOT_COUNT):
                cell = ws.cell(row=row, column=5 + slot)
                if slot < used_slots:
                    cell.fill = styles["used_fill"]
                else:
                    cell.fill = remnant_fill
                cell.border = slot_border

            pieces = " | ".join(f"{piece.demand_length:.0f}({piece.source})" for piece in bar.pieces)
            ws.cell(row=row, column=5 + VISUAL_SLOT_COUNT, value=pieces)
            row += 1

    last_row = max(row - 1, 5)
    if last_row >= 5:
        add_color_scale(ws, f"C5:C{last_row}", "util")
    from openpyxl.utils import get_column_letter
    for col in range(5, 5 + VISUAL_SLOT_COUNT):
        ws.column_dimensions[get_column_letter(col)].width = 2.2
    _set_widths(ws, [28, 10, 10, 12])
    ws.column_dimensions[get_column_letter(5 + VISUAL_SLOT_COUNT)].width = 54
    ws.freeze_panes = "A5"
    set_print_layout(ws, title_rows=None, area=f"A1:{get_column_letter(5 + VISUAL_SLOT_COUNT)}{last_row}", footer_title="下料圖示")



# ══════════════════════════════════════════════════════════════════════
#  重量明細表 Sheet  (Pivot-friendly flat table)
# ══════════════════════════════════════════════════════════════════════
