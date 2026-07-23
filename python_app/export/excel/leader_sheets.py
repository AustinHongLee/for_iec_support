"""Renderers and rules for leader-facing support classification sheets."""

from core.parser import get_lookup_value, get_part
from core.project_aggregation import ProjectAnalysisResult

from .headers import LEADER_DETAIL_HEADERS
from .models import LeaderHitDetail, LeaderStatRow
from .styles import (
    NUMFMT,
    apply_report_table,
    apply_status_fill,
    set_print_layout,
    _set_widths,
    _setup_sheet,
    _styles,
)


def _leader_stat_template() -> list[LeaderStatRow]:
    return [
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band <= 6" 熱浸鍍鋅', "組", "uband_hdg_le6", '明細含 U-Bolt/Band，管徑 <= 6"，材質非 SUS304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band >= 8" 熱浸鍍鋅', "組", "uband_hdg_ge8", '明細含 U-Bolt/Band，管徑 >= 8"，材質非 SUS304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band <= 6" (SUS304)', "組", "uband_304_le6", '明細含 U-Bolt/Band，管徑 <= 6"，材質含 304'),
        LeaderStatRow("U-Bolt / Band", 'U-Bolt & Band >= 8" (SUS304)', "組", "uband_304_ge8", '明細含 U-Bolt/Band，管徑 >= 8"，材質含 304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE <= 4" 熱浸鍍鋅', "組", "shoe_hdg_le4", 'Type 52/53/54/55/66/67/80/85，管徑 <= 4"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE 5"~10" 熱浸鍍鋅', "組", "shoe_hdg_5_10", 'Type 52/53/54/55/66/67/80/85，管徑 5"~10"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE 12"~24" 熱浸鍍鋅', "組", "shoe_hdg_12_24", 'Type 52/53/54/55/66/67/80/85，管徑 12"~24"，整組不含 SUS304'),
        LeaderStatRow("Pipe Shoe", 'PIPE SHOE >= 26" 熱浸鍍鋅', "組", "shoe_hdg_ge26", 'Type 52/53/54/55/66/67/80/85，管徑 >= 26"，整組不含 SUS304'),
        LeaderStatRow("Cold Support", "保冷支撐座（長春帶料）", "組", "cold_support", "Type 代碼尾碼為 C 的保冷支撐"),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE <= 4" (SUS304)', "組", "shoe_304_le4", 'Type 52/53/54/55/66/67/80/85，管徑 <= 4"，整組任一明細材質含 304'),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE 5"~10" (SUS304)', "組", "shoe_304_5_10", 'Type 52/53/54/55/66/67/80/85，管徑 5"~10"，整組任一明細材質含 304'),
        LeaderStatRow("Pipe Shoe SUS304", 'PIPE SHOE 12"~24" (SUS304)', "組", "shoe_304_12_24", 'Type 52/53/54/55/66/67/80/85，管徑 12"~24"，整組任一明細材質含 304'),
        LeaderStatRow("CS Support", "CS 管支撐製裝 <= 15 kg/組", "組", "cs_support_le15", "製裝分類不因 SUS304 另分；單組總重 <= 15 kg，按支撐組數統計"),
        LeaderStatRow("CS Support", "CS 管支撐製裝 > 15 kg/組", "KG", "cs_support_gt15", "製裝分類不因 SUS304 另分；單組總重 > 15 kg，按總重量統計"),
    ]


def _parse_designation_type(designation: str) -> str:
    return (get_part(designation, 1) or "").strip()


def _parse_designation_pipe_size(designation: str) -> float | None:
    token = (get_part(designation, 2) or "").strip()
    if not token:
        return None
    token = token.split("(")[0].replace('"', "").strip()
    if token.upper().endswith("B"):
        token = token[:-1]
    try:
        return float(get_lookup_value(token))
    except (TypeError, ValueError):
        return None


def _is_304_material(material: str) -> bool:
    return "304" in str(material or "").upper().replace(" ", "")


def _is_ubolt_or_band_entry(name: str) -> bool:
    upper = str(name or "").upper().replace(" ", "")
    return "U-BOLT" in upper or "UBOLT" in upper or "U-BAND" in upper or "UBAND" in upper


def _support_has_304_material(row_result) -> bool:
    return any(_is_304_material(entry.material) for entry in row_result.scaled_result.entries)


def _is_cold_support_type(type_id: str) -> bool:
    return type_id.endswith("C") and type_id[:-1].isdigit()


def _leader_size_bucket(size: float | None, buckets: tuple[tuple[str, float, float], ...]) -> str | None:
    if size is None:
        return None
    for key, lower, upper in buckets:
        if lower <= size <= upper:
            return key
    return None


def _leader_procurement_stats(
    project: ProjectAnalysisResult,
) -> tuple[dict[str, float], dict[str, list[str]], list[LeaderHitDetail]]:
    template = _leader_stat_template()
    template_by_key = {row.key: row for row in template}
    stats = {row.key: 0.0 for row in template}
    sources = {row.key: [] for row in template}
    details: list[LeaderHitDetail] = []
    pipe_shoe_types = {"52", "53", "54", "55", "66", "67", "80", "85"}
    ubolt_contract_types = {"57", "58"}
    current_drawing_line_number = ""
    current_single_weight = 0.0
    current_project_weight = 0.0

    def material_label(is_304: bool) -> str:
        return "SUS304" if is_304 else "HDG/CS"

    def pipe_size_label(size: float | None) -> str:
        return "" if size is None else f'{size:g}"'

    def add_detail(
        *,
        status: str,
        key: str,
        serial: str,
        designation: str,
        project_qty: int,
        source_unit: str,
        pipe_size: float | None,
        amount: float,
        unit: str,
        matched_detail: str,
        material_basis: str,
        note: str = "",
        drawing_line_number: str = "",
    ) -> None:
        stat_row = template_by_key.get(key)
        if stat_row is None:
            return
        if status != "命中":
            claim_calculation = note or "未計入；需先確認分類"
        elif key == "cs_support_gt15":
            claim_calculation = (
                f"{current_single_weight:.3f} kg/組 × {project_qty:g} 組"
                f" = {amount:.3f} kg"
            )
        elif key == "cs_support_le15":
            claim_calculation = (
                f"單組 {current_single_weight:.3f} kg ≤ 15 kg；"
                f"1 組/支撐 × {project_qty:g} = {amount:g} 組"
            )
        elif key.startswith("uband_"):
            per_support = amount / project_qty if project_qty else 0.0
            claim_calculation = (
                f"單組命中 {per_support:g} {unit} × {project_qty:g} 組"
                f" = {amount:g} {unit}"
            )
        else:
            claim_calculation = f"1 組/支撐 × {project_qty:g} = {amount:g} {unit}"
        details.append(
            LeaderHitDetail(
                stat_key=stat_row.key,
                status=status,
                category=stat_row.item,
                label=stat_row.label,
                drawing_line_number=drawing_line_number or current_drawing_line_number,
                serial=serial,
                source_unit=source_unit or "組",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                amount=amount,
                unit=unit,
                matched_detail=matched_detail,
                material_basis=material_basis,
                criteria=stat_row.criteria,
                note=note,
                single_weight=current_single_weight,
                project_weight=current_project_weight,
                claim_calculation=claim_calculation,
            )
        )

    def add_stat(key: str, amount: float, source: str, **detail_kwargs) -> None:
        stats[key] += amount
        if source and source not in sources[key]:
            sources[key].append(source)
        add_detail(status="命中", key=key, amount=amount, **detail_kwargs)

    def add_issue(
        *,
        key: str,
        serial: str,
        designation: str,
        project_qty: int,
        source_unit: str,
        pipe_size: float | None,
        matched_detail: str,
        material_basis: str,
        note: str,
    ) -> None:
        add_detail(
            status="需確認",
            key=key,
            serial=serial,
            designation=designation,
            project_qty=project_qty,
            source_unit=source_unit,
            pipe_size=pipe_size,
            amount=0.0,
            unit="",
            matched_detail=matched_detail,
            material_basis=material_basis,
            note=note,
        )

    def add_unmatched(
        *,
        designation: str,
        project_qty: int,
        serial: str,
        source_unit: str,
        pipe_size: float | None,
        row_result,
    ) -> None:
        has_304 = _support_has_304_material(row_result)
        entry_names = "、".join(entry.name for entry in row_result.scaled_result.entries[:5])
        if len(row_result.scaled_result.entries) > 5:
            entry_names += "..."
        details.append(
            LeaderHitDetail(
                stat_key="unmatched",
                status="未納入",
                category="未納入支撐分類統計",
                label="未命中摘要規則",
                drawing_line_number=current_drawing_line_number,
                serial=serial,
                source_unit=source_unit or "組",
                designation=designation,
                project_qty=project_qty,
                pipe_size=pipe_size,
                amount=0.0,
                unit="",
                matched_detail=entry_names or "無材料明細",
                material_basis="整組含 SUS304" if has_304 else "整組不含 SUS304",
                criteria="目前摘要統計 U-Bolt/Band、Pipe Shoe、Cold Support、CS 管支撐製裝",
                note=(
                    "整組含 SUS304，但未符合目前摘要規則，請確認是否需要新增採購/製裝分類。"
                    if has_304
                    else "未符合目前支撐分類統計規則，請確認是否需要新增採購/製裝分類。"
                ),
                single_weight=current_single_weight,
                project_weight=current_project_weight,
                claim_calculation="未計入；需先確認是否屬於合約請款項目",
            )
        )

    for row_result in project.rows:
        source_designation = row_result.input_row.designation
        designation = row_result.input_row.display_designation or source_designation
        project_qty = row_result.input_row.quantity
        current_drawing_line_number = row_result.input_row.drawing_line_number
        source_serial = row_result.input_row.serial
        source_unit = row_result.input_row.unit or "組"
        current_single_weight = row_result.single_result.total_weight
        current_project_weight = row_result.scaled_result.total_weight
        type_id = _parse_designation_type(source_designation)
        pipe_size = _parse_designation_pipe_size(source_designation)
        detail_count_before = len(details)
        is_separate_contract_item = (
            type_id in ubolt_contract_types
            or type_id in pipe_shoe_types
            or _is_cold_support_type(type_id)
        )

        if row_result.single_result.error:
            add_issue(
                key="cs_support_le15",
                serial=source_serial,
                designation=designation,
                project_qty=project_qty,
                source_unit=source_unit,
                pipe_size=pipe_size,
                matched_detail="分析失敗",
                material_basis="",
                note=row_result.single_result.error,
            )
            continue

        for entry in row_result.scaled_result.entries:
            if not _is_ubolt_or_band_entry(entry.name):
                continue
            entry_is_304 = _is_304_material(entry.material)
            material_key = "304" if entry_is_304 else "hdg"
            bucket = _leader_size_bucket(pipe_size, (
                ("le6", 0.0, 6.0),
                ("ge8", 8.0, 999.0),
            ))
            if bucket:
                key = f"uband_{material_key}_{bucket}"
                add_stat(
                    key,
                    entry.quantity,
                    f"{designation} ×{project_qty}: {entry.name} {entry.quantity:g}{entry.unit}",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    unit=entry.unit,
                    matched_detail=(
                        f"項次{entry.item_no} {entry.name} {entry.display_spec}"
                        f" ×{entry.quantity:g}"
                    ),
                    material_basis=f"{entry.material} -> {material_label(entry_is_304)}",
                )
            else:
                add_issue(
                    key=f"uband_{material_key}_le6",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    matched_detail=f"項次{entry.item_no} {entry.name} {entry.display_spec}",
                    material_basis=f"{entry.material} -> {material_label(entry_is_304)}",
                    note="U-Bolt/Band 命中，但管徑無法落入 <=6 或 >=8 統計區間。",
                )

        if type_id in pipe_shoe_types:
            support_is_304 = _support_has_304_material(row_result)
            material_key = "304" if support_is_304 else "hdg"
            bucket = _leader_size_bucket(pipe_size, (
                ("le4", 0.0, 4.0),
                ("5_10", 5.0, 10.0),
                ("12_24", 12.0, 24.0),
                ("ge26", 26.0, 999.0),
            ))
            if bucket and f"shoe_{material_key}_{bucket}" in stats:
                key = f"shoe_{material_key}_{bucket}"
                add_stat(
                    key,
                    project_qty,
                    f"{designation}: {project_qty}組",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    unit="組",
                    matched_detail=f"Type {type_id} Pipe Shoe，{pipe_size_label(pipe_size)}",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                )
            elif bucket:
                add_issue(
                    key="shoe_hdg_ge26",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    matched_detail=f"Type {type_id} Pipe Shoe，{pipe_size_label(pipe_size)}",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                    note=f"Pipe Shoe 命中 {bucket} 區間，但摘要表尚無對應統計列。",
                )
            else:
                add_issue(
                    key="shoe_hdg_le4",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    matched_detail=f"Type {type_id} Pipe Shoe",
                    material_basis=f"整組材質 -> {material_label(support_is_304)}",
                    note="Pipe Shoe Type 命中，但管徑無法落入統計區間。",
                )

        if _is_cold_support_type(type_id):
            add_stat(
                "cold_support",
                project_qty,
                f"{designation}: {project_qty}組",
                serial=source_serial,
                designation=designation,
                project_qty=project_qty,
                source_unit=source_unit,
                pipe_size=pipe_size,
                unit="組",
                matched_detail=f"Type {type_id} 保冷支撐",
                material_basis="Type 代碼尾碼 C",
            )

        if not is_separate_contract_item:
            support_is_304 = _support_has_304_material(row_result)
            material_prefix = "cs"
            material_basis = (
                "整組含 SUS304；依本批業主口徑併入 CS 管支撐製裝"
                if support_is_304
                else "整組不含 SUS304"
            )
            single_weight = row_result.single_result.total_weight
            scaled_weight = row_result.scaled_result.total_weight
            if single_weight <= 15:
                add_stat(
                    f"{material_prefix}_support_le15",
                    project_qty,
                    f"{designation}: {project_qty}組，單組 {single_weight:.2f}kg",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    unit="組",
                    matched_detail=f"單組總重 {single_weight:.2f}kg <= 15kg",
                    material_basis=material_basis,
                )
            else:
                add_stat(
                    f"{material_prefix}_support_gt15",
                    scaled_weight,
                    f"{designation}: {scaled_weight:.2f}kg，單組 {single_weight:.2f}kg",
                    serial=source_serial,
                    designation=designation,
                    project_qty=project_qty,
                    source_unit=source_unit,
                    pipe_size=pipe_size,
                    unit="KG",
                    matched_detail=f"單組總重 {single_weight:.2f}kg > 15kg",
                    material_basis=material_basis,
                )

        if len(details) == detail_count_before:
            add_unmatched(
                designation=designation,
                project_qty=project_qty,
                serial=source_serial,
                source_unit=source_unit,
                pipe_size=pipe_size,
                row_result=row_result,
            )

    return stats, sources, details


def _ordered_leader_details(details: list[LeaderHitDetail]) -> list[LeaderHitDetail]:
    """Keep every contract item's evidence contiguous for one-click claim review."""
    order = {item.key: index for index, item in enumerate(_leader_stat_template())}
    return sorted(
        details,
        key=lambda detail: (
            order.get(detail.stat_key, len(order)),
            detail.drawing_line_number or "",
            str(detail.serial or ""),
            detail.designation or "",
            detail.status or "",
        ),
    )


def _boss_summary_rows(stats: dict[str, float]) -> list[dict]:
    """Fixed leader-facing summary rows. Details remain in 查核-支撐明細."""

    def qty(key: str) -> float:
        return stats.get(key, 0.0)

    return [
        {
            "group": "1",
            "key": "",
            "label": "防靜電片(兩端附銅壓端子披覆型跨接線)廠商提供",
            "unit": "組",
            "qty": 0,
            "note": "含SUS鋼片(3t)及導線",
        },
        {"group": "2", "key": "uband_hdg_le6", "label": 'U-Bolt & Band ≦ 6" 熱浸鍍鋅', "unit": "組", "qty": qty("uband_hdg_le6"), "note": ""},
        {"group": "2", "key": "uband_hdg_ge8", "label": 'U-Bolt & Band ≧ 8" 熱浸鍍鋅', "unit": "組", "qty": qty("uband_hdg_ge8"), "note": ""},
        {"group": "2", "key": "uband_304_le6", "label": 'U-Bolt & Band ≦ 6"(SUS 304)', "unit": "組", "qty": qty("uband_304_le6"), "note": ""},
        {"group": "2", "key": "uband_304_ge8", "label": 'U-Bolt & Band ≧ 8"(SUS 304)', "unit": "組", "qty": qty("uband_304_ge8"), "note": ""},
        {"group": "3", "key": "shoe_hdg_le4", "label": '管鞋(PIPE SHOE)≦4"', "unit": "組", "qty": qty("shoe_hdg_le4"), "note": "依長春規範"},
        {"group": "3", "key": "shoe_hdg_5_10", "label": '管鞋(PIPE SHOE) 5"~10"', "unit": "組", "qty": qty("shoe_hdg_5_10"), "note": "依長春規範"},
        {"group": "3", "key": "shoe_hdg_12_24", "label": '管鞋(PIPE SHOE) 12"~24"', "unit": "組", "qty": qty("shoe_hdg_12_24"), "note": "依長春規範"},
        {"group": "3", "key": "shoe_hdg_ge26", "label": '管鞋(PIPE SHOE)≧26"', "unit": "組", "qty": qty("shoe_hdg_ge26"), "note": "依長春規範"},
        {"group": "3", "key": "cold_support", "label": "保冷支撐座(長春帶料)", "unit": "組", "qty": qty("cold_support"), "note": "依長春規範"},
        {"group": "4", "key": "shoe_304_le4", "label": '管鞋(PIPE SHOE)≦4"', "unit": "組", "qty": qty("shoe_304_le4"), "note": "CLAMP"},
        {"group": "4", "key": "shoe_304_5_10", "label": '管鞋(PIPE SHOE) 5"~10"', "unit": "組", "qty": qty("shoe_304_5_10"), "note": "CLAMP"},
        {"group": "4", "key": "shoe_304_12_24", "label": '管鞋(PIPE SHOE) 12"~24"', "unit": "組", "qty": qty("shoe_304_12_24"), "note": "CLAMP"},
        {
            "group": "5",
            "key": "cs_support_le15",
            "label": "CS(熱鍍鋅)管支撐(Pipe Support)製裝<=15Kg",
            "unit": "組",
            "qty": qty("cs_support_le15"),
            "note": "熱浸鍍鋅，依長春規範",
        },
        {
            "group": "5",
            "key": "cs_support_gt15",
            "label": "CS(熱鍍鋅)管支撐(Pipe Support)製裝>15Kg",
            "unit": "KG",
            "qty": qty("cs_support_gt15"),
            "note": "熱浸鍍鋅，依長春規範",
        },
        {"group": "6", "key": "", "label": "管夾", "unit": "組", "qty": 0, "note": "CLAMP"},
        {"group": "7", "key": "", "label": "管支撐小基礎制裝工料(一樓)", "unit": "組", "qty": 0, "note": "依長春規範"},
    ]


def _write_leader_procurement_sheet(ws, project: ProjectAnalysisResult):
    from openpyxl.styles import Alignment, Font, PatternFill

    styles = _styles()
    stats, _, details = _leader_procurement_stats(project)
    template_by_key = {item.key: item for item in _leader_stat_template()}
    detail_rows_by_key: dict[str, list[tuple[int, LeaderHitDetail]]] = {}
    ordered_details = _ordered_leader_details(details)
    for detail_index, detail in enumerate(ordered_details, start=4):
        detail_rows_by_key.setdefault(detail.stat_key, []).append((detail_index, detail))

    _setup_sheet(ws, "支撐分類統計", "I1")
    ws.cell(
        row=2,
        column=1,
        value=(
            "請款查核表：每個合約名稱旁直接列出判定規則與本批命中範例；"
            "需要逐筆舉證時，點「開啟明細」查看完整來源。"
        ),
    )
    ws.cell(row=2, column=1).font = styles["section_font"]
    ws.merge_cells("A2:I2")

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    title = ws.cell(row=row, column=1, value="二、管支撐(連工帶料，含油漆)")
    title.font = Font(name="Microsoft JhengHei", bold=True, size=13, color="000000")
    title.fill = PatternFill("solid", fgColor="FFFFFF")
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.border = styles["border"]
    ws.row_dimensions[row].height = 24
    row += 1

    boss_rows = _boss_summary_rows(stats)
    headers = ["類", "合約名稱", "單位", "本批數量", "合約名稱怎麼來的", "本批命中型號例", "來源圖號 / 流水號", "逐筆舉證", "備註"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.row_dimensions[row].height = 30
    row += 1

    for item in boss_rows:
        key = item["key"]
        stat = template_by_key.get(key)
        hit_rows = [pair for pair in detail_rows_by_key.get(key, []) if pair[1].status == "命中"]
        first_detail_row, example = hit_rows[0] if hit_rows else (None, None)
        ws.cell(row=row, column=1, value=item["group"])
        ws.cell(row=row, column=2, value=item["label"])
        ws.cell(row=row, column=3, value=item["unit"])
        ws.cell(row=row, column=4, value=round(item["qty"], 2) if item["unit"] == "KG" else int(item["qty"]))
        ws.cell(
            row=row,
            column=5,
            value=stat.criteria if stat else "目前沒有自動判定規則；數量固定為 0，需依合約另行確認。",
        )
        ws.cell(row=row, column=6, value=example.designation if example else "本批未命中")
        ws.cell(
            row=row,
            column=7,
            value=(
                f"{example.drawing_line_number or '—'} / {example.serial or '—'}"
                if example else "—"
            ),
        )
        trace_cell = ws.cell(
            row=row,
            column=8,
            value=f"查看 {len(hit_rows):,} 筆來源" if first_detail_row else "本批無明細",
        )
        if first_detail_row:
            trace_cell.hyperlink = f"#'查核-支撐明細'!A{first_detail_row}"
            trace_cell.style = "Hyperlink"
        ws.cell(row=row, column=9, value=item["note"])
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            if col != 8 or not first_detail_row:
                cell.font = Font(name="Microsoft JhengHei", size=10, color="000000")
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if col in {1, 3, 4, 8} else "left")
            if row % 2 == 0:
                cell.fill = styles["zebra_fill"]
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4).number_format = NUMFMT["WEIGHT_KG"] if item["unit"] == "KG" else NUMFMT["QTY_INT"]
        ws.row_dimensions[row].height = 46
        row += 1

    last_row = row - 1
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:I{last_row}"
    _set_widths(ws, [6, 34, 8, 12, 50, 24, 22, 12, 22])
    set_print_layout(ws, orientation="landscape", title_rows="5:5", area=f"A1:I{last_row}", footer_title="支撐分類統計")


def _write_leader_detail_sheet(ws, project: ProjectAnalysisResult):
    from openpyxl.utils import get_column_letter

    styles = _styles()
    stats, _, details = _leader_procurement_stats(project)
    details = _ordered_leader_details(details)
    contract_label_by_key = {
        item["key"]: item["label"] for item in _boss_summary_rows(stats) if item["key"]
    }
    col_idx = {header: index + 1 for index, header in enumerate(LEADER_DETAIL_HEADERS)}
    contract_total_col = col_idx["合約總量"]
    project_qty_col = col_idx["專案組數"]
    single_weight_col = col_idx["支撐單組總重(kg)"]
    amount_col = col_idx["本列計入"]
    status_col = col_idx["狀態 / 備註"]
    numeric_cols = {contract_total_col, project_qty_col, single_weight_col, amount_col}

    last_col_letter = get_column_letter(len(LEADER_DETAIL_HEADERS))
    _setup_sheet(ws, "請款分類來源（逐筆舉證）", f"{last_col_letter}1")
    ws.cell(
        row=2,
        column=1,
        value=(
            "用法：從「長官-支撐分類」點開來源，或直接篩選合約名稱。"
            "同一合約的來源會連續排列；合約總量、單組重、判定門檻、本列算式與計入量都在同一列。"
        ),
    )
    ws.cell(row=2, column=1).font = styles["section_font"]
    ws.merge_cells(f"A2:{last_col_letter}2")

    row = 4
    if not details:
        ws.cell(row=row, column=1, value="無命中資料")
        ws.cell(row=row, column=1).border = styles["border"]
        row += 1
    else:
        for detail in details:
            contract_total = stats.get(detail.stat_key, "")
            status_note = detail.status if not detail.note else f"{detail.status}：{detail.note}"
            evidence = "；".join(
                part for part in (detail.matched_detail, detail.material_basis) if part
            )
            row_values = {
                "合約名稱": contract_label_by_key.get(detail.stat_key, detail.label),
                "合約總量": round(contract_total, 3) if contract_total != "" else "",
                "計價單位": detail.unit,
                "來源圖號": detail.drawing_line_number,
                "流水號": detail.serial,
                "型號": detail.designation,
                "專案組數": detail.project_qty,
                "支撐單組總重(kg)": round(detail.single_weight, 3),
                "分類門檻 / 原因": detail.criteria,
                "本列請款計算": detail.claim_calculation,
                "本列計入": round(detail.amount, 3) if detail.unit == "KG" else int(detail.amount),
                "計入單位": detail.unit,
                "命中材料 / 零件": evidence,
                "狀態 / 備註": status_note,
            }
            values = [row_values[header] for header in LEADER_DETAIL_HEADERS]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = styles["border"]
                cell.alignment = styles["wrap"]
                if col in numeric_cols:
                    cell.alignment = styles["right"]
                if col == status_col:
                    apply_status_fill(cell, detail.status)
            ws.cell(row=row, column=project_qty_col).number_format = NUMFMT["QTY_INT"]
            ws.cell(row=row, column=single_weight_col).number_format = NUMFMT["WEIGHT_KG3"]
            ws.cell(row=row, column=contract_total_col).number_format = (
                NUMFMT["WEIGHT_KG3"] if detail.unit == "KG" else NUMFMT["QTY_INT"]
            )
            ws.cell(row=row, column=amount_col).number_format = NUMFMT["WEIGHT_KG3"] if detail.unit == "KG" else NUMFMT["QTY_INT"]
            ws.row_dimensions[row].height = 42
            row += 1

    last_row = max(row - 1, 3)
    width_by_header = {
        "合約名稱": 38,
        "合約總量": 14,
        "計價單位": 10,
        "來源圖號": 22,
        "流水號": 10,
        "型號": 24,
        "專案組數": 10,
        "支撐單組總重(kg)": 18,
        "分類門檻 / 原因": 48,
        "本列請款計算": 42,
        "本列計入": 14,
        "計入單位": 10,
        "命中材料 / 零件": 42,
        "狀態 / 備註": 38,
    }
    apply_report_table(
        ws,
        3,
        LEADER_DETAIL_HEADERS,
        4,
        last_row,
        col_formats={
            project_qty_col: NUMFMT["QTY_INT"],
            single_weight_col: NUMFMT["WEIGHT_KG3"],
        },
        widths=[width_by_header[header] for header in LEADER_DETAIL_HEADERS],
    )
    for data_row in range(4, last_row + 1):
        status = str(ws.cell(row=data_row, column=status_col).value or "").split("：", 1)[0]
        apply_status_fill(ws.cell(row=data_row, column=status_col), status)
    ws.freeze_panes = "D4"
    set_print_layout(ws, title_rows="3:3", area=f"A1:{last_col_letter}{last_row}", footer_title="請款分類來源")
