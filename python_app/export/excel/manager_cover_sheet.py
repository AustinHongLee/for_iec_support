"""Minimal manager-facing cover sheet for project workbook exports."""

from core.material_summary import MaterialSummary
from core.parser import get_type_code
from core.project_aggregation import ProjectAnalysisResult

from .confidence_summary import (
    format_confidence_counts,
    project_confidence_counts,
    review_required_count,
    worst_confidence_level,
)
from .leader_sheets import _boss_summary_rows, _leader_procurement_stats
from .navigation import workbook_navigation
from .styles import (
    COLORS,
    FONT_CJK,
    NUMFMT,
    apply_confidence_fill,
    _set_widths,
    _setup_sheet,
    _styles,
    set_print_layout,
)


def _project_type_count(project: ProjectAnalysisResult) -> int:
    types = {
        get_type_code(row_result.input_row.designation) or "未解析"
        for row_result in project.rows
    }
    return len(types)


def _cover_support_rows(stats: dict[str, float], details: list) -> list[dict]:
    boss_rows = _boss_summary_rows(stats)

    def sum_groups(*groups: str) -> float:
        return sum(float(row["qty"] or 0) for row in boss_rows if row["group"] in groups)

    def qty(key: str) -> float:
        return float(stats.get(key, 0.0))

    rows = [
        {
            "item": "管支撐製裝 <=15Kg",
            "qty": qty("cs_support_le15"),
            "unit": "組",
            "where": "長官-支撐分類",
            "note": "合約項目第 5 類",
        },
        {
            "item": "管支撐製裝 >15Kg",
            "qty": qty("cs_support_gt15"),
            "unit": "KG",
            "where": "長官-支撐分類",
            "note": "合約項目第 5 類",
        },
        {
            "item": "U-Bolt & Band",
            "qty": sum_groups("2"),
            "unit": "組",
            "where": "長官-支撐分類",
            "note": "依管徑與材質分段",
        },
        {
            "item": "Pipe Shoe / 保冷支撐座",
            "qty": sum_groups("3", "4"),
            "unit": "組",
            "where": "長官-支撐分類",
            "note": "依管徑與材質分段",
        },
    ]
    confirm_count = sum(1 for detail in details if detail.status == "需確認")
    unmatched_count = sum(1 for detail in details if detail.status == "未納入")
    if confirm_count:
        rows.append(
            {
                "item": "需人工確認分類",
                "qty": confirm_count,
                "unit": "筆",
                "where": "查核-支撐明細",
                "note": "需確認是否新增分類規則",
            }
        )
    if unmatched_count:
        rows.append(
            {
                "item": "未納入分類",
                "qty": unmatched_count,
                "unit": "筆",
                "where": "查核-支撐明細",
                "note": "目前不列入長官摘要統計",
            }
        )
    return [row for row in rows if row["qty"] or row["item"].startswith("管支撐製裝")]


def _sheet_available(sheet_name: str, available_sheets: tuple[str, ...] | None) -> bool:
    return available_sheets is None or sheet_name in available_sheets


def _package_navigation(available_sheets: tuple[str, ...] | None) -> list:
    items = workbook_navigation()
    if available_sheets is None:
        return items
    return [item for item in items if item.sheet in available_sheets]


def _write_link_cell(
    ws,
    row: int,
    column: int,
    sheet_name: str,
    label: str | None = None,
    *,
    available_sheets: tuple[str, ...] | None = None,
) -> None:
    cell = ws.cell(row=row, column=column, value=label or sheet_name)
    if not _sheet_available(sheet_name, available_sheets):
        cell.value = f"{label or sheet_name}（見完整活頁簿）"
        return
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.style = "Hyperlink"


def _write_manager_cover_sheet(
    ws,
    project: ProjectAnalysisResult,
    summary: MaterialSummary,
    *,
    available_sheets: tuple[str, ...] | None = None,
    export_context: dict | None = None,
) -> None:
    """長官摘要 - one-page reading path and high-level support quantities."""
    import datetime as _dt
    from openpyxl.styles import Alignment, Font, PatternFill

    styles = _styles()
    stats, _, details = _leader_procurement_stats(project)
    type_count = _project_type_count(project)
    confirm_count = sum(1 for detail in details if detail.status == "需確認")
    unmatched_count = sum(1 for detail in details if detail.status == "未納入")
    cover_rows = [
        {
            "item": "合約名稱怎麼來的？",
            "qty": len(_boss_summary_rows(stats)),
            "unit": "項",
            "where": "長官-支撐分類",
            "note": "每列直接顯示判定規則、本批命中型號例與逐筆舉證連結。",
        },
        {
            "item": "某支撐計入哪個合約？",
            "qty": len(details),
            "unit": "筆",
            "where": "查核-支撐明細",
            "note": "用來源圖號、流水號或 OPEN / Type 型號篩選，查看完整判定鏈。",
        },
        {
            "item": "一組支撐本身多重？",
            "qty": len(project.rows),
            "unit": "型號列",
            "where": "單組重量明細",
            "note": "直接查型號與單組重量；不展開材料，專案組數不會乘入。",
        },
        {
            "item": "全案請款重量是多少？",
            "qty": summary.total_weight,
            "unit": "KG",
            "where": "重量明細表",
            "note": "本頁保留專案組數乘算後的材料與重量。",
        },
        {
            "item": "哪些資料不能直接說明？",
            "qty": confirm_count + unmatched_count,
            "unit": "筆",
            "where": "查核-支撐明細",
            "note": f"需確認 {confirm_count} 筆；未納入 {unmatched_count} 筆。請款前先處理。",
        },
    ]

    ws.title = "長官-摘要"
    subtitle = (
        f"製表日期 {_dt.date.today():%Y-%m-%d}    "
        f"支撐 {project.total_support_count:,} 組    "
        f"使用 Type {type_count:,} 種    "
        f"全案總重 {summary.total_weight:,.2f} kg"
    )
    _setup_sheet(ws, "請款分類查核入口", "H1", subtitle=subtitle, audience="主管 / 請款 / 業主", freeze_title=False)
    _set_widths(ws, [22, 14, 10, 18, 22, 22, 14, 14])
    ws.sheet_view.zoomScale = 105

    ws.merge_cells("A4:H4")
    intro = ws["A4"]
    if _sheet_available("查核-支撐明細", available_sheets):
        intro.value = "不要從摘要猜數字：先選對方問的問題，再點詳細位置取得可展示的規則、型號與來源。"
    else:
        intro.value = "本頁是請款問答入口；完整規則與逐筆來源請使用完整活頁簿或採購材料包。"
    intro.font = Font(name=FONT_CJK, size=11, color=COLORS["ink"])
    intro.fill = styles["subtitle_fill"]
    intro.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[4].height = 28

    if export_context:
        assumption_count = int(export_context.get("assumption_count") or 0)
        mode = export_context.get("mode")
        reason = str(export_context.get("exception_reason") or "").strip()
        if mode == "final" and assumption_count:
            banner_text = (
                f"精算版例外放行：仍含 {assumption_count} 筆假設值；"
                f"放行原因：{reason}"
            )
            banner_color = "F8CBAD"
        elif mode == "final":
            banner_text = "精算版：未檢出假設值，可作為發包前查核版本。"
            banner_color = "E2F0D9"
        else:
            banner_text = (
                f"概算版：含 {assumption_count} 筆假設值，不得作為發包依據"
            )
            banner_color = "FFF2CC"
        ws.merge_cells("A5:H5")
        banner = ws["A5"]
        banner.value = banner_text
        banner.font = Font(name=FONT_CJK, size=12, bold=True, color=COLORS["ink"])
        banner.fill = PatternFill("solid", fgColor=banner_color)
        banner.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[5].height = 30

    header_row = 6
    headers = ["對方可能會問", "可查資料量", "單位", "直接看哪裡", "可以怎麼回答"]
    for col in range(1, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=8)
    ws.row_dimensions[header_row].height = 24

    row = header_row + 1
    for item in cover_rows:
        ws.cell(row=row, column=1, value=item["item"])
        ws.cell(row=row, column=2, value=round(item["qty"], 2) if item["unit"] == "KG" else int(item["qty"]))
        ws.cell(row=row, column=3, value=item["unit"])
        _write_link_cell(ws, row, 4, item["where"], available_sheets=available_sheets)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        ws.cell(row=row, column=5, value=item["note"])
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.font = Font(name=FONT_CJK, size=12 if col in (1, 2) else 10, color=COLORS["ink"])
            cell.alignment = Alignment(
                horizontal="right" if col == 2 else "center" if col in (3, 4) else "left",
                vertical="center",
                wrap_text=True,
                indent=1 if col in (1, 5) else 0,
            )
        ws.cell(row=row, column=4).font = Font(name=FONT_CJK, size=10, color="0563C1", underline="single")
        ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=16, color=COLORS["accent"])
        ws.cell(row=row, column=2).number_format = NUMFMT["WEIGHT_KG"] if item["unit"] == "KG" else NUMFMT["QTY_INT"]
        if item.get("confidence_level"):
            apply_confidence_fill(ws.cell(row=row, column=5), item["confidence_level"])
        if (row - header_row) % 2 == 0:
            for col in range(1, 9):
                if not (item.get("confidence_level") and col == 5):
                    ws.cell(row=row, column=col).fill = styles["zebra_fill"]
        ws.row_dimensions[row].height = 30
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    guide = ws.cell(row=row, column=1, value="分頁索引")
    guide.font = styles["section_font"]
    guide.fill = styles["section_fill"]
    guide.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="分頁")
    ws.cell(row=row, column=3, value="適合查看")
    ws.cell(row=row, column=7, value="角色 / 列印")
    ws.row_dimensions[row].height = 22
    row += 1

    for nav_item in _package_navigation(available_sheets):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            if (row - header_row) % 2 == 0:
                cell.fill = styles["zebra_fill"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        _write_link_cell(ws, row, 1, nav_item.sheet, available_sheets=available_sheets)
        ws.cell(row=row, column=3, value=nav_item.purpose)
        ws.cell(row=row, column=7, value=f"{nav_item.audience} / {nav_item.print_note}")
        ws.cell(row=row, column=1).font = Font(name=FONT_CJK, size=10, color="0563C1", underline="single")
        ws.cell(row=row, column=3).font = Font(name=FONT_CJK, size=9, color=COLORS["text_mute"])
        ws.cell(row=row, column=7).font = Font(name=FONT_CJK, size=9, color=COLORS["text_mute"])
        ws.row_dimensions[row].height = 28
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    note_text = "請款原則：先展示合約分類規則，再以來源圖號 / 流水號 / 型號逐筆舉證；不要只提供總量摘要。"
    note = ws.cell(row=row, column=1, value=note_text)
    note.font = Font(name=FONT_CJK, size=10, italic=True, color=COLORS["text_mute"])
    note.fill = PatternFill("solid", fgColor="FFFFFF")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 24

    set_print_layout(ws, orientation="portrait", title_rows=None, area=f"A1:H{row}", footer_title="請款分類查核入口")
