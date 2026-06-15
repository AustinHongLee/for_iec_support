"""Renderers for project weight detail sheets."""

from core.parser import get_type_code
from core.project_aggregation import ProjectAnalysisResult

from .headers import PROJECT_HEADERS
from .styles import (
    NUMFMT,
    add_color_scale,
    apply_report_table,
    apply_status_fill,
    set_print_layout,
    _setup_sheet,
    _styles,
)


def _write_project_weight_sheet(ws, project: ProjectAnalysisResult):
    from openpyxl.utils import get_column_letter

    styles = _styles()
    last_col_letter = get_column_letter(len(PROJECT_HEADERS))
    _setup_sheet(
        ws,
        "重量分析明細",
        f"{last_col_letter}1",
        subtitle=(
            f"工程審查明細    支撐 {project.total_support_count} 組    "
            f"型號列 {len(project.rows)}    全案總重 {project.total_weight:,.2f} kg"
        ),
        audience="工程 / 審查",
    )
    error_rows: list[int] = []

    row = 4
    for row_result in project.rows:
        input_row = row_result.input_row
        single_result = row_result.single_result
        scaled_result = row_result.scaled_result

        if single_result.error:
            # 與正常資料一致的順序：型號優先，來源放最後
            values = [
                input_row.designation,
                get_type_code(input_row.designation),
                "錯誤",
                single_result.error,
            ] + [""] * (len(PROJECT_HEADERS) - 4)
            # 來源放最後
            values[-4] = input_row.drawing_line_number
            values[-3] = input_row.serial
            values[-2] = input_row.quantity
            values[-1] = input_row.unit or "組"
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            error_rows.append(row)
            row += 1
            continue

        for single_entry, scaled_entry in zip(single_result.entries, scaled_result.entries):
            # 型號為主角，來源資訊移到最右側作為配角
            values = [
                input_row.designation,
                get_type_code(input_row.designation),
                single_entry.item_no,
                single_entry.name,
                single_entry.display_spec,
                single_entry.material,
                single_entry.length,
                single_entry.width if single_entry.width else "",
                single_entry.quantity,                                   # 單件數量
                scaled_entry.quantity,                                  # 總數量
                single_entry.weight_output,                             # 單組重(kg)
                scaled_entry.weight_output,                             # 總重(kg)
                single_entry.category,
                single_entry.item_class,
                single_entry.manufacturing_type,
                single_entry.part_key,
                single_entry.stock_id,
                single_entry.display_remark,
                input_row.drawing_line_number,                           # 來源圖號 (配角)
                input_row.serial,                                        # 流水號 (唯一值，配角)
                input_row.quantity,
                input_row.unit or "組",
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    last_row = max(row - 1, 3)
    apply_report_table(
        ws,
        3,
        PROJECT_HEADERS,
        4,
        last_row,
        col_formats={
            7: NUMFMT["LEN_MM"],
            8: NUMFMT["LEN_MM"],
            9: NUMFMT["QTY_INT"],
            10: NUMFMT["QTY_INT"],
            21: NUMFMT["QTY_INT"],
            11: NUMFMT["WEIGHT_KG"],
            12: NUMFMT["WEIGHT_KG"],
        },
        widths=[
            22, 6, 6, 16, 20, 12,   # 型號 Type 項次 品名 規格 材質
            9, 9,                   # 長 寬
            8, 8, 10, 10, 10,       # 單件 總數 單組重 總重 屬性
            10, 10, 12, 10, 34,     # 類別 製造 零件ID 庫存 計算說明
            16, 10, 8, 6,           # 來源圖號 流水號 輸入數量 輸入單位
        ],
    )
    for error_row in error_rows:
        for col in range(1, len(PROJECT_HEADERS) + 1):
            cell = ws.cell(row=error_row, column=col)
            cell.fill = styles["bad_fill"]
            cell.border = styles["border"]
        # 錯誤標記現在在第 3 欄（項次位置）
        apply_status_fill(ws.cell(row=error_row, column=3), "錯誤", set_font=True)
    if last_row >= 4:
        total_weight_col = get_column_letter(PROJECT_HEADERS.index("總重(kg)") + 1)
        add_color_scale(ws, f"{total_weight_col}4:{total_weight_col}{last_row}", "weight")
    set_print_layout(ws, title_rows="3:3", area=f"A1:{last_col_letter}{last_row}", footer_title="重量分析")
