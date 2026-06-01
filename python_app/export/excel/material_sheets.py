"""Renderer for the material summary/procurement sheet."""

from core.material_summary import MaterialSummary

from .headers import SUMMARY_HEADERS
from .styles import (
    NUMFMT,
    add_color_scale,
    add_bar_chart,
    apply_report_table,
    set_print_layout,
    write_grand_total_band,
    _setup_sheet,
)


def _write_material_summary_sheet(ws, summary: MaterialSummary):
    from openpyxl.chart import Reference
    from openpyxl.utils import get_column_letter

    _setup_sheet(
        ws,
        "材料合計與採購清單",
        "M1",
        subtitle=f"採購/製造清單    材料 {len(summary.lines)} 項    全案總重 {summary.total_weight:,.2f} kg",
        audience="採購 / 製造",
    )

    row = 4
    for ln in summary.lines:
        length_value = round(ln.total_length_mm, 1) if ln.aggregate_type == "linear" else ""
        qty_value = ln.piece_count if ln.aggregate_type == "linear" else ln.total_qty
        stock_length = round(ln.stock_length, 0) if ln.stock_length else ""
        sources = ", ".join(ln.source_fullstrings[:8])
        if len(ln.source_fullstrings) > 8:
            sources += f" ...+{len(ln.source_fullstrings) - 8}"
        values = [
            ln.name,
            ln.spec,
            ln.material,
            ln.category,
            ln.item_class,
            ln.manufacturing_type,
            length_value,
            qty_value,
            round(ln.total_weight, 2),
            stock_length,
            ln.purchase_qty,
            ln.purchase_unit,
            sources,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    last_row = max(row - 1, 3)
    apply_report_table(
        ws,
        3,
        SUMMARY_HEADERS,
        4,
        last_row,
        col_formats={
            7: NUMFMT["LEN_MM1"],
            8: NUMFMT["QTY_INT"],
            9: NUMFMT["WEIGHT_KG"],
            10: NUMFMT["LEN_MM"],
            11: NUMFMT["QTY_INT"],
        },
        widths=[16, 22, 16, 10, 14, 14, 14, 12, 12, 12, 12, 8, 46],
    )
    if last_row >= 4:
        add_color_scale(ws, f"I4:I{last_row}", "weight")
        chart_start = 4
        label_col = 27
        value_col = 28
        ws.cell(row=chart_start, column=label_col, value="材料")
        ws.cell(row=chart_start, column=value_col, value="總重")
        top_lines = sorted(summary.lines, key=lambda ln: ln.total_weight, reverse=True)[:8]
        for offset, line in enumerate(top_lines, start=1):
            ws.cell(row=chart_start + offset, column=label_col, value=f"{line.name} {line.spec}")
            ws.cell(row=chart_start + offset, column=value_col, value=round(line.total_weight, 2))
        add_bar_chart(
            ws,
            Reference(ws, min_col=label_col, min_row=chart_start + 1, max_row=chart_start + len(top_lines)),
            Reference(ws, min_col=value_col, min_row=chart_start, max_row=chart_start + len(top_lines)),
            "O4",
            "teal",
            "材料重量 Top 8 (kg)",
            horizontal=True,
        )
        for col in (label_col, value_col):
            ws.column_dimensions[get_column_letter(col)].hidden = True

    total_row = row
    write_grand_total_band(
        ws,
        total_row,
        len(SUMMARY_HEADERS),
        "合計總重",
        9,
        round(summary.total_weight, 2),
        fmt=NUMFMT["WEIGHT_KG"],
        label_col=8,
    )
    set_print_layout(ws, title_rows="3:3", area=f"A1:Y{total_row}", footer_title="材料合計")
