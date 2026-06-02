"""Legacy single-result and flat project Excel exports."""

from typing import List

from core.models import AnalysisResult
from core.parser import get_type_code
from core.project_aggregation import ProjectAnalysisResult

from .headers import HEADERS, PROJECT_HEADERS
from .styles import _format_sheet


def export_to_excel(results: List[AnalysisResult], filepath: str):
    """匯出分析結果至 Excel"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weight_Analysis"

    _format_sheet(ws, HEADERS)

    # 資料
    row = 2
    for result in results:
        # 寫入原始字串到 A 欄
        ws.cell(row=row, column=1, value=result.fullstring)

        if result.error:
            ws.cell(row=row, column=2, value="Error")
            ws.cell(row=row, column=3, value=result.error)
            row += 1
            continue

        for entry in result.entries:
            ws.cell(row=row, column=1, value=result.fullstring if entry.item_no == 1 else "")
            ws.cell(row=row, column=2, value=entry.item_no)
            ws.cell(row=row, column=3, value=entry.name)
            ws.cell(row=row, column=4, value=entry.display_spec)
            ws.cell(row=row, column=5, value=entry.length)
            ws.cell(row=row, column=6, value=entry.width if entry.width else "")
            ws.cell(row=row, column=7, value=entry.material)
            ws.cell(row=row, column=8, value=entry.quantity)
            ws.cell(row=row, column=9, value=entry.weight_per_unit if entry.weight_per_unit else "")
            ws.cell(row=row, column=10, value=entry.unit_weight)
            ws.cell(row=row, column=11, value=entry.total_weight)
            ws.cell(row=row, column=12, value=entry.unit)
            ws.cell(row=row, column=13, value=entry.factor)
            ws.cell(row=row, column=14, value=entry.length_subtotal if entry.length_subtotal else "")
            ws.cell(row=row, column=15, value=entry.qty_subtotal if entry.qty_subtotal else "")
            ws.cell(row=row, column=16, value=entry.weight_output)
            ws.cell(row=row, column=17, value=entry.category)
            ws.cell(row=row, column=18, value=entry.item_class)
            ws.cell(row=row, column=19, value=entry.manufacturing_type)
            ws.cell(row=row, column=20, value=entry.part_key)
            ws.cell(row=row, column=21, value=entry.stock_id)
            ws.cell(row=row, column=22, value=entry.display_remark)
            row += 1

    _format_sheet(ws, HEADERS)

    wb.save(filepath)


def export_project_to_excel(project: ProjectAnalysisResult, filepath: str):
    """匯出 project-aware 分析結果（平坦表格，每列填滿型號/組數）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project_Weight_Analysis"

    _format_sheet(ws, PROJECT_HEADERS)

    row = 2
    for row_result in project.rows:
        input_row = row_result.input_row
        single_result = row_result.single_result
        scaled_result = row_result.scaled_result

        if single_result.error:
            values = [
                input_row.drawing_line_number,
                input_row.serial,
                input_row.quantity,
                input_row.unit or "組",
                input_row.designation,
                get_type_code(input_row.designation),
                "Error",
                single_result.error,
            ] + [""] * (len(PROJECT_HEADERS) - 8)
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
            continue

        for single_entry, scaled_entry in zip(single_result.entries, scaled_result.entries):
            values = [
                input_row.drawing_line_number,
                input_row.serial,
                input_row.quantity,
                input_row.unit or "組",
                input_row.designation,
                get_type_code(input_row.designation),
                single_entry.item_no,
                single_entry.name,
                single_entry.display_spec,
                single_entry.material,
                single_entry.length,
                single_entry.width if single_entry.width else "",
                single_entry.quantity,
                scaled_entry.quantity,
                single_entry.weight_output,
                scaled_entry.weight_output,
                single_entry.category,
                single_entry.item_class,
                single_entry.manufacturing_type,
                single_entry.part_key,
                single_entry.stock_id,
                single_entry.display_remark,
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    _format_sheet(ws, PROJECT_HEADERS)

    wb.save(filepath)
