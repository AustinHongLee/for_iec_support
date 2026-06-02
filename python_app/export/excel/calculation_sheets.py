"""Renderers for calculation-basis and reference sheets."""

from core.parser import get_type_code
from core.project_aggregation import ProjectAnalysisResult

from .headers import _CALC_BASIS_HEADERS, _STANDARDS_TABLE
from .styles import (
    NUMFMT,
    add_color_scale,
    apply_confidence_fill,
    freeze_and_filter,
    set_print_layout,
    write_grand_total_band,
    _section_header,
    _set_widths,
    _setup_sheet,
    _styles,
    _write_headers,
)


def _weight_formula_str(entry) -> str:
    """從 AnalysisEntry 欄位重建人類可讀的計算式。"""
    qty = entry.quantity
    unit_weight = entry.unit_weight
    line_total = entry.weight_output

    if entry.unit == "M" and entry.weight_per_unit and entry.weight_per_unit > 0:
        wpm = entry.weight_per_unit
        return (
            f"{qty}件 × {entry.length:.0f}mm ÷ 1,000"
            f" × {wpm:.3f} kg/m"
            f" = {line_total:.3f} kg"
        )

    if entry.unit == "PC" and entry.width and entry.width > 0:
        try:
            t = float(entry.spec)
            density = 7.93 if "304" in entry.material else 7.82 if entry.material == "AS" else 7.85
            geometry = getattr(entry, "geometry", None)
            net_area = getattr(geometry, "net_area_mm2", 0) if geometry else 0
            formula = getattr(geometry, "formula", "") if geometry else ""
            if net_area and net_area != entry.length * entry.width:
                area_label = formula or f"{entry.length:.0f}×{entry.width:.0f}"
                return (
                    f"{qty}件 × [{area_label} = {net_area:.0f} mm2]"
                    f" × t{t:.0f} × {density:.2f} ÷ 1,000,000"
                    f" = {entry.weight_output:.3f} kg"
                )
            return (
                f"{qty}件 × {entry.length:.0f}×{entry.width:.0f}×{t:.0f}mm"
                f" × {density:.2f} t/m³"
                f" = {entry.weight_output:.3f} kg"
            )
        except (ValueError, TypeError):
            return (
                f"{qty}件 × {entry.length:.0f}×{entry.width:.0f}mm"
                f" × t×ρ = {line_total:.3f} kg"
            )

    if entry.unit in ("SET", "EA", "KG"):
        return (
            f"{qty} {entry.unit} × {unit_weight:.3f} kg/{entry.unit}"
            f" = {line_total:.3f} kg"
        )

    return f"{qty} × {unit_weight:.3f} kg = {line_total:.3f} kg"


def _confidence_label(meta: dict) -> str:
    return meta.get("truth_level", "未知")


def _write_calculation_basis_sheet(ws, project: ProjectAnalysisResult):
    """重量明細表 — Flat 表，含小計與全案合計，適合審查與樞紐分析。"""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    styles = _styles()
    qty_fill = PatternFill("solid", fgColor="EAF2F8")
    subtotal_fill = styles["subtotal_fill"]
    alt_fill = styles["zebra_fill"]
    error_fill = PatternFill("solid", fgColor="FCE4D6")

    n_cols = len(_CALC_BASIS_HEADERS)
    last_col_letter = get_column_letter(n_cols)

    subtitle = (
        f"支撐總組數 {project.total_support_count} 組    "
        f"成功項目 {len(project.rows) - len(project.errors)}    "
        f"錯誤項目 {len(project.errors)}    "
        f"全案總重 {project.total_weight:,.3f} kg"
    )
    _setup_sheet(
        ws,
        "IEC 管架支撐 - 重量明細表",
        f"{last_col_letter}1",
        subtitle=subtitle,
        audience="工程 / 採購",
    )

    HEADER_ROW = 3
    _write_headers(ws, HEADER_ROW, _CALC_BASIS_HEADERS)
    ws.row_dimensions[HEADER_ROW].height = 28

    data_row = HEADER_ROW + 1

    for row_result in project.rows:
        inp = row_result.input_row
        single = row_result.single_result
        scaled = row_result.scaled_result

        if single.error:
            vals = [
                inp.drawing_line_number,
                inp.serial,
                inp.quantity,
                inp.unit or "組",
                inp.designation,
                get_type_code(inp.designation),
                "錯誤",
                single.error,
            ] + [""] * (n_cols - 8)
            vals[-1] = "明細"
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.fill = error_fill
                cell.border = styles["border"]
                cell.alignment = styles["wrap"]
            data_row += 1
            continue

        group_start_row = data_row
        for s_entry, sc_entry in zip(single.entries, scaled.entries):
            formula_str = _weight_formula_str(s_entry)
            single_unit_w = round(s_entry.unit_weight, 3)
            single_group_w = round(s_entry.weight_output, 3)
            total_w = round(sc_entry.weight_output, 3)

            vals = [
                inp.drawing_line_number, inp.serial, inp.quantity, inp.unit or "組",
                inp.designation, get_type_code(inp.designation), s_entry.item_no, s_entry.name,
                s_entry.display_spec, s_entry.material,
                s_entry.length if s_entry.length else "",
                s_entry.width if s_entry.width else "",
                getattr(s_entry, "category", ""),
                s_entry.quantity, sc_entry.quantity,
                single_unit_w, single_group_w, total_w,
                formula_str,
                getattr(s_entry, "item_class", ""),
                getattr(s_entry, "manufacturing_type", ""),
                "明細",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = styles["border"]
                if data_row % 2 == 0:
                    cell.fill = alt_fill
                if col in (3, 14, 15):
                    cell.fill = qty_fill
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="right" if col in (3, 11, 12, 14, 15, 16, 17, 18) else "left",
                    wrap_text=(col == 19),
                    indent=1 if col in (5, 8, 9, 10, 13) else 0,
                )
                if col in (3, 14, 15):
                    cell.number_format = NUMFMT["QTY_INT"]
                elif col in (11, 12):
                    cell.number_format = NUMFMT["LEN_MM"]
                elif col in (16, 17, 18):
                    cell.number_format = NUMFMT["WEIGHT_KG3"]
            ws.row_dimensions[data_row].height = 16
            data_row += 1

        if data_row > group_start_row:
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=data_row, column=col)
                cell.fill = subtotal_fill
                cell.border = styles["border"]
                cell.font = styles["bold_font"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=data_row, column=1, value=inp.drawing_line_number)
            ws.cell(row=data_row, column=2, value=inp.serial)
            ws.cell(row=data_row, column=3, value=inp.quantity)
            ws.cell(row=data_row, column=4, value=inp.unit or "組")
            ws.cell(row=data_row, column=5, value=f"小計 {inp.designation}")
            ws.cell(row=data_row, column=6, value=get_type_code(inp.designation))
            ws.cell(row=data_row, column=18, value=round(scaled.total_weight, 3))
            ws.cell(row=data_row, column=3).number_format = NUMFMT["QTY_INT"]
            ws.cell(row=data_row, column=18).number_format = NUMFMT["WEIGHT_KG3"]
            ws.cell(row=data_row, column=n_cols, value="小計")
            ws.row_dimensions[data_row].height = 18
            data_row += 1

    filter_last_row = max(data_row - 1, HEADER_ROW)
    write_grand_total_band(
        ws,
        data_row,
        n_cols,
        "全案合計",
        18,
        round(project.total_weight, 3),
        fmt=NUMFMT["WEIGHT_KG3"],
    )
    ws.cell(row=data_row, column=3, value=project.total_support_count)
    ws.cell(row=data_row, column=3).number_format = NUMFMT["QTY_INT"]
    ws.cell(row=data_row, column=n_cols, value="合計")
    last_data_row = data_row

    freeze_and_filter(ws, HEADER_ROW, filter_last_row, last_col_letter, autofilter=True)
    if filter_last_row >= HEADER_ROW + 1:
        add_color_scale(ws, f"R{HEADER_ROW + 1}:R{filter_last_row}", "weight")

    _set_widths(ws, [18, 12, 8, 7, 20, 8, 7, 18, 22, 14, 10, 10, 10, 10, 10, 12, 13, 13, 50, 14, 14, 8])
    set_print_layout(ws, title_rows="3:3", area=f"A1:{last_col_letter}{last_data_row}", footer_title="重量明細表")


def _write_calc_reference_sheet(ws, project: ProjectAnalysisResult):
    """計算標準與假設 — 給長官或客戶看的靜態說明頁。"""
    from openpyxl.styles import Alignment, PatternFill

    styles = _styles()

    _setup_sheet(
        ws,
        "計算標準與假設",
        "F1",
        subtitle="本頁列示重量計算依據與資料狀態判讀依據",
        audience="主管 / 業主 / 客戶",
    )

    row = 4
    _section_header(ws, row, "計算標準與假設", span_cols=6)
    row += 1

    std_headers = ["計算項目", "引用標準 / 依據", "計算方式說明"]
    for col, h in enumerate(std_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = styles["subheader_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws.cell(row, 3).value = "計算方式說明"
    ws.cell(row, 3).fill = styles["subheader_fill"]
    ws.cell(row, 3).font = styles["bold_font"]
    ws.cell(row, 3).alignment = styles["center"]
    ws.cell(row, 3).border = styles["border"]
    ws.row_dimensions[row].height = 24
    row += 1

    for item, standard, desc in _STANDARDS_TABLE:
        for col, val in enumerate([item, standard, desc], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3),
                                       horizontal="left", indent=1)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    _section_header(ws, row, "資料狀態圖例", span_cols=6)
    row += 1
    legends = [
        ("精確 — 直接查表", "E2EFDA"),
        ("推導 — 公式計算", "DDEBF7"),
        ("估算 — 工程假設", "FFF2CC"),
        ("未知 — 需複核",   "FCE4D6"),
    ]
    for col_off, (label, color) in enumerate(legends, 1):
        cell = ws.cell(row=row, column=col_off + 1, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.border = styles["border"]
        cell.alignment = styles["center"]
        cell.font = styles["bold_font"]
    ws.row_dimensions[row].height = 22
    row += 2

    _section_header(ws, row, "Type 計算資料狀態彙整", span_cols=6)
    row += 1

    summ_headers = ["Type", "型號列數", "支撐組數", "資料狀態", "合計重(kg)", "代表型號 / 備註"]
    for col, h in enumerate(summ_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = styles["subheader_fill"]
        cell.font = styles["bold_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]
    ws.row_dimensions[row].height = 24
    row += 1

    def status_label(counts: dict[str, int]) -> str:
        active = [f"{key} {value}" for key, value in counts.items() if value]
        return " / ".join(active) if active else "未知"

    def worst_status(counts: dict[str, int]) -> str:
        for key in ("錯誤", "未知", "估算", "推導", "精確"):
            if counts.get(key):
                return key
        return "未知"

    type_stats: dict[str, dict] = {}

    for row_result in project.rows:
        inp = row_result.input_row
        single = row_result.single_result
        scaled = row_result.scaled_result
        type_id = get_type_code(inp.designation) or "未解析"
        meta = single.meta or {}
        confidence = _confidence_label(meta)
        stat = type_stats.setdefault(type_id, {
            "rows": 0,
            "support_count": 0,
            "total_weight": 0.0,
            "status_counts": {"精確": 0, "推導": 0, "估算": 0, "未知": 0, "錯誤": 0},
            "examples": [],
            "errors": [],
        })
        stat["rows"] += 1
        stat["support_count"] += inp.quantity
        if inp.designation not in stat["examples"] and len(stat["examples"]) < 3:
            stat["examples"].append(inp.designation)
        if single.error:
            stat["status_counts"]["錯誤"] += 1
            stat["errors"].append(single.error)
        else:
            scaled_total = round(scaled.total_weight, 3)
            stat["total_weight"] += scaled_total
            stat["status_counts"][confidence if confidence in stat["status_counts"] else "未知"] += 1

    def type_sort_key(item: tuple[str, dict]) -> tuple[int, str]:
        key = item[0].rstrip("C")
        return (int(key) if key.isdigit() else 9999, item[0])

    for type_id, stat in sorted(type_stats.items(), key=type_sort_key):
        status = status_label(stat["status_counts"])
        note = "、".join(stat["examples"])
        if stat["rows"] > len(stat["examples"]):
            note += f" 等 {stat['rows']} 列"
        if stat["errors"]:
            note = f"{note}；錯誤：{stat['errors'][0]}"
        vals = [
            f"Type {type_id}",
            stat["rows"],
            stat["support_count"],
            status,
            round(stat["total_weight"], 3),
            note,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if col in (2, 3, 5) else "left",
                wrap_text=(col == 6),
                indent=1 if col in (1, 4, 6) else 0,
            )
            if col == 4:
                worst = worst_status(stat["status_counts"])
                if worst == "錯誤":
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
                else:
                    apply_confidence_fill(cell, worst)
            if col == 5:
                cell.number_format = NUMFMT["WEIGHT_KG3"]
        ws.row_dimensions[row].height = 24
        row += 1

    write_grand_total_band(
        ws,
        row,
        6,
        "■ 全案合計總重",
        5,
        round(project.total_weight, 3),
        fmt=NUMFMT["WEIGHT_KG3"],
    )

    col_widths = [14, 10, 10, 18, 14, 54]
    _set_widths(ws, col_widths)
    set_print_layout(ws, orientation="portrait", title_rows=None, area=f"A1:F{row}", footer_title="計算標準與假設")
