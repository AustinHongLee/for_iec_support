"""
材料合計表 + 下料表 Excel 匯出

Sheet 1: 材料合計表 (採購用)
Sheet 2: 下料計算表 (現場裁切用)
"""
from typing import List
from core.material_summary import MaterialSummary, SummaryLine
from core.cutting_optimizer import CuttingPlan, optimize_from_summary


def export_summary_and_cutting(
    summary: MaterialSummary,
    filepath: str,
    cutting_plans: List[CuttingPlan] = None,
):
    """
    匯出材料合計表 + 下料表到同一個 Excel 檔

    Parameters
    ----------
    summary : MaterialSummary
    filepath : str
    cutting_plans : list of CuttingPlan, optional
        若未提供，自動對 linear 材料產生
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()

    # ── Sheet 1: 材料合計表 ──
    _write_summary_sheet(wb.active, summary)

    # ── 下料方案 ──
    if cutting_plans is None:
        cutting_plans = []
        for ln in summary.get_linear_lines():
            plan = optimize_from_summary(ln)
            if plan and plan.total_pieces > 0:
                cutting_plans.append(plan)

    # ── Sheet 2: 下料計算表 ──
    if cutting_plans:
        ws2 = wb.create_sheet("下料計算表")
        _write_cutting_sheet(ws2, cutting_plans)

    wb.save(filepath)


# ═══════════════════════════════════════════════════════
#  Sheet 1: 材料合計表
# ═══════════════════════════════════════════════════════

_SUMMARY_HEADERS = [
    "品名", "規格", "材質", "屬性",
    "需求總長(mm)", "需求件數", "總重(kg)",
    "原料長度(mm)", "建議採購量", "單位",
    "來源編碼",
]


def _write_summary_sheet(ws, summary: MaterialSummary):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    ws.title = "材料合計表"

    # 標題
    ws.merge_cells("A1:K1")
    title_cell = ws.cell(row=1, column=1, value="材料合計表 — 採購清單")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 表頭
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(_SUMMARY_HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 資料
    row = 4
    current_category = ""

    for ln in summary.lines:
        # 類別分隔行
        if ln.category != current_category:
            current_category = ln.category
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_SUMMARY_HEADERS))
            cat_cell = ws.cell(row=row, column=1, value=f"── {current_category} ──")
            cat_cell.font = Font(bold=True, size=11)
            cat_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            row += 1

        ws.cell(row=row, column=1, value=ln.name)
        ws.cell(row=row, column=2, value=ln.spec)
        ws.cell(row=row, column=3, value=ln.material)
        ws.cell(row=row, column=4, value=ln.category)

        if ln.aggregate_type == "linear":
            ws.cell(row=row, column=5, value=round(ln.total_length_mm, 1))
            ws.cell(row=row, column=6, value=ln.piece_count)
        elif ln.aggregate_type == "plate":
            ws.cell(row=row, column=5, value="-")
            ws.cell(row=row, column=6, value=ln.total_qty)
        else:
            ws.cell(row=row, column=5, value="-")
            ws.cell(row=row, column=6, value=ln.total_qty)

        ws.cell(row=row, column=7, value=round(ln.total_weight, 2))

        if ln.stock_length > 0:
            ws.cell(row=row, column=8, value=round(ln.stock_length, 0))
        else:
            ws.cell(row=row, column=8, value="-")

        ws.cell(row=row, column=9, value=ln.purchase_qty)
        ws.cell(row=row, column=10, value=ln.purchase_unit)

        sources = ", ".join(ln.source_fullstrings[:5])
        if len(ln.source_fullstrings) > 5:
            sources += f" ...+{len(ln.source_fullstrings) - 5}"
        ws.cell(row=row, column=11, value=sources)
        row += 1

    # 合計行
    row += 1
    ws.cell(row=row, column=6, value="合計總重:").font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(summary.total_weight, 2)).font = Font(bold=True)

    # 欄寬
    widths = [14, 18, 12, 8, 14, 10, 10, 12, 10, 6, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


# ═══════════════════════════════════════════════════════
#  Sheet 2: 下料計算表
# ═══════════════════════════════════════════════════════

def _write_cutting_sheet(ws, plans: List[CuttingPlan]):
    from openpyxl.styles import Font, Alignment, PatternFill

    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value="下料計算表 — 現場裁切方案")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3

    for plan in plans:
        # 材料標題
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        mat_cell = ws.cell(
            row=row, column=1,
            value=f"{plan.name}  {plan.spec}  ({plan.material})"
        )
        mat_cell.font = Font(bold=True, size=12)
        mat_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        mat_cell.font = Font(bold=True, size=12, color="FFFFFF")
        row += 1

        # 摘要
        ws.cell(row=row, column=1, value="需求段數:")
        ws.cell(row=row, column=2, value=plan.total_pieces)
        ws.cell(row=row, column=3, value="需求總長:")
        ws.cell(row=row, column=4, value=f"{plan.total_demand_length:.0f} mm")
        ws.cell(row=row, column=5, value="原料根數:")
        ws.cell(row=row, column=6, value=plan.total_bars)
        ws.cell(row=row, column=7, value="平均使用率:")
        ws.cell(row=row, column=8, value=f"{plan.avg_utilization:.1f}%")
        row += 1

        # 表頭
        cut_headers = ["原料 #", "切割段", "需求長(mm)", "含損耗(mm)", "累計(mm)", "餘料(mm)", "使用率", "備註", "用於"]
        header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        for col, h in enumerate(cut_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # 每根原料
        for bar_idx, bar in enumerate(plan.bars):
            first_row = True
            cumulative = 0.0

            for piece_idx, piece in enumerate(bar.pieces):
                cumulative += piece.cut_length

                if first_row:
                    ws.cell(row=row, column=1, value=f"#{bar_idx + 1}")
                    first_row = False
                else:
                    ws.cell(row=row, column=1, value="")

                ws.cell(row=row, column=2, value=f"段 {piece_idx + 1}")
                ws.cell(row=row, column=3, value=round(piece.demand_length, 1))
                ws.cell(row=row, column=4, value=round(piece.cut_length, 1))
                ws.cell(row=row, column=5, value=round(cumulative, 1))
                ws.cell(row=row, column=9, value=piece.source)
                row += 1

            # 餘料行
            remnant = bar.remnant
            note = ""
            if remnant < 100:
                note = "廢料"
            elif remnant < 300:
                note = "短料"

            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value="─ 餘料 ─")
            ws.cell(row=row, column=6, value=round(remnant, 1))
            ws.cell(row=row, column=7, value=f"{bar.utilization:.1f}%")
            ws.cell(row=row, column=8, value=note)

            # 餘料行底色
            if note == "廢料":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif note == "短料":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = fill
            row += 1

        row += 1  # 材料間空行

    # 欄寬
    widths = [10, 10, 12, 12, 12, 12, 10, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
